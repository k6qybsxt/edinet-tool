import os
import requests
from io import StringIO
import logging
from logging.handlers import RotatingFileHandler
import certifi

os.environ["SSL_CERT_FILE"] = certifi.where()

import openpyxl
from bs4 import BeautifulSoup
import yfinance as yf
import re
from datetime import datetime, timedelta
import glob
from tkinter import Tk, filedialog
import pandas as pd
import json
from openpyxl.cell.cell import Cell
import shutil
from time import perf_counter

# ===== skipped_files: 構造化（原因コード化）=====
from dataclasses import dataclass, asdict
from enum import Enum
import traceback
from lxml import etree



class SkipCode(str, Enum):
    EXCEL_NOT_FOUND   = "EXCEL_NOT_FOUND"      # 作業Excelが見つからない
    FILE1_PARSE_ERROR = "FILE1_PARSE_ERROR"    # 半期XBRL解析失敗
    FILE2_NOT_FOUND   = "FILE2_NOT_FOUND"      # 最新有報XBRLなし
    FILE2_ERROR       = "FILE2_ERROR"          # 最新有報 解析/書込失敗
    FILE3_YEAR_MISS   = "FILE3_YEAR_MISS"      # 過去有報 期末年取れず
    FILE3_ERROR       = "FILE3_ERROR"          # 過去有報 解析/書込失敗
    HALF_WRITE_ERROR  = "HALF_WRITE_ERROR"     # 半期 書込失敗
    NO_SECURITY_CODE  = "NO_SECURITY_CODE"     # 証券コードが取れない（株価等で必要）
    RENAME_ERROR      = "RENAME_ERROR"         # リネーム失敗（非致命）
    UNKNOWN           = "UNKNOWN"

@dataclass
class SkipItem:
    code: str
    phase: str               # "excel_select" / "file1" / "file2" / "file3" / "half" / "stock" / "rename"
    slot: int | None         # loop["slot"] を入れる（将来トレースしやすい）
    excel: str | None
    xbrl: str | None
    message: str
    exc_type: str | None = None
    exc_msg: str | None = None

def add_skip(skipped_files: list, *, code: SkipCode, phase: str, loop: dict | None,
             excel: str | None, xbrl: str | None, message: str, exc: Exception | None = None):
    item = SkipItem(
        code=code.value if isinstance(code, SkipCode) else str(code),
        phase=phase,
        slot=(loop.get("slot") if loop else None),
        excel=excel,
        xbrl=xbrl,
        message=message,
        exc_type=(type(exc).__name__ if exc else None),
        exc_msg=(str(exc) if exc else None),
    )
    skipped_files.append(asdict(item))

def log_skip_summary(logger: logging.Logger, skipped_files: list):
    logger.info("--- skipped summary ---")
    if not skipped_files:
        logger.info("skipped=0")
        return

    # code別件数
    counts: dict[str, int] = {}
    for s in skipped_files:
        c = s.get("code", "UNKNOWN")
        counts[c] = counts.get(c, 0) + 1

    # 1行サマリ（運用向け）
    parts = [f"{k}={counts[k]}" for k in sorted(counts.keys())]
    logger.warning("[skipped summary] " + " ".join(parts))

    # 詳細（最大30件だけ表示：ログ肥大化防止）
    logger.info("--- skipped details (first 30) ---")
    for s in skipped_files[:30]:
        logger.warning(
            "skip "
            f"code={s.get('code')} phase={s.get('phase')} slot={s.get('slot')} "
            f"excel={s.get('excel')} xbrl={s.get('xbrl')} msg={s.get('message')} "
            f"exc={s.get('exc_type')}:{s.get('exc_msg')}"
        )

# =========================
# 1) METRICS（あなたの現行をそのまま）
# =========================
METRICS = {
    # ---------- PL / CF（duration） ----------
    "NetSales": {
        "tags": [
            # ★半期サマリー（最優先）
            "jpcrp_cor:NetSalesSummaryOfBusinessResults",
            # 通常PL
            "jppfs_cor:NetSales",
            # 保険（売上相当）
            "jpcrp_cor:RevenuesFromExternalCustomers",
            # IFRS
            "jpigp_cor:RevenueIFRS",
            "jpigp_cor:NetSalesIFRS",
        ],
        "kind": "duration",
        "unit": "millions",
    },
    "CostOfSales": {
        "tags": ["jppfs_cor:CostOfSales", "jpigp_cor:CostOfSalesIFRS"],
        "kind": "duration",
        "unit": "millions",
    },
    "GrossProfit": {
        "tags": ["jppfs_cor:GrossProfit", "jpigp_cor:GrossProfitIFRS"],
        "kind": "duration",
        "unit": "millions",
    },
    "SellingExpenses": {
        "tags": [
            "jppfs_cor:SellingGeneralAndAdministrativeExpenses",
            "jpigp_cor:SellingGeneralAndAdministrativeExpensesIFRS",
        ],
        "kind": "duration",
        "unit": "millions",
    },
    "OperatingIncome": {
        "tags": ["jppfs_cor:OperatingIncome", "jpigp_cor:OperatingProfitLossIFRS"],
        "kind": "duration",
        "unit": "millions",
    },
    "OrdinaryIncome": {
        "tags": [
            # ★半期サマリー（最優先）
            "jpcrp_cor:OrdinaryIncomeLossSummaryOfBusinessResults",
            # J-GAAP
            "jppfs_cor:OrdinaryIncome",
            # IFRS（経常相当）
            "jpigp_cor:ProfitLossBeforeTaxIFRS",
            "jpigp_cor:ProfitLossBeforeIncomeTaxesIFRS",
        ],
        "kind": "duration",
        "unit": "millions",
    },
    "ProfitLoss": {
        "tags": [
            # ★半期サマリー（最優先）
            "jpcrp_cor:ProfitLossAttributableToOwnersOfParentSummaryOfBusinessResults",
            # 日本基準（半期）
            "jppfs_cor:ProfitLossAttributableToOwnersOfParent",
            # 日本基準（通期）
            "jppfs_cor:ProfitLoss",
            # IFRS
            "jpigp_cor:ProfitLossAttributableToOwnersOfParentIFRS",
            "jpigp_cor:ProfitLossIFRS",
        ],
        "kind": "duration",
        "unit": "millions",
    },
    #営業CF
    "OperatingCash": {
        "tags": [
            # ★半期サマリー（最優先）
            "jpcrp_cor:NetCashProvidedByUsedInOperatingActivitiesSummaryOfBusinessResults",
            # 通常CF
            "jppfs_cor:NetCashProvidedByUsedInOperatingActivities",
            # IFRS
            "jpigp_cor:NetCashProvidedByUsedInOperatingActivitiesIFRS",
        ],
        "kind": "duration",
        "unit": "millions",
    },
    #投資CF
    "InvestmentCash": {
        "tags": [
            # ★半期サマリー（最優先）
            "jpcrp_cor:NetCashProvidedByUsedInInvestingActivitiesSummaryOfBusinessResults",
            # 通常CF（表記ゆれ対策で2つ）
            "jppfs_cor:NetCashProvidedByUsedInInvestmentActivities",
            "jppfs_cor:NetCashProvidedByUsedInInvestingActivities",
            # IFRS
            "jpigp_cor:NetCashProvidedByUsedInInvestingActivitiesIFRS",
        ],
        "kind": "duration",
        "unit": "millions",
    },
    #財務CF
    "FinancingCash": {
        "tags": [
            # ★半期サマリー（最優先）
            "jpcrp_cor:NetCashProvidedByUsedInFinancingActivitiesSummaryOfBusinessResults",
            # 通常CF
            "jppfs_cor:NetCashProvidedByUsedInFinancingActivities",
            # IFRS
            "jpigp_cor:NetCashProvidedByUsedInFinancingActivitiesIFRS",
        ],
        "kind": "duration",
        "unit": "millions",
    },

    # ---------- BS（instant） ----------
    "TotalAssets": {
        "tags": [
            "jpcrp_cor:TotalAssetsSummaryOfBusinessResults",
            "jpcrp_cor:TotalAssetsIFRSSummaryOfBusinessResults",
            "jpigp_cor:AssetsIFRS",
        ],
        "kind": "instant_num",
        "unit": "millions",
    },
    "NetAssets": {
        "tags": [
            "jpcrp_cor:NetAssetsSummaryOfBusinessResults",
            "jpcrp_cor:EquityAttributableToOwnersOfParentIFRSSummaryOfBusinessResults",
            "jpigp_cor:EquityAttributableToOwnersOfParentIFRS",
        ],
        "kind": "instant_num",
        "unit": "millions",
    },
    "CashAndCashEquivalents": {
        "tags": [
            "jppfs_cor:CashAndCashEquivalents",
            "jpigp_cor:CashAndCashEquivalentsIFRS",
            "jpcrp_cor:CashAndCashEquivalentsSummaryOfBusinessResults",
        ],
        "kind": "instant_num",
        "unit": "millions",
    },

    # ---------- 株数（instant）：「材料」を拾って差分でTotalNumberに統一 ----------
    # 発行済株式総数（優先順を整理）
    "IssuedShares": {
        "tags": [
            # ★期末ベース（最優先：最も安定しやすい）
            "jpcrp_cor:NumberOfIssuedSharesAsOfFiscalYearEndIssuedSharesTotalNumberOfSharesEtc",
            # ★提出日ベース（次点）
            "jpcrp_cor:NumberOfIssuedSharesAsOfFilingDateIssuedSharesTotalNumberOfSharesEtc",

            # 代表的（株式等の状況）
            "jpcrp_cor:TotalNumberOfIssuedSharesIssuedSharesTotalNumberOfSharesEtc",
            # サマリー系（会社により出る）
            "jpcrp_cor:TotalNumberOfIssuedSharesSummaryOfBusinessResults",
            # 普通株/普通株式名義系（会社により出る）
            "jpcrp_cor:TotalNumberOfIssuedSharesCommonStockIssuedSharesTotalNumberOfSharesEtc",
            "jpcrp_cor:TotalNumberOfIssuedSharesOrdinaryShareIssuedSharesTotalNumberOfSharesEtc",
            "jpcrp_cor:NumberOfIssuedSharesAsOfFiscalYearEndIssuedSharesTotalNumberOfSharesEtc",

            # ★議決権表など“表系”は最後（会社によっては出るが優先しない）
            "jpcrp_cor:NumberOfSharesIssuedSharesVotingRights",
        ],
        "kind": "instant_num",
        "unit": "ones",
    },

    # 自己株式数（優先順を整理）
    "TreasuryShares": {
        "tags": [
            # ★自己株式数（合計）が最優先
            "jpcrp_cor:TotalNumberOfSharesHeldTreasurySharesEtc",
            # ★自己名義（合計が取れない場合の次点）
            "jpcrp_cor:NumberOfSharesHeldInOwnNameTreasurySharesEtc",

            # 表記ゆれ・補助（会社により出る）
            "jpcrp_cor:TotalNumberOfSharesHeldInTheNameOfOthersTreasurySharesEtc",
            "jpcrp_cor:TotalNumberOfSharesHeldInOwnNameTreasurySharesEtc",
            "jpcrp_cor:TotalNumberOfTreasurySharesSummaryOfBusinessResults",
            "jpcrp_cor:NumberOfTreasurySharesAsOfFiscalYearEndIssuedSharesTotalNumberOfSharesEtc",
            "jpcrp_cor:TreasurySharesAtTheEndOfFiscalYearIssuedSharesTotalNumberOfSharesEtc",
        ],
        "kind": "instant_num",
        "unit": "ones",
    },
    }

def normalize_security_code(raw: object) -> str | None:
    """
    4桁の証券コード(数字)に正規化して返す。取れなければNone。
    例: "2206", "2206.0", "2206-T", "2206.T", "2206 " -> "2206"
    """
    if raw is None:
        return None

    s = str(raw).strip()
    if not s:
        return None

    # 小数っぽい "2206.0" 対策
    if re.fullmatch(r"\d+\.0+", s):
        s = s.split(".", 1)[0]

    # "2206.T" / "2206-T" / "2206（東証）" みたいな混在を数字だけに寄せる
    m = re.search(r"(\d{4})", s)
    if not m:
        return None

    code = m.group(1)
    return code

def pick_security_code(*candidates: object) -> str | None:
    """
    候補を左から順に見て、最初に正規化できた4桁コードを返す。
    """
    for c in candidates:
        code = normalize_security_code(c)
        if code:
            return code
    return None

def _append_jsonl(path: str, obj: dict) -> None:
    os.makedirs(os.path.dirname(path), exist_ok=True)
    with open(path, "a", encoding="utf-8") as f:
        f.write(json.dumps(obj, ensure_ascii=False) + "\n")

def ensure_security_code(x2: dict | None, parsed_code: object, x1: dict | None = None) -> str | None:
    """
    file2成功時の security_code を必ず確定させる（優先順位固定）
    """
    if x2 is None:
        return None

    # 優先順位（固定）
    # 1) parse_xbrl_data の戻り値
    # 2) DEI の SecurityCode
    # 3) よくある別名キー（念のため）
    # 4) file1側（半期）があるなら最後の保険
    return pick_security_code(
        parsed_code,
        x2.get("SecurityCodeDEI"),
        x2.get("SecurityCode"),
        x2.get("SecurityCodeCoverPage"),
        (x1.get("SecurityCodeDEI") if x1 else None),
        (x1.get("SecurityCode") if x1 else None),
    )

# ===== Logging Policy (固定ルール) =====
# CRITICAL: 続行不可。設定不備・テンプレ汚染防止ロック等。→ SystemExit で終了
# ERROR   : 想定外例外。続行はするが必ず調査対象（logger.exception含む）
# WARNING : 非致命だが結果に影響する可能性（欠損/スキップ/見つからない等）
# INFO    : 進捗と結果サマリ（開始/終了、選択した決算期、各フェーズwritten/missing 等）
# DEBUG   : 詳細（XBRLファイル一覧、キー一覧、内部データの中身など）
def setup_logger(debug: bool = False, log_dir: str | None = None) -> logging.Logger:
    """
    - 入口で1回だけ呼ぶ
    - console: INFO以上（debug=TrueならDEBUG）
    - file   : DEBUG以上を全部保存（運用証跡）
    - 時刻は秒まで（ミリ秒無し）
    - ハンドラ重複を確実に防ぐ
    """
    logger = logging.getLogger("edinet")
    logger.setLevel(logging.DEBUG)  # ロガー本体は常にDEBUG、出力はハンドラ側で制御
    logger.propagate = False

    # 既存ハンドラがあれば全削除（再実行/REPLでも二重出力しない）
    if logger.handlers:
        for h in list(logger.handlers):
            logger.removeHandler(h)

    datefmt = "%Y-%m-%d %H:%M:%S"
    fmt = logging.Formatter(
        "%(asctime)s [%(levelname)s] %(message)s",
        datefmt=datefmt
    )

    # --- Console handler ---
    ch = logging.StreamHandler()
    ch.setFormatter(fmt)
    ch.setLevel(logging.DEBUG if debug else logging.INFO)
    logger.addHandler(ch)

    # --- File handler (証跡) ---
    if log_dir is None:
        log_dir = os.path.join(os.getcwd(), "logs")
    os.makedirs(log_dir, exist_ok=True)

    log_path = os.path.join(log_dir, "run.log")
    fh = RotatingFileHandler(
        log_path,
        maxBytes=2_000_000,  # 2MB
        backupCount=5,
        encoding="utf-8"
    )
    fh.setFormatter(fmt)
    fh.setLevel(logging.DEBUG)  # ファイルは常に全部残す
    logger.addHandler(fh)

    logger.debug(f"logger initialized: debug={debug}, log_path={log_path}")
    return logger

logger = None


# ===== DEBUG設定 =====
DEBUG = False  # True にすると詳細ログ

def log(*args, **kwargs):
    if DEBUG:
        logger.debug(" ".join(map(str, args)))
        
# スクリプトのあるフォルダに作業ディレクトリを変更
script_dir = os.path.dirname(os.path.abspath(__file__))  # スクリプトのあるフォルダを取得
os.chdir(script_dir)  # 作業ディレクトリをスクリプトのフォルダに変更

# フォルダを選択する関数
def choose_directory():
    root = Tk()
    root.withdraw()              # メインウィンドウ非表示
    root.update()                # これが効く環境が多い
    folder_path = filedialog.askdirectory(title="XBRLフォルダを選択してください")
    root.destroy()               # ★これ重要：Tkを閉じる
    return folder_path

def trim_value(value, unit='millions'):
    """
    EDINETのfactは文字列で来ることが多いので、まず int 化。
    株数(unit='ones')は割らずにそのまま返す。
    """
    try:
        s = str(value).replace(",", "").strip()
        v = int(s)

        if unit == "ones":
            return v

        factor = {'millions': 1_000_000, 'thousands': 1_000, 'ten': 10}[unit]
        return v // factor
    except Exception:
        return 'データなし'

SLOTS = ("Quarter", "YTD", "Current", "Prior1", "Prior2", "Prior3", "Prior4")

def split_out_key(out_key: str) -> tuple[str, str] | None:
    # 例: NetSalesPrior1 -> ("NetSales","Prior1")
    #     TotalAssetsQuarter -> ("TotalAssets","Quarter")
    for slot in SLOTS:
        if out_key.endswith(slot):
            return out_key[:-len(slot)], slot
    return None

RAW_COLS = [
    "company_code",
    "doc_id",
    "doc_type",
    "consolidation",
    "metric_key",
    "time_slot",
    "period_start",
    "period_end",
    "period_kind",
    "value",
    "unit",
    "tag_used",
    "tag_rank",
    "status",
    "dup_key",
    "run_id",
    "source_file",
]
def build_raw_rows(*, company_code: str, doc_id: str, doc_type: str,
                   out: dict, out_meta: dict) -> list[dict]:
    rows = []
    for out_key, val in out.items():
        sp = split_out_key(out_key)
        if sp is None:
            continue
        metric_key, time_slot = sp

        meta = out_meta.get(out_key, {})

        # valueは数値のみ採用（"データなし"などはNone）
        value_num = val if isinstance(val, (int, float)) else None

        rows.append({
            "company_code": company_code,
            "doc_id": doc_id,
            "doc_type": doc_type,  # "annual" or "half"
            "consolidation": meta.get("consolidation"),
            "metric_key": metric_key,
            "time_slot": time_slot,
            "period_start": meta.get("period_start"),
            "period_end": meta.get("period_end"),
            "period_kind": meta.get("period_kind"),
            "value": value_num,
            "unit": meta.get("unit"),
            "tag_used": meta.get("tag_used"),
            "tag_rank": meta.get("tag_rank"),
            "status": meta.get("status", "OK" if value_num is not None else "MISSING"),
        })
    return rows



# XBRLデータの解析（完成版）
def parse_xbrl_data(xbrl_file, mode="full"):
    from datetime import datetime, timedelta
    from lxml import etree
    from collections import defaultdict, deque

    def _attr_any(el, *names):
        """
        属性を name候補から探して返す。
        lxmlはQName属性になることがあるので、末尾一致でも拾う。
        """
        # 1) まず通常の get
        for n in names:
            v = el.get(n)
            if v is not None:
                return v

        # 2) QName属性対策： {uri}contextRef のようなキーが来る
        if el.attrib:
            for k, v in el.attrib.items():
                kk = k.split("}")[-1]  # localname
                if kk in names:
                    return v
        return None

    # ===== util =====
    def parse_ymd(s):
        try:
            return datetime.strptime(s, "%Y-%m-%d")
        except Exception:
            return None

    def months_diff(s, e):
        if not s or not e:
            return None
        m = (e.year - s.year) * 12 + (e.month - s.month)
        if e.day < s.day:
            m -= 1
        return m

    def duration_bucket_months(start_ymd, end_ymd):
        s = parse_ymd(start_ymd)
        e = parse_ymd(end_ymd)
        md = months_diff(s, e)
        if md is None:
            return None
        if 5 <= md <= 7:
            return 6
        if 11 <= md <= 13:
            return 12
        return None

    def dim_rank(dim):
        return 0 if dim == "Consolidated" else 1

    def best_update(store, key, cand):
        cur = store.get(key)
        if cur is None:
            store[key] = cand
            return
        if (dim_rank(cand["dim"]), cand["tag_priority"]) < (dim_rank(cur["dim"]), cur["tag_priority"]):
            store[key] = cand

    # ===== outputs =====
    out = {}
    out_meta = {}

    # ===== nsmap -> QName helper =====
    nsmap = {}
    for ev, el0 in etree.iterparse(xbrl_file, events=("start",), recover=True, huge_tree=True):
        nsmap = dict(el0.nsmap or {})
        break

    def _qname(tag_prefixed: str) -> str:
        pref, local = tag_prefixed.split(":", 1)
        uri = nsmap.get(pref)
        return f"{{{uri}}}{local}" if uri else f"{{}}{local}"

    # === PATCH: METRIC判定を QName ではなく localname で行う（prefix差を吸収）===
    METRIC_L = {}  # localname -> (metric, tag_priority, kind, unit, tag_used)

    for metric, meta in METRICS.items():
        for tag_priority, tag in enumerate(meta["tags"]):
            local = tag.split(":", 1)[1] if ":" in tag else tag
            # 同じlocalが複数出る場合は、先に入った(=優先度高い)を残す
            METRIC_L.setdefault(local, (metric, tag_priority, meta["kind"], meta["unit"], tag))

    DEI_TAGS = {
        "jpdei_cor:CurrentFiscalYearStartDateDEI": "CurrentFiscalYearStartDateDEI",
        "jpdei_cor:CurrentPeriodEndDateDEI": "CurrentPeriodEndDateDEI",
        "jpdei_cor:TypeOfCurrentPeriodDEI": "TypeOfCurrentPeriodDEI",
        "jpdei_cor:CurrentFiscalYearEndDateDEI": "CurrentFiscalYearEndDateDEI",
        "jpdei_cor:SecurityCodeDEI": "SecurityCodeDEI",
        "jpdei_cor:FilerNameInJapaneseDEI": "CompanyNameCoverPage",
        "jpdei_cor:FilerNameDEI": "CompanyNameCoverPage",
    }
    DEI_Q = {_qname(k): v for k, v in DEI_TAGS.items()}

    # ===== state =====
    contexts = {}  # ctx_id -> {start,end,instant,dim}

    ctxref_missing = 0
    ctxref_notfound = 0
    ctxref_found = 0
    ctxref_samples = []

    # factがcontextより先に来たときの保険（ctxref -> list[factinfo]）
    pending = defaultdict(list)

    # best stores
    dur_best = {}
    inst_best = {}
    fy_end_candidates = set()

    # index（後段を高速に）
    dur_ends_by_metric_months = defaultdict(set)  # (metric, months) -> set(end)
    inst_ends_by_metric = defaultdict(set)        # metric -> set(end)

    fy_start_dei = None
    period_end_dei = None
    period_type = None
    security_code = None

    metric_hit = 0
    metric_hit_nonempty = 0
    metric_hit_sample = []

    seen_locals = set()
    seen_local_sample = []



    # ===== 1-pass iterparse =====
    for event, el in etree.iterparse(xbrl_file, events=("end",), recover=True, huge_tree=True):

        local = el.tag.split("}")[-1]

        if len(seen_local_sample) < 10 and local not in seen_locals:
            seen_locals.add(local)
            seen_local_sample.append(local)

        # --- context ---
        if local == "context":
            ctx_id = (el.get("id") or "").strip() or None
            if ctx_id:
                start_s = end_s = inst_s = None

                for p in el.iter():
                    pl = p.tag.split("}")[-1]
                    if pl == "startDate":
                        start_s = (p.text or "").strip() or None
                    elif pl == "endDate":
                        end_s = (p.text or "").strip() or None
                    elif pl == "instant":
                        inst_s = (p.text or "").strip() or None

                members = []
                for m in el.iter():
                    if m.tag.split("}")[-1] == "explicitMember":
                        t = (m.text or "").strip()
                        if t:
                            members.append(t)

                is_noncon = any("NonConsolidatedMember" in t for t in members)
                dim = "NonConsolidated" if is_noncon else "Consolidated"

                ctx = {"start": start_s, "end": end_s, "instant": inst_s, "dim": dim}
                contexts[ctx_id] = ctx

                if ctx_id in pending:
                    for finfo in pending.pop(ctx_id):
                        metric, tag_priority, kind, unit, tag_used, txt = finfo

                        if kind == "duration":
                            if ctx["start"] and ctx["end"]:
                                months = duration_bucket_months(ctx["start"], ctx["end"])
                                if months in (6, 12):
                                    end_date = ctx["end"]

                                    cand = {
                                        "value": txt,
                                        "start": ctx["start"],
                                        "end": end_date,
                                        "months": months,
                                        "dim": ctx["dim"],
                                        "tag_priority": tag_priority,
                                        "tag_used": tag_used,
                                    }

                                    best_update(dur_best, (metric, end_date, months), cand)
                                    dur_ends_by_metric_months[(metric, months)].add(end_date)

                                    if months == 12 and parse_ymd(end_date):
                                        fy_end_candidates.add(end_date)

                        else:
                            if ctx["instant"]:
                                end_date = ctx["instant"]

                                cand = {
                                    "value": txt,
                                    "start": None,
                                    "end": end_date,
                                    "dim": ctx["dim"],
                                    "tag_priority": tag_priority,
                                    "tag_used": tag_used,
                                }

                                best_update(inst_best, (metric, end_date), cand)
                                inst_ends_by_metric[metric].add(end_date)

            el.clear()
            continue

        # --- DEI ---
        k = DEI_Q.get(el.tag)

        if k:
            v = (el.text or "").strip()

            if v:
                if k == "CompanyNameCoverPage":
                    if "CompanyNameCoverPage" not in out:
                        out["CompanyNameCoverPage"] = v
                else:
                    out[k] = v

                if k == "CurrentFiscalYearStartDateDEI":
                    fy_start_dei = v

                elif k == "CurrentPeriodEndDateDEI":
                    period_end_dei = v

                elif k == "TypeOfCurrentPeriodDEI":
                    period_type = v

                elif k == "SecurityCodeDEI":
                    if v.isdigit() and len(v) >= 2:
                        security_code = v[:-1]
                        out["SecurityCodeDEI"] = security_code

            el.clear()
            continue

        # --- FACT ---
        info = METRIC_L.get(local)

        if info:

            metric, tag_priority, kind, unit, tag_used = info

            metric_hit += 1

            txt = (el.text or "").strip()

            if txt:
                metric_hit_nonempty += 1

            if len(metric_hit_sample) < 8:
                metric_hit_sample.append((local, tag_used))

            ctxref = (_attr_any(el, "contextRef") or "").strip() or None

            # --- ctxrefメーター（最初の10件だけサンプル）---
            if len(ctxref_samples) < 10:
                ctxref_samples.append((local, ctxref, ctxref in contexts if ctxref else None))

            if not ctxref:
                ctxref_missing += 1
            else:
                if ctxref in contexts:
                    ctxref_found += 1
                else:
                    ctxref_notfound += 1

            if metric_hit <= 5:
                logger.warning(
                    "[ctxref sample] local=%s ctxref=%s has_ctx=%s",
                    local,
                    ctxref,
                    (ctxref in contexts) if ctxref else None
                )

            if ctxref:

                txt = (el.text or "").strip()

                if txt:

                    ctx = contexts.get(ctxref)

                    if ctx is None:
                        pending[ctxref].append(
                            (metric, tag_priority, kind, unit, tag_used, txt)
                        )

                    else:

                        if kind == "duration":

                            if ctx["start"] and ctx["end"]:

                                months = duration_bucket_months(ctx["start"], ctx["end"])

                                if months in (6, 12):

                                    end_date = ctx["end"]

                                    cand = {
                                        "value": txt,
                                        "start": ctx["start"],
                                        "end": end_date,
                                        "months": months,
                                        "dim": ctx["dim"],
                                        "tag_priority": tag_priority,
                                        "tag_used": tag_used,
                                    }

                                    best_update(dur_best, (metric, end_date, months), cand)

                                    dur_ends_by_metric_months[(metric, months)].add(end_date)

                                    if months == 12 and parse_ymd(end_date):
                                        fy_end_candidates.add(end_date)

                        else:

                            if ctx["instant"]:

                                end_date = ctx["instant"]

                                cand = {
                                    "value": txt,
                                    "start": None,
                                    "end": end_date,
                                    "dim": ctx["dim"],
                                    "tag_priority": tag_priority,
                                    "tag_used": tag_used,
                                }

                                best_update(inst_best, (metric, end_date), cand)

                                inst_ends_by_metric[metric].add(end_date)

        el.clear()

    logger.warning(
        "[ctxref meter] mode=%s missing=%d notfound=%d found=%d samples=%s",
        mode, ctxref_missing, ctxref_notfound, ctxref_found, ctxref_samples
    )

    # ===== DEBUG METER (temporary) =====
    try:

        dur_n = len(dur_best)
        inst_n = len(inst_best)

        dur_6 = sum(1 for (m, end, months) in dur_best.keys() if months == 6)
        dur_12 = sum(1 for (m, end, months) in dur_best.keys() if months == 12)

        logger.warning(
            "[parse debug] mode=%s contexts=%d dur_best=%d(inst=%d) dur6=%d dur12=%d ns_has_jppfs=%s ns_has_jpcrp=%s",
            mode,
            len(contexts),
            dur_n,
            inst_n,
            dur_6,
            dur_12,
            ("jppfs_cor" in nsmap),
            ("jpcrp_cor" in nsmap),
        )

    except Exception:
        pass


    logger.warning("[local sample] mode=%s %s", mode, seen_local_sample)

    logger.warning(
        "[fact meter] mode=%s metric_hit=%d metric_nonempty=%d sample=%s",
        mode,
        metric_hit,
        metric_hit_nonempty,
        metric_hit_sample
    )

    # ===== DEI half key =====
    if mode == "half" and period_end_dei:
        out["HalfPeriodEndDateDEI"] = period_end_dei

    # ===== base_year / fy_start =====
    fy_ends = sorted({d for d in fy_end_candidates if parse_ymd(d)}, reverse=True)
    base_fy_end = fy_ends[0] if fy_ends else None
    base_dt = parse_ymd(base_fy_end) if base_fy_end else None
    base_year = base_dt.year if base_dt else None

    fy_start = None
    if fy_start_dei and parse_ymd(fy_start_dei):
        fy_start = fy_start_dei
    if fy_start is None:
        fy_end_dei = out.get("CurrentFiscalYearEndDateDEI")
        if fy_end_dei:
            fy_end_dt = parse_ymd(fy_end_dei)
            if fy_end_dt:
                prev_fy_end = fy_end_dt.replace(year=fy_end_dt.year - 1)
                fy_start = (prev_fy_end + timedelta(days=1)).strftime("%Y-%m-%d")
    if fy_start is None and base_fy_end:
        # base_fy_endに一致する12ヶ月 cand の start を拾う
        starts = []
        for (m, end, months), cand in dur_best.items():
            if months == 12 and end == base_fy_end and cand.get("start") and parse_ymd(cand["start"]):
                starts.append(cand["start"])
        starts = sorted(set(starts), key=lambda s: parse_ymd(s))
        if starts:
            fy_start = starts[0]

    period_end = period_end_dei

    # ===== OUTPUT: duration =====
    for metric, meta in METRICS.items():
        if meta["kind"] != "duration":
            continue

        # YTD(6)
        best = None
        if mode == "half" and period_end and parse_ymd(period_end):
            cand = dur_best.get((metric, period_end, 6))
            if cand and ((fy_start is None) or (cand.get("start") == fy_start)):
                best = cand

        if best is None and fy_start:
            ends = sorted({e for e in dur_ends_by_metric_months.get((metric, 6), set())
                           if parse_ymd(e) and dur_best.get((metric, e, 6), {}).get("start") == fy_start},
                          reverse=True)
            if ends:
                best = dur_best.get((metric, ends[0], 6))

        if best is None:
            ends = sorted({e for e in dur_ends_by_metric_months.get((metric, 6), set()) if parse_ymd(e)}, reverse=True)
            if ends:
                best = dur_best.get((metric, ends[0], 6))

        key = f"{metric}YTD"
        if best:
            out[key] = trim_value(best["value"], meta["unit"])
            out_meta[key] = {
                "period_start": best.get("start"),
                "period_end": best.get("end"),
                "period_kind": "duration",
                "unit": meta["unit"],
                "consolidation": "C" if best.get("dim") == "Consolidated" else "N",
                "tag_used": best.get("tag_used"),
                "tag_rank": (best.get("tag_priority") or 0) + 1,
                "status": "OK",
            }
        else:
            out[key] = None
            out_meta[key] = {
                "period_start": None,
                "period_end": None,
                "period_kind": "duration",
                "unit": meta["unit"],
                "consolidation": None,
                "tag_used": None,
                "tag_rank": None,
                "status": "MISSING",
            }

        # 12ヶ月 Current/Prior（fullのみ）
        if mode != "half":
            ends_12 = sorted({e for e in dur_ends_by_metric_months.get((metric, 12), set()) if parse_ymd(e)}, reverse=True)
            for end_date in ends_12:
                if base_year is None:
                    break
                dt = parse_ymd(end_date)
                if not dt:
                    continue
                diff = base_year - dt.year
                if diff < 0 or diff > 4:
                    continue

                suffix = "Current" if diff == 0 else f"Prior{diff}"
                best12 = dur_best.get((metric, end_date, 12))
                if best12:
                    k2 = f"{metric}{suffix}"
                    out[k2] = trim_value(best12["value"], meta["unit"])
                    out_meta[k2] = {
                        "period_start": best12.get("start"),
                        "period_end": best12.get("end"),
                        "period_kind": "duration",
                        "unit": meta["unit"],
                        "consolidation": "C" if best12.get("dim") == "Consolidated" else "N",
                        "tag_used": best12.get("tag_used"),
                        "tag_rank": (best12.get("tag_priority") or 0) + 1,
                        "status": "OK",
                    }

    # ===== OUTPUT: instant =====
    for metric, meta in METRICS.items():
        if meta["kind"] == "duration":
            continue

        inst_ends = sorted({e for e in inst_ends_by_metric.get(metric, set()) if parse_ymd(e)}, reverse=True)

        # Quarter
        if inst_ends:
            chosen_end = inst_ends[0]
            target_dt = parse_ymd(period_end) if period_end else None
            if target_dt:
                if period_end in inst_ends:
                    chosen_end = period_end
                else:
                    fy_start_dt = parse_ymd(fy_start) if fy_start else None
                    inst_dts = []
                    for e in inst_ends:
                        dt = parse_ymd(e)
                        if not dt:
                            continue
                        if fy_start_dt and dt < fy_start_dt:
                            continue
                        inst_dts.append(dt)
                    if not inst_dts:
                        inst_dts = [parse_ymd(e) for e in inst_ends if parse_ymd(e)]
                    if inst_dts:
                        inst_dts.sort(key=lambda dt: abs((dt - target_dt).days))
                        chosen_end = inst_dts[0].strftime("%Y-%m-%d")

            best_q = inst_best.get((metric, chosen_end))
            if best_q:
                key = f"{metric}Quarter"
                out[key] = trim_value(best_q["value"], meta["unit"])
                out_meta[key] = {
                    "period_start": None,
                    "period_end": best_q.get("end"),
                    "period_kind": "instant",
                    "unit": meta["unit"],
                    "consolidation": "C" if best_q.get("dim") == "Consolidated" else "N",
                    "tag_used": best_q.get("tag_used"),
                    "tag_rank": (best_q.get("tag_priority") or 0) + 1,
                    "status": "OK",
                }

        # Current/Prior（fullのみ）
        if mode != "half":
            for end_date in inst_ends:
                if base_year is None:
                    break
                dt = parse_ymd(end_date)
                if not dt:
                    continue
                diff = base_year - dt.year
                if diff < 0 or diff > 4:
                    continue

                suffix = "Current" if diff == 0 else f"Prior{diff}"
                best_i = inst_best.get((metric, end_date))
                if best_i:
                    key = f"{metric}{suffix}"
                    out[key] = trim_value(best_i["value"], meta["unit"])
                    out_meta[key] = {
                        "period_start": None,
                        "period_end": best_i.get("end"),
                        "period_kind": "instant",
                        "unit": meta["unit"],
                        "consolidation": "C" if best_i.get("dim") == "Consolidated" else "N",
                        "tag_used": best_i.get("tag_used"),
                        "tag_rank": (best_i.get("tag_priority") or 0) + 1,
                        "status": "OK",
                    }

    # ===== TotalNumber =====
    suffixes = ["Quarter"] if mode == "half" else ["Current", "Prior1", "Prior2", "Prior3", "Prior4", "Quarter"]
    for suffix in suffixes:
        issued = out.get(f"IssuedShares{suffix}")
        treasury = out.get(f"TreasuryShares{suffix}")
        if isinstance(issued, int) and isinstance(treasury, int):
            key = f"TotalNumber{suffix}"
            out[key] = issued - treasury
            out_meta[key] = {
                "period_start": None,
                "period_end": (out_meta.get(f"IssuedShares{suffix}", {}) or {}).get("period_end"),
                "period_kind": "instant",
                "unit": "ones",
                "consolidation": (out_meta.get(f"IssuedShares{suffix}", {}) or {}).get("consolidation"),
                "tag_used": "CALC(IssuedShares-TreasuryShares)",
                "tag_rank": 0,
                "status": "OK",
            }

    return out, security_code, out_meta

def _candidate_names(base_key: str):
    """
    base_key: 'NetSales_YTD' のような「Scope無し」のキーを想定
    返り値: ['NetSales_C_YTD', 'NetSales_N_YTD'] のように優先順で返す
    """
    parts = base_key.split("_")
    if len(parts) >= 3:
        # すでに NetSales_C_YTD のようにScope入りならそのまま
        return [base_key]

    # 例: NetSales_YTD → NetSales_C_YTD / NetSales_N_YTD
    if len(parts) == 2:
        metric, period = parts
        return [f"{metric}_C_{period}", f"{metric}_N_{period}"]

    # 例外：変換できない形はそのまま
    return [base_key]

# requests session を1回だけ作って使い回す（毎回作るより安定＆高速）
_YF_SESSION = None

def _get_yf_session():
    global _YF_SESSION
    if _YF_SESSION is None:
        s = requests.Session()
        s.verify = certifi.where()  # ← CA を固定
        _YF_SESSION = s
    return _YF_SESSION

# 株価データの取得
def _to_stooq_symbol(stock_code: str) -> str:
    """
    '2206.T' / '2206' どちらが来ても stooq用に変換する
    stooqの日本株は '2206.JP' 形式
    """
    code = stock_code.strip().upper()
    if code.endswith(".T"):
        code = code[:-2]
    # 数字4桁だけ抽出したい場合はここで調整してOK
    return f"{code}.JP"

def get_stock_price(stock_code, target_date, backup_date, buffer_days=3):
    """
    stooq から日足CSVを取得して、target_date→見つからなければ過去に遡ってCloseを返す
    """
    # 取得期間（余裕を持たせる）
    start_date = (datetime.strptime(backup_date, "%Y-%m-%d") - timedelta(days=buffer_days)).date()
    end_date   = (datetime.strptime(target_date, "%Y-%m-%d") + timedelta(days=1)).date()

    symbol = _to_stooq_symbol(stock_code)
    url = f"https://stooq.com/q/d/l/?s={symbol}&i=d"

    r = requests.get(url, timeout=20)
    r.raise_for_status()

    df = pd.read_csv(StringIO(r.text))
    if df.empty or "Date" not in df.columns or "Close" not in df.columns:
        return None

    df["Date"] = pd.to_datetime(df["Date"]).dt.date
    # 期間で絞る
    df = df[(df["Date"] >= start_date) & (df["Date"] <= end_date)]
    if df.empty:
        return None

    # target_dateから遡って探す
    check = datetime.strptime(target_date, "%Y-%m-%d").date()
    start = start_date

    close_map = dict(zip(df["Date"], df["Close"]))

    while check >= start:
        if check in close_map and pd.notna(close_map[check]):
            return float(close_map[check])
        check -= timedelta(days=1)

    return None

def validate_stock_date_pairs(date_pairs):
    """
    JSONのdate_pairsが「name前提」になっているか検証する。
    不整合があれば早期に気づけるように ValueError を投げる。
    """
    bad = []
    for i, item in enumerate(date_pairs):
        # cell が残っていたらNG（完全NamedRange化の契約違反）
        if "cell" in item:
            bad.append((i, "cell_is_not_allowed", item))
            continue
        # name / target_date / backup_date は必須
        if not item.get("name") or not item.get("target_date") or not item.get("backup_date"):
            bad.append((i, "missing_required_keys", item))

    if bad:
        msg_lines = ["株価date_pairsが name前提になっていません（JSONを修正してください）。"]
        for i, reason, item in bad[:10]:  # 長くなりすぎないように先頭だけ
            msg_lines.append(f"  - index={i}, reason={reason}, item={item}")
        raise ValueError("\n".join(msg_lines))

#NamedRangeに書き込むユーティリティ
def _set_value_to_namedrange(workbook, range_name: str, value) -> bool:
    """
    NamedRangeが存在すれば、その参照先セル(群)に value を書く。成功なら True。
    1セル想定だが、範囲でも全セルに同じ値を書ける。
    """
    dn = workbook.defined_names.get(range_name)
    if dn is None:
        return False

    for sheet_name, ref in dn.destinations:
        if sheet_name not in workbook.sheetnames:
            continue
        ws = workbook[sheet_name]

        obj = ws[ref]
        # ws["A1"] は Cell、ws["A1:B2"] は tuple(tuple(Cell))
        if isinstance(obj, openpyxl.cell.cell.Cell):
            obj.value = value
        else:
            for row in obj:
                for cell in row:
                    cell.value = value
    return True

# 株価データをExcelに書き込む
def write_stock_data_to_excel(excel_file, stock_code, date_pairs):
    """
    株価は NamedRange 専用（必須）。
    date_pairs の各要素は必ず:
      {"name": "...", "target_date": "YYYY-MM-DD", "backup_date": "YYYY-MM-DD"}
    を満たすこと。
    """
    workbook = openpyxl.load_workbook(excel_file)

    result = {
        "written": 0,
        "miss": 0,          # 株価データが取得できなかった（非致命）
        "errors": 0,        # 例外など（非致命だが要調査）
        "missing_name": 0,  # NamedRange が見つからない（テンプレ不整合）
        "bad_input": 0,     # JSON要素の不足（設計ミス）
    }

    for item in date_pairs:
        name = item.get("name")
        target_date = item.get("target_date")
        backup_date = item.get("backup_date")

        # 入力仕様違反（name前提）
        if not name or not target_date or not backup_date:
            logger.warning(f"[WARNING] 株価date_pairsの要素が不完全なのでスキップ: {item}")
            result["bad_input"] += 1
            continue

        # 取得
        try:
            price = get_stock_price(stock_code, target_date, backup_date)
        except Exception as e:
            logger.exception(
                f"株価取得で想定外エラー（続行） "
                f"code={stock_code} target={target_date} backup={backup_date}"
            )
            result["errors"] += 1
            continue

        if price is None:
            logger.warning(f"{target_date} の株価が取得できませんでした（続行）: name={name} code={stock_code}")
            result["miss"] += 1
            continue

        v = float(price)

        # 書き込み（NamedRangeのみ）
        wrote = _set_value_to_namedrange(workbook, name, v)
        if not wrote:
            logger.warning(f"NamedRangeが見つからず書けませんでした: {name} ({target_date})")
            result["missing_name"] += 1
            continue

        result["written"] += 1
        logger.debug(f"{target_date} の株価を書き込みました: {name}")

    # --- 株価処理サマリ（必ず1回出す：運用の核） ---
    logger.info(
        f"[stock summary] "
        f"written={result['written']} "
        f"miss={result['miss']} "
        f"errors={result['errors']} "
        f"missing_name={result['missing_name']} "
        f"bad_input={result['bad_input']} "
        f"(code={stock_code})"
    )

    # --- 問題があれば警告（要調査だが続行） ---
    if result["errors"] > 0 or result["missing_name"] > 0 or result["bad_input"] > 0:
        logger.warning(
            f"[stock warning] issues detected "
            f"errors={result['errors']} "
            f"missing_name={result['missing_name']} "
            f"bad_input={result['bad_input']} "
            f"(code={stock_code})"
        )

    workbook.save(excel_file)
    return result

# Excelファイルのリネーム（安全版）
def safe_filename(s: str) -> str:
    """
    Windowsでファイル名に使えない文字を置換して安全にする
    """
    if s is None:
        return ""
    s = str(s).strip()
    # Windows NG: \ / : * ? " < > |
    s = re.sub(r'[\\/:*?"<>|]', '_', s)
    # 末尾のドット/スペースもNG
    s = s.rstrip(". ").strip()
    return s

def rename_excel_file(original_path, security_code, company_name, period_end_date):
    """
    original_path を「<code>_<name>_<date>.xlsx」にリネームし、パスを返す。
    同名があれば _1, _2... を付ける。
    """
    dir_path = os.path.dirname(original_path)

    code = safe_filename(security_code)
    name = safe_filename(company_name)
    date = safe_filename(period_end_date)

    base_name = f"{code}_{name}_{date}".strip("_")
    if not base_name:
        raise ValueError("リネーム用の情報が不足しています（code/name/date が空です）。")

    new_file_path = os.path.join(dir_path, f"{base_name}.xlsx")

    counter = 1
    while os.path.exists(new_file_path):
        new_file_path = os.path.join(dir_path, f"{base_name}_{counter}.xlsx")
        counter += 1

    try:
        os.rename(original_path, new_file_path)
    except PermissionError as e:
        # Excelで開いている、または別プロセスが掴んでいる可能性が高い
        raise PermissionError(
            f"リネームできません（ファイルが開かれている可能性があります）。\n"
            f"対象: {original_path}\n"
            f"対策: Excelで該当ファイルを閉じてから再実行してください。"
        ) from e

    logger.info(f"Excelファイルがリネームされました: {new_file_path}")
    return new_file_path

cell_map_annual = {
    # 売上
    'NetSalesPrior4': 'D5',
    'NetSalesPrior3': 'G5',
    'NetSalesPrior2': 'J5',
    'NetSalesPrior1': 'M5',
    'NetSalesCurrent': 'P36',   # 半期なしのときだけ使う

    # 売上原価〜営業利益（当期P6〜P9 は半期なしのときだけ）
    'CostOfSalesPrior4': 'D6', 'CostOfSalesPrior3': 'G6', 'CostOfSalesPrior2': 'J6', 'CostOfSalesPrior1': 'M6', 'CostOfSalesCurrent': 'P6',
    'GrossProfitPrior4': 'D7', 'GrossProfitPrior3': 'G7', 'GrossProfitPrior2': 'J7', 'GrossProfitPrior1': 'M7', 'GrossProfitCurrent': 'P7',
    'SellingExpensesPrior4': 'D8','SellingExpensesPrior3': 'G8','SellingExpensesPrior2': 'J8','SellingExpensesPrior1': 'M8','SellingExpensesCurrent': 'P8',
    'OperatingIncomePrior4': 'D9','OperatingIncomePrior3': 'G9','OperatingIncomePrior2': 'J9','OperatingIncomePrior1': 'M9','OperatingIncomeCurrent': 'P9',

    # 経常・純利益
    'OrdinaryIncomePrior4': 'D10',
    'OrdinaryIncomePrior3': 'G10',
    'OrdinaryIncomePrior2': 'J10',
    'OrdinaryIncomePrior1': 'M10',
    'OrdinaryIncomeCurrent': 'P37',  # 半期なしのみ

    'ProfitLossPrior4': 'D11',
    'ProfitLossPrior3': 'G11',
    'ProfitLossPrior2': 'J11',
    'ProfitLossPrior1': 'M11',
    'ProfitLossCurrent': 'P38',      # 半期なしのみ

    # 発行株式数（自己株控除後）
    'TotalNumberPrior4': 'D13',
    'TotalNumberPrior3': 'G13',
    'TotalNumberPrior2': 'J13',
    'TotalNumberPrior1': 'M13',
    'TotalNumberCurrent': 'P40',     # 半期なしのみ

    # BS
    'TotalAssetsPrior4': 'C17',
    'TotalAssetsPrior3': 'F17',
    'TotalAssetsPrior2': 'I17',
    'TotalAssetsPrior1': 'L17',
    'TotalAssetsCurrent': 'O44',     # 半期なしのみ（あなたの運用に合わせたセル）

    'NetAssetsPrior4': 'D17',
    'NetAssetsPrior3': 'G17',
    'NetAssetsPrior2': 'J17',
    'NetAssetsPrior1': 'M17',
    'NetAssetsCurrent': 'P44',       # 半期なしのみ

    # CF
    'OperatingCashPrior4': 'C21',
    'OperatingCashPrior3': 'F21',
    'OperatingCashPrior2': 'I21',
    'OperatingCashPrior1': 'L21',
    'OperatingCashCurrent': 'O48',   # 半期なしのみ

    'InvestmentCashPrior4': 'D21',
    'InvestmentCashPrior3': 'G21',
    'InvestmentCashPrior2': 'J21',
    'InvestmentCashPrior1': 'M21',
    'InvestmentCashCurrent': 'P48',  # 半期なしのみ

    'FinancingCashPrior4': 'E21',
    'FinancingCashPrior3': 'H21',
    'FinancingCashPrior2': 'K21',
    'FinancingCashPrior1': 'N21',
    'FinancingCashCurrent': 'Q48',   # 半期なしのみ

    'CashAndCashEquivalentsPrior4': 'D22',
    'CashAndCashEquivalentsPrior3': 'G22',
    'CashAndCashEquivalentsPrior2': 'J22',
    'CashAndCashEquivalentsPrior1': 'M22',
    'CashAndCashEquivalentsCurrent': 'P49',  # 半期なしのみ

    # 表紙
    'SecurityCodeDEI': 'K2',
    'CompanyNameCoverPage': 'L2',
    'CurrentFiscalYearEndDateDEIyear': 'N2',
    'CurrentFiscalYearEndDateDEImonth': 'O2',
}

cell_map_half = {
    'NetSalesYTD': 'J36',
    'OrdinaryIncomeYTD': 'J37',
    'ProfitLossYTD': 'J38',

    'TotalNumberQuarter': 'J40',
    'TotalAssetsQuarter': 'I44',
    'NetAssetsQuarter': 'J44',

    'OperatingCashYTD': 'I48',
    'InvestmentCashYTD': 'J48',
    'FinancingCashYTD': 'K48',

    'CashAndCashEquivalentsQuarter': 'J49',
}

# ファイルパスを順番に確認する関数
def find_available_excel_file(base_path, file_name, max_copies=30):
    """
    方針：
      1) まずコピー系（- コピー / - コピー (n)）を探し、あれば「更新日時が最新」を返す
      2) コピーが無ければ、オリジナル（file_name.xlsx/.xlsm）を探す
      3) オリジナルが見つかったら「新しいコピーを作って」そのパスを返す
      4) それも無ければワイルドカードで最後の救済
    """
    import shutil

    exts = ["xlsx", "xlsm"]

    def newest(paths):
        paths = [p for p in paths if os.path.exists(p)]
        if not paths:
            return None
        paths.sort(key=lambda p: os.path.getmtime(p), reverse=True)
        return paths[0]

    def make_next_copy(original_path):
        # 例：決算分析シート-1.xlsx → 決算分析シート-1 - コピー.xlsx / (2) / (3)...
        root, ext = os.path.splitext(original_path)
        base = root  # 既に拡張子無しのフルパス

        # まず「 - コピー.ext」が空いていればそこへ
        candidate0 = f"{base} - コピー{ext}"
        if not os.path.exists(candidate0):
            shutil.copy2(original_path, candidate0)
            return candidate0

        # 次に (2)〜 を探す
        for i in range(2, max_copies + 2):
            cand = f"{base} - コピー ({i}){ext}"
            if not os.path.exists(cand):
                shutil.copy2(original_path, cand)
                return cand

        raise RuntimeError("コピー上限に達しました（max_copies を増やしてください）")

    # --- 0) 名前ゆれ吸収（_ と - の両方を試す） ---
    variants = {file_name}
    variants.add(file_name.replace("決算分析シート-", "決算分析シート_"))
    variants.add(file_name.replace("決算分析シート_", "決算分析シート-"))

    # --- 1) まずコピーを探す（厳密） ---
    copy_hits = []
    for v in variants:
        for ext in exts:
            copy_hits.append(os.path.join(base_path, f"{v} - コピー.{ext}"))
            for i in range(2, max_copies + 2):
                copy_hits.append(os.path.join(base_path, f"{v} - コピー ({i}).{ext}"))

    hit = newest(copy_hits)
    if hit:
        return hit

    # --- 2) コピーが無い → オリジナルを探す ---
    originals = []
    for v in variants:
        for ext in exts:
            originals.append(os.path.join(base_path, f"{v}.{ext}"))

    orig = newest(originals)
    if orig:
        # ★重要：オリジナルを返さず、必ずコピーを作って返す
        try:
            new_copy = make_next_copy(orig)
            logger.info(f"テンプレから作業用コピーを作成しました: {new_copy}")
            return new_copy
        except Exception:
            logger.exception("作業用コピー作成に失敗しました")
            raise

    # --- 3) 最終フォールバック：ワイルドカード（コピー表記のゆれ吸収） ---
    patterns = []
    for v in variants:
        patterns += [
            os.path.join(base_path, f"{v}*コピー*.xlsx"),
            os.path.join(base_path, f"{v}*コピー*.xlsm"),
        ]
    hits = []
    for pat in patterns:
        hits.extend(glob.glob(pat))

    hit2 = newest(hits)
    if hit2:
        return hit2

    logger.warning(f"{file_name} のいずれのバージョンも見つかりませんでした。")
    return None

#「どのキーをどの書類が書くか」を分けるフィルタ
def filter_keys_for_file(data: dict, file_kind: str) -> dict:
    """
    file_kind:
      - "half"  : file1（半期最新）→ YTD と Quarter と DEIだけ
      - "annual": file2（最新有報）→ Current/Prior* と DEIだけ（YTD/Quarterは触らない）
      - "annual_old": file3（過去有報）→ Prior* だけ（Currentは触らない）
    """
    out = {}

    for k, v in data.items():
        if file_kind == "half":
            if k.endswith("YTD") or k.endswith("Quarter") or k in ("SecurityCodeDEI", "CurrentFiscalYearEndDateDEI"):
                out[k] = v

        elif file_kind == "annual":
            if (k.endswith("Current") or k.endswith(("Prior1","Prior2","Prior3","Prior4"))
                or k in ("SecurityCodeDEI", "CurrentFiscalYearEndDateDEI")):
                # YTD/Quarterは annual 側では書かない
                if not (k.endswith("YTD") or k.endswith("Quarter")):
                    out[k] = v

        elif file_kind == "annual_old":
            # 古い有報はPriorだけ埋めたい
            if k.endswith(("Prior1", "Prior2", "Prior3", "Prior4")):
                out[k] = v
            # DEIは触らなくてOK（必要なら入れても良いが基本不要）

    return out

# ===============================
# 有報（XBRL）の期末日から「期末年（西暦）」を取得する関数
# 例：
#   "2025-03-31" → 2025
# file2（最新有報）の基準年を決めるために使用する
# ===============================
def get_fy_end_year(xbrl_data: dict) -> int | None:
    """
    CurrentFiscalYearEndDateDEI (YYYY-MM-DD) から期末年(西暦)を取り出す
    """
    s = xbrl_data.get("CurrentFiscalYearEndDateDEI")
    if isinstance(s, str):
        try:
            return datetime.strptime(s, "%Y-%m-%d").year
        except Exception:
            return None
    return None

def shift_suffixes_by_yeargap(data: dict, year_gap: int) -> dict:
    """
    file2の期末年を基準に、古い有報(file3)の Current/Prior を Prior側へずらす。

    year_gap = base_year(file2) - year(file3)
    例：file2=2025, file3=2023 → year_gap=2
        file3の NetSalesCurrent → NetSalesPrior2
        file3の NetSalesPrior1 → NetSalesPrior3
    """
    if year_gap <= 0:
        return data

    out = {}
    for k, v in data.items():
        m = re.match(r"^(.*?)(Current|Prior(\d+))$", k)
        if not m:
            out[k] = v
            continue

        prefix = m.group(1)
        suf = m.group(2)

        if suf == "Current":
            n = year_gap
        else:
            n = year_gap + int(m.group(3))

        # Prior1..4 の範囲だけ採用（あなたのExcelが5年枠の想定のため）
        if 1 <= n <= 4:
            out[f"{prefix}Prior{n}"] = v

    return out

def shift_with_keep(data: dict, year_gap: int,
                    keep_keys=("SecurityCodeDEI","CurrentFiscalYearEndDateDEI","CompanyNameCoverPage")) -> dict:
    """
    shift_suffixes_by_yeargap は Current/Prior 系しか返さないため、
    DEIなど「残したいキー」を shift後に戻すラッパー
    """
    shifted = shift_suffixes_by_yeargap(data, year_gap)
    for k in keep_keys:
        if k in data:
            shifted[k] = data[k]
    return shifted

def filter_for_annual(data: dict, use_half: bool = False) -> dict:
    """
    通期テンプレ（旧 cell_map_annual 相当）に存在するキーだけ返す。
    use_half=True（半期あり）のときは *_Current を書かない運用を維持。
    """
    allow = {
        # 売上
        "NetSalesPrior4", "NetSalesPrior3", "NetSalesPrior2", "NetSalesPrior1", "NetSalesCurrent",

        # 売上原価〜営業利益
        "CostOfSalesPrior4", "CostOfSalesPrior3", "CostOfSalesPrior2", "CostOfSalesPrior1", "CostOfSalesCurrent",
        "GrossProfitPrior4", "GrossProfitPrior3", "GrossProfitPrior2", "GrossProfitPrior1", "GrossProfitCurrent",
        "SellingExpensesPrior4", "SellingExpensesPrior3", "SellingExpensesPrior2", "SellingExpensesPrior1", "SellingExpensesCurrent",
        "OperatingIncomePrior4", "OperatingIncomePrior3", "OperatingIncomePrior2", "OperatingIncomePrior1", "OperatingIncomeCurrent",

        # 経常・純利益
        "OrdinaryIncomePrior4", "OrdinaryIncomePrior3", "OrdinaryIncomePrior2", "OrdinaryIncomePrior1", "OrdinaryIncomeCurrent",
        "ProfitLossPrior4", "ProfitLossPrior3", "ProfitLossPrior2", "ProfitLossPrior1", "ProfitLossCurrent",

        # 発行株式数（自己株控除後）
        "TotalNumberPrior4", "TotalNumberPrior3", "TotalNumberPrior2", "TotalNumberPrior1", "TotalNumberCurrent",

        # BS
        "TotalAssetsPrior4", "TotalAssetsPrior3", "TotalAssetsPrior2", "TotalAssetsPrior1", "TotalAssetsCurrent",
        "NetAssetsPrior4", "NetAssetsPrior3", "NetAssetsPrior2", "NetAssetsPrior1", "NetAssetsCurrent",

        # CF
        "OperatingCashPrior4", "OperatingCashPrior3", "OperatingCashPrior2", "OperatingCashPrior1", "OperatingCashCurrent",
        "InvestmentCashPrior4", "InvestmentCashPrior3", "InvestmentCashPrior2", "InvestmentCashPrior1", "InvestmentCashCurrent",
        "FinancingCashPrior4", "FinancingCashPrior3", "FinancingCashPrior2", "FinancingCashPrior1", "FinancingCashCurrent",

        # 現金等
        "CashAndCashEquivalentsPrior4", "CashAndCashEquivalentsPrior3", "CashAndCashEquivalentsPrior2",
        "CashAndCashEquivalentsPrior1", "CashAndCashEquivalentsCurrent",

        # 表紙系（必要なら）
        "SecurityCodeDEI",
        "CompanyNameCoverPage",
    }

    out = {k: v for k, v in data.items() if k in allow}

    # 半期あり運用：通期の *_Current は使わない（今まで通り）
    if use_half:
        out = {k: v for k, v in out.items() if not k.endswith("Current")}

    return out

def filter_for_annual_old(data: dict) -> dict:
    """
    過去有報（file3）は Prior 側だけを使う。
    さらにテンプレに入力欄がある指標だけに絞る。
    """
    allow = {
        # 売上
        "NetSalesPrior4", "NetSalesPrior3", "NetSalesPrior2", "NetSalesPrior1",

        # 売上原価〜営業利益
        "CostOfSalesPrior4", "CostOfSalesPrior3", "CostOfSalesPrior2", "CostOfSalesPrior1",
        "GrossProfitPrior4", "GrossProfitPrior3", "GrossProfitPrior2", "GrossProfitPrior1",
        "SellingExpensesPrior4", "SellingExpensesPrior3", "SellingExpensesPrior2", "SellingExpensesPrior1",
        "OperatingIncomePrior4", "OperatingIncomePrior3", "OperatingIncomePrior2", "OperatingIncomePrior1",

        # 経常・純利益
        "OrdinaryIncomePrior4", "OrdinaryIncomePrior3", "OrdinaryIncomePrior2", "OrdinaryIncomePrior1",
        "ProfitLossPrior4", "ProfitLossPrior3", "ProfitLossPrior2", "ProfitLossPrior1",

        # 発行株式数（自己株控除後）
        "TotalNumberPrior4", "TotalNumberPrior3", "TotalNumberPrior2", "TotalNumberPrior1",

        # BS
        "TotalAssetsPrior4", "TotalAssetsPrior3", "TotalAssetsPrior2", "TotalAssetsPrior1",
        "NetAssetsPrior4", "NetAssetsPrior3", "NetAssetsPrior2", "NetAssetsPrior1",

        # CF
        "OperatingCashPrior4", "OperatingCashPrior3", "OperatingCashPrior2", "OperatingCashPrior1",
        "InvestmentCashPrior4", "InvestmentCashPrior3", "InvestmentCashPrior2", "InvestmentCashPrior1",
        "FinancingCashPrior4", "FinancingCashPrior3", "FinancingCashPrior2", "FinancingCashPrior1",

        # 現金等
        "CashAndCashEquivalentsPrior4", "CashAndCashEquivalentsPrior3",
        "CashAndCashEquivalentsPrior2", "CashAndCashEquivalentsPrior1",
    }

    return {k: v for k, v in data.items() if k in allow}

def filter_for_half(data: dict) -> dict:
    """
    半期テンプレ（旧 cell_map_half 相当）に存在するキーだけ返す。
    NamedRange移行後も、テンプレに無い項目は出力しない。
    """
    allow = {
        # PL（YTD）
        "NetSalesYTD",
        "OrdinaryIncomeYTD",
        "ProfitLossYTD",

        # 株式・BS（Quarter）
        "TotalNumberQuarter",
        "TotalAssetsQuarter",
        "NetAssetsQuarter",

        # CF（YTD）
        "OperatingCashYTD",
        "InvestmentCashYTD",
        "FinancingCashYTD",

        # 現金等（Quarter）
        "CashAndCashEquivalentsQuarter",

        # 表紙（必要なら）
        "SecurityCodeDEI",
        "CompanyNameCoverPage",
        "CurrentFiscalYearEndDateDEIyear",
        "CurrentFiscalYearEndDateDEImonth",
    }

    out = {}
    for k, v in data.items():
        if k in allow:
            out[k] = v
    return out

def make_annual_map_for_use_half(use_half: bool, base_map: dict) -> dict:
    """
    半期ありの場合：
      - 通期Current（P列やO44/P44/O48…）はテンプレ仕様上「書かない」
      - Prior側（D/G/J/M）は埋める
    """
    m = dict(base_map)
    if use_half:
        for k in list(m.keys()):
            if k.endswith("Current") and k not in ("SecurityCodeDEI","CompanyNameCoverPage",
                                                  "CurrentFiscalYearEndDateDEIyear","CurrentFiscalYearEndDateDEImonth"):
                m.pop(k, None)
    return m

# =========================
# NamedRange Writer + キー変換（Scopeなし版）
# =========================
# 変換対象の「語尾」一覧（あなたの out キー体系に合わせる）
_SUFFIXES = [
    "YTD",
    "Quarter",
    "Current",
    "Prior1", "Prior2", "Prior3", "Prior4",
]

_suffix_pat = re.compile(rf"^(.+?)({'|'.join(_SUFFIXES)})$")

def to_namedrange_key(key: str) -> str:
    """
    例：
      NetSalesYTD        -> NetSales_YTD
      TotalAssetsQuarter -> TotalAssets_Quarter
      ProfitLossPrior2   -> ProfitLoss_Prior2
      SecurityCodeDEI    -> SecurityCodeDEI（そのまま）
    """
    if not isinstance(key, str) or not key:
        return key

    # DEI系や既に '_' を含むものは基本そのまま（必要なら後で個別対応）
    if "_" in key:
        return key
    if key.endswith("DEI"):
        return key

    m = _suffix_pat.match(key)
    if not m:
        return key

    metric = m.group(1)
    suffix = m.group(2)
    return f"{metric}_{suffix}"


def transform_keys_for_namedranges(data: dict) -> dict:
    """
    data のキーを NamedRange 名に合う形へ変換して返す
    """
    out = {}
    for k, v in data.items():
        nk = to_namedrange_key(k)
        out[nk] = v
    return out


def _iter_namedrange_cells(workbook: openpyxl.Workbook, range_name: str):
    """
    NamedRangeが指すセル（1セル/複数セル）を返す
    - ref が "A1" だと ws[ref] は Cell を返す
    - ref が "A1:B2" だと ws[ref] は 2次元タプルを返す
    """
    dn = workbook.defined_names.get(range_name)
    if dn is None:
        return

    for sheet_name, ref in dn.destinations:
        if sheet_name not in workbook.sheetnames:
            continue
        ws = workbook[sheet_name]

        obj = ws[ref]

        # 1セル（Cell）ケース
        if isinstance(obj, Cell):
            yield obj
            continue

        # 範囲（2次元のタプル）ケース
        # obj は ((Cell, Cell, ...), (Cell, Cell, ...), ...)
        for row in obj:
            for cell in row:
                yield cell

# raw_edinet 用：キーを (metric_key, time_slot) に分解
_TIME_SLOTS = ("YTD", "Quarter", "Current", "Prior1", "Prior2", "Prior3", "Prior4")

def split_metric_timeslot(key: str) -> tuple[str, str | None]:
    for ts in _TIME_SLOTS:
        if key.endswith(ts):
            return key[: -len(ts)], ts
    return key, None

def build_raw_rows_from_out(*,
                            company_code: str | None,
                            doc_id: str,
                            doc_type: str,
                            out: dict,
                            out_meta: dict) -> list[dict]:
    """
    out/out_meta から raw_rows（RAW_COLSに一致するdict配列）を作る
    - out_meta を主として回す（MISSING行を残すため）
    """
    rows: list[dict] = []

    for key, meta in (out_meta or {}).items():
        metric_key, time_slot = split_metric_timeslot(key)

        rows.append({
            "company_code": company_code,
            "doc_id": doc_id,
            "doc_type": doc_type,
            "consolidation": meta.get("consolidation"),
            "metric_key": metric_key,
            "time_slot": time_slot,
            "period_start": meta.get("period_start"),
            "period_end": meta.get("period_end"),
            "period_kind": meta.get("period_kind"),
            "value": out.get(key),
            "unit": meta.get("unit"),
            "tag_used": meta.get("tag_used"),
            "tag_rank": meta.get("tag_rank"),
            "status": meta.get("status"),
        })

    return rows

def _split_key(key: str):
    """
    key例:
      NetSalesYTD / NetSalesCurrent / NetSalesPrior1 / TotalAssetsQuarter
    戻り:
      metric_key="NetSales", time_slot="YTD" など
    """
    if key.endswith("YTD"):
        return key[:-3], "YTD"
    if key.endswith("Quarter"):
        return key[:-7], "Quarter"
    if key.endswith("Current"):
        return key[:-7], "Current"
    if key.startswith("Prior") and False:
        pass
    # Prior1..4
    for n in ("1","2","3","4"):
        suf = f"Prior{n}"
        if key.endswith(suf):
            return key[:-len(suf)], suf
    return key, ""


def append_rows_from_meta(raw_rows: list, *, company_code: str, doc_id: str, doc_type: str, out: dict, out_meta: dict):
    """
    out_meta を raw_rows に展開して追記
    """
    for key, meta in (out_meta or {}).items():
        metric_key, time_slot = _split_key(key)

        raw_rows.append({
            "company_code": company_code,
            "doc_id": doc_id,
            "doc_type": doc_type,
            "consolidation": meta.get("consolidation"),
            "metric_key": metric_key,
            "time_slot": time_slot,
            "period_start": meta.get("period_start"),
            "period_end": meta.get("period_end"),
            "period_kind": meta.get("period_kind"),
            "value": out.get(key),                 # out の値
            "unit": meta.get("unit"),
            "tag_used": meta.get("tag_used"),
            "tag_rank": meta.get("tag_rank"),
            "status": meta.get("status", "OK"),
        })


def append_missing_annual_ytd_rows(raw_rows: list, *, company_code: str, doc_id: str, out_meta: dict, duration_metric_keys: list[str]):
    """
    annual の YTD を「必ず行として残す」
    - out_meta に {metric}YTD が無いものは MISSING 行を追加
    """
    for metric in duration_metric_keys:
        key = f"{metric}YTD"
        if (out_meta or {}).get(key):
            continue

        raw_rows.append({
            "company_code": company_code,
            "doc_id": doc_id,
            "doc_type": "annual",
            "consolidation": None,
            "metric_key": metric,
            "time_slot": "YTD",
            "period_start": None,
            "period_end": None,
            "period_kind": "duration",
            "value": None,
            "unit": None,
            "tag_used": None,
            "tag_rank": None,
            "status": "MISSING",
        })

def write_rows_to_raw_sheet(excel_file: str, rows: list[dict], *, sheet_name: str = "raw_edinet"):
    """
    raw_edinet シートに rows を書き込む（ヘッダ1行は残し、2行目以降を再生成）
    重要：
      - period_start / period_end は Excel 側の MAXIFS 等で「日付」として扱える必要がある
      - よって "YYYY-MM-DD" 文字列 → datetime.date に変換して書く
    """
    import datetime as _dt

    def _to_excel_date(v):
        # すでに date/datetime ならそのまま
        if isinstance(v, _dt.datetime):
            return v.date()
        if isinstance(v, _dt.date):
            return v

        # "YYYY-MM-DD" を date に変換
        if isinstance(v, str):
            s = v.strip()
            try:
                return _dt.datetime.strptime(s, "%Y-%m-%d").date()
            except Exception:
                return v  # 変換できないものはそのまま（監査優先）
        return v

    wb = openpyxl.load_workbook(
        excel_file,
        keep_vba=excel_file.lower().endswith(".xlsm")
    )
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"sheet not found: {sheet_name}")

    ws = wb[sheet_name]

    # 1) 2行目以降を全消し
    if ws.max_row >= 2:
        ws.delete_rows(2, ws.max_row - 1)

    # 2) rowsを書き込み（型を整える）
    r = 2
    for row in rows:
        for c, col in enumerate(RAW_COLS, start=1):
            v = row.get(col)

            # ★ここが肝：日付は Excel 日付型へ
            if col in ("period_start", "period_end"):
                v = _to_excel_date(v)

            ws.cell(row=r, column=c).value = v
        r += 1

    wb.save(excel_file)

# === PATCH: out_meta の suffix shift（file3用）===
def shift_out_meta_by_yeargap(out_meta: dict, year_gap: int) -> dict:
    """
    out_meta のキー（NetSalesCurrent / NetSalesPrior1..）を year_gap 分ずらす
    値(meta dict)はそのまま
    """
    if year_gap <= 0:
        return out_meta

    shifted = {}
    for k, meta in (out_meta or {}).items():
        m = re.match(r"^(.*?)(Current|Prior(\d+))$", k)
        if not m:
            shifted[k] = meta
            continue

        prefix = m.group(1)
        suf = m.group(2)

        if suf == "Current":
            n = year_gap
        else:
            n = year_gap + int(m.group(3))

        if 1 <= n <= 4:
            shifted[f"{prefix}Prior{n}"] = meta

    return shifted

def snapshot_namedranges(wb, values: dict[str, object]) -> dict[str, object]:
    """
    values: {range_name: new_value} の形（これから書く対象だけ）
    戻り値: {range_name: old_value} （復旧用）
    """
    old = {}
    for name in values.keys():
        try:
            old[name] = wb.names[name].refers_to_range.value
        except Exception:
            old[name] = None  # 存在しない/参照できない場合
    return old

def restore_namedranges(wb, old_values: dict[str, object]) -> None:
    for name, v in old_values.items():
        try:
            wb.names[name].refers_to_range.value = v
        except Exception:
            pass

def write_data_to_excel_namedranges(excel_file: str, data: dict, *,
                                   transform_keys: bool = True,
                                   skip_if_formula: bool = True,
                                   skip_values=("データなし", "", None),
                                   dry_run: bool = False) -> dict:
    """
    data の key と同名の NamedRange に書き込む（cell_map不要）
    - transform_keys=True のとき NetSalesYTD -> NetSales_YTD のように変換してから書く
    - 数式セルは上書きしない（あなたの既存挙動を踏襲）
    戻り値:
      {
        "written": [(name, "Sheet!A1"), ...],
        "skipped": [(name, reason), ...],
        "missing":  [name, ...]
      }
    """
    wb = openpyxl.load_workbook(excel_file, keep_vba=excel_file.lower().endswith(".xlsm"))
    result = {"written": [], "skipped": [], "missing": []}

    payload = transform_keys_for_namedranges(data) if transform_keys else dict(data)

    for name, value in payload.items():
        # None / 空 / データなし は書かない
        if value in skip_values or (isinstance(value, str) and value.strip() in skip_values):
            result["skipped"].append((name, "empty"))
            continue

        cells = list(_iter_namedrange_cells(wb, name))
        if not cells:
            result["missing"].append(name)
            continue

        for cell in cells:
            if skip_if_formula and isinstance(cell.value, str) and cell.value.startswith("="):
                result["skipped"].append((name, f"formula@{cell.coordinate}"))
                continue
            cell.value = value
            result["written"].append((name, f"{cell.parent.title}!{cell.coordinate}"))

    if not dry_run:
        wb.save(excel_file)

    return result

# 処理するファイル数を選択
def choose_file_count():
    while True:
        try:
            count = int(input("処理するファイルの数を選択してください（1～50）: "))
            if 1 <= count <= 50:
                return count
            else:
                print("1から50の範囲で入力してください。")
        except ValueError:
            print("数字を入力してください。")

def load_config(file_path):
            with open(file_path, "r", encoding="utf-8") as f:
                return json.load(f)

# 各ループで処理するファイルパスをリスト化
def build_loops(base_dir, template_dir, max_n=50):
    """
    現在：決算分析シート-{n}.xlsx を対象に loops を作る
    将来：会社コード_会社名 方式に切替予定（この関数だけ差し替えればOK）
    """
    loops = []
    for n in range(1, max_n + 1):
        file1 = glob.glob(os.path.join(base_dir, f"{n}-2*.xbrl"))
        file2 = glob.glob(os.path.join(base_dir, f"{n}-4*.xbrl"))
        file3 = glob.glob(os.path.join(base_dir, f"{n}-5*.xbrl"))

        # デバッグ（必要な時だけ）
        logger.debug(f"[{n}] file1: {file1}")
        logger.debug(f"[{n}] file2: {file2}")
        logger.debug(f"[{n}] file3: {file3}")

        excel_file_path = os.path.join(template_dir, f"決算分析シート_{n}.xlsx")

        loops.append({
            "xbrl_file_paths": {"file1": file1, "file2": file2, "file3": file3},
            "excel_file_path": excel_file_path,
            "slot": n,  # 将来の拡張用（ログ/追跡用）
        })

    return loops

# XBRLデータの取得、証券コードの取得、Excelへの書き込み、株価データ取得までをループ処理に含める
def process_one_loop(loop, date_pairs, skipped_files):
    
    # === ANCHOR: LOOP_START === 
    parsed_docs = []   # ★ここに file1/2/3 の parse結果を溜める（raw書込は最後に1回）
    raw_rows = []

    # === PATCH: out_buffer put（上書き検知つき）===
    out_buffer_src = {}          # key -> src label
    out_buffer_collisions = []   # (key, old_src, new_src)

    def buffer_put(key: str, value, src_label: str):
        if value is None or value == "":
            return

        prio = {"file1_half": 3, "file2_annual": 2, "file3_annual": 1}
        new_p = prio.get(src_label, 0)

        if key in out_buffer:
            old_src = out_buffer_src.get(key, "?")
            old_p = prio.get(old_src, 0)

            # half優先：halfが既に入っていれば annual は上書きしない
            if old_src == "file1_half" and src_label in ("file2_annual", "file3_annual"):
                out_buffer_collisions.append((key, old_src, src_label))
                return

            # それ以外は優先度で弱い方は上書き禁止
            if new_p < old_p:
                out_buffer_collisions.append((key, old_src, src_label))
                return

            out_buffer_collisions.append((key, old_src, src_label))

        out_buffer[key] = value
        out_buffer_src[key] = src_label

    out_buffer: dict[str, object] = {}
    
    run_id = datetime.now().strftime("%Y%m%d_%H%M%S_%f")

    t0 = perf_counter()

    loop_event = {
        "ts": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "run_id": run_id,
        "slot": loop.get("slot"),
        "excel": os.path.basename(excel_file_path) if "excel_file_path" in locals() else None,
        "security_code": None,
        "phases": {},        # {"file1_parse": {"ok": True, "sec": 0.12}, ...}
        "counts": {},        # {"raw_rows": 123, "excel_ranges": 45, ...}
        "errors": [],        # 例外など（ここに短い文字列で）
    }

    # Excel探索用
    excel_base_name = os.path.basename(loop["excel_file_path"]).replace(".xlsx", "")
    excel_directory = os.path.dirname(loop["excel_file_path"])

    # --- Excel選択失敗 ---
    selected_file = find_available_excel_file(excel_directory, excel_base_name)
    if not selected_file:
        add_skip(
            skipped_files,
            code=SkipCode.EXCEL_NOT_FOUND,
            phase="excel_select",
            loop=loop,
            excel=excel_base_name,
            xbrl=None,
            message="使用するExcelが見つからない"
        )
        logger.warning("使用するファイルが見つかりませんでした。次のループを実行します。")
        return

    logger.info(f"使用するExcelファイル（元）: {selected_file}")

    # ★安全装置：テンプレ（元）をコピーして作業ファイルに書く（原本保護）
    base_no_ext, ext = os.path.splitext(selected_file)
    work_excel_path = f"{base_no_ext}_work_{run_id}{ext}"
    shutil.copy(selected_file, work_excel_path)

    # 以降は作業ファイルへ書く（loopにも反映）
    loop["excel_file_path"] = work_excel_path
    excel_file_path = work_excel_path
    loop_event["excel"] = os.path.basename(excel_file_path)
    logger.info(f"使用するExcelファイル（作業）: {excel_file_path}")

    # ★安全ロック：work以外に書き込ませない
    if "_work_" not in os.path.splitext(excel_file_path)[0]:
        logger.critical(f"安全ロック発動：workではないExcelに書き込もうとしました: {excel_file_path}")
        raise SystemExit(1)

    rename_info = None
    xbrl_file_paths = loop["xbrl_file_paths"]
    security_code = None
    base_year = None

    # -------------------------
    # 0) file1（半期）があれば先に読む（base_year決定）
    # -------------------------
    x1 = None
    use_half = bool(xbrl_file_paths.get("file1") and xbrl_file_paths["file1"])

    # --- file1（半期）解析失敗 ---
    if use_half:
        try:
            t = perf_counter()

            path1 = xbrl_file_paths["file1"][0]
            x1, sc1, meta1 = parse_xbrl_data(path1, mode="half")

            base_year = get_fy_end_year(x1)

            parsed_docs.append({
                "doc_id": os.path.basename(path1),
                "doc_type": "half",
                "out": x1,
                "out_meta": meta1,
                "parsed_code": sc1,
            })

            logger.info(f"[parse bench] mode=half xbrl={os.path.basename(path1)} out={len(x1)} meta={len(meta1)} sec={round(perf_counter()-t,3)}")

            # raw_rows（監査DB）へ追加
            company_code = ensure_security_code(x1, sc1, None)

            # ★成功記録（file1）
            loop_event["phases"]["file1_parse"] = {"ok": True, "sec": round(perf_counter() - t, 3)}

        except Exception as e:
            loop_event["phases"]["file1_parse"] = {"ok": False, "sec": None}  # ★ここ（exceptの最初）
            loop_event["errors"].append("file1_parse_error")

            add_skip(
                skipped_files,
                code=SkipCode.FILE1_PARSE_ERROR,
                phase="file1",
                loop=loop,
                excel=excel_file_path,
                xbrl=(xbrl_file_paths["file1"][0] if xbrl_file_paths.get("file1") else None),
                message="file1(半期) 解析エラー",
                exc=e
            )
            x1 = None
            use_half = False

    # -------------------------
    # 1) file2（最新有報）
    # -------------------------
    x2 = None
    meta2 = None
    path2 = None

    if xbrl_file_paths.get("file2") and xbrl_file_paths["file2"]:
        try:
            t = perf_counter()
            
            path2 = xbrl_file_paths["file2"][0]
            x2, parsed_security_code, meta2 = parse_xbrl_data(path2, mode="full")

            # === ANCHOR: AFTER_PARSE file2(annual) ===
            parsed_docs.append({
                "doc_id": os.path.basename(path2),
                "doc_type": "annual",
                "out": x2,
                "out_meta": meta2,
                "parsed_code": parsed_security_code,
            })
            
            logger.info(f"[parse bench] mode=full xbrl={os.path.basename(path2)} out={len(x2)} meta={len(meta2)} sec={round(perf_counter()-t,3)}")

            # ★ここで必ず確定（優先順位固定）
            security_code = ensure_security_code(x2, parsed_security_code, x1)
            
            # ★成功記録（file2）
            loop_event["phases"]["file2_parse"] = {"ok": True, "sec": round(perf_counter() - t, 3)}

            # file2 の Excel反映（annual）
            out2_write = filter_for_annual(x2, use_half=use_half)  # use_half=TrueならCurrentは書かない

            logger.warning(f"[buffer debug] file2_annual keys={sorted(list(out2_write.keys()))}")
            logger.warning(f"[buffer debug] file2_annual nonempty={sum(1 for v in out2_write.values() if v not in (None, ''))}")

            for k, v in out2_write.items():
                buffer_put(k, v, "file2_annual")

        except Exception as e:
            loop_event["phases"]["file2_parse"] = {"ok": False, "sec": None} 
            loop_event["errors"].append("file2_error")

            add_skip(
                skipped_files,
                code=SkipCode.FILE2_ERROR,
                phase="file2",
                loop=loop,
                excel=excel_file_path,
                xbrl=path2,
                message="file2(最新有報) 解析/書込エラー",
                exc=e
            )
    else:
        add_skip(
            skipped_files,
            code=SkipCode.FILE2_NOT_FOUND,
            phase="file2",
            loop=loop,
            excel=excel_file_path,
            xbrl=None,
            message="file2(最新有報) が見つからない"
        )

    # -------------------------
    # 2) file3（過去有報）→ Prior補完
    # -------------------------
    # --- file3（過去有報）年取れず / 失敗 ---
    if base_year is not None and xbrl_file_paths.get("file3") and xbrl_file_paths["file3"]:
        try:
            t = perf_counter()

            path3 = xbrl_file_paths["file3"][0]

            # === ANCHOR: AFTER_PARSE file3(annual) ===
            x3, sc3, meta3 = parse_xbrl_data(path3, mode="full")

            parsed_docs.append({
                "doc_id": os.path.basename(path3),
                "doc_type": "annual",
                "out": x3,
                "out_meta": meta3,
                "parsed_code": sc3,
            })

            # 期末年は1回だけ
            y3 = get_fy_end_year(x3)

            # raw_rows（監査DB）へ追加（file3も残す）
            company_code = security_code or ensure_security_code(x3, sc3, x1) or ""

            # === ANCHOR: AFTER_PARSE file3(annual) ===
            # ★成功記録（file3）
            loop_event["phases"]["file3_parse"] = {"ok": True, "sec": round(perf_counter() - t, 3)}

            if y3 is None:
                add_skip(
                    skipped_files,
                    code=SkipCode.FILE3_YEAR_MISS,
                    phase="file3",
                    loop=loop,
                    excel=excel_file_path,
                    xbrl=path3,
                    message="file3 期末年が取れない"
                )
            else:
                # ① year_gap を計算（例：base_year - y3）
                year_gap = base_year - y3

                # === PATCH: year_gap 安全チェック ===
                if not (1 <= year_gap <= 4):
                    add_skip(
                        skipped_files,
                        code=SkipCode.FILE3_YEAR_MISS,
                        phase="file3",
                        loop=loop,
                        excel=excel_file_path,
                        xbrl=path3,
                        message=f"file3 year_gap abnormal base={base_year} y3={y3} gap={year_gap}"
                    )
                else:
                    # ② x3 を Prior 側にずらす
                    x3_shifted = shift_with_keep(x3, year_gap)

                    # meta3 もずらす
                    meta3_shifted = shift_out_meta_by_yeargap(meta3, year_gap)

                    # parsed_docs を更新
                    parsed_docs[-1]["out"] = x3_shifted
                    parsed_docs[-1]["out_meta"] = meta3_shifted

                    # ③ ずらした結果を Excel書込みバッファへ
                    out3_write = filter_for_annual_old(x3_shifted)  # ここはあなたの関数名でOK

                    logger.warning(f"[buffer debug] file3_annual keys={sorted(list(out3_write.keys()))}")
                    for k, v in out3_write.items():
                        buffer_put(k, v, "file3_annual")
                    # （中略：あなたの既存処理）
                    pass

        except Exception as e:
            loop_event["phases"]["file3_parse"] = {"ok": False, "sec": None} 
            loop_event["errors"].append("file3_error")

            add_skip(
                skipped_files,
                code=SkipCode.FILE3_ERROR,
                phase="file3",
                loop=loop,
                excel=excel_file_path,
                xbrl=(xbrl_file_paths["file3"][0] if xbrl_file_paths.get("file3") else None),
                message="file3(過去有報) 解析/書込エラー",
                exc=e
            )

    # -------------------------
    # 3) 半期ありなら最後にYTD/Quarter確定
    # -------------------------
    # --- half（半期）書込失敗 ---
    if use_half and x1 is not None:
        try:
            t = perf_counter()

            out_half = filter_for_half(x1)

            logger.warning(f"[buffer debug] half_final keys={sorted(list(out_half.keys()))}")
            logger.warning(f"[buffer debug] half_final nonempty={sum(1 for v in out_half.values() if v not in (None, ''))}")

            for k, v in out_half.items():
                buffer_put(k, v, "half_final")

            loop_event["phases"]["half_buffer"] = {"ok": True, "sec": round(perf_counter() - t, 3)}

            logger.info(f"[half finalize bench] sec={round(perf_counter()-t,3)}")

        except Exception as e:
            loop_event["phases"]["half_buffer"] = {"ok": False, "sec": None}
            loop_event["errors"].append("half_buffer_error")

            add_skip(
                skipped_files,
                code=SkipCode.HALF_WRITE_ERROR,
                phase="half",
                loop=loop,
                excel=excel_file_path,
                xbrl=(xbrl_file_paths["file1"][0] if xbrl_file_paths.get("file1") else None),
                message="half(半期) 書込エラー",
                exc=e
            )

    for r in raw_rows:
        r["run_id"] = run_id
        r["source_file"] = r.get("doc_id")

    # === ANCHOR: BEFORE_EXCEL_WRITE ===
    if out_buffer_collisions:
        logger.warning("[excel buffer] collisions=%d", len(out_buffer_collisions))
        for k, old_src, new_src in out_buffer_collisions[:50]:
            winner = out_buffer_src.get(k, "?")
            logger.warning(" overwrite: %s  %s -> %s (winner=%s)", k, old_src, new_src, winner)
    if out_buffer:
        t = perf_counter() 

        write_data_to_excel_namedranges(excel_file_path, out_buffer)

        loop_event["phases"]["excel_write"] = {"ok": True, "sec": round(perf_counter() - t, 3)}  # ★ここ
        logger.info(f"[excel write] ranges={len(out_buffer)}")
    else:
        loop_event["phases"]["excel_write"] = {"ok": True, "sec": 0.0}

    # === ANCHOR: BEFORE_RAW_BUILD ===
    # company_code（raw用）はここで1回だけ確定
    company_code_for_raw = security_code or ""
    if not company_code_for_raw:
        # 最後の保険：parsed_docs に parsed_code があれば拾う
        for d in parsed_docs:
            pc = d.get("parsed_code")
            if pc:
                company_code_for_raw = pc
                break
    company_code_for_raw = company_code_for_raw or ""

    # out_meta 主導で raw を作る（MISSING行を残せる方式）
    for d in parsed_docs:
        raw_rows.extend(
            build_raw_rows_from_out(
                company_code=company_code_for_raw,
                doc_id=d["doc_id"],
                doc_type=d["doc_type"],
                out=d["out"],
                out_meta=d["out_meta"],
            )
        )

    # annual の YTD(MISSING) を必ず残す（docごと）
    duration_metric_keys = [
        "NetSales",
        "CostOfSales",
        "GrossProfit",
        "SellingExpenses",
        "OperatingIncome",
        "OrdinaryIncome",
        "ProfitLoss",
        "OperatingCash",
        "InvestmentCash",
        "FinancingCash",
    ]
    for d in parsed_docs:
        if d["doc_type"] == "annual":
            append_missing_annual_ytd_rows(
                raw_rows,
                company_code=company_code_for_raw,
                doc_id=d["doc_id"],
                out_meta=d["out_meta"],
                duration_metric_keys=duration_metric_keys,
            )

    # === PATCH: raw_rows 最終一意化（重複ゼロ化）===

    def _raw_key(row: dict):
        # 「同一判定キー」：raw異常検知のロジックと合わせる
        # まずは監査の粒度（会社×書類×doc_type×metric×slot×連結×期間×unit）で固める
        return (
            row.get("company_code", ""),
            row.get("doc_id", ""),
            row.get("doc_type", ""),
            row.get("consolidation", ""),
            row.get("metric_key", ""),
            row.get("time_slot", ""),
            row.get("period_start", ""),
            row.get("period_end", ""),
            row.get("period_kind", ""),   # duration/instant
            row.get("unit", ""),
        )

    # === PATCH: raw_rows 最終一意化（テンプレの重複判定に寄せる）===

    def _raw_key_for_template(row: dict):
        # テンプレの重複判定に合わせる（期間は見ない想定）
        # ★ここが重要：period_start/end を入れない
        return (
            row.get("company_code", ""),
            row.get("doc_type", ""),
            row.get("consolidation", ""),
            row.get("metric_key", ""),
            row.get("time_slot", ""),
            row.get("period_kind", ""),   # duration/instant を区別したい場合だけ残す
        )

    def dedupe_raw_rows_keep_best(rows: list[dict]) -> tuple[list[dict], int]:
        """
        同一キー（テンプレ基準）を1行に統合する。
        優先順位:
        OK > MISSING > ERROR
        tag_rank が小さいほど良い（あれば）
        period_end が新しい方を優先（あれば）
        """
        import datetime as _dt

        def _to_date(x):
            if isinstance(x, _dt.datetime): return x.date()
            if isinstance(x, _dt.date): return x
            return None

        def score(r: dict):
            status = (r.get("status") or "").upper()
            status_score = {"OK": 3, "MISSING": 2, "ERROR": 1}.get(status, 0)

            tr = r.get("tag_rank")
            try:
                tag_rank_score = -int(tr)    # 小さいほど良い -> マイナスで大きいほど良いにする
            except Exception:
                tag_rank_score = -999999

            has_value = 1 if (r.get("value") not in (None, "")) else 0
            has_unit  = 1 if (r.get("unit")  not in (None, "")) else 0

            pe = _to_date(r.get("period_end"))
            period_score = pe.toordinal() if pe else -1

            return (status_score, has_value, has_unit, tag_rank_score, period_score)

        best = {}
        dup_count = 0
        for row in rows:
            k = _raw_key_for_template(row)
            if k not in best:
                best[k] = row
            else:
                dup_count += 1
                if score(row) > score(best[k]):
                    best[k] = row

        # 元順をなるべく維持
        seen = set()
        out = []
        for row in rows:
            k = _raw_key_for_template(row)
            if k in seen:
                continue
            out.append(best[k])
            seen.add(k)

        return out, dup_count
    
    from collections import Counter
    cnt = Counter(_raw_key_for_template(r) for r in raw_rows)
    dup_keys = [k for k,v in cnt.items() if v > 1]
    if dup_keys:
        logger.warning("[raw dup still] groups=%d (show top 10)", len(dup_keys))
        for k in dup_keys[:10]:
            logger.warning(" dup_key=%s count=%d", k, cnt[k])

    raw_rows, deduped = dedupe_raw_rows_keep_best(raw_rows)
    if deduped:
        logger.warning("[raw dedupe] removed_duplicates=%d final_rows=%d", deduped, len(raw_rows)) 

    # run_id / source_file は raw 確定後に一括付与
    for r in raw_rows:
        r["run_id"] = run_id
        r["source_file"] = r.get("doc_id")

    # === ANCHOR: BEFORE_RAW_WRITE ===
    write_rows_to_raw_sheet(excel_file_path, raw_rows, sheet_name="raw_edinet")
    logger.info(f"[raw] written rows={len(raw_rows)} sheet=raw_edinet")

    # === ANCHOR: BEFORE_RAW_WRITE ===
    # 1) company_code を確定
    company_code_for_raw = security_code
    if not company_code_for_raw:
        for d in parsed_docs:
            if d.get("parsed_code"):
                company_code_for_raw = d["parsed_code"]
                break
    company_code_for_raw = company_code_for_raw or ""

    # 3) run_id/source_file を付与（ここがベスト）
    for r in raw_rows:
        r["run_id"] = run_id
        r["source_file"] = r.get("doc_id")

    # === ANCHOR: LOOP_SUMMARY ===

    loop_event["security_code"] = security_code
    loop_event["counts"]["raw_rows"] = len(raw_rows)
    loop_event["counts"]["excel_ranges"] = len(out_buffer)
    loop_event["counts"]["skipped_in_loop"] = sum(
        1 for s in skipped_files if s.get("slot") == loop.get("slot")
    )

    loop_event["phases"]["loop_total"] = {
        "ok": True,
        "sec": round(perf_counter() - t0, 3)
    }

    # JSONLログへ保存
    jsonl_path = os.path.join(os.getcwd(), "logs", "loop_summary.jsonl")
    os.makedirs(os.path.dirname(jsonl_path), exist_ok=True)

    with open(jsonl_path, "a", encoding="utf-8") as f:
        f.write(json.dumps(loop_event, ensure_ascii=False) + "\n")

    # 人間が読む用ログ
    logger.info(
        f"[loop summary] slot={loop.get('slot')} "
        f"code={security_code} "
        f"excel_ranges={loop_event['counts']['excel_ranges']} "
        f"raw_rows={loop_event['counts']['raw_rows']} "
        f"sec={loop_event['phases']['loop_total']['sec']}"
    )

    # -------------------------
    # 4) 株価（非致命）
    # -------------------------
    if security_code:
        logger.info(f"取得した証券コード: {security_code}")
        stock_code = f"{security_code}.T"
        logger.debug(f"[stock check] security_code={security_code} stock_code={stock_code}")
        try:
            stock_result = write_stock_data_to_excel(excel_file_path, stock_code, date_pairs)
            if stock_result:
                logger.debug(
                    f"[stock] written={stock_result.get('written',0)} "
                    f"miss={stock_result.get('miss',0)} "
                    f"errors={stock_result.get('errors',0)} "
                    f"missing_name={stock_result.get('missing_name',0)} "
                    f"bad_input={stock_result.get('bad_input',0)}"
                )

        except Exception:
            logger.exception("株価データの書き込みで想定外エラー（続行）")
    # --- 証券コード無し（株価の前提） ---
    else:
        add_skip(
            skipped_files,
            code=SkipCode.NO_SECURITY_CODE,
            phase="stock",
            loop=loop,
            excel=excel_file_path,
            xbrl=None,
            message="証券コードが取得できない"
        )
        logger.warning("証券コードが取得できませんでした。")

    # -------------------------
    # 5) リネーム（非致命）
    # -------------------------
    if rename_info:
        try:
            rename_excel_file(
                excel_file_path,
                rename_info[0],  # security_code
                rename_info[1],  # company_name
                rename_info[2],  # period_end_date
            )
        except Exception:
            logger.exception("リネーム中にエラーが発生しました（続行）")


def main():
    # 作業ディレクトリを表示
    logger.info(f"作業ディレクトリ: {os.getcwd()}")

    # === XBRLフォルダ設定 ===
    base_dir = r"C:\Users\silve\OneDrive\PC\開発\test\Python\XBRL"
    template_dir = r"C:\Users\silve\OneDrive\PC\EDINET\決算分析シート"

    logger.info(f"XBRLフォルダ（固定）: {base_dir}")
    if not os.path.isdir(base_dir):
        logger.critical(f"base_dir が存在しません。パスを確認してください: {base_dir}")
        raise SystemExit(1)

    # === 件数入力 ===
    file_count = choose_file_count()

    # === スキップ一覧（ローカル）===
    skipped_files = []

    # === loops生成 ===
    loops = build_loops(base_dir, template_dir, max_n=50)

    # === 決算期を最初に1回だけ選択 ===
    try:
        config = load_config("決算期_KANPE.json")
        chosen_period = input("決算期を選択してください（例 25-1）: ")

        if chosen_period not in config:
            logger.critical("無効な選択です。プログラムを終了します。")
            raise SystemExit(1)

        date_pairs = config[chosen_period]
        validate_stock_date_pairs(date_pairs)
        logger.info(f"選択された決算期: {chosen_period}")
        logger.debug(f"決算期データ: {date_pairs}")
    except Exception:
        logger.exception("決算期設定の読み込み/選択に失敗しました")
        raise SystemExit(1)

    # === ★ここが必要：1件ずつ処理する（呼び出し）===
    for i in range(min(file_count, len(loops))):
        try:
            process_one_loop(loops[i], date_pairs, skipped_files)
        except SystemExit:
            raise
        except Exception:
            logger.exception(f"1件処理で想定外エラー（続行）: index={i}")

    # === スキップ一覧表示（最後）===
    log_skip_summary(logger, skipped_files)

if __name__ == "__main__":
    logger = setup_logger(debug=DEBUG)
    try:
        logger.info("===== プログラム開始 =====")
        main()
        logger.info("===== 正常終了 =====")
    except SystemExit:
        raise
    except Exception:
        logger.exception("致命的エラーで終了しました")
        raise