from __future__ import annotations

import shutil
import sqlite3
import sys
import unittest
import uuid
from pathlib import Path

from openpyxl import load_workbook


ROOT_DIR = Path(__file__).resolve().parents[2]
SRC_DIR = ROOT_DIR / "src"
TMP_ROOT = ROOT_DIR / "tests" / "_tmp_data_quality_report_service"
if str(SRC_DIR) not in sys.path:
    sys.path.insert(0, str(SRC_DIR))

from edinet_monitor.db.schema import create_tables, get_connection  # noqa: E402
from edinet_monitor.services.data_quality_report_service import (  # noqa: E402
    DataQualityReportOptions,
    export_data_quality_report,
)


def _insert_fixture(conn: sqlite3.Connection) -> None:
    conn.execute(
        """
        INSERT INTO issuer_master (
            edinet_code, security_code, company_name, market, industry_33,
            industry_17, is_listed, exchange, listing_category_raw,
            listing_source, updated_at
        ) VALUES (
            'E00001', '1111', 'Test Corp', 'Prime', 'Information',
            'IT', 1, 'TSE', 'Prime', 'fixture', '2026-05-25'
        )
        """
    )
    conn.execute(
        """
        INSERT INTO filings (
            doc_id, edinet_code, security_code, form_type, period_end,
            submit_date, amendment_flag, doc_info_edit_status, legal_status,
            accounting_standard, document_display_unit, zip_path, xbrl_path,
            xbrl_member_name, download_status, parse_status, created_at, updated_at
        ) VALUES (
            'S100AAAA', 'E00001', '1111', '030000', '2025-03-31',
            '2025-06-30', 0, NULL, NULL,
            'JGAAP', 'yen', 'zip/S100AAAA.zip', 'xbrl/S100AAAA.xbrl',
            'member', 'downloaded', 'derived_metrics_saved', '2026-05-25', '2026-05-25'
        )
        """
    )
    conn.execute(
        """
        INSERT INTO raw_facts (
            doc_id, tag_name, context_ref, unit_ref, period_type, period_end,
            consolidation, value_text, created_at
        ) VALUES (
            'S100AAAA', 'NetSales', 'CurrentYearDuration', 'JPY',
            'duration', '2025-03-31', 'consolidated', '100000000', '2026-05-25'
        )
        """
    )
    conn.execute(
        """
        INSERT INTO normalized_metrics (
            doc_id, edinet_code, security_code, metric_key, fiscal_year,
            period_end, value_num, source_tag, consolidation, rule_version,
            created_at, updated_at
        ) VALUES (
            'S100AAAA', 'E00001', '1111', 'NetSalesCurrent', 2025,
            '2025-03-31', 100000000, 'NetSales', 'consolidated', 'v1',
            '2026-05-25', '2026-05-25'
        )
        """
    )
    conn.executemany(
        """
        INSERT INTO derived_metrics (
            doc_id, edinet_code, security_code, metric_key, metric_base,
            metric_group, fiscal_year, period_end, period_scope, period_offset,
            consolidation, accounting_standard, document_display_unit, value_num,
            value_unit, calc_status, formula_name, source_detail_json,
            rule_version, created_at, updated_at
        ) VALUES (
            'S100AAAA', 'E00001', '1111', ?, ?, 'fixture', 2025,
            '2025-03-31', 'current', 0, 'consolidated', 'JGAAP', 'yen',
            ?, ?, 'ok', 'fixture', '{}', 'v1', '2026-05-25', '2026-05-25'
        )
        """,
        [
            ("NetSalesCurrent", "NetSales", 100000000.0, "yen"),
            ("EquityRatioCurrent", "EquityRatio", 10.0, "ratio"),
        ],
    )
    conn.commit()


class DataQualityReportServiceTest(unittest.TestCase):
    def setUp(self) -> None:
        self.tmp_path = TMP_ROOT / uuid.uuid4().hex
        self.tmp_path.mkdir(parents=True, exist_ok=True)
        self.db_path = self.tmp_path / "monitor.db"
        self.output_dir = self.tmp_path / "reports"
        create_tables(self.db_path)
        self.conn = get_connection(self.db_path)
        _insert_fixture(self.conn)

    def tearDown(self) -> None:
        self.conn.close()
        shutil.rmtree(self.tmp_path, ignore_errors=True)

    def test_export_data_quality_report_writes_excel_and_persists_run(self) -> None:
        result = export_data_quality_report(
            self.conn,
            options=DataQualityReportOptions(
                date_from="2025-01-01",
                date_to="2025-12-31",
                codes=("1111",),
                output_dir=self.output_dir,
            ),
        )

        self.assertTrue(result.excel_path.exists())
        self.assertGreater(result.issue_count, 0)
        self.assertTrue(any(item.check_name == "ratio_metric_extreme_value" for item in result.items))
        run_count = self.conn.execute("SELECT COUNT(*) FROM data_quality_report_runs").fetchone()[0]
        item_count = self.conn.execute("SELECT COUNT(*) FROM data_quality_report_items").fetchone()[0]
        self.assertEqual(run_count, 1)
        self.assertEqual(item_count, len(result.items))

        workbook = load_workbook(result.excel_path, read_only=True)
        try:
            self.assertEqual(
                set(workbook.sheetnames),
                {"Summary", "Diff", "EDINET_Status", "Metric_Coverage", "JQuants_Quality", "Issues"},
            )
        finally:
            workbook.close()

    def test_second_report_uses_previous_run_for_diff_values(self) -> None:
        first = export_data_quality_report(
            self.conn,
            options=DataQualityReportOptions(
                date_from="2025-01-01",
                date_to="2025-12-31",
                codes=("1111",),
                output_dir=self.output_dir,
            ),
        )
        second = export_data_quality_report(
            self.conn,
            options=DataQualityReportOptions(
                date_from="2025-01-01",
                date_to="2025-12-31",
                codes=("1111",),
                output_dir=self.output_dir,
            ),
        )

        self.assertEqual(second.previous_run_id, first.run_id)
        filing_count = next(item for item in second.items if item.check_name == "filing_count")
        self.assertEqual(filing_count.previous_value, filing_count.current_value)
        self.assertEqual(filing_count.delta_value, 0)

    def test_report_history_keeps_latest_twenty_runs_and_items(self) -> None:
        results = []
        for _ in range(21):
            results.append(
                export_data_quality_report(
                    self.conn,
                    options=DataQualityReportOptions(
                        date_from="2025-01-01",
                        date_to="2025-12-31",
                        codes=("1111",),
                        output_dir=self.output_dir,
                    ),
                )
            )

        run_rows = self.conn.execute(
            """
            SELECT run_id
            FROM data_quality_report_runs
            ORDER BY generated_at DESC, id DESC
            """
        ).fetchall()
        self.assertEqual(len(run_rows), 20)
        remaining_run_ids = {str(row["run_id"]) for row in run_rows}
        self.assertNotIn(results[0].run_id, remaining_run_ids)
        stale_items = self.conn.execute(
            """
            SELECT COUNT(*)
            FROM data_quality_report_items
            WHERE run_id = ?
            """,
            (results[0].run_id,),
        ).fetchone()[0]
        self.assertEqual(stale_items, 0)


if __name__ == "__main__":
    unittest.main()
