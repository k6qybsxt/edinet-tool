from __future__ import annotations

import json
import shutil
import sqlite3
import sys
import unittest
import uuid
from pathlib import Path
from unittest.mock import patch

from openpyxl import load_workbook


ROOT_DIR = Path(__file__).resolve().parents[2]
SRC_DIR = ROOT_DIR / "src"
TMP_ROOT = ROOT_DIR / "tests" / "_tmp_metric_excel_audit_service"
if str(SRC_DIR) not in sys.path:
    sys.path.insert(0, str(SRC_DIR))

from edinet_monitor.services.metric_excel_audit_service import (  # noqa: E402
    HEADER_DECISION,
    HEADER_METRIC,
    HEADER_ROW_KIND,
    HEADER_VALUE_KIND,
    ExcelAuditOptions,
    audit_metric_excel,
    read_metric_excel_summary_options,
)
from edinet_monitor.services.metric_excel_export_service import (  # noqa: E402
    GENERAL_SHEET,
    MetricExcelCondition,
    MetricExcelRow,
    PERIOD_LABEL_BY_OFFSET,
    ROW_KIND_DETAIL,
    VALUE_KIND_BASE,
    build_metric_excel_rows,
    write_metric_excel,
)
from tests.edinet_monitor.test_metric_excel_export_service import (  # noqa: E402
    _create_schema,
    _insert_company,
)


def _write_target_config(path: Path, *codes: str) -> None:
    path.write_text(
        json.dumps(
            {
                "version": 1,
                "target_sets": {
                    "normal": [
                        {"security_code": code, "features": ["test"]}
                        for code in codes
                    ],
                    "known_issue": [],
                },
            },
            ensure_ascii=False,
        ),
        encoding="utf-8",
    )


def _header_index(ws, header: str) -> int:
    headers = [cell.value for cell in ws[1]]
    return headers.index(header) + 1


def _set_first_metric_cell(path: Path, *, metric_label: str, header: str, value: object) -> None:
    workbook = load_workbook(path)
    try:
        ws = workbook[GENERAL_SHEET]
        metric_col = _header_index(ws, HEADER_METRIC)
        target_col = _header_index(ws, header)
        for row_idx in range(2, ws.max_row + 1):
            if ws.cell(row_idx, metric_col).value == metric_label:
                ws.cell(row_idx, target_col).value = value
                workbook.save(path)
                return
    finally:
        workbook.close()
    raise AssertionError(f"metric row not found: {metric_label}")


def _insert_current_derived_metric(
    conn: sqlite3.Connection,
    *,
    metric_base: str,
    value_num: float,
    value_unit: str,
    document_display_unit: str = "",
) -> None:
    conn.execute(
        """
        INSERT INTO derived_metrics (
            doc_id, edinet_code, security_code, metric_key, metric_base, metric_group,
            fiscal_year, period_end, period_scope, period_offset, consolidation,
            accounting_standard, document_display_unit, value_num, value_unit, calc_status,
            formula_name, source_detail_json, rule_version, created_at, updated_at
        ) VALUES (
            'E00001_0', 'E00001', '11110', ?, ?, 'test',
            2025, '2026-03-31', 'annual', 0, 'consolidated',
            'Japan GAAP', ?, ?, ?, 'ok',
            'test', '{}', 'v1', '2026-04-24', '2026-04-24'
        )
        """,
        (
            f"{metric_base}Current",
            metric_base,
            document_display_unit,
            value_num,
            value_unit,
        ),
    )


def _insert_current_market_metric(
    conn: sqlite3.Connection,
    *,
    metric_base: str,
    value_num: float,
    value_unit: str,
) -> None:
    conn.execute(
        """
        INSERT INTO market_derived_metrics (
            source_type, source_id, edinet_code, security_code, period_scope, period_key,
            quarter_type, fiscal_year, period_end, metric_key, metric_base, metric_group,
            value_num, value_unit, calc_status, formula_name, source_detail_json,
            rule_version, created_at, updated_at
        ) VALUES (
            'edinet', 'E00001_0', 'E00001', '1111', 'annual', 'annual:FY',
            NULL, 2026, '2026-03-31', ?, ?, 'market',
            ?, ?, 'ok', 'test', '{}',
            'v1', '2026-04-24', '2026-04-24'
        )
        """,
        (f"{metric_base}Current", metric_base, value_num, value_unit),
    )


class MetricExcelAuditServiceTest(unittest.TestCase):
    def setUp(self) -> None:
        self.tmp_path = TMP_ROOT / uuid.uuid4().hex
        self.tmp_path.mkdir(parents=True, exist_ok=True)
        self.target_config_path = self.tmp_path / "targets.json"
        _write_target_config(self.target_config_path, "1111")
        self.conn = sqlite3.connect(":memory:")
        self.conn.row_factory = sqlite3.Row
        _create_schema(self.conn)
        _insert_company(
            self.conn,
            edinet_code="E00001",
            security_code="1111",
            company_name="A\u793e",
            industry_33="\u5316\u5b66",
            net_sales_values=[100_000_000.0, 80_000_000.0, 70_000_000.0],
        )
        self.conn.commit()

    def tearDown(self) -> None:
        self.conn.close()
        shutil.rmtree(self.tmp_path, ignore_errors=True)

    def _options(
        self,
        excel_path: Path,
        *,
        period_offsets: tuple[int, ...] = (0,),
        period_scopes: tuple[str, ...] = ("annual",),
    ) -> ExcelAuditOptions:
        return ExcelAuditOptions(
            excel_path=excel_path,
            target_set="normal",
            target_config_path=self.target_config_path,
            output_dir=self.tmp_path / "reports",
            period_offsets=period_offsets,
            period_scopes=period_scopes,
        )

    def _write_expected_workbook(
        self,
        path: Path,
        *,
        period_offsets: list[int] | None = None,
        period_scopes: list[str] | None = None,
    ) -> list[MetricExcelRow]:
        condition = MetricExcelCondition(
            security_codes=["1111"],
            period_scopes=period_scopes or ["annual"],
            period_offsets=period_offsets or [0],
        )
        rows, errors, _warnings, _preview, target_companies = build_metric_excel_rows(self.conn, condition)
        self.assertEqual(errors, [])
        write_metric_excel(
            rows=rows,
            condition=condition,
            output_path=path,
            db_path=":memory:",
            errors=[],
            warnings=[],
            target_companies=target_companies,
        )
        return rows

    def test_missing_expected_row_is_critical(self) -> None:
        output_path = self.tmp_path / "empty.xlsx"
        write_metric_excel(
            rows=[],
            condition=MetricExcelCondition(security_codes=["1111"], period_offsets=[0]),
            output_path=output_path,
            db_path=":memory:",
            errors=[],
            warnings=[],
            target_companies=1,
        )

        result = audit_metric_excel(self.conn, self._options(output_path))

        self.assertTrue(
            any(
                issue.severity == "critical" and issue.check_name == "missing_expected_row"
                for issue in result.issues
            )
        )

    def test_summary_periods_are_used_when_offsets_are_not_explicit(self) -> None:
        output_path = self.tmp_path / "summary_periods.xlsx"
        self._write_expected_workbook(output_path, period_offsets=[1])

        offsets, segment_mode, warnings = read_metric_excel_summary_options(output_path)
        result = audit_metric_excel(
            self.conn,
            ExcelAuditOptions(
                excel_path=output_path,
                target_set="normal",
                target_config_path=self.target_config_path,
                output_dir=self.tmp_path / "reports",
                period_scopes=("annual",),
            ),
        )

        self.assertEqual(offsets, (1,))
        self.assertEqual(segment_mode, "none")
        self.assertEqual(warnings, [])
        self.assertFalse(any(issue.period_label == PERIOD_LABEL_BY_OFFSET[0] for issue in result.issues))

    def test_summary_segment_mode_is_used_when_not_explicit(self) -> None:
        output_path = self.tmp_path / "summary_segment.xlsx"
        row = MetricExcelRow(
            sheet_name=GENERAL_SHEET,
            security_code="1111",
            company_name="A\u793e",
            industry_33="\u5316\u5b66",
            market="Prime",
            period_scope="annual",
            current_period_end="2026-03-31",
            metric_base="NetSales",
            metric_label="4Q \u58f2\u4e0a\u9ad8 <\u5408\u8a08>",
            periods_by_offset={1: "\u901a\u671f 2025-03"},
            values_by_offset={1: 80.0},
            units_by_offset={1: "\u767e\u4e07\u5186"},
            ratios_by_offset={1: None},
            row_kind=ROW_KIND_DETAIL,
            segment_kind="\u5408\u8a08",
            segment_name="\u5408\u8a08",
        )
        condition = MetricExcelCondition(
            security_codes=["1111"],
            period_offsets=[1],
            segment_mode="all",
        )
        write_metric_excel(
            rows=[row],
            condition=condition,
            output_path=output_path,
            db_path=":memory:",
            errors=[],
            warnings=[],
            target_companies=1,
        )
        captured = {}

        def fake_build_metric_excel_rows(conn, condition, **kwargs):
            captured["period_offsets"] = tuple(condition.period_offsets)
            captured["segment_mode"] = condition.segment_mode
            return [row], [], [], [], 1

        with patch(
            "edinet_monitor.services.metric_excel_audit_service.build_metric_excel_rows",
            fake_build_metric_excel_rows,
        ):
            result = audit_metric_excel(
                self.conn,
                ExcelAuditOptions(
                    excel_path=output_path,
                    target_set="normal",
                    target_config_path=self.target_config_path,
                    output_dir=self.tmp_path / "reports",
                    period_scopes=("annual",),
                ),
            )

        self.assertEqual(captured["period_offsets"], (1,))
        self.assertEqual(captured["segment_mode"], "all")
        self.assertEqual(result.issue_count, 0)

    def test_legacy_row_kind_header_and_values_are_normalized(self) -> None:
        output_path = self.tmp_path / "legacy_row_kind.xlsx"
        self._write_expected_workbook(output_path)
        workbook = load_workbook(output_path)
        try:
            ws = workbook[GENERAL_SHEET]
            decision_col = _header_index(ws, HEADER_DECISION)
            value_kind_col = _header_index(ws, HEADER_VALUE_KIND)
            ws.cell(1, value_kind_col).value = HEADER_ROW_KIND
            for row_idx in range(2, ws.max_row + 1):
                if ws.cell(row_idx, decision_col).value == "4Q":
                    ws.cell(row_idx, decision_col).value = "\u901a\u671f"
                if ws.cell(row_idx, value_kind_col).value == VALUE_KIND_BASE:
                    ws.cell(row_idx, value_kind_col).value = ROW_KIND_DETAIL
            workbook.save(output_path)
        finally:
            workbook.close()

        result = audit_metric_excel(self.conn, self._options(output_path))

        self.assertEqual(result.issue_count, 0)

    def test_blank_value_cell_does_not_create_value_mismatch(self) -> None:
        output_path = self.tmp_path / "blank_value.xlsx"
        rows = self._write_expected_workbook(output_path, period_offsets=[1])
        net_sales = next(row for row in rows if row.metric_base == "NetSales" and row.security_code == "1111")
        _set_first_metric_cell(
            output_path,
            metric_label=net_sales.metric_label,
            header=f"{PERIOD_LABEL_BY_OFFSET[1]}_\u6570\u5024",
            value=None,
        )

        result = audit_metric_excel(self.conn, self._options(output_path, period_offsets=(1,)))

        self.assertFalse(any(issue.check_name == "value_mismatch" for issue in result.issues))

    def test_nonblank_value_mismatch_is_critical(self) -> None:
        output_path = self.tmp_path / "value_mismatch.xlsx"
        rows = self._write_expected_workbook(output_path, period_offsets=[1])
        net_sales = next(row for row in rows if row.metric_base == "NetSales" and row.security_code == "1111")
        _set_first_metric_cell(
            output_path,
            metric_label=net_sales.metric_label,
            header=f"{PERIOD_LABEL_BY_OFFSET[1]}_\u6570\u5024",
            value=999999,
        )

        result = audit_metric_excel(self.conn, self._options(output_path, period_offsets=(1,)))

        self.assertTrue(
            any(
                issue.severity == "critical" and issue.check_name == "value_mismatch"
                for issue in result.issues
            )
        )

    def test_market_capitalization_and_outstanding_shares_units_match_export(self) -> None:
        _insert_current_market_metric(
            self.conn,
            metric_base="MarketCapitalization",
            value_num=12_345_678_900.0,
            value_unit="yen",
        )
        _insert_current_derived_metric(
            self.conn,
            metric_base="OutstandingShares",
            value_num=1_234_567.0,
            value_unit="shares",
        )
        self.conn.commit()
        output_path = self.tmp_path / "market_units.xlsx"
        self._write_expected_workbook(output_path)

        result = audit_metric_excel(self.conn, self._options(output_path))

        unit_mismatches = [issue for issue in result.issues if issue.check_name == "unit_mismatch"]
        self.assertEqual(unit_mismatches, [])

    def test_2q_period_in_annual_row_is_period_scope_mixing(self) -> None:
        output_path = self.tmp_path / "period_mixing.xlsx"
        rows = self._write_expected_workbook(output_path, period_offsets=[1])
        net_sales = next(row for row in rows if row.metric_base == "NetSales" and row.security_code == "1111")
        _set_first_metric_cell(
            output_path,
            metric_label=net_sales.metric_label,
            header=f"{PERIOD_LABEL_BY_OFFSET[1]}_\u671f\u9593",
            value="2Q 2026-09",
        )

        result = audit_metric_excel(self.conn, self._options(output_path, period_offsets=(1,)))

        self.assertTrue(
            any(
                issue.severity == "critical" and issue.check_name == "period_scope_mixing"
                for issue in result.issues
            )
        )

    def test_half_disabled_metric_present_is_critical(self) -> None:
        output_path = self.tmp_path / "half_disabled.xlsx"
        row = MetricExcelRow(
            sheet_name=GENERAL_SHEET,
            security_code="1111",
            company_name="A\u793e",
            industry_33="\u5316\u5b66",
            market="Prime",
            period_scope="quarter:2Q",
            current_period_end="2025-09-30",
            metric_base="OutstandingSharesGrowthRate",
            metric_label="2Q OutstandingSharesGrowthRate",
            periods_by_offset={0: "2Q 2025-09"},
            values_by_offset={0: 0.1},
            units_by_offset={0: "%"},
            ratios_by_offset={0: None},
            row_kind=ROW_KIND_DETAIL,
        )
        write_metric_excel(
            rows=[row],
            condition=MetricExcelCondition(security_codes=["1111"], period_scopes=["quarter"], period_offsets=[0]),
            output_path=output_path,
            db_path=":memory:",
            errors=[],
            warnings=[],
            target_companies=1,
        )

        result = audit_metric_excel(self.conn, self._options(output_path, period_scopes=("quarter",)))

        self.assertTrue(
            any(
                issue.severity == "critical"
                and issue.check_name == "half_disabled_metric_present"
                for issue in result.issues
            )
        )

    def test_ifrs_ordinary_income_expected_row_matches_profit_before_tax_label(self) -> None:
        self.conn.execute(
            "UPDATE filings SET accounting_standard = 'IFRS' WHERE doc_id = 'E00001_0'"
        )
        _insert_current_derived_metric(
            self.conn,
            metric_base="ProfitBeforeTax",
            value_num=42_000_000.0,
            value_unit="yen",
            document_display_unit="\u767e\u4e07\u5186",
        )
        self.conn.commit()
        output_path = self.tmp_path / "ifrs.xlsx"
        self._write_expected_workbook(output_path)

        result = audit_metric_excel(self.conn, self._options(output_path))

        self.assertFalse(
            any(
                issue.metric_base == "OrdinaryIncome"
                and issue.check_name in {"missing_expected_row", "value_mismatch"}
                for issue in result.issues
            )
        )

    def test_jquants_quarter_growth_row_can_be_audited(self) -> None:
        self.conn.executemany(
            """
            INSERT INTO jquants_financial_metrics (
                disclosure_number, local_code, security_code, edinet_code, metric_kind,
                period_scope, period_key, quarter_type, forecast_target, forecast_stage,
                fiscal_year, period_start, period_end, disclosed_date, disclosed_time,
                metric_key, metric_base, metric_group, value_num, value_unit, calc_status,
                source_field, source_detail_json, rule_version, created_at, updated_at
            ) VALUES (
                ?, '11110', '1111', 'E00001', 'actual',
                'quarter', 'actual:1Q', '1Q', NULL, NULL,
                ?, '', ?, ?, '12:00',
                'NetSalesCurrent', 'NetSales', 'sales', ?, 'yen', 'ok',
                'NetSales', '{}', 'v1', '2026-04-24', '2026-04-24'
            )
            """,
            [
                ("DISC2026", 2026, "2026-06-30", "2026-07-31", 200_000_000.0),
                ("DISC2025", 2025, "2025-06-30", "2025-07-31", 100_000_000.0),
            ],
        )
        self.conn.commit()
        output_path = self.tmp_path / "jquants_growth.xlsx"
        condition = MetricExcelCondition(
            security_codes=["1111"],
            metric_labels=["NetSalesGrowthRate"],
            period_scopes=["quarter"],
            period_offsets=[0],
        )
        rows, errors, _warnings, _preview, target_companies = build_metric_excel_rows(self.conn, condition)
        self.assertEqual(errors, [])
        write_metric_excel(
            rows=rows,
            condition=condition,
            output_path=output_path,
            db_path=":memory:",
            errors=[],
            warnings=[],
            target_companies=target_companies,
        )

        result = audit_metric_excel(self.conn, self._options(output_path, period_scopes=("quarter",)))

        self.assertFalse(any(issue.check_name == "value_mismatch" for issue in result.issues))


if __name__ == "__main__":
    unittest.main()
