from __future__ import annotations

import sqlite3
import shutil
import sys
import unittest
import uuid
from pathlib import Path

from openpyxl import Workbook, load_workbook


ROOT_DIR = Path(__file__).resolve().parents[2]
SRC_DIR = ROOT_DIR / "src"
TMP_ROOT = ROOT_DIR / "tests" / "_tmp_metric_excel_export"
if str(SRC_DIR) not in sys.path:
    sys.path.insert(0, str(SRC_DIR))

from edinet_monitor.services.metric_excel_export_service import (  # noqa: E402
    GENERAL_SHEET,
    MetricExcelCondition,
    MetricExcelRow,
    ROW_KIND_AVERAGE,
    ROW_KIND_DETAIL,
    ROW_KIND_MEDIAN,
    SUMMARY_SHEET,
    VERTICAL_DATA_SHEET,
    build_metric_excel_rows,
    export_metric_excel,
    read_metric_excel_condition,
    write_metric_excel,
)


def _create_condition_workbook(path: Path, rows: list[tuple[str, str]]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "条件"
    for key, value in rows:
        ws.append([key, value])
    wb.save(path)


def _create_schema(conn: sqlite3.Connection) -> None:
    conn.executescript(
        """
        CREATE TABLE issuer_master (
            edinet_code TEXT PRIMARY KEY,
            security_code TEXT,
            company_name TEXT NOT NULL,
            market TEXT,
            industry_33 TEXT,
            industry_17 TEXT,
            is_listed INTEGER NOT NULL DEFAULT 1,
            exchange TEXT,
            listing_category_raw TEXT,
            listing_source TEXT,
            updated_at TEXT NOT NULL
        );

        CREATE TABLE filings (
            doc_id TEXT PRIMARY KEY,
            edinet_code TEXT NOT NULL,
            security_code TEXT,
            form_type TEXT NOT NULL,
            period_end TEXT,
            submit_date TEXT,
            amendment_flag INTEGER NOT NULL DEFAULT 0,
            doc_info_edit_status TEXT,
            legal_status TEXT,
            accounting_standard TEXT,
            document_display_unit TEXT,
            zip_path TEXT,
            xbrl_path TEXT,
            download_status TEXT NOT NULL,
            parse_status TEXT NOT NULL,
            created_at TEXT NOT NULL,
            updated_at TEXT NOT NULL
        );

        CREATE TABLE normalized_metrics (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            doc_id TEXT NOT NULL,
            edinet_code TEXT NOT NULL,
            security_code TEXT,
            metric_key TEXT NOT NULL,
            fiscal_year INTEGER,
            period_end TEXT,
            value_num REAL,
            source_tag TEXT,
            consolidation TEXT,
            rule_version TEXT NOT NULL,
            created_at TEXT NOT NULL,
            updated_at TEXT NOT NULL
        );

        CREATE TABLE derived_metrics (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            doc_id TEXT NOT NULL,
            edinet_code TEXT NOT NULL,
            security_code TEXT,
            metric_key TEXT NOT NULL,
            metric_base TEXT NOT NULL,
            metric_group TEXT NOT NULL,
            fiscal_year INTEGER,
            period_end TEXT,
            period_scope TEXT NOT NULL,
            period_offset INTEGER NOT NULL DEFAULT 0,
            consolidation TEXT,
            accounting_standard TEXT,
            document_display_unit TEXT,
            value_num REAL,
            value_unit TEXT NOT NULL,
            calc_status TEXT NOT NULL,
            formula_name TEXT NOT NULL,
            source_detail_json TEXT,
            rule_version TEXT NOT NULL,
            created_at TEXT NOT NULL,
            updated_at TEXT NOT NULL
        );

        CREATE TABLE industry_aggregate_metrics (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            industry_33 TEXT NOT NULL,
            period_scope TEXT NOT NULL,
            fiscal_year INTEGER NOT NULL,
            period_bucket_start TEXT,
            period_bucket_end TEXT,
            metric_key TEXT NOT NULL,
            metric_base TEXT NOT NULL,
            metric_group TEXT NOT NULL,
            value_num REAL,
            value_unit TEXT NOT NULL,
            calc_status TEXT NOT NULL,
            formula_name TEXT NOT NULL,
            source_company_count INTEGER NOT NULL DEFAULT 0,
            source_detail_json TEXT,
            rule_version TEXT NOT NULL,
            created_at TEXT NOT NULL,
            updated_at TEXT NOT NULL
        );

        CREATE UNIQUE INDEX uq_industry_aggregate_metrics_scope
        ON industry_aggregate_metrics(industry_33, period_scope, fiscal_year, metric_key);

        CREATE TABLE market_derived_metrics (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            source_type TEXT NOT NULL,
            source_id TEXT NOT NULL,
            edinet_code TEXT,
            security_code TEXT NOT NULL,
            period_scope TEXT NOT NULL,
            period_key TEXT NOT NULL,
            quarter_type TEXT,
            fiscal_year INTEGER,
            period_end TEXT NOT NULL,
            metric_key TEXT NOT NULL,
            metric_base TEXT NOT NULL,
            metric_group TEXT NOT NULL,
            value_num REAL,
            value_unit TEXT NOT NULL,
            calc_status TEXT NOT NULL,
            formula_name TEXT NOT NULL,
            source_detail_json TEXT,
            rule_version TEXT NOT NULL,
            created_at TEXT NOT NULL,
            updated_at TEXT NOT NULL
        );

        CREATE TABLE jquants_financial_metrics (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            disclosure_number TEXT NOT NULL,
            local_code TEXT NOT NULL,
            security_code TEXT,
            edinet_code TEXT,
            metric_kind TEXT NOT NULL,
            period_scope TEXT NOT NULL,
            period_key TEXT NOT NULL,
            quarter_type TEXT,
            forecast_target TEXT,
            forecast_stage TEXT,
            fiscal_year INTEGER,
            period_start TEXT,
            period_end TEXT,
            disclosed_date TEXT,
            disclosed_time TEXT,
            metric_key TEXT NOT NULL,
            metric_base TEXT NOT NULL,
            metric_group TEXT NOT NULL,
            value_num REAL,
            value_unit TEXT NOT NULL,
            calc_status TEXT NOT NULL,
            source_field TEXT NOT NULL,
            source_detail_json TEXT,
            rule_version TEXT NOT NULL,
            created_at TEXT NOT NULL,
            updated_at TEXT NOT NULL
        );

        CREATE TABLE quarter_standalone_metrics (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            security_code TEXT NOT NULL,
            edinet_code TEXT,
            fiscal_year INTEGER NOT NULL,
            quarter_type TEXT NOT NULL,
            period_end TEXT,
            metric_key TEXT NOT NULL,
            metric_base TEXT NOT NULL,
            metric_group TEXT NOT NULL,
            value_num REAL,
            value_unit TEXT NOT NULL,
            calc_status TEXT NOT NULL,
            formula_name TEXT NOT NULL,
            source_detail_json TEXT,
            rule_version TEXT NOT NULL,
            created_at TEXT NOT NULL,
            updated_at TEXT NOT NULL
        );

        CREATE TABLE segment_metrics (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            doc_id TEXT NOT NULL,
            edinet_code TEXT NOT NULL,
            security_code TEXT,
            form_type TEXT NOT NULL,
            period_scope TEXT NOT NULL,
            quarter_type TEXT,
            fiscal_year INTEGER,
            period_start TEXT,
            period_end TEXT,
            segment_kind TEXT NOT NULL,
            segment_name TEXT NOT NULL,
            axis_qname TEXT,
            member_qname TEXT,
            metric_base TEXT NOT NULL,
            metric_key TEXT NOT NULL,
            value_kind TEXT NOT NULL,
            value_num REAL,
            value_unit TEXT NOT NULL,
            source_tag TEXT,
            tag_qname TEXT,
            context_ref TEXT,
            decimals TEXT,
            calc_status TEXT NOT NULL,
            source_detail_json TEXT,
            rule_version TEXT NOT NULL,
            created_at TEXT NOT NULL,
            updated_at TEXT NOT NULL
        );
        """
    )


def _detail_rows(rows):
    return [row for row in rows if row.row_kind == ROW_KIND_DETAIL]


def _insert_industry_aggregate_metric(
    conn: sqlite3.Connection,
    *,
    industry_33: str,
    fiscal_year: int,
    metric_base: str,
    value_num: float | None,
    value_unit: str = "yen",
    calc_status: str = "ok",
) -> None:
    conn.execute(
        """
        INSERT INTO industry_aggregate_metrics (
            industry_33, period_scope, fiscal_year, period_bucket_start, period_bucket_end,
            metric_key, metric_base, metric_group, value_num, value_unit, calc_status,
            formula_name, source_company_count, source_detail_json, rule_version,
            created_at, updated_at
        ) VALUES (
            ?, 'annual', ?, ?, ?, ?, ?, 'industry', ?, ?, ?,
            'test', 2, '{}', 'v1', '2026-04-24', '2026-04-24'
        )
        """,
        (
            industry_33,
            fiscal_year,
            f"{fiscal_year}-01-01",
            f"{fiscal_year}-12-31",
            f"{metric_base}Current",
            metric_base,
            value_num,
            value_unit,
            calc_status,
        ),
    )


def _insert_company(
    conn: sqlite3.Connection,
    *,
    edinet_code: str,
    security_code: str,
    company_name: str,
    industry_33: str,
    net_sales_values: list[float],
) -> None:
    conn.execute(
        """
        INSERT INTO issuer_master (
            edinet_code, security_code, company_name, market, industry_33, industry_17,
            is_listed, exchange, listing_category_raw, listing_source, updated_at
        ) VALUES (?, ?, ?, 'Prime', ?, ?, 1, 'TSE', 'Prime', 'csv', '2026-04-24')
        """,
        (edinet_code, f"{security_code}0", company_name, industry_33, industry_33),
    )
    period_ends = ["2026-03-31", "2025-03-31", "2024-03-31"]
    for offset, period_end in enumerate(period_ends):
        doc_id = f"{edinet_code}_{offset}"
        conn.execute(
            """
            INSERT INTO filings (
                doc_id, edinet_code, security_code, form_type, period_end, submit_date,
                amendment_flag, doc_info_edit_status, legal_status, accounting_standard,
                document_display_unit, zip_path, xbrl_path, download_status, parse_status,
                created_at, updated_at
            ) VALUES (?, ?, ?, '030000', ?, ?, 0, '0', '1', 'Japan GAAP', '百万円',
                      'zip', 'xbrl', 'downloaded', 'derived_metrics_saved',
                      '2026-04-24', '2026-04-24')
            """,
            (doc_id, edinet_code, f"{security_code}0", period_end, f"{period_end} 12:00"),
        )
        conn.executemany(
            """
            INSERT INTO normalized_metrics (
                doc_id, edinet_code, security_code, metric_key, fiscal_year, period_end,
                value_num, source_tag, consolidation, rule_version, created_at, updated_at
            ) VALUES (?, ?, ?, ?, 2025, ?, ?, 'tag', 'consolidated', 'v1', '2026-04-24', '2026-04-24')
            """,
            [
                (doc_id, edinet_code, f"{security_code}0", "NetSalesCurrent", period_end, net_sales_values[offset]),
                (doc_id, edinet_code, f"{security_code}0", "CostOfSalesCurrent", period_end, 60_000_000.0),
                (doc_id, edinet_code, f"{security_code}0", "NetAssetsCurrent", period_end, 50_000_000.0),
            ],
        )
        conn.executemany(
            """
            INSERT INTO derived_metrics (
                doc_id, edinet_code, security_code, metric_key, metric_base, metric_group,
                fiscal_year, period_end, period_scope, period_offset, consolidation,
                accounting_standard, document_display_unit, value_num, value_unit, calc_status,
                formula_name, source_detail_json, rule_version, created_at, updated_at
            ) VALUES (?, ?, ?, ?, ?, 'profitability', 2025, ?, 'annual', 0, 'consolidated',
                      'Japan GAAP', '百万円', ?, 'ratio', 'ok',
                      'test', '{}', 'v1', '2026-04-24', '2026-04-24')
            """,
            [
                (doc_id, edinet_code, f"{security_code}0", "CostOfSalesRatioCurrent", "CostOfSalesRatio", period_end, 0.6),
                (doc_id, edinet_code, f"{security_code}0", "EquityRatioCurrent", "EquityRatio", period_end, 0.5),
            ],
        )


class MetricExcelExportServiceTest(unittest.TestCase):
    def setUp(self) -> None:
        self.tmp_path = TMP_ROOT / uuid.uuid4().hex
        self.tmp_path.mkdir(parents=True, exist_ok=True)
        self.conn = sqlite3.connect(":memory:")
        self.conn.row_factory = sqlite3.Row
        _create_schema(self.conn)
        _insert_company(
            self.conn,
            edinet_code="E00001",
            security_code="1111",
            company_name="A社",
            industry_33="化学",
            net_sales_values=[100_000_000.0, 80_000_000.0, 70_000_000.0],
        )
        _insert_company(
            self.conn,
            edinet_code="E00002",
            security_code="2222",
            company_name="B社",
            industry_33="化学",
            net_sales_values=[70_000_000.0, 80_000_000.0, 100_000_000.0],
        )
        _insert_company(
            self.conn,
            edinet_code="E00003",
            security_code="3333",
            company_name="C銀行",
            industry_33="銀行業",
            net_sales_values=[100_000_000.0, 80_000_000.0, 70_000_000.0],
        )
        _insert_company(
            self.conn,
            edinet_code="E00004",
            security_code="4444",
            company_name="D証券",
            industry_33="証券、商品先物取引業",
            net_sales_values=[100_000_000.0, 80_000_000.0, 70_000_000.0],
        )
        _insert_company(
            self.conn,
            edinet_code="E00005",
            security_code="5555",
            company_name="E保険",
            industry_33="保険業",
            net_sales_values=[100_000_000.0, 80_000_000.0, 70_000_000.0],
        )
        self.conn.commit()

    def tearDown(self) -> None:
        self.conn.close()
        shutil.rmtree(self.tmp_path, ignore_errors=True)

    def test_read_metric_excel_condition_parses_all_and_period_range(self) -> None:
        path = self.tmp_path / "condition.xlsx"
        _create_condition_workbook(
            path,
            [
                ("業種", ""),
                ("証券コード", "1111,2222"),
                ("指標", "売上高,売上原価率"),
                ("期間", "3年前-前期"),
                ("増減判定", "increase"),
                ("増減判定指標", "売上高"),
            ],
        )

        condition = read_metric_excel_condition(path)

        self.assertEqual(condition.industries, [])
        self.assertEqual(condition.security_codes, ["1111", "2222"])
        self.assertEqual(condition.metric_labels, ["売上高", "売上原価率"])
        self.assertEqual(condition.period_offsets, [3, 2, 1])
        self.assertEqual(condition.trend, "increase")

    def test_read_metric_excel_condition_parses_segment_mode(self) -> None:
        path = self.tmp_path / "condition.xlsx"
        _create_condition_workbook(path, [("\u30bb\u30b0\u30e1\u30f3\u30c8", "\u5730\u57df")])

        condition = read_metric_excel_condition(path)

        self.assertEqual(condition.segment_mode, "region")

    def test_write_metric_excel_uses_decision_column_for_segments(self) -> None:
        output_path = self.tmp_path / "segment.xlsx"
        rows = [
            MetricExcelRow(
                sheet_name=GENERAL_SHEET,
                security_code="4613",
                company_name="\u95a2\u897f\u30da\u30a4\u30f3\u30c8",
                industry_33="\u5316\u5b66",
                market="Prime",
                period_scope="quarter:2Q",
                current_period_end="2025-09-30",
                metric_base="ProfitBeforeTax",
                metric_label="\u7d4c\u5e38\u5229\u76ca",
                periods_by_offset={0: "2Q 2025-09"},
                values_by_offset={0: 123.0},
                units_by_offset={0: "\u767e\u4e07\u5186"},
                ratios_by_offset={0: None},
                segment_kind="\u5730\u57df\u5225",
                segment_name="\u65e5\u672c",
            )
        ]

        write_metric_excel(
            rows=rows,
            condition=MetricExcelCondition(period_offsets=[0], segment_mode="region"),
            output_path=output_path,
            db_path=":memory:",
            errors=[],
            warnings=[],
            target_companies=1,
        )

        workbook = load_workbook(output_path)
        ws = workbook[GENERAL_SHEET]
        headers = [cell.value for cell in ws[1]]
        self.assertNotIn("\u30bb\u30b0\u30e1\u30f3\u30c8\u533a\u5206", headers)
        self.assertNotIn("\u30bb\u30b0\u30e1\u30f3\u30c8\u540d", headers)
        self.assertEqual(ws.cell(2, headers.index("\u6c7a\u7b97\u7a2e\u5225") + 1).value, "\u5730\u57df\u5225")

    def test_segment_rows_align_offsets_to_company_latest_period(self) -> None:
        long_other = (
            "Operating Segments Not Included In Reportable Segments "
            "And Other Revenue Generating Business Activities"
        )
        self.conn.executemany(
            """
            INSERT INTO segment_metrics (
                doc_id, edinet_code, security_code, form_type, period_scope, quarter_type,
                fiscal_year, period_start, period_end, segment_kind, segment_name,
                axis_qname, member_qname, metric_base, metric_key, value_kind,
                value_num, value_unit, source_tag, tag_qname, context_ref, decimals,
                calc_status, source_detail_json, rule_version, created_at, updated_at
            ) VALUES (
                ?, 'E00001', '1111', '043A00', 'quarter', '2Q',
                ?, '2025-04-01', ?, 'business', ?,
                'axis', 'member', 'NetSales', 'NetSalesCurrent', 'external',
                ?, 'yen', 'RevenuesFromExternalCustomers', 'tag', 'context', '-6',
                'ok', '{}', 'test', 'now', 'now'
            )
            """,
            [
                ("old_segment", 2025, "2024-09-30", long_other, 10_000_000.0),
                ("current_segment", 2026, "2025-09-30", "\u305d\u306e\u4ed6", 20_000_000.0),
            ],
        )
        self.conn.commit()
        condition = MetricExcelCondition(
            security_codes=["1111"],
            metric_labels=["\u58f2\u4e0a\u9ad8"],
            period_scopes=["quarter"],
            period_offsets=[1, 0],
            segment_mode="business",
        )

        rows, errors, _warnings, _preview, _target = build_metric_excel_rows(self.conn, condition)

        self.assertEqual(errors, [])
        segment_row = next(row for row in _detail_rows(rows) if row.segment_kind)
        self.assertEqual(segment_row.segment_name, "\u305d\u306e\u4ed6")
        self.assertEqual(segment_row.current_period_end, "2025-09-30")
        self.assertEqual(segment_row.periods_by_offset[0], "2Q 2025-09-30")
        self.assertEqual(segment_row.values_by_offset[0], 20.0)
        self.assertEqual(segment_row.periods_by_offset[1], "2Q 2024-09-30")
        self.assertEqual(segment_row.values_by_offset[1], 10.0)

    def test_segment_rows_sort_by_linkbase_order(self) -> None:
        self.conn.executemany(
            """
            INSERT INTO segment_metrics (
                doc_id, edinet_code, security_code, form_type, period_scope, quarter_type,
                fiscal_year, period_start, period_end, segment_kind, segment_name,
                axis_qname, member_qname, metric_base, metric_key, value_kind,
                value_num, value_unit, source_tag, tag_qname, context_ref, decimals,
                calc_status, source_detail_json, rule_version, created_at, updated_at
            ) VALUES (
                ?, 'E00001', '1111', '043A00', 'quarter', '2Q',
                2026, '2025-04-01', '2025-09-30', 'business', ?,
                'axis', ?, 'NetSales', 'NetSalesCurrent', 'external',
                ?, 'yen', 'RevenuesFromExternalCustomers', 'tag', 'context', '-6',
                'ok', ?, 'test', 'now', 'now'
            )
            """,
            [
                ("second_segment", "後順位", "member2", 20_000_000.0, '{"segment_order": 20}'),
                ("first_segment", "先順位", "member1", 10_000_000.0, '{"segment_order": 10}'),
            ],
        )
        self.conn.commit()
        condition = MetricExcelCondition(
            security_codes=["1111"],
            metric_labels=["売上高"],
            period_scopes=["quarter"],
            period_offsets=[0],
            segment_mode="business",
        )

        rows, errors, _warnings, _preview, _target = build_metric_excel_rows(self.conn, condition)

        self.assertEqual(errors, [])
        segment_rows = [row for row in _detail_rows(rows) if row.segment_kind]
        self.assertEqual([row.segment_name for row in segment_rows], ["先順位", "後順位"])

    def test_read_metric_excel_condition_defaults_to_nine_years(self) -> None:
        path = self.tmp_path / "condition.xlsx"
        _create_condition_workbook(path, [("指標", "売上高")])

        condition = read_metric_excel_condition(path)

        self.assertEqual(condition.period_offsets, [10, 9, 8, 7, 6, 5, 4, 3, 2, 1, 0])

    def test_read_metric_excel_condition_parses_percent_filter(self) -> None:
        path = self.tmp_path / "condition.xlsx"
        _create_condition_workbook(
            path,
            [
                ("％条件指標", "売上高増収率(５年)"),
                ("％下限", "75%"),
                ("％上限", "150"),
            ],
        )

        condition = read_metric_excel_condition(path)

        self.assertEqual(condition.percent_filter_metric_labels, ["売上高増収率(５年)"])
        self.assertEqual(condition.percent_filter_min, 0.75)
        self.assertEqual(condition.percent_filter_max, 1.5)

    def test_read_metric_excel_condition_keeps_excel_percent_formatted_threshold(self) -> None:
        path = self.tmp_path / "condition.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.title = "条件"
        ws.append(["％条件指標", "売上高増収率(５年)"])
        ws.append(["％下限", 5])
        ws["B2"].number_format = "0%"
        wb.save(path)

        condition = read_metric_excel_condition(path)

        self.assertEqual(condition.percent_filter_metric_labels, ["売上高増収率(５年)"])
        self.assertEqual(condition.percent_filter_min, 5.0)

    def test_read_metric_excel_condition_parses_trend_thresholds(self) -> None:
        path = self.tmp_path / "condition.xlsx"
        _create_condition_workbook(
            path,
            [
                ("増減判定", "none"),
                ("増減判定指標", "売上高増収率"),
                ("増減判定期間", "3年前-前期"),
                ("増減判定下限", "115%"),
                ("増減判定上限", "200%"),
            ],
        )

        condition = read_metric_excel_condition(path)

        self.assertEqual(condition.trend_metric_labels, ["売上高増収率"])
        self.assertEqual(condition.trend_period_offsets, [3, 2, 1])
        self.assertEqual(condition.trend_min, 1.15)
        self.assertEqual(condition.trend_max, 2.0)

    def test_read_metric_excel_condition_accepts_clear_filter_names(self) -> None:
        path = self.tmp_path / "condition.xlsx"
        _create_condition_workbook(
            path,
            [
                ("連続増減", "increase"),
                ("連続増減指標", "売上高"),
                ("連続増減期間", "前期-当期"),
                ("連続増減下限", "115%"),
                ("連続増減上限", "200%"),
                ("比率条件", "売上高増収率(５年)"),
                ("比率条件期間", "前期-当期"),
                ("比率条件下限", "500%"),
                ("比率条件上限", "1000%"),
            ],
        )

        condition = read_metric_excel_condition(path)

        self.assertEqual(condition.trend, "increase")
        self.assertEqual(condition.trend_metric_labels, ["売上高"])
        self.assertEqual(condition.trend_period_offsets, [1, 0])
        self.assertEqual(condition.trend_min, 1.15)
        self.assertEqual(condition.trend_max, 2.0)
        self.assertEqual(condition.percent_filter_metric_labels, ["売上高増収率(５年)"])
        self.assertEqual(condition.percent_filter_period_offsets, [1, 0])
        self.assertEqual(condition.percent_filter_min, 5.0)
        self.assertEqual(condition.percent_filter_max, 10.0)

    def test_read_metric_excel_condition_accepts_start_end_period_fields(self) -> None:
        path = self.tmp_path / "condition.xlsx"
        _create_condition_workbook(
            path,
            [
                ("期間\u3000Start", "2期前"),
                ("期間\u3000End", "当期"),
                ("連続増減期間\u3000Start", "前期"),
                ("連続増減期間\u3000End", "当期"),
                ("比率条件期間\u3000Start", "5期前"),
                ("比率条件期間\u3000End", "当期"),
            ],
        )

        condition = read_metric_excel_condition(path)

        self.assertEqual(condition.period_offsets, [2, 1, 0])
        self.assertEqual(condition.trend_period_offsets, [1, 0])
        self.assertEqual(condition.percent_filter_period_offsets, [5, 4, 3, 2, 1, 0])
        self.assertEqual(condition.period_scopes, ["annual", "quarter", "quarter_standalone", "forecast"])

    def test_read_metric_excel_condition_accepts_securities_industry_alias(self) -> None:
        path = self.tmp_path / "condition.xlsx"
        _create_condition_workbook(
            path,
            [
                ("業種", "証券・商品先物取引業"),
                ("指標", "売上高"),
            ],
        )

        condition = read_metric_excel_condition(path)

        self.assertEqual(condition.industries, ["証券、商品先物取引業"])

    def test_read_metric_excel_condition_keeps_official_securities_industry_name(self) -> None:
        path = self.tmp_path / "condition.xlsx"
        _create_condition_workbook(
            path,
            [
                ("業種", "証券、商品先物取引業"),
                ("指標", "売上高"),
            ],
        )

        condition = read_metric_excel_condition(path)

        self.assertEqual(condition.industries, ["証券、商品先物取引業"])

    def test_build_rows_absorbs_ratio_metrics_into_ratio_columns(self) -> None:
        path = self.tmp_path / "condition.xlsx"
        _create_condition_workbook(
            path,
            [
                ("証券コード", "1111"),
                ("指標", "売上高,売上原価率,自己資本比率"),
                ("期間", "3年前-前期"),
            ],
        )
        condition = read_metric_excel_condition(path)

        rows, errors, _warnings, _preview, target_companies = build_metric_excel_rows(
            self.conn,
            condition,
        )

        self.assertEqual(errors, [])
        self.assertEqual(target_companies, 1)
        by_metric = {row.metric_base: row for row in _detail_rows(rows)}
        self.assertEqual(set(by_metric), {"NetSales", "CostOfSales", "NetAssets"})
        self.assertEqual(by_metric["NetSales"].values_by_offset[1], 100.0)
        self.assertEqual(by_metric["NetSales"].ratios_by_offset[1], 1.0)
        self.assertEqual(by_metric["CostOfSales"].ratios_by_offset[1], 0.6)
        self.assertEqual(by_metric["NetAssets"].ratios_by_offset[1], 0.5)

    def test_trend_filter_keeps_only_continuous_increase_companies(self) -> None:
        path = self.tmp_path / "condition.xlsx"
        _create_condition_workbook(
            path,
            [
                ("業種", "化学"),
                ("指標", "売上高"),
                ("期間", "2年前-当期"),
                ("増減判定", "increase"),
                ("増減判定指標", "売上高"),
                ("増減判定期間", "3年前-前期"),
            ],
        )
        condition = read_metric_excel_condition(path)

        rows, errors, _warnings, _preview, target_companies = build_metric_excel_rows(
            self.conn,
            condition,
        )

        self.assertEqual(errors, [])
        self.assertEqual(target_companies, 1)
        self.assertEqual({row.security_code for row in _detail_rows(rows)}, {"1111"})

    def test_trend_threshold_keeps_companies_above_lower_bound_every_period(self) -> None:
        self.conn.executemany(
            """
            INSERT INTO derived_metrics (
                doc_id, edinet_code, security_code, metric_key, metric_base, metric_group,
                fiscal_year, period_end, period_scope, period_offset, consolidation,
                accounting_standard, document_display_unit, value_num, value_unit, calc_status,
                formula_name, source_detail_json, rule_version, created_at, updated_at
            ) VALUES (?, ?, ?, 'NetSalesGrowthRateCurrent',
                      'NetSalesGrowthRate', 'growth', 2025, ?, 'annual',
                      0, 'consolidated', 'Japan GAAP', '百万円', ?, 'ratio',
                      'ok', 'test', '{}', 'v1', '2026-04-24', '2026-04-24')
            """,
            [
                ("E00001_0", "E00001", "11110", "2026-03-31", 1.2),
                ("E00001_1", "E00001", "11110", "2025-03-31", 1.18),
                ("E00001_2", "E00001", "11110", "2024-03-31", 1.15),
                ("E00002_0", "E00002", "22220", "2026-03-31", 1.3),
                ("E00002_1", "E00002", "22220", "2025-03-31", 1.1),
                ("E00002_2", "E00002", "22220", "2024-03-31", 1.4),
            ],
        )
        self.conn.commit()
        path = self.tmp_path / "condition.xlsx"
        _create_condition_workbook(
            path,
            [
                ("業種", "化学"),
                ("指標", "売上高"),
                ("期間", "前期"),
                ("増減判定", "none"),
                ("増減判定指標", "売上高増収率"),
                ("増減判定期間", "3年前-前期"),
                ("増減判定下限", "115%"),
            ],
        )
        condition = read_metric_excel_condition(path)

        rows, errors, _warnings, _preview, target_companies = build_metric_excel_rows(
            self.conn,
            condition,
        )

        self.assertEqual(errors, [])
        self.assertEqual(target_companies, 1)
        self.assertEqual({row.security_code for row in _detail_rows(rows)}, {"1111"})

    def test_build_rows_uses_fixed_metric_order(self) -> None:
        path = self.tmp_path / "condition.xlsx"
        _create_condition_workbook(
            path,
            [
                ("証券コード", "1111"),
                ("指標", "純資産,売上原価率,売上高,EPS,発行株数"),
                ("期間", "前期"),
            ],
        )
        condition = read_metric_excel_condition(path)

        rows, errors, _warnings, _preview, _target_companies = build_metric_excel_rows(
            self.conn,
            condition,
        )

        self.assertEqual(errors, [])
        detail_rows = _detail_rows(rows)
        self.assertEqual(
            [row.metric_base for row in detail_rows],
            ["NetSales", "CostOfSales", "NetAssets", "OutstandingShares", "EPS"],
        )

    def test_build_rows_separates_annual_and_half_when_period_scope_all(self) -> None:
        self.conn.execute(
            """
            INSERT INTO filings (
                doc_id, edinet_code, security_code, form_type, period_end, submit_date,
                amendment_flag, doc_info_edit_status, legal_status, accounting_standard,
                document_display_unit, zip_path, xbrl_path, download_status, parse_status,
                created_at, updated_at
            ) VALUES ('E00001_HALF_0', 'E00001', '11110', '043A00',
                      '2026-09-30', '2026-11-14 12:00', 0, '0', '1',
                      'Japan GAAP', '逋ｾ荳・・', 'zip', 'xbrl',
                      'downloaded', 'derived_metrics_saved',
                      '2026-04-24', '2026-04-24')
            """
        )
        self.conn.execute(
            """
            INSERT INTO normalized_metrics (
                doc_id, edinet_code, security_code, metric_key, fiscal_year, period_end,
                value_num, source_tag, consolidation, rule_version, created_at, updated_at
            ) VALUES ('E00001_HALF_0', 'E00001', '11110', 'NetSalesCurrent',
                      2026, '2026-09-30', 55.0, 'tag',
                      'consolidated', 'v1', '2026-04-24', '2026-04-24')
            """
        )
        self.conn.commit()
        path = self.tmp_path / "condition.xlsx"
        _create_condition_workbook(
            path,
            [
                ("\u8a3c\u5238\u30b3\u30fc\u30c9", "1111"),
                ("\u6307\u6a19", "\u58f2\u4e0a\u9ad8"),
                ("\u671f\u9593", "1\u5e74\u524d-\u5f53\u671f"),
                ("\u6c7a\u7b97\u7a2e\u5225", "ALL"),
            ],
        )
        condition = read_metric_excel_condition(path)

        rows, errors, _warnings, _preview, target_companies = build_metric_excel_rows(
            self.conn,
            condition,
        )

        detail_rows = _detail_rows(rows)
        self.assertEqual(errors, [])
        self.assertEqual(target_companies, 1)
        self.assertEqual([row.period_scope for row in detail_rows], ["quarter:2Q", "annual"])
        self.assertEqual(
            [row.metric_label for row in detail_rows],
            ["2Q \u58f2\u4e0a\u9ad8", "4Q \u58f2\u4e0a\u9ad8"],
        )
        self.assertEqual(
            [row.periods_by_offset[0] for row in detail_rows],
            ["2Q 2026-09", ""],
        )
        self.assertEqual(
            [row.periods_by_offset[1] for row in detail_rows],
            ["", "\u901a\u671f 2026-03"],
        )
        self.assertEqual([row.values_by_offset[0] for row in detail_rows], [55.0, None])
        self.assertEqual([row.values_by_offset[1] for row in detail_rows], [None, 100.0])

    def test_build_rows_buckets_half_with_corresponding_annual_period(self) -> None:
        self.conn.executemany(
            """
            INSERT INTO filings (
                doc_id, edinet_code, security_code, form_type, period_end, submit_date,
                amendment_flag, doc_info_edit_status, legal_status, accounting_standard,
                document_display_unit, zip_path, xbrl_path, download_status, parse_status,
                created_at, updated_at
            ) VALUES (?, 'E00001', '11110', '043A00', ?, ?, 0, '0', '1',
                      'Japan GAAP', '逋ｾ荳・・', 'zip', 'xbrl',
                      'downloaded', 'derived_metrics_saved',
                      '2026-04-24', '2026-04-24')
            """,
            [
                ("E00001_HALF_0", "2025-09-30", "2025-11-14 12:00"),
                ("E00001_HALF_1", "2024-09-30", "2024-11-14 12:00"),
            ],
        )
        self.conn.executemany(
            """
            INSERT INTO normalized_metrics (
                doc_id, edinet_code, security_code, metric_key, fiscal_year, period_end,
                value_num, source_tag, consolidation, rule_version, created_at, updated_at
            ) VALUES (?, 'E00001', '11110', 'NetSalesCurrent',
                      ?, ?, ?, 'tag', 'consolidated', 'v1', '2026-04-24', '2026-04-24')
            """,
            [
                ("E00001_HALF_0", 2025, "2025-09-30", 55.0),
                ("E00001_HALF_1", 2024, "2024-09-30", 44.0),
            ],
        )
        self.conn.commit()
        path = self.tmp_path / "condition.xlsx"
        _create_condition_workbook(
            path,
            [
                ("\u8a3c\u5238\u30b3\u30fc\u30c9", "1111"),
                ("\u6307\u6a19", "\u58f2\u4e0a\u9ad8"),
                ("\u671f\u9593", "\u524d\u671f-\u5f53\u671f"),
                ("\u6c7a\u7b97\u7a2e\u5225", "ALL"),
            ],
        )
        condition = read_metric_excel_condition(path)

        rows, errors, _warnings, _preview, target_companies = build_metric_excel_rows(
            self.conn,
            condition,
        )

        detail_rows = _detail_rows(rows)
        self.assertEqual(errors, [])
        self.assertEqual(target_companies, 1)
        self.assertEqual([row.period_scope for row in detail_rows], ["quarter:2Q", "annual"])
        self.assertEqual(
            [row.periods_by_offset[0] for row in detail_rows],
            ["2Q 2025-09", ""],
        )
        self.assertEqual(
            [row.periods_by_offset[1] for row in detail_rows],
            ["2Q 2024-09", "\u901a\u671f 2026-03"],
        )
        self.assertEqual([row.values_by_offset[0] for row in detail_rows], [55.0, None])
        self.assertEqual([row.values_by_offset[1] for row in detail_rows], [44.0, 100.0])

    def test_build_rows_uses_industry_specific_labels_and_order(self) -> None:
        cases = [
            (
                "銀行業",
                "3333",
                ["売上高", "費用合計", "├資金調達費用", "└営業経費", "役務取引等収益"],
            ),
            (
                "証券、商品先物取引業",
                "4444",
                ["売上高", "純収益", "費用合計", "├金融費用", "└販管費", "　├ 一般管理費", "　└ 販売費"],
            ),
            (
                "保険業",
                "5555",
                ["売上高", "売上総利益", "費用合計", "├保険金等支払金", "├責任準備金等繰入額", "├資産運用費用", "└ 事業費"],
            ),
        ]
        for industry, security_code, expected_labels in cases:
            with self.subTest(industry=industry):
                path = self.tmp_path / f"{security_code}.xlsx"
                _create_condition_workbook(
                    path,
                    [
                        ("業種", industry),
                        ("証券コード", security_code),
                        ("指標", "ALL"),
                        ("期間", "当期"),
                    ],
                )
                condition = read_metric_excel_condition(path)

                rows, errors, _warnings, _preview, _target_companies = build_metric_excel_rows(
                    self.conn,
                    condition,
                )

                self.assertEqual(errors, [])
                detail_rows = _detail_rows(rows)
                self.assertEqual(
                    [row.metric_label for row in detail_rows[: len(expected_labels)]],
                    [f"4Q {label}" for label in expected_labels],
                )

    def test_cash_balance_metric_label_is_registered(self) -> None:
        path = self.tmp_path / "condition.xlsx"
        _create_condition_workbook(
            path,
            [
                ("証券コード", "1111"),
                ("指標", "期末残"),
                ("期間", "前期"),
                ("％条件期間", "前期"),
            ],
        )
        condition = read_metric_excel_condition(path)

        rows, errors, _warnings, _preview, _target_companies = build_metric_excel_rows(
            self.conn,
            condition,
        )

        self.assertEqual(errors, [])
        detail_rows = _detail_rows(rows)
        self.assertEqual([row.metric_base for row in detail_rows], ["CashAndCashEquivalents"])
        self.assertEqual(detail_rows[0].metric_label, "4Q 期末残")

    def test_beginning_cash_balance_metric_label_is_registered(self) -> None:
        path = self.tmp_path / "condition.xlsx"
        _create_condition_workbook(
            path,
            [
                ("\u8a3c\u5238\u30b3\u30fc\u30c9", "1111"),
                ("\u6307\u6a19", "\u671f\u9996\u6b8b\u9ad8"),
                ("\u671f\u9593", "\u524d\u671f"),
            ],
        )
        condition = read_metric_excel_condition(path)

        rows, errors, _warnings, _preview, _target_companies = build_metric_excel_rows(
            self.conn,
            condition,
        )

        self.assertEqual(errors, [])
        detail_rows = _detail_rows(rows)
        self.assertEqual([row.metric_base for row in detail_rows], ["BeginningCashBalance"])
        self.assertEqual(detail_rows[0].metric_label, "4Q \u671f\u9996\u6b8b\u9ad8")

    def test_eps_growth_metric_label_uses_prior_period_suffix(self) -> None:
        path = self.tmp_path / "condition.xlsx"
        _create_condition_workbook(
            path,
            [
                ("証券コード", "1111"),
                ("指標", "EPS増加率"),
                ("期間", "前期"),
            ],
        )
        condition = read_metric_excel_condition(path)

        rows, errors, _warnings, _preview, _target_companies = build_metric_excel_rows(
            self.conn,
            condition,
        )

        self.assertEqual(errors, [])
        detail_rows = _detail_rows(rows)
        self.assertEqual([row.metric_base for row in detail_rows], ["EPSGrowthRate"])
        self.assertEqual(detail_rows[0].metric_label, "4Q EPS増加率")

    def test_export_metric_excel_writes_percent_format(self) -> None:
        condition_path = self.tmp_path / "condition.xlsx"
        output_path = self.tmp_path / "export.xlsx"
        _create_condition_workbook(
            condition_path,
            [
                ("証券コード", "1111"),
                ("指標", "売上原価率"),
                ("期間", "前期"),
            ],
        )

        result = export_metric_excel(
            self.conn,
            condition_xlsx=condition_path,
            output_path=output_path,
            db_path=":memory:",
        )

        self.assertEqual(result.errors, [])
        workbook = load_workbook(output_path)
        ws = workbook[GENERAL_SHEET]
        self.assertEqual(ws["E2"].value, "Prime")
        self.assertEqual(ws["F2"].value, "通期")
        self.assertEqual(ws["G2"].value, "明細")
        self.assertEqual(ws["I2"].value, "4Q ├売上原価")
        self.assertEqual(ws["J2"].value, "通期 2026-03")
        self.assertEqual(ws["K2"].value, 60.0)
        self.assertEqual(ws["L2"].value, "百万円")
        self.assertEqual(ws["M2"].value, 0.6)
        self.assertEqual(ws["M2"].number_format, "0.0%")

    def test_export_metric_excel_formats_growth_rates_as_percent(self) -> None:
        self.conn.execute(
            """
            INSERT INTO derived_metrics (
                doc_id, edinet_code, security_code, metric_key, metric_base, metric_group,
                fiscal_year, period_end, period_scope, period_offset, consolidation,
                accounting_standard, document_display_unit, value_num, value_unit, calc_status,
                formula_name, source_detail_json, rule_version, created_at, updated_at
            ) VALUES (
                'E00001_0', 'E00001', '11110', 'NetSalesGrowthRateCurrent',
                'NetSalesGrowthRate', 'growth', 2025, '2026-03-31', 'annual',
                0, 'consolidated', 'Japan GAAP', '百万円', 1.25, 'ratio',
                'ok', 'test', '{}', 'v1', '2026-04-24', '2026-04-24'
            )
            """
        )
        self.conn.commit()
        condition_path = self.tmp_path / "condition.xlsx"
        output_path = self.tmp_path / "growth.xlsx"
        _create_condition_workbook(
            condition_path,
            [
                ("証券コード", "1111"),
                ("指標", "売上高増収率"),
                ("期間", "前期"),
            ],
        )

        result = export_metric_excel(
            self.conn,
            condition_xlsx=condition_path,
            output_path=output_path,
            db_path=":memory:",
        )

        self.assertEqual(result.errors, [])
        workbook = load_workbook(output_path)
        ws = workbook[GENERAL_SHEET]
        self.assertEqual(ws["I2"].value, "4Q 売上高増収率(前期比)")
        self.assertEqual(ws["J2"].value, "通期 2026-03")
        self.assertEqual(ws["K2"].value, 1.25)
        self.assertEqual(ws["L2"].value, "%")
        self.assertEqual(ws["K2"].number_format, "0.0%")

    def test_percent_filter_keeps_matching_companies(self) -> None:
        self.conn.executemany(
            """
            INSERT INTO derived_metrics (
                doc_id, edinet_code, security_code, metric_key, metric_base, metric_group,
                fiscal_year, period_end, period_scope, period_offset, consolidation,
                accounting_standard, document_display_unit, value_num, value_unit, calc_status,
                formula_name, source_detail_json, rule_version, created_at, updated_at
            ) VALUES (?, ?, ?, 'NetSalesGrowthRate5YearCurrent',
                      'NetSalesGrowthRate5Year', 'growth', 2025, ?, 'annual',
                      0, 'consolidated', 'Japan GAAP', '百万円', ?, 'ratio',
                      'ok', 'test', '{}', 'v1', '2026-04-24', '2026-04-24')
            """,
            [
                ("E00001_0", "E00001", "11110", "2026-03-31", 0.8),
                ("E00002_0", "E00002", "22220", "2026-03-31", 0.7),
            ],
        )
        self.conn.commit()
        path = self.tmp_path / "condition.xlsx"
        _create_condition_workbook(
            path,
            [
                ("業種", "化学"),
                ("指標", "売上高"),
                ("期間", "前期"),
                ("％条件指標", "売上高増収率(５年)"),
                ("％条件期間", "前期"),
                ("％下限", "75%"),
            ],
        )
        condition = read_metric_excel_condition(path)

        rows, errors, _warnings, _preview, target_companies = build_metric_excel_rows(
            self.conn,
            condition,
        )

        self.assertEqual(errors, [])
        self.assertEqual(target_companies, 1)
        self.assertEqual({row.security_code for row in _detail_rows(rows)}, {"1111"})

    def test_percent_filter_period_requires_each_selected_period(self) -> None:
        self.conn.executemany(
            """
            INSERT INTO derived_metrics (
                doc_id, edinet_code, security_code, metric_key, metric_base, metric_group,
                fiscal_year, period_end, period_scope, period_offset, consolidation,
                accounting_standard, document_display_unit, value_num, value_unit, calc_status,
                formula_name, source_detail_json, rule_version, created_at, updated_at
            ) VALUES (?, ?, ?, 'NetSalesGrowthRateCurrent',
                      'NetSalesGrowthRate', 'growth', 2025, ?, 'annual',
                      0, 'consolidated', 'Japan GAAP', '逋ｾ荳・・', ?, 'ratio',
                      'ok', 'test', '{}', 'v1', '2026-04-24', '2026-04-24')
            """,
            [
                ("E00001_0", "E00001", "11110", "2026-03-31", 1.2),
                ("E00001_1", "E00001", "11110", "2025-03-31", 1.1),
                ("E00002_0", "E00002", "22220", "2026-03-31", 1.2),
                ("E00002_1", "E00002", "22220", "2025-03-31", 0.9),
            ],
        )
        self.conn.commit()
        path = self.tmp_path / "condition.xlsx"
        _create_condition_workbook(
            path,
            [
                ("業種", "化学"),
                ("指標", "売上高"),
                ("期間", "2年前-前期"),
                ("比率条件", "売上高増収率"),
                ("比率条件期間", "2年前-前期"),
                ("比率条件下限", "100%"),
            ],
        )
        condition = read_metric_excel_condition(path)

        rows, errors, _warnings, _preview, target_companies = build_metric_excel_rows(
            self.conn,
            condition,
        )

        self.assertEqual(errors, [])
        self.assertEqual(target_companies, 1)
        self.assertEqual({row.security_code for row in _detail_rows(rows)}, {"1111"})

    def test_industry_only_aggregates_additive_metrics_in_oku_yen(self) -> None:
        _insert_industry_aggregate_metric(
            self.conn,
            industry_33="化学",
            fiscal_year=2026,
            metric_base="NetSales",
            value_num=170_000_000.0,
        )
        _insert_industry_aggregate_metric(
            self.conn,
            industry_33="化学",
            fiscal_year=2026,
            metric_base="CostOfSales",
            value_num=120_000_000.0,
        )
        _insert_industry_aggregate_metric(
            self.conn,
            industry_33="化学",
            fiscal_year=2026,
            metric_base="CostOfSalesRatio",
            value_num=120_000_000 / 170_000_000,
            value_unit="ratio",
        )
        path = self.tmp_path / "condition.xlsx"
        _create_condition_workbook(
            path,
            [
                ("\u696d\u7a2e", "\u696d\u7a2e\u306e\u307f"),
                ("\u6307\u6a19", "\u58f2\u4e0a\u9ad8,\u58f2\u4e0a\u539f\u4fa1"),
                ("\u671f\u9593", "\u524d\u671f"),
                ("\u6c7a\u7b97\u7a2e\u5225", "\u901a\u671f"),
            ],
        )
        condition = read_metric_excel_condition(path)

        rows, errors, _warnings, _preview, target_companies = build_metric_excel_rows(
            self.conn,
            condition,
        )

        self.assertEqual(errors, [])
        self.assertTrue(condition.industry_only)
        self.assertGreaterEqual(target_companies, 1)
        detail_rows = _detail_rows(rows)
        net_sales_row = next(
            row
            for row in detail_rows
            if row.metric_base == "NetSales" and row.values_by_offset[1] == 1.7
        )
        by_base = {
            row.metric_base: row
            for row in detail_rows
            if row.industry_33 == net_sales_row.industry_33
        }
        self.assertEqual(set(by_base), {"NetSales", "CostOfSales"})
        self.assertEqual(by_base["NetSales"].security_code, "")
        self.assertEqual(by_base["NetSales"].company_name, "")
        self.assertAlmostEqual(by_base["NetSales"].values_by_offset[1], 1.7)
        self.assertEqual(by_base["NetSales"].units_by_offset[1], "\u5104\u5186")
        self.assertAlmostEqual(by_base["CostOfSales"].values_by_offset[1], 1.2)
        self.assertAlmostEqual(by_base["CostOfSales"].ratios_by_offset[1], 120_000_000 / 170_000_000)

    def test_build_rows_adds_average_median_and_rank_rows(self) -> None:
        path = self.tmp_path / "condition.xlsx"
        _create_condition_workbook(
            path,
            [
                ("\u8a3c\u5238\u30b3\u30fc\u30c9", "1111,2222"),
                ("\u6307\u6a19", "\u58f2\u4e0a\u9ad8"),
                ("\u671f\u9593", "\u524d\u671f"),
                ("\u6c7a\u7b97\u7a2e\u5225", "\u901a\u671f"),
            ],
        )
        condition = read_metric_excel_condition(path)

        rows, errors, _warnings, _preview, target_companies = build_metric_excel_rows(
            self.conn,
            condition,
        )

        self.assertEqual(errors, [])
        self.assertEqual(target_companies, 2)
        self.assertEqual([row.row_kind for row in rows], [ROW_KIND_DETAIL, ROW_KIND_DETAIL, ROW_KIND_AVERAGE, ROW_KIND_MEDIAN])
        self.assertEqual([row.ranks_by_offset.get(1) for row in rows[:2]], ["1/2", "2/2"])
        average_row = next(row for row in rows if row.row_kind == ROW_KIND_AVERAGE)
        median_row = next(row for row in rows if row.row_kind == ROW_KIND_MEDIAN)
        self.assertEqual(average_row.metric_label, "4Q \u58f2\u4e0a\u9ad8")
        self.assertEqual(median_row.metric_label, "4Q \u58f2\u4e0a\u9ad8")
        self.assertAlmostEqual(average_row.values_by_offset[1], 85.0)
        self.assertAlmostEqual(median_row.values_by_offset[1], 85.0)
        self.assertEqual(average_row.units_by_offset[1], "\u767e\u4e07\u5186")

    def test_export_metric_excel_writes_row_kind_rank_and_vertical_data(self) -> None:
        condition_path = self.tmp_path / "condition.xlsx"
        output_path = self.tmp_path / "rank.xlsx"
        _create_condition_workbook(
            condition_path,
            [
                ("\u8a3c\u5238\u30b3\u30fc\u30c9", "1111,2222"),
                ("\u6307\u6a19", "\u58f2\u4e0a\u9ad8"),
                ("\u671f\u9593", "\u524d\u671f"),
                ("\u6c7a\u7b97\u7a2e\u5225", "\u901a\u671f"),
            ],
        )

        result = export_metric_excel(
            self.conn,
            condition_xlsx=condition_path,
            output_path=output_path,
            db_path=":memory:",
        )

        self.assertEqual(result.errors, [])
        workbook = load_workbook(output_path)
        ws = workbook[GENERAL_SHEET]
        self.assertEqual(ws["C1"].value, "\u30c6\u30f3\u30d0\u30ac\u30fc")
        self.assertEqual(ws["G1"].value, "\u884c\u7a2e\u5225")
        self.assertEqual(ws["N1"].value, "\u524d\u671f_\u9806\u4f4d")
        self.assertEqual(ws["G2"].value, ROW_KIND_DETAIL)
        self.assertEqual(ws["N2"].value, "1/2")
        self.assertEqual(ws["J1"].fill.fgColor.rgb, "00EAF4FF")
        self.assertEqual(ws["J2"].border.left.style, "thin")
        vertical = workbook[VERTICAL_DATA_SHEET]
        self.assertEqual(vertical["C1"].value, "\u30c6\u30f3\u30d0\u30ac\u30fc")
        self.assertEqual(vertical["G1"].value, "\u884c\u7a2e\u5225")
        self.assertEqual(vertical["M1"].value, "\u9806\u4f4d")
        self.assertEqual(vertical["G2"].value, ROW_KIND_DETAIL)
        self.assertEqual(vertical["M2"].value, "1/2")

    def test_export_metric_excel_marks_tenbagger_learning_security(self) -> None:
        _insert_company(
            self.conn,
            edinet_code="E06920",
            security_code="6920",
            company_name="\u30ec\u30fc\u30b6\u30fc\u30c6\u30c3\u30af",
            industry_33="\u96fb\u6c17\u6a5f\u5668",
            net_sales_values=[100_000_000.0, 90_000_000.0, 80_000_000.0],
        )
        self.conn.commit()
        condition_path = self.tmp_path / "condition.xlsx"
        output_path = self.tmp_path / "tenbagger.xlsx"
        _create_condition_workbook(
            condition_path,
            [
                ("\u8a3c\u5238\u30b3\u30fc\u30c9", "6920"),
                ("\u6307\u6a19", "\u58f2\u4e0a\u9ad8"),
                ("\u671f\u9593", "\u524d\u671f"),
                ("\u6c7a\u7b97\u7a2e\u5225", "\u901a\u671f"),
            ],
        )

        result = export_metric_excel(
            self.conn,
            condition_xlsx=condition_path,
            output_path=output_path,
            db_path=":memory:",
        )

        self.assertEqual(result.errors, [])
        workbook = load_workbook(output_path)
        ws = workbook[GENERAL_SHEET]
        self.assertEqual(ws["A2"].value, "6920")
        self.assertEqual(ws["C2"].value, "\u3007")

    def test_half_progress_metrics_are_suppressed_in_excel_rows(self) -> None:
        path = self.tmp_path / "condition.xlsx"
        _create_condition_workbook(
            path,
            [
                ("\u8a3c\u5238\u30b3\u30fc\u30c9", "1111"),
                ("\u6307\u6a19", "HalfNetSalesProgressRate"),
                ("\u671f\u9593", "\u524d\u671f"),
                ("\u6c7a\u7b97\u7a2e\u5225", "ALL"),
            ],
        )
        condition = read_metric_excel_condition(path)

        rows, errors, warnings, _preview, target_companies = build_metric_excel_rows(
            self.conn,
            condition,
        )

        self.assertEqual(errors, [])
        self.assertEqual(rows, [])
        self.assertEqual(target_companies, 0)
        self.assertTrue(any("excel_suppressed_metrics" in warning for warning in warnings))

    def test_industry_only_forces_annual_calendar_year_and_adds_summary_note(self) -> None:
        _insert_industry_aggregate_metric(
            self.conn,
            industry_33="化学",
            fiscal_year=2026,
            metric_base="NetSales",
            value_num=170_000_000.0,
        )
        self.conn.execute(
            """
            INSERT INTO filings (
                doc_id, edinet_code, security_code, form_type, period_end, submit_date,
                amendment_flag, doc_info_edit_status, legal_status, accounting_standard,
                document_display_unit, zip_path, xbrl_path, download_status, parse_status,
                created_at, updated_at
            ) VALUES ('E00001_HALF_0', 'E00001', '11110', '043A00',
                      '2026-09-30', '2026-11-14 12:00', 0, '0', '1',
                      'Japan GAAP', '逋ｾ荳・・', 'zip', 'xbrl',
                      'downloaded', 'derived_metrics_saved',
                      '2026-04-24', '2026-04-24')
            """
        )
        self.conn.execute(
            """
            INSERT INTO normalized_metrics (
                doc_id, edinet_code, security_code, metric_key, fiscal_year, period_end,
                value_num, source_tag, consolidation, rule_version, created_at, updated_at
            ) VALUES ('E00001_HALF_0', 'E00001', '11110', 'NetSalesCurrent',
                      2026, '2026-09-30', 55_000_000.0, 'tag',
                      'consolidated', 'v1', '2026-04-24', '2026-04-24')
            """
        )
        self.conn.commit()
        condition_path = self.tmp_path / "industry_condition.xlsx"
        output_path = self.tmp_path / "industry.xlsx"
        _create_condition_workbook(
            condition_path,
            [
                ("\u696d\u7a2e", "\u696d\u7a2e\u306e\u307f"),
                ("\u6307\u6a19", "\u58f2\u4e0a\u9ad8"),
                ("\u671f\u9593", "\u524d\u671f"),
                ("\u6c7a\u7b97\u7a2e\u5225", "ALL"),
            ],
        )
        condition = read_metric_excel_condition(condition_path)

        rows, errors, warnings, _preview, _target = build_metric_excel_rows(
            self.conn,
            condition,
        )

        detail_rows = _detail_rows(rows)
        self.assertEqual(errors, [])
        self.assertTrue(all(row.period_scope == "annual" for row in detail_rows))
        self.assertTrue(all(row.periods_by_offset[1].startswith("\u901a\u671f 2026") for row in detail_rows))
        self.assertTrue(any(warning == "industry_only_mode_forced_to_annual" for warning in warnings))
        self.assertTrue(any(row.ranks_by_offset.get(1) for row in detail_rows))

        result = export_metric_excel(
            self.conn,
            condition_xlsx=condition_path,
            output_path=output_path,
            db_path=":memory:",
        )
        self.assertEqual(result.errors, [])
        workbook = load_workbook(output_path)
        summary = workbook[SUMMARY_SHEET]
        summary_values = {summary.cell(row=i, column=1).value: summary.cell(row=i, column=2).value for i in range(1, summary.max_row + 1)}
        self.assertIn("\u96c6\u8a08\u57fa\u6e96:", summary_values["aggregation_basis_note"])

    def test_export_metric_excel_formats_per_share_and_theoretical_values(self) -> None:
        self.conn.executemany(
            """
            INSERT INTO derived_metrics (
                doc_id, edinet_code, security_code, metric_key, metric_base, metric_group,
                fiscal_year, period_end, period_scope, period_offset, consolidation,
                accounting_standard, document_display_unit, value_num, value_unit, calc_status,
                formula_name, source_detail_json, rule_version, created_at, updated_at
            ) VALUES (
                'E00001_0', 'E00001', '11110', ?, ?, 'valuation', 2025,
                '2026-03-31', 'annual', 0, 'consolidated', 'Japan GAAP',
                '百万円', ?, 'number', 'ok', 'test', '{}', 'v1',
                '2026-04-24', '2026-04-24'
            )
            """,
            [
                ("EPSCurrent", "EPS", 123.456),
                ("BPSCurrent", "BPS", 234.567),
                ("TheoreticalPBRCurrent", "TheoreticalPBR", 1.234),
                ("AssetsPerShareCurrent", "AssetsPerShare", 3456.789),
                ("TheoreticalSharePriceCurrent", "TheoreticalSharePrice", 4567.89),
            ],
        )
        self.conn.commit()
        condition_path = self.tmp_path / "condition.xlsx"
        output_path = self.tmp_path / "formats.xlsx"
        _create_condition_workbook(
            condition_path,
            [
                ("証券コード", "1111"),
                ("指標", "EPS,BPS,理論PBR,1株資産,理論株価"),
                ("期間", "前期"),
            ],
        )

        result = export_metric_excel(
            self.conn,
            condition_xlsx=condition_path,
            output_path=output_path,
            db_path=":memory:",
        )

        self.assertEqual(result.errors, [])
        workbook = load_workbook(output_path)
        ws = workbook[GENERAL_SHEET]
        formats_by_label = {
            ws.cell(row=row_index, column=9).value: ws.cell(row=row_index, column=11).number_format
            for row_index in range(2, ws.max_row + 1)
        }
        self.assertEqual(formats_by_label["4Q EPS"], "#,##0.0")
        self.assertEqual(formats_by_label["4Q BPS"], "#,##0.0")
        self.assertEqual(formats_by_label["4Q 理論PBR"], "#,##0.0")
        self.assertEqual(formats_by_label["4Q 1株資産"], "#,##0")
        self.assertEqual(formats_by_label["4Q 理論株価"], "#,##0")

    def test_market_derived_metrics_are_loaded_with_date_point_period(self) -> None:
        self.conn.executemany(
            """
            INSERT INTO market_derived_metrics (
                source_type, source_id, edinet_code, security_code, period_scope,
                period_key, quarter_type, fiscal_year, period_end, metric_key,
                metric_base, metric_group, value_num, value_unit, calc_status,
                formula_name, source_detail_json, rule_version, created_at, updated_at
            ) VALUES (
                'edinet', 'E00001_0', 'E00001', '1111', 'annual',
                'annual:FY', NULL, 2026, '2026-03-31', ?,
                ?, 'market', ?, 'ratio', 'ok',
                'fixture', '{}', 'v1', '2026-05-08', '2026-05-08'
            )
            """,
            [
                ("StockPriceCurrent", "StockPrice", 1200.25),
                ("PBRCurrent", "PBR", 1.2),
                ("PCFRCurrent", "PCFR", 6.0),
            ],
        )
        self.conn.commit()
        condition = MetricExcelCondition(
            security_codes=["1111"],
            metric_labels=["StockPrice", "PBR", "PCFR"],
            period_scopes=["annual"],
            period_offsets=[1],
        )

        rows, errors, warnings, _preview, _target = build_metric_excel_rows(self.conn, condition)

        self.assertEqual(errors, [])
        self.assertNotIn("market_derived_metrics_not_ready", warnings)
        by_base = {row.metric_base: row for row in _detail_rows(rows)}
        self.assertEqual(by_base["StockPrice"].values_by_offset[1], 1200.25)
        self.assertEqual(by_base["StockPrice"].units_by_offset[1], "\u5186")
        self.assertEqual(by_base["PBR"].values_by_offset[1], 1.2)
        self.assertEqual(by_base["PCFR"].values_by_offset[1], 6.0)
        self.assertEqual(by_base["StockPrice"].periods_by_offset[1], "\u901a\u671f 2026-03-31\u6642\u70b9")

    def test_jquants_market_derived_metrics_are_loaded_for_quarter(self) -> None:
        self.conn.execute(
            """
            INSERT INTO market_derived_metrics (
                source_type, source_id, edinet_code, security_code, period_scope,
                period_key, quarter_type, fiscal_year, period_end, metric_key,
                metric_base, metric_group, value_num, value_unit, calc_status,
                formula_name, source_detail_json, rule_version, created_at, updated_at
            ) VALUES (
                'jquants', 'DISC1', 'E00001', '1111', 'quarter',
                'actual:1Q', '1Q', 2026, '2026-06-30', 'StockPriceCurrent',
                'StockPrice', 'market', 900.5, 'yen_per_share', 'ok',
                'fixture', '{}', 'v1', '2026-05-08', '2026-05-08'
            )
            """
        )
        self.conn.commit()
        condition = MetricExcelCondition(
            security_codes=["1111"],
            metric_labels=["StockPrice"],
            period_scopes=["quarter"],
            period_offsets=[0],
        )

        rows, errors, warnings, _preview, target = build_metric_excel_rows(self.conn, condition)

        self.assertEqual(errors, [])
        self.assertEqual(target, 1)
        self.assertNotIn("jquants_metrics_not_found", warnings)
        row = next(row for row in _detail_rows(rows) if row.period_scope == "quarter:1Q")
        self.assertEqual(row.metric_label, "1Q \u682a\u4fa1")
        self.assertEqual(row.values_by_offset[0], 900.5)
        self.assertEqual(row.periods_by_offset[0], "1Q 2026-06-30\u6642\u70b9")

    def test_jquants_quarter_rows_use_forecast_progress_ratio(self) -> None:
        self.conn.executemany(
            """
            INSERT INTO jquants_financial_metrics (
                disclosure_number, local_code, security_code, edinet_code, metric_kind,
                period_scope, period_key, quarter_type, forecast_target, fiscal_year,
                period_start, period_end, disclosed_date, disclosed_time, metric_key,
                metric_base, metric_group, value_num, value_unit, calc_status,
                source_field, source_detail_json, rule_version, created_at, updated_at
            ) VALUES (?, '11110', '1111', 'E00001', ?, ?, ?, ?, ?, 2026,
                      '2025-04-01', '2026-03-31', ?, '15:00', 'NetSalesCurrent',
                      'NetSales', 'sales', ?, 'yen', 'ok', 'field', '{}', 'v1',
                      '2026-05-06', '2026-05-06')
            """,
            [
                ("forecast1", "forecast", "forecast", "forecast:FY", None, "FY", "2026-05-01", 400_000_000.0),
                ("actual1", "actual", "quarter", "actual:1Q", "1Q", None, "2026-05-02", 100_000_000.0),
            ],
        )
        self.conn.commit()
        path = self.tmp_path / "condition.xlsx"
        output_path = self.tmp_path / "jquants.xlsx"
        _create_condition_workbook(
            path,
            [
                ("\u8a3c\u5238\u30b3\u30fc\u30c9", "1111"),
                ("\u6307\u6a19", "NetSales"),
                ("\u671f\u9593", "\u5f53\u671f"),
                ("\u6c7a\u7b97\u7a2e\u5225", "\u56db\u534a\u671f"),
            ],
        )
        condition = read_metric_excel_condition(path)

        rows, errors, warnings, _preview, target = build_metric_excel_rows(self.conn, condition)

        self.assertEqual(errors, [])
        self.assertEqual(target, 1)
        quarter_row = next(row for row in _detail_rows(rows) if row.period_scope == "quarter:1Q")
        self.assertEqual(quarter_row.metric_label, "1Q \u58f2\u4e0a\u9ad8")
        self.assertEqual(quarter_row.values_by_offset[0], 100.0)
        self.assertEqual(quarter_row.units_by_offset[0], "\u767e\u4e07\u5186")
        self.assertAlmostEqual(quarter_row.ratios_by_offset[0], 0.25)
        self.assertEqual(quarter_row.ratio_kinds_by_offset[0], "forecast_progress")
        self.assertIn("quarter_ratio_cells_show_latest_forecast_progress", warnings)

        result = export_metric_excel(
            self.conn,
            condition_xlsx=path,
            output_path=output_path,
            db_path=":memory:",
        )
        self.assertEqual(result.errors, [])
        workbook = load_workbook(output_path)
        ws = workbook[GENERAL_SHEET]
        self.assertEqual(ws["M2"].value, 0.25)
        self.assertIsNotNone(ws["M2"].comment)
        self.assertEqual(ws["M2"].fill.fgColor.rgb, "00FFF2CC")

    def test_jquants_quarter_bps_falls_back_to_net_assets_per_share(self) -> None:
        self.conn.executemany(
            """
            INSERT INTO jquants_financial_metrics (
                disclosure_number, local_code, security_code, edinet_code, metric_kind,
                period_scope, period_key, quarter_type, forecast_target, fiscal_year,
                period_start, period_end, disclosed_date, disclosed_time, metric_key,
                metric_base, metric_group, value_num, value_unit, calc_status,
                source_field, source_detail_json, rule_version, created_at, updated_at
            ) VALUES (?, '11110', '1111', 'E00001', 'actual',
                      'quarter', 'actual:2Q', '2Q', NULL, 2026,
                      '2025-04-01', '2025-09-30', '2025-11-01', '15:00', ?,
                      ?, ?, ?, ?, ?, 'field', '{}', 'v1',
                      '2026-05-06', '2026-05-06')
            """,
            [
                ("bps_missing", "BPSCurrent", "BPS", "per_share", None, "yen", "missing"),
                ("net_assets", "NetAssetsCurrent", "NetAssets", "balance", 500_000_000.0, "yen", "ok"),
                ("shares", "OutstandingSharesCurrent", "OutstandingShares", "per_share", 100_000_000.0, "shares", "ok"),
            ],
        )
        self.conn.commit()
        condition = MetricExcelCondition(
            security_codes=["1111"],
            metric_labels=["BPS"],
            period_scopes=["quarter"],
            period_offsets=[0],
        )

        rows, errors, warnings, _preview, target = build_metric_excel_rows(self.conn, condition)

        self.assertEqual(errors, [])
        self.assertEqual(target, 1)
        self.assertNotIn("jquants_metrics_not_found", warnings)
        row = next(row for row in _detail_rows(rows) if row.period_scope == "quarter:2Q")
        self.assertEqual(row.metric_label, "2Q BPS")
        self.assertEqual(row.values_by_offset[0], 5.0)
        self.assertEqual(row.units_by_offset[0], "円")

    def test_quarter_standalone_rows_are_loaded(self) -> None:
        self.conn.execute(
            """
            INSERT INTO quarter_standalone_metrics (
                security_code, edinet_code, fiscal_year, quarter_type, period_end,
                metric_key, metric_base, metric_group, value_num, value_unit,
                calc_status, formula_name, source_detail_json, rule_version,
                created_at, updated_at
            )
            VALUES ('1111', 'E00001', 2026, '2Q', '2025-09-30',
                    'NetSalesCurrent', 'NetSales', 'sales', 120000000.0, 'yen',
                    'ok', 'fixture', '{}', 'test', 'now', 'now')
            """
        )
        self.conn.commit()
        path = self.tmp_path / "condition.xlsx"
        _create_condition_workbook(
            path,
            [
                ("\u8a3c\u5238\u30b3\u30fc\u30c9", "1111"),
                ("\u6307\u6a19", "\u58f2\u4e0a\u9ad8"),
                ("\u671f\u9593", "\u5f53\u671f"),
                ("\u6c7a\u7b97\u7a2e\u5225", "\u56db\u534a\u671f\u5358\u72ec"),
            ],
        )
        condition = read_metric_excel_condition(path)

        rows, errors, warnings, _preview, target = build_metric_excel_rows(self.conn, condition)

        self.assertEqual(errors, [])
        self.assertEqual(target, 1)
        self.assertNotIn("quarter_standalone_metrics_not_ready", warnings)
        row = next(row for row in _detail_rows(rows) if row.period_scope == "quarter_standalone:2Q")
        self.assertEqual(row.metric_label, "2Q \u58f2\u4e0a\u9ad8 \u5358\u72ec")
        self.assertEqual(row.values_by_offset[0], 120.0)
        self.assertEqual(row.units_by_offset[0], "\u767e\u4e07\u5186")

    def test_quarter_standalone_infers_4q_current_period_end(self) -> None:
        self.conn.executemany(
            """
            INSERT INTO quarter_standalone_metrics (
                security_code, edinet_code, fiscal_year, quarter_type, period_end,
                metric_key, metric_base, metric_group, value_num, value_unit,
                calc_status, formula_name, source_detail_json, rule_version,
                created_at, updated_at
            )
            VALUES ('1111', 'E00001', 2026, ?, ?,
                    'NetSalesCurrent', 'NetSales', 'sales', ?, 'yen',
                    'ok', 'fixture', '{}', 'test', 'now', 'now')
            """,
            [
                ("3Q", "2025-12-31", 300_000_000.0),
                ("4Q", "", None),
            ],
        )
        self.conn.commit()
        condition = MetricExcelCondition(
            security_codes=["1111"],
            metric_labels=["\u58f2\u4e0a\u9ad8"],
            period_scopes=["quarter_standalone"],
            period_offsets=[0],
        )

        rows, errors, _warnings, _preview, _target = build_metric_excel_rows(self.conn, condition)

        self.assertEqual(errors, [])
        row = next(
            row
            for row in _detail_rows(rows)
            if row.period_scope == "quarter_standalone:4Q" and row.metric_base == "NetSales"
        )
        self.assertEqual(row.current_period_end, "2026-03-31")
        self.assertEqual(row.periods_by_offset[0], "4Q 2026-03-31")

    def test_forecast_revision_values_use_font_color_only(self) -> None:
        self.conn.executemany(
            """
            INSERT INTO jquants_financial_metrics (
                disclosure_number, local_code, security_code, edinet_code, metric_kind,
                period_scope, period_key, quarter_type, forecast_target, forecast_stage,
                fiscal_year, period_start, period_end, disclosed_date, disclosed_time,
                metric_key, metric_base, metric_group, value_num, value_unit, calc_status,
                source_field, source_detail_json, rule_version, created_at, updated_at
            ) VALUES (?, '11110', '1111', 'E00001', 'forecast',
                      'forecast', 'forecast:FY', NULL, 'FY', ?,
                      2026, '2025-04-01', '2026-03-31', ?, '15:00',
                      'NetSalesCurrent', 'NetSales', 'sales', ?, 'yen', 'ok',
                      'field', '{}', 'v1', '2026-05-06', '2026-05-06')
            """,
            [
                ("forecast0", "initial", "2026-05-01", 100_000_000.0),
                ("forecast1", "1Q", "2026-08-01", 120_000_000.0),
                ("forecast2", "2Q", "2026-11-01", 110_000_000.0),
            ],
        )
        self.conn.commit()
        condition_path = self.tmp_path / "condition.xlsx"
        output_path = self.tmp_path / "forecast_revision.xlsx"
        _create_condition_workbook(
            condition_path,
            [
                ("\u8a3c\u5238\u30b3\u30fc\u30c9", "1111"),
                ("\u6307\u6a19", "\u58f2\u4e0a\u9ad8"),
                ("\u671f\u9593", "\u5f53\u671f"),
                ("\u6c7a\u7b97\u7a2e\u5225", "\u4e88\u60f3"),
            ],
        )

        result = export_metric_excel(
            self.conn,
            condition_xlsx=condition_path,
            output_path=output_path,
            db_path=":memory:",
        )

        self.assertEqual(result.errors, [])
        workbook = load_workbook(output_path)
        ws = workbook[GENERAL_SHEET]
        value_cells_by_label = {
            ws.cell(row=row_index, column=9).value: ws.cell(row=row_index, column=11)
            for row_index in range(2, ws.max_row + 1)
            if ws.cell(row=row_index, column=7).value == ROW_KIND_DETAIL
        }
        up_color = value_cells_by_label["1Q \u58f2\u4e0a\u9ad8 \u4e88\u60f3"].font.color.rgb
        down_color = value_cells_by_label["2Q \u58f2\u4e0a\u9ad8 \u4e88\u60f3"].font.color.rgb
        self.assertTrue(str(up_color).endswith("FF0000"))
        self.assertTrue(str(down_color).endswith("00B0F0"))


if __name__ == "__main__":
    unittest.main()
