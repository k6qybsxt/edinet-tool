from __future__ import annotations

import shutil
import sqlite3
import sys
import unittest
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.workbook.defined_name import DefinedName

ROOT_DIR = Path(__file__).resolve().parents[2]
SRC_DIR = ROOT_DIR / "src"
if str(SRC_DIR) not in sys.path:
    sys.path.insert(0, str(SRC_DIR))

from edinet_pipeline.services.db_excel_export_service import (
    INPUT_SHEET_NAME,
    export_analysis_workbooks_from_db,
    read_analysis_workbook_condition,
)


def _add_name(workbook: openpyxl.Workbook, name: str, cell_ref: str) -> None:
    workbook.defined_names.add(
        DefinedName(name=name, attr_text=f"'{INPUT_SHEET_NAME}'!${cell_ref[0]}${cell_ref[1:]}")
    )


class DbExcelExportServiceTest(unittest.TestCase):
    def setUp(self) -> None:
        self.tmp = ROOT_DIR / "tests" / "_tmp_db_excel_export"
        if self.tmp.exists():
            shutil.rmtree(self.tmp)
        self.tmp.mkdir(parents=True, exist_ok=True)
        self.conn = sqlite3.connect(":memory:")
        self.conn.row_factory = sqlite3.Row
        self._create_schema()
        self.template = self.tmp / "template.xlsx"
        self._create_template(self.template)

    def tearDown(self) -> None:
        self.conn.close()
        shutil.rmtree(self.tmp)

    def _create_schema(self) -> None:
        self.conn.executescript(
            """
            CREATE TABLE issuer_master (
                edinet_code TEXT PRIMARY KEY,
                security_code TEXT,
                company_name TEXT,
                is_listed INTEGER
            );
            CREATE TABLE filings (
                doc_id TEXT PRIMARY KEY,
                edinet_code TEXT,
                security_code TEXT,
                form_type TEXT,
                period_end TEXT,
                submit_date TEXT,
                document_display_unit TEXT
            );
            CREATE TABLE normalized_metrics (
                doc_id TEXT,
                metric_key TEXT,
                value_num REAL
            );
            CREATE TABLE derived_metrics (
                doc_id TEXT,
                metric_key TEXT,
                value_num REAL,
                calc_status TEXT
            );
            CREATE TABLE market_derived_metrics (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                source_type TEXT,
                source_id TEXT,
                security_code TEXT,
                metric_base TEXT,
                value_num REAL,
                calc_status TEXT,
                updated_at TEXT
            );
            CREATE TABLE jquants_financial_metrics (
                disclosure_number TEXT,
                local_code TEXT,
                security_code TEXT,
                metric_kind TEXT,
                period_key TEXT,
                forecast_stage TEXT,
                fiscal_year INTEGER,
                period_end TEXT,
                disclosed_date TEXT,
                disclosed_time TEXT,
                metric_base TEXT,
                value_num REAL,
                calc_status TEXT
            );
            """
        )
        self.conn.execute(
            "INSERT INTO issuer_master VALUES ('E00893', '4613', 'Kansai Paint', 1)"
        )
        annual_periods = [
            ("DOC2025", "2025-03-31", 100_000_000),
            ("DOC2024", "2024-03-31", 90_000_000),
            ("DOC2023", "2023-03-31", 80_000_000),
            ("DOC2022", "2022-03-31", 70_000_000),
            ("DOC2021", "2021-03-31", 60_000_000),
        ]
        for index, (doc_id, period_end, net_sales) in enumerate(annual_periods):
            self.conn.execute(
                "INSERT INTO filings VALUES (?, 'E00893', '4613', '030000', ?, ?, '百万円')",
                (doc_id, period_end, f"{period_end} 15:00"),
            )
            self.conn.execute(
                "INSERT INTO normalized_metrics VALUES (?, 'NetSalesCurrent', ?)",
                (doc_id, net_sales),
            )
            self.conn.execute(
                "INSERT INTO derived_metrics VALUES (?, 'OutstandingSharesCurrent', ?, 'ok')",
                (doc_id, 10_000_000 + index),
            )
        self.conn.execute(
            "INSERT INTO filings VALUES ('HALF2026', 'E00893', '4613', '043A00', '2025-09-30', '2025-11-01 15:00', '百万円')"
        )
        self.conn.execute(
            "INSERT INTO normalized_metrics VALUES ('HALF2026', 'NetSalesCurrent', 50_000_000)"
        )
        self.conn.execute(
            """
            INSERT INTO market_derived_metrics
              (source_type, source_id, security_code, metric_base, value_num, calc_status, updated_at)
            VALUES ('edinet', 'DOC2026', '4613', 'StockPrice', 1200, 'ok', '2026-05-01')
            """
        )
        self.conn.execute(
            """
            INSERT INTO jquants_financial_metrics
            VALUES ('DISC1Q', '4613', '4613', 'actual', 'actual:1Q', NULL, 2026,
                    '2025-06-30', '2025-08-01', '15:00', 'NetSales', 20_000_000, 'ok')
            """
        )
        self.conn.execute(
            """
            INSERT INTO jquants_financial_metrics
            VALUES ('DISC3Q', '4613', '4613', 'actual', 'actual:3Q', NULL, 2026,
                    '2025-12-31', '2026-02-01', '15:00', 'NetSales', 75_000_000, 'ok')
            """
        )
        for stage, value in (("initial", 110_000_000), ("2Q", 115_000_000)):
            self.conn.execute(
                """
                INSERT INTO jquants_financial_metrics
                VALUES (?, '4613', '4613', 'forecast', 'forecast:FY', ?, 2026,
                        '2026-03-31', '2025-05-01', '15:00', 'NetSales', ?, 'ok')
                """,
                (f"FC_{stage}", stage, value),
            )
        self.conn.commit()

    def _create_template(self, path: Path) -> None:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = INPUT_SHEET_NAME
        wb.create_sheet("raw_edinet")
        wb.create_sheet("raw異常検知メーター")
        for name, cell in {
            "NetSales_Current": "P36",
            "NetSales_Prior1": "M5",
            "NetSales_Prior2": "J5",
            "NetSales_Prior3": "G5",
            "NetSales_Prior4": "D5",
            "StockPrice_Q4": "P53",
        }.items():
            _add_name(wb, name, cell)
        wb.save(path)

    def _create_condition(self, start: str, end: str) -> Path:
        path = self.tmp / "condition.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "条件"
        ws["A2"] = "証券コード"
        ws["B2"] = "4613"
        ws["A4"] = "期間 Start"
        ws["B4"] = start
        ws["A5"] = "期間 End"
        ws["B5"] = end
        wb.save(path)
        return path

    def test_read_condition_normalizes_periods(self) -> None:
        condition = read_analysis_workbook_condition(self._create_condition("3期前", "最新"))

        self.assertEqual(condition.security_codes, ["4613"])
        self.assertEqual(condition.period_start, "Prior3")
        self.assertEqual(condition.period_end, "Latest")

    def test_export_respects_selected_period_range(self) -> None:
        result = export_analysis_workbooks_from_db(
            self.conn,
            condition_xlsx=self._create_condition("3期前", "1期前"),
            template_path=self.template,
            output_dir=self.tmp / "out",
        )

        self.assertEqual(result.errors, [])
        wb = openpyxl.load_workbook(result.output_paths[0], data_only=False)
        try:
            ws = wb[INPUT_SHEET_NAME]
            self.assertIsNone(ws["D5"].value)
            self.assertEqual(ws["G5"].value, 80)
            self.assertEqual(ws["J5"].value, 90)
            self.assertEqual(ws["M5"].value, 100)
            self.assertIsNone(ws["P36"].value)
            self.assertEqual(ws["J2"].value, "百万円")
            self.assertEqual(ws["K2"].value, "4613")
            self.assertEqual(ws["L2"].value, "Kansai Paint")
            self.assertEqual(ws["N2"].value, "2026")
            self.assertEqual(ws["O2"].value, "03")
            self.assertEqual(ws["P2"].value, datetime.now().strftime("%Y-%m-%d"))
            self.assertNotIn("raw_edinet", wb.sheetnames)
            self.assertNotIn("raw異常検知メーター", wb.sheetnames)
        finally:
            wb.close()

    def test_export_writes_current_quarter_forecast_and_stock_values(self) -> None:
        result = export_analysis_workbooks_from_db(
            self.conn,
            security_codes=["4613"],
            template_path=self.template,
            output_dir=self.tmp / "out",
        )

        self.assertEqual(result.errors, [])
        wb = openpyxl.load_workbook(result.output_paths[0], data_only=False)
        try:
            ws = wb[INPUT_SHEET_NAME]
            self.assertIsNone(ws["P36"].value)
            self.assertEqual(ws["M5"].value, 100)
            self.assertEqual(ws["G36"].value, 20)
            self.assertEqual(ws["J36"].value, 50)
            self.assertEqual(ws["M36"].value, 75)
            self.assertEqual(ws["D30"].value, 110)
            self.assertEqual(ws["J30"].value, 115)
            self.assertIsNone(ws["P53"].value)
        finally:
            wb.close()


if __name__ == "__main__":
    unittest.main()
