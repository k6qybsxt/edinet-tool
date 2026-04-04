from __future__ import annotations

import shutil
import sys
import unittest
from pathlib import Path

from openpyxl import Workbook
from openpyxl.workbook.defined_name import DefinedName


ROOT_DIR = Path(__file__).resolve().parents[2]
SRC_DIR = ROOT_DIR / "src"
if str(SRC_DIR) not in sys.path:
    sys.path.insert(0, str(SRC_DIR))

from edinet_pipeline.domain.output_buffer import OutputBuffer  # noqa: E402
from edinet_pipeline.services import loop_stage_service as loop_stage_service_module  # noqa: E402
from edinet_pipeline.services.loop_stage_service import (  # noqa: E402
    build_excel_not_found_result,
    build_excel_write_inputs_stage,
    build_excel_output_payload,
    build_raw_rows_stage,
    build_stock_write_context,
    close_workbook_quietly,
    execute_stock_write_stage,
    finalize_company_result_stage,
    finalize_output_excel,
    open_workbook_stage,
    pick_company_name,
    pick_period_end,
    prepare_excel_stage,
    resolve_document_display_unit,
    resolve_runtime_flags,
    run_parse_stages,
    run_workbook_output_stages,
    save_workbook_stage,
    write_named_range_stage,
    write_raw_sheet_stage,
)


class _FakeParsedDocument:
    def __init__(self, document_display_unit: str) -> None:
        self.document_display_unit = document_display_unit


class _FakeParseCache:
    def __init__(self, *documents: _FakeParsedDocument) -> None:
        self._documents = list(documents)
        self.calls: list[str] = []

    def get_or_create(self, path, parser_func):
        self.calls.append(path)
        if self._documents:
            return self._documents.pop(0)
        return parser_func(path)


class _DummyLogger:
    def info(self, *args, **kwargs):
        return None

    def warning(self, *args, **kwargs):
        return None

    def debug(self, *args, **kwargs):
        return None


class _CloseTrackingWorkbook:
    def __init__(self) -> None:
        self.closed = False
        self._archive = None
        self.vba_archive = None

    def close(self) -> None:
        self.closed = True


class LoopStageServiceTest(unittest.TestCase):
    def _build_workbook_with_named_range(self) -> Workbook:
        workbook = Workbook()
        input_sheet = workbook.active
        input_sheet.title = "決算入力"
        raw_sheet = workbook.create_sheet("raw_edinet")
        raw_sheet["A1"] = "header"
        input_sheet["A1"] = ""
        workbook.defined_names["NetSales_Current"] = DefinedName(
            "NetSales_Current",
            attr_text="'決算入力'!$A$1",
        )
        return workbook

    def test_build_excel_output_payload_adds_dates_and_filters_quarters(self) -> None:
        payload = build_excel_output_payload(
            {
                "NetSalesCurrent": 100,
                "NetSalesQuarter": 10,
            },
            x1={
                "CurrentFiscalYearEndDateDEI": "2026/03/31",
                "CurrentPeriodEndDateDEI": "2026/03/31",
                "CurrentFiscalYearStartDateDEI": "2025/04/01",
            },
            use_half=False,
        )

        self.assertEqual(payload["NetSalesCurrent"], 100)
        self.assertNotIn("NetSalesQuarter", payload)
        self.assertEqual(payload["CurrentFiscalYearEndDateDEI"], "2026-03-31")
        self.assertEqual(payload["CurrentFiscalYearEndDateDEIyear"], "2026")
        self.assertEqual(payload["CurrentFiscalYearEndDateDEImonth"], "03")
        self.assertEqual(payload["CurrentPeriodEndDateDEI"], "2026-03-31")
        self.assertEqual(payload["CurrentFiscalYearStartDateDEI"], "2025-04-01")

    def test_resolve_document_display_unit_uses_parse_cache_result(self) -> None:
        fake_cache = _FakeParseCache(_FakeParsedDocument("千円"))

        result = resolve_document_display_unit(
            xbrl_file_paths={"file1": ["sample1.xbrl"], "file2": []},
            x1={"DocumentDisplayUnit": "百万円"},
            x2=None,
            use_half=False,
            parse_cache=fake_cache,
            logger=_DummyLogger(),
            parse_document_func=lambda path, mode, logger: _FakeParsedDocument("百万円"),
        )

        self.assertEqual(result, "千円")
        self.assertEqual(fake_cache.calls, ["sample1.xbrl"])

    def test_build_stock_write_context_shifts_half_mode_year(self) -> None:
        context = build_stock_write_context(
            out_buffer_dict={"CurrentFiscalYearEndDateDEI": "2026-03-31"},
            x1=None,
            use_half=True,
            security_code="1234",
        )

        self.assertEqual(context["stock_code"], "1234.T")
        self.assertEqual(context["fiscal_year_end"], "2025-03-31")
        self.assertEqual(context["stock_date_pairs"][-1]["name"], "StockPrice_Q4")
        self.assertEqual(context["stock_date_pairs"][-1]["target_date"], "2025-03-31")

    def test_pick_helpers_fall_back_to_xbrl_metadata(self) -> None:
        period_end = pick_period_end(
            None,
            {"CurrentFiscalYearEndDateDEI": "2026/03/31"},
            None,
        )
        company_name = pick_company_name(
            None,
            {"CompanyNameInJapaneseDEI": "テスト株式会社"},
            None,
            None,
        )

        self.assertEqual(period_end, "2026-03-31")
        self.assertEqual(company_name, "テスト株式会社")

    def test_finalize_output_excel_moves_file_under_output_root(self) -> None:
        temp_root = ROOT_DIR / "tests" / "_tmp_loop_stage"
        source_file = temp_root / "source.xlsm"
        temp_root.mkdir(parents=True, exist_ok=True)
        source_file.write_bytes(b"dummy")

        try:
            final_path = finalize_output_excel(
                excel_file_path=str(source_file),
                output_root=str(temp_root),
                security_code="12340",
                company_name="テスト株式会社",
                period_end_date="2026-03-31",
                logger=_DummyLogger(),
            )

            self.assertTrue(Path(final_path).exists())
            self.assertEqual(Path(final_path).parent, temp_root / "excel")
            self.assertFalse(source_file.exists())
        finally:
            if temp_root.exists():
                shutil.rmtree(temp_root)

    def test_build_excel_not_found_result_returns_failed_shape(self) -> None:
        result = build_excel_not_found_result(
            slot=1,
            company_code="12340",
            company_name="テスト株式会社",
        )

        self.assertEqual(result["status"], "failed")
        self.assertIsNone(result["stock_status"])
        self.assertIsNone(result["output_excel"])

    def test_run_parse_stages_returns_parsed_bundle(self) -> None:
        out_buffer = OutputBuffer()
        call_order: list[str] = []

        def fake_parse_half_doc(**kwargs):
            call_order.append("file1")
            return {"HalfMetric": 1}, 2026, False

        def fake_append_initial(buffer, x1):
            call_order.append("append")
            buffer.put("UseHalfModeFlag", 0, "file1_annual")

        def fake_parse_latest_annual_doc(**kwargs):
            call_order.append("file2")
            kwargs["out_buffer"].put("NetSalesCurrent", 100, "file2_annual")
            return {"NetSalesCurrent": 100}, {"CompanyNameDEI": "Test"}, "file2.xbrl", "12340", 2026

        def fake_parse_old_annual_doc(**kwargs):
            call_order.append("file3")
            kwargs["out_buffer"].put("NetSalesPrior2", 80, "file3_annual")

        def fake_finalize_half_buffer(**kwargs):
            call_order.append("half")

        result = run_parse_stages(
            loop={"slot": 1},
            xbrl_file_paths={"file1": ["f1"], "file2": ["f2"], "file3": ["f3"]},
            excel_file_path="sample.xlsm",
            parsed_docs=[],
            skipped_files=[],
            loop_event={"phases": {}, "counts": {}, "errors": []},
            out_buffer=out_buffer,
            logger=_DummyLogger(),
            perf_counter=lambda: 1.0,
            parse_cache=object(),
            parse_half_doc_func=fake_parse_half_doc,
            parse_latest_annual_doc_func=fake_parse_latest_annual_doc,
            parse_old_annual_doc_func=fake_parse_old_annual_doc,
            finalize_half_buffer_func=fake_finalize_half_buffer,
            append_initial_annual_output_func=fake_append_initial,
        )

        self.assertEqual(call_order, ["file1", "append", "file2", "file3", "half"])
        self.assertEqual(result["security_code"], "12340")
        self.assertFalse(result["use_half"])
        self.assertEqual(out_buffer.to_dict()["NetSalesCurrent"], 100)
        self.assertEqual(out_buffer.to_dict()["NetSalesPrior2"], 80)

    def test_build_excel_write_inputs_stage_builds_payload_and_display_unit(self) -> None:
        out_buffer = OutputBuffer()
        out_buffer.put("NetSalesCurrent", 100, "file2_annual")
        out_buffer.put("NetSalesCurrent", 120, "half_final")
        out_buffer.put("CurrentFiscalYearEndDateDEI", "2026/03/31", "file1_annual")
        out_buffer.put("CurrentFiscalYearStartDateDEI", "2025/04/01", "file1_annual")
        out_buffer.put("CurrentPeriodEndDateDEI", "2026/03/31", "file1_annual")

        out_buffer_dict, display_unit = build_excel_write_inputs_stage(
            out_buffer=out_buffer,
            xbrl_file_paths={"file1": [], "file2": []},
            x1={
                "CurrentFiscalYearEndDateDEI": "2026/03/31",
                "CurrentFiscalYearStartDateDEI": "2025/04/01",
                "CurrentPeriodEndDateDEI": "2026/03/31",
                "DocumentDisplayUnit": loop_stage_service_module._VALID_DISPLAY_UNITS[0],
            },
            x2=None,
            use_half=False,
            loop={"slot": 1},
            company_code="12340",
            security_code="12340",
            company_name="テスト株式会社",
            parse_cache=None,
            logger=_DummyLogger(),
        )

        self.assertEqual(out_buffer_dict["NetSalesCurrent"], 120)
        self.assertEqual(out_buffer_dict["CurrentFiscalYearEndDateDEIyear"], "2026")
        self.assertEqual(display_unit, loop_stage_service_module._VALID_DISPLAY_UNITS[0])

    def test_write_named_range_stage_updates_counts_and_logs_optional_missing_only(self) -> None:
        workbook = self._build_workbook_with_named_range()
        loop_event = {"phases": {}, "counts": {}}

        try:
            write_named_range_stage(
                workbook=workbook,
                out_buffer_dict={
                    "NetSalesCurrent": 1_000_000,
                    "CurrentPeriodEndDateDEI": "2026-03-31",
                },
                display_unit="百万円",
                loop_event=loop_event,
                loop={"slot": 1},
                company_code="12340",
                security_code="12340",
                company_name="テスト株式会社",
                logger=_DummyLogger(),
                perf_counter=lambda: 10.0,
                optional_output_names={"CurrentPeriodEndDateDEI"},
            )

            self.assertEqual(loop_event["counts"]["named_ranges_written"], 1)
            self.assertEqual(loop_event["counts"]["named_ranges_missing"], 0)
            self.assertEqual(workbook["決算入力"]["A1"].value, 1)
        finally:
            workbook.close()

    def test_write_raw_sheet_stage_skips_when_disabled(self) -> None:
        workbook = self._build_workbook_with_named_range()
        raw_sheet = workbook["raw_edinet"]
        raw_sheet["A2"] = "keep"
        loop_event = {"phases": {}, "counts": {}}

        try:
            write_raw_sheet_stage(
                workbook=workbook,
                raw_rows=[{"col1": "value"}],
                raw_cols=["col1"],
                write_raw_sheet=False,
                loop_event=loop_event,
                loop={"slot": 1},
                company_code="12340",
                security_code="12340",
                logger=_DummyLogger(),
                perf_counter=lambda: 5.0,
            )

            self.assertEqual(loop_event["phases"]["raw_write"]["sec"], 0.0)
            self.assertEqual(raw_sheet["A2"].value, "keep")
        finally:
            workbook.close()

    def test_execute_stock_write_stage_returns_disabled_before_network_call(self) -> None:
        status = execute_stock_write_stage(
            workbook=object(),
            stock_code="1234.T",
            stock_date_pairs=[{"name": "StockPrice_Q4", "target_date": "2026-03-31"}],
            enable_stock=False,
            loop_event={"phases": {}, "counts": {}},
            loop={"slot": 1},
            company_code="12340",
            security_code="12340",
            company_name="テスト株式会社",
            logger=_DummyLogger(),
            perf_counter=lambda: 5.0,
            write_stock_func=lambda *args, **kwargs: (_ for _ in ()).throw(AssertionError("should not be called")),
        )

        self.assertEqual(status, "disabled")

    def test_save_and_close_workbook_stage(self) -> None:
        temp_root = ROOT_DIR / "tests" / "_tmp_loop_stage_save"
        target_file = temp_root / "saved.xlsx"
        temp_root.mkdir(parents=True, exist_ok=True)

        workbook = Workbook()
        workbook.active["A1"] = "saved"
        loop_event = {"phases": {}, "counts": {}}

        try:
            save_workbook_stage(
                workbook=workbook,
                excel_file_path=str(target_file),
                loop_event=loop_event,
                loop={"slot": 1},
                company_code="12340",
                security_code="12340",
                logger=_DummyLogger(),
                perf_counter=lambda: 7.0,
            )
            self.assertTrue(target_file.exists())
            self.assertIn("workbook_save", loop_event["phases"])

            tracking_workbook = _CloseTrackingWorkbook()
            close_workbook_quietly(tracking_workbook)
            self.assertTrue(tracking_workbook.closed)
        finally:
            workbook.close()
            if temp_root.exists():
                shutil.rmtree(temp_root)

    def test_build_raw_rows_stage_records_phase(self) -> None:
        loop_event = {"phases": {}, "counts": {}}

        raw_rows = build_raw_rows_stage(
            parsed_docs=["doc1"],
            security_code="12340",
            run_id="run1",
            loop_event=loop_event,
            loop={"slot": 1},
            company_code="12340",
            logger=_DummyLogger(),
            perf_counter=lambda: 3.0,
            build_raw_rows_func=lambda **kwargs: [{"col1": "value"}],
        )

        self.assertEqual(raw_rows, [{"col1": "value"}])
        self.assertIn("raw_build", loop_event["phases"])

    def test_open_workbook_stage_records_phase(self) -> None:
        temp_root = ROOT_DIR / "tests" / "_tmp_loop_stage_open"
        target_file = temp_root / "open.xlsx"
        temp_root.mkdir(parents=True, exist_ok=True)

        workbook = Workbook()
        workbook.save(target_file)
        workbook.close()
        loop_event = {"phases": {}, "counts": {}}

        try:
            opened = open_workbook_stage(
                excel_file_path=str(target_file),
                loop_event=loop_event,
                loop={"slot": 1},
                company_code="12340",
                security_code="12340",
                logger=_DummyLogger(),
                perf_counter=lambda: 4.0,
                load_workbook_func=__import__("openpyxl").load_workbook,
            )
            try:
                self.assertIn("workbook_open", loop_event["phases"])
                self.assertEqual(opened.active["A1"].value, None)
            finally:
                opened.close()
        finally:
            if temp_root.exists():
                shutil.rmtree(temp_root)

    def test_prepare_excel_stage_returns_failed_result_when_template_missing(self) -> None:
        skipped_files: list[dict] = []

        result = prepare_excel_stage(
            loop={"slot": 1, "company_code": "12340", "company_name": "Test"},
            run_id="run1",
            skipped_files=skipped_files,
            logger=_DummyLogger(),
            prepare_workbook_func=lambda loop, run_id, logger: (None, None, "missing.xlsm"),
        )

        self.assertIsNotNone(result["failed_result"])
        self.assertEqual(result["failed_result"]["status"], "failed")
        self.assertEqual(len(skipped_files), 1)

    def test_resolve_runtime_flags_uses_defaults_and_runtime_values(self) -> None:
        class _Runtime:
            write_raw_sheet = False
            enable_stock = True

        self.assertEqual(
            resolve_runtime_flags(None),
            {"write_raw_sheet": True, "enable_stock": True},
        )
        self.assertEqual(
            resolve_runtime_flags(_Runtime()),
            {"write_raw_sheet": False, "enable_stock": True},
        )

    def test_run_workbook_output_stages_returns_stock_status(self) -> None:
        temp_root = ROOT_DIR / "tests" / "_tmp_loop_stage_output"
        target_file = temp_root / "dummy.xlsm"
        temp_root.mkdir(parents=True, exist_ok=True)
        workbook = self._build_workbook_with_named_range()
        original_close = workbook.close
        workbook.closed = False

        def tracked_close():
            workbook.closed = True
            original_close()

        workbook.close = tracked_close

        def fake_load_workbook(*args, **kwargs):
            return workbook

        try:
            result = run_workbook_output_stages(
                excel_file_path=str(target_file),
                out_buffer_dict={"NetSalesCurrent": 1_000_000},
                display_unit="百万円",
                raw_rows=[{"status": "OK"}],
                raw_cols=["status"],
                x1={"CurrentFiscalYearEndDateDEI": "2026-03-31"},
                use_half=False,
                security_code="12340",
                company_code="12340",
                company_name="Test",
                loop_event={"phases": {}, "counts": {}},
                loop={"slot": 1},
                logger=_DummyLogger(),
                perf_counter=lambda: 5.0,
                optional_output_names=set(),
                write_raw_sheet=False,
                enable_stock=False,
                load_workbook_func=fake_load_workbook,
                write_stock_func=lambda *args, **kwargs: {"written": 0, "miss": 0, "errors": 0},
            )

            self.assertEqual(result["stock_status"], "disabled")
            self.assertTrue(workbook.closed)
            self.assertTrue(target_file.exists())
        finally:
            if temp_root.exists():
                shutil.rmtree(temp_root)

    def test_finalize_company_result_stage_updates_loop_and_returns_success(self) -> None:
        temp_root = ROOT_DIR / "tests" / "_tmp_loop_stage_finalize"
        source_file = temp_root / "source.xlsm"
        temp_root.mkdir(parents=True, exist_ok=True)
        source_file.write_bytes(b"dummy")
        loop = {"slot": 1}
        summary_calls: list[dict] = []

        try:
            result = finalize_company_result_stage(
                loop=loop,
                loop_event={"phases": {}, "counts": {}, "has_half": False},
                x1={"CompanyNameDEI": "Test", "CurrentFiscalYearEndDateDEI": "2026/03/31"},
                x2=None,
                meta2=None,
                use_half=False,
                security_code="12340",
                company_code="12340",
                company_name=None,
                excel_file_path=str(source_file),
                output_root=str(temp_root),
                stock_status="success",
                raw_rows=[{"status": "OK"}],
                out_buffer_dict={"NetSalesCurrent": 100},
                skipped_files=[],
                t0=0.0,
                perf_counter=lambda: 1.0,
                logger=_DummyLogger(),
                write_loop_summary_func=lambda **kwargs: summary_calls.append(kwargs),
            )

            self.assertEqual(result["status"], "success")
            self.assertEqual(result["company_code"], "12340")
            self.assertTrue(Path(result["output_excel"]).exists())
            self.assertEqual(loop["final_excel_file_path"], result["output_excel"])
            self.assertEqual(len(summary_calls), 1)
        finally:
            if temp_root.exists():
                shutil.rmtree(temp_root)


if __name__ == "__main__":
    unittest.main()
