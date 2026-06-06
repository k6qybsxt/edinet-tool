from __future__ import annotations

from dataclasses import dataclass, field
from datetime import datetime
import json
import math
from pathlib import Path
import sqlite3
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill

from edinet_monitor.config.settings import OPERATION_LOG_ROOT, PROJECT_ROOT
from edinet_monitor.services.jquants.mapper import normalize_security_code
from edinet_monitor.services.metric_excel_export_service import (
    ALL_PERIOD_SCOPES,
    CONDITION_SHEET,
    GENERAL_SHEET,
    HALF_DISABLED_BASES,
    MARKET_METRIC_BASES,
    PERIOD_LABEL_BY_OFFSET,
    ROW_KIND_AVERAGE,
    ROW_KIND_MEDIAN,
    SUMMARY_SHEET,
    MetricExcelCondition,
    MetricExcelRow,
    _build_label_to_base_map,
    _decision_label_for_row,
    _normalize_text,
    build_metric_excel_rows,
)


DEFAULT_TARGET_CONFIG_PATH = PROJECT_ROOT / "config" / "excel" / "metric_excel_audit_targets.json"
DEFAULT_OUTPUT_DIR = OPERATION_LOG_ROOT / "excel_audit"
SEVERITY_ORDER = {"critical": 0, "warning": 1, "info": 2}

HEADER_SECURITY_CODE = "\u8a3c\u5238\u30b3\u30fc\u30c9"
HEADER_COMPANY_NAME = "\u4f01\u696d\u540d"
HEADER_INDUSTRY = "\u696d\u7a2e"
HEADER_MARKET = "\u5e02\u5834\u533a\u5206"
HEADER_DECISION = "\u6c7a\u7b97\u7a2e\u5225"
HEADER_ROW_KIND = "\u884c\u7a2e\u5225"
HEADER_CURRENT_PERIOD_END = "\u671f\u672b\u5e74\u6708\u65e5_\u5f53\u671f"
HEADER_METRIC = "\u6307\u6a19"
PERIOD_VALUE_SUFFIX = "\u6570\u5024"
PERIOD_UNIT_SUFFIX = "\u5358\u4f4d"
PERIOD_RATIO_SUFFIX = "\u6bd4\u7387"
PERIOD_RANK_SUFFIX = "\u9806\u4f4d"
PERIOD_PERIOD_SUFFIX = "\u671f\u9593"
DECISION_ANNUAL = "\u901a\u671f"
DECISION_QUARTER = "\u56db\u534a\u671f"
DECISION_QUARTER_STANDALONE = "\u56db\u534a\u671f\u5358\u72ec"
DECISION_FORECAST = "\u4e88\u60f3"


@dataclass(frozen=True)
class ExcelAuditTarget:
    security_code: str
    target_set: str
    features: tuple[str, ...] = ()


@dataclass
class ExcelAuditIssue:
    severity: str
    category: str
    check_name: str
    security_code: str = ""
    company_name: str = ""
    sheet_name: str = ""
    period_scope: str = ""
    metric_base: str = ""
    metric_label: str = ""
    row_kind: str = ""
    period_label: str = ""
    expected_value: Any = ""
    actual_value: Any = ""
    expected_unit: str = ""
    actual_unit: str = ""
    source_table: str = ""
    message: str = ""
    detail: dict[str, Any] = field(default_factory=dict)

    def to_dict(self) -> dict[str, Any]:
        return {
            "severity": self.severity,
            "category": self.category,
            "check_name": self.check_name,
            "security_code": self.security_code,
            "company_name": self.company_name,
            "sheet_name": self.sheet_name,
            "period_scope": self.period_scope,
            "metric_base": self.metric_base,
            "metric_label": self.metric_label,
            "row_kind": self.row_kind,
            "period_label": self.period_label,
            "expected_value": self.expected_value,
            "actual_value": self.actual_value,
            "expected_unit": self.expected_unit,
            "actual_unit": self.actual_unit,
            "source_table": self.source_table,
            "message": self.message,
            "detail": self.detail,
        }


@dataclass(frozen=True)
class ExcelAuditOptions:
    excel_path: Path
    db_path: Path | None = None
    target_set: str = "normal"
    target_config_path: Path = DEFAULT_TARGET_CONFIG_PATH
    output_dir: Path = DEFAULT_OUTPUT_DIR
    period_offsets: tuple[int, ...] = tuple(range(10, -1, -1))
    period_scopes: tuple[str, ...] = tuple(ALL_PERIOD_SCOPES)
    value_tolerance: float = 1e-6


@dataclass(frozen=True)
class ExcelAuditResult:
    audit_id: str
    generated_at: str
    excel_path: Path
    json_path: Path
    report_excel_path: Path
    target_set: str
    targets: list[ExcelAuditTarget]
    expected_rows: int
    actual_rows: int
    issues: list[ExcelAuditIssue]
    errors: list[str]
    warnings: list[str]
    counts_by_severity: dict[str, int]

    @property
    def issue_count(self) -> int:
        return self.counts_by_severity.get("critical", 0) + self.counts_by_severity.get("warning", 0)


@dataclass(frozen=True)
class _AuditRowKey:
    sheet_name: str
    security_code: str
    decision_label: str
    row_kind: str
    current_period_end: str
    metric_label: str


@dataclass
class _ActualExcelRow:
    sheet_name: str
    row_number: int
    security_code: str
    company_name: str
    industry_33: str
    market: str
    decision_label: str
    row_kind: str
    current_period_end: str
    metric_label: str
    metric_base: str
    periods_by_offset: dict[int, str]
    values_by_offset: dict[int, Any]
    units_by_offset: dict[int, str]
    ratios_by_offset: dict[int, Any]
    ranks_by_offset: dict[int, str]


def _json_dumps(value: Any) -> str:
    return json.dumps(value, ensure_ascii=False, sort_keys=True, separators=(",", ":"))


def _normalized_code(value: Any) -> str:
    return normalize_security_code(str(value or "").strip())


def _clean_text(value: Any) -> str:
    return str(value or "").strip()


def _is_blank(value: Any) -> bool:
    return value is None or str(value).strip() == ""


def _to_float_or_none(value: Any) -> float | None:
    if _is_blank(value):
        return None
    if isinstance(value, (int, float)):
        if isinstance(value, float) and math.isnan(value):
            return None
        return float(value)
    text = str(value).replace(",", "").strip()
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def _same_value(expected: Any, actual: Any, tolerance: float) -> bool:
    expected_num = _to_float_or_none(expected)
    actual_num = _to_float_or_none(actual)
    if expected_num is not None and actual_num is not None:
        return abs(expected_num - actual_num) <= tolerance
    return _clean_text(expected) == _clean_text(actual)


def load_excel_audit_targets(
    target_config_path: str | Path = DEFAULT_TARGET_CONFIG_PATH,
    *,
    target_set: str = "normal",
) -> list[ExcelAuditTarget]:
    path = Path(target_config_path)
    payload = json.loads(path.read_text(encoding="utf-8"))
    sets = payload.get("target_sets")
    if not isinstance(sets, dict):
        raise ValueError(f"target config must contain target_sets: {path}")

    selected_names = ["normal", "known_issue"] if target_set == "all" else [target_set]
    targets: list[ExcelAuditTarget] = []
    seen_codes: set[str] = set()
    for set_name in selected_names:
        raw_targets = sets.get(set_name)
        if not isinstance(raw_targets, list):
            raise ValueError(f"unknown target set: {set_name}")
        for item in raw_targets:
            if not isinstance(item, dict):
                raise ValueError(f"target item must be an object in set={set_name}")
            code = _normalized_code(item.get("security_code"))
            if not code:
                raise ValueError(f"target item is missing security_code in set={set_name}")
            if code in seen_codes:
                continue
            seen_codes.add(code)
            features = tuple(str(value) for value in item.get("features", []) if str(value).strip())
            targets.append(ExcelAuditTarget(security_code=code, target_set=set_name, features=features))
    if not targets:
        raise ValueError(f"target set is empty: {target_set}")
    return targets


def _expected_key(row: MetricExcelRow) -> _AuditRowKey:
    return _AuditRowKey(
        sheet_name=row.sheet_name,
        security_code=_normalized_code(row.security_code),
        decision_label=_decision_label_for_row(row),
        row_kind=_clean_text(row.row_kind),
        current_period_end=_clean_text(row.current_period_end),
        metric_label=_normalize_text(row.metric_label),
    )


def _actual_key(row: _ActualExcelRow) -> _AuditRowKey:
    return _AuditRowKey(
        sheet_name=row.sheet_name,
        security_code=_normalized_code(row.security_code),
        decision_label=_clean_text(row.decision_label),
        row_kind=_clean_text(row.row_kind),
        current_period_end=_clean_text(row.current_period_end),
        metric_label=_normalize_text(row.metric_label),
    )


def _key_without_sheet(key: _AuditRowKey) -> tuple[str, str, str, str, str]:
    return (
        key.security_code,
        key.decision_label,
        key.row_kind,
        key.current_period_end,
        key.metric_label,
    )


def _key_without_decision(key: _AuditRowKey) -> tuple[str, str, str, str, str]:
    return (
        key.sheet_name,
        key.security_code,
        key.row_kind,
        key.current_period_end,
        key.metric_label,
    )


def _source_table_for_expected(row: MetricExcelRow) -> str:
    if row.segment_kind:
        return "segment_metrics"
    if row.row_kind in {ROW_KIND_AVERAGE, ROW_KIND_MEDIAN}:
        return "derived_metrics"
    if row.period_scope.startswith("quarter_standalone"):
        return "quarter_standalone_metrics"
    if row.period_scope.startswith("forecast"):
        return "jquants_financial_metrics"
    if row.metric_base in MARKET_METRIC_BASES:
        return "market_derived_metrics"
    if row.period_scope.startswith("quarter:") and row.period_scope != "quarter:2Q":
        return "jquants_financial_metrics"
    return "derived_metrics"


def _issue_for_expected(
    row: MetricExcelRow,
    *,
    severity: str,
    category: str,
    check_name: str,
    period_label: str = "",
    expected_value: Any = "",
    actual_value: Any = "",
    expected_unit: str = "",
    actual_unit: str = "",
    message: str,
    detail: dict[str, Any] | None = None,
) -> ExcelAuditIssue:
    return ExcelAuditIssue(
        severity=severity,
        category=category,
        check_name=check_name,
        security_code=_normalized_code(row.security_code),
        company_name=row.company_name,
        sheet_name=row.sheet_name,
        period_scope=row.period_scope,
        metric_base=row.metric_base,
        metric_label=row.metric_label,
        row_kind=row.row_kind,
        period_label=period_label,
        expected_value=expected_value,
        actual_value=actual_value,
        expected_unit=expected_unit,
        actual_unit=actual_unit,
        source_table=_source_table_for_expected(row),
        message=message,
        detail=detail or {},
    )


def _issue_for_actual(
    row: _ActualExcelRow,
    *,
    severity: str,
    category: str,
    check_name: str,
    message: str,
    detail: dict[str, Any] | None = None,
) -> ExcelAuditIssue:
    return ExcelAuditIssue(
        severity=severity,
        category=category,
        check_name=check_name,
        security_code=_normalized_code(row.security_code),
        company_name=row.company_name,
        sheet_name=row.sheet_name,
        period_scope=row.decision_label,
        metric_base=row.metric_base,
        metric_label=row.metric_label,
        row_kind=row.row_kind,
        message=message,
        detail={"row_number": row.row_number, **(detail or {})},
    )


def _header_map(header_row: tuple[Any, ...]) -> dict[str, int]:
    return {_clean_text(value): index for index, value in enumerate(header_row) if _clean_text(value)}


def _period_columns(headers: dict[str, int]) -> dict[int, dict[str, int]]:
    columns: dict[int, dict[str, int]] = {}
    for offset, label in PERIOD_LABEL_BY_OFFSET.items():
        prefix = f"{label}_"
        mapping: dict[str, int] = {}
        for key, suffix in [
            ("period", PERIOD_PERIOD_SUFFIX),
            ("value", PERIOD_VALUE_SUFFIX),
            ("unit", PERIOD_UNIT_SUFFIX),
            ("ratio", PERIOD_RATIO_SUFFIX),
            ("rank", PERIOD_RANK_SUFFIX),
        ]:
            header = f"{prefix}{suffix}"
            if header in headers:
                mapping[key] = headers[header]
        if mapping:
            columns[offset] = mapping
    return columns


def _value_at(values: tuple[Any, ...], index: int | None) -> Any:
    if index is None or index >= len(values):
        return ""
    return values[index]


def _infer_metric_base(sheet_name: str, metric_label: str) -> str:
    try:
        mapping = _build_label_to_base_map(sheet_name)
    except KeyError:
        mapping = _build_label_to_base_map(GENERAL_SHEET)
    normalized = _normalize_text(metric_label)
    if normalized in mapping:
        return mapping[normalized]
    for prefix in ("1Q", "2Q", "3Q", "4Q"):
        if str(metric_label).strip().startswith(f"{prefix} "):
            inner = str(metric_label).strip()[len(prefix) + 1 :]
            return mapping.get(_normalize_text(inner), "")
    return ""


def read_metric_excel_rows(
    excel_path: str | Path,
    *,
    target_codes: set[str],
) -> tuple[list[_ActualExcelRow], list[str]]:
    workbook = load_workbook(excel_path, data_only=True, read_only=True)
    rows: list[_ActualExcelRow] = []
    warnings: list[str] = []
    required_headers = {
        HEADER_SECURITY_CODE,
        HEADER_COMPANY_NAME,
        HEADER_INDUSTRY,
        HEADER_MARKET,
        HEADER_DECISION,
        HEADER_ROW_KIND,
        HEADER_CURRENT_PERIOD_END,
        HEADER_METRIC,
    }
    try:
        for sheet_name in workbook.sheetnames:
            if sheet_name in {SUMMARY_SHEET, CONDITION_SHEET}:
                continue
            ws = workbook[sheet_name]
            iterator = ws.iter_rows(values_only=True)
            try:
                header_row = next(iterator)
            except StopIteration:
                continue
            headers = _header_map(header_row)
            if not required_headers.issubset(headers):
                continue
            period_cols = _period_columns(headers)
            if not period_cols:
                warnings.append(f"no_period_columns sheet={sheet_name}")
                continue
            for row_number, values in enumerate(iterator, start=2):
                security_code = _normalized_code(_value_at(values, headers[HEADER_SECURITY_CODE]))
                if target_codes and security_code and security_code not in target_codes:
                    continue
                if target_codes and not security_code:
                    # Keep average/median rows, but drop fully blank trailing rows.
                    metric_label = _clean_text(_value_at(values, headers[HEADER_METRIC]))
                    row_kind = _clean_text(_value_at(values, headers[HEADER_ROW_KIND]))
                    if not metric_label and not row_kind:
                        continue
                metric_label = _clean_text(_value_at(values, headers[HEADER_METRIC]))
                if not metric_label:
                    continue
                periods_by_offset: dict[int, str] = {}
                values_by_offset: dict[int, Any] = {}
                units_by_offset: dict[int, str] = {}
                ratios_by_offset: dict[int, Any] = {}
                ranks_by_offset: dict[int, str] = {}
                for offset, cols in period_cols.items():
                    periods_by_offset[offset] = _clean_text(_value_at(values, cols.get("period")))
                    values_by_offset[offset] = _value_at(values, cols.get("value"))
                    units_by_offset[offset] = _clean_text(_value_at(values, cols.get("unit")))
                    ratios_by_offset[offset] = _value_at(values, cols.get("ratio"))
                    ranks_by_offset[offset] = _clean_text(_value_at(values, cols.get("rank")))
                rows.append(
                    _ActualExcelRow(
                        sheet_name=sheet_name,
                        row_number=row_number,
                        security_code=security_code,
                        company_name=_clean_text(_value_at(values, headers[HEADER_COMPANY_NAME])),
                        industry_33=_clean_text(_value_at(values, headers[HEADER_INDUSTRY])),
                        market=_clean_text(_value_at(values, headers[HEADER_MARKET])),
                        decision_label=_clean_text(_value_at(values, headers[HEADER_DECISION])),
                        row_kind=_clean_text(_value_at(values, headers[HEADER_ROW_KIND])),
                        current_period_end=_clean_text(
                            _value_at(values, headers[HEADER_CURRENT_PERIOD_END])
                        ),
                        metric_label=metric_label,
                        metric_base=_infer_metric_base(sheet_name, metric_label),
                        periods_by_offset=periods_by_offset,
                        values_by_offset=values_by_offset,
                        units_by_offset=units_by_offset,
                        ratios_by_offset=ratios_by_offset,
                        ranks_by_offset=ranks_by_offset,
                    )
                )
    finally:
        workbook.close()
    return rows, warnings


def _index_actual_rows(
    actual_rows: list[_ActualExcelRow],
) -> tuple[
    dict[_AuditRowKey, list[_ActualExcelRow]],
    dict[tuple[str, str, str, str, str], list[_ActualExcelRow]],
    dict[tuple[str, str, str, str, str], list[_ActualExcelRow]],
]:
    by_key: dict[_AuditRowKey, list[_ActualExcelRow]] = {}
    by_without_sheet: dict[tuple[str, str, str, str, str], list[_ActualExcelRow]] = {}
    by_without_decision: dict[tuple[str, str, str, str, str], list[_ActualExcelRow]] = {}
    for row in actual_rows:
        key = _actual_key(row)
        by_key.setdefault(key, []).append(row)
        by_without_sheet.setdefault(_key_without_sheet(key), []).append(row)
        by_without_decision.setdefault(_key_without_decision(key), []).append(row)
    return by_key, by_without_sheet, by_without_decision


def _has_2q_period(row: _ActualExcelRow) -> bool:
    return any(_clean_text(value).startswith("2Q ") for value in row.periods_by_offset.values())


def _add_missing_or_classification_issue(
    issues: list[ExcelAuditIssue],
    expected: MetricExcelRow,
    *,
    by_without_sheet: dict[tuple[str, str, str, str, str], list[_ActualExcelRow]],
    by_without_decision: dict[tuple[str, str, str, str, str], list[_ActualExcelRow]],
    explained_actual_keys: set[_AuditRowKey],
) -> None:
    key = _expected_key(expected)
    sheet_candidates = by_without_sheet.get(_key_without_sheet(key), [])
    if sheet_candidates:
        actual = sheet_candidates[0]
        explained_actual_keys.add(_actual_key(actual))
        issues.append(
            _issue_for_expected(
                expected,
                severity="warning",
                category="classification",
                check_name="sheet_mismatch",
                actual_value=actual.sheet_name,
                message=(
                    "expected row was found on a different sheet: "
                    f"expected={expected.sheet_name} actual={actual.sheet_name}"
                ),
                detail={"actual_row_number": actual.row_number},
            )
        )
        return

    decision_candidates = by_without_decision.get(_key_without_decision(key), [])
    if decision_candidates:
        actual = decision_candidates[0]
        explained_actual_keys.add(_actual_key(actual))
        is_period_mixing = expected.period_scope == "annual" and (
            actual.decision_label == DECISION_QUARTER or _has_2q_period(actual)
        )
        issues.append(
            _issue_for_expected(
                expected,
                severity="critical" if is_period_mixing else "warning",
                category="period_scope" if is_period_mixing else "classification",
                check_name="period_scope_mixing" if is_period_mixing else "decision_label_mismatch",
                actual_value=actual.decision_label,
                message=(
                    "expected row was found with a different decision label: "
                    f"expected={_decision_label_for_row(expected)} actual={actual.decision_label}"
                ),
                detail={"actual_row_number": actual.row_number},
            )
        )
        return

    issues.append(
        _issue_for_expected(
            expected,
            severity="critical",
            category="row_presence",
            check_name="missing_expected_row",
            message="expected metric row was not found in Excel output",
        )
    )


def _compare_period_cells(
    issues: list[ExcelAuditIssue],
    *,
    expected: MetricExcelRow,
    actual: _ActualExcelRow,
    offset: int,
    tolerance: float,
) -> None:
    period_label = PERIOD_LABEL_BY_OFFSET.get(offset, str(offset))
    expected_period = _clean_text(expected.periods_by_offset.get(offset, ""))
    actual_period = _clean_text(actual.periods_by_offset.get(offset, ""))
    if expected_period and actual_period and expected_period != actual_period:
        is_period_mixing = expected.period_scope == "annual" and actual_period.startswith("2Q ")
        issues.append(
            _issue_for_expected(
                expected,
                severity="critical" if is_period_mixing else "warning",
                category="period_scope" if is_period_mixing else "classification",
                check_name="period_scope_mixing" if is_period_mixing else "period_label_mismatch",
                period_label=period_label,
                expected_value=expected_period,
                actual_value=actual_period,
                message=(
                    "period label differs from expected: "
                    f"expected={expected_period} actual={actual_period}"
                ),
                detail={"offset": offset, "actual_row_number": actual.row_number},
            )
        )

    expected_unit = _clean_text(expected.units_by_offset.get(offset, ""))
    actual_unit = _clean_text(actual.units_by_offset.get(offset, ""))
    if expected_unit and actual_unit != expected_unit:
        issues.append(
            _issue_for_expected(
                expected,
                severity="critical",
                category="unit",
                check_name="unit_mismatch",
                period_label=period_label,
                expected_unit=expected_unit,
                actual_unit=actual_unit,
                message=f"unit differs from expected: expected={expected_unit} actual={actual_unit}",
                detail={"offset": offset, "actual_row_number": actual.row_number},
            )
        )

    actual_value = actual.values_by_offset.get(offset)
    if not _is_blank(actual_value):
        expected_value = expected.values_by_offset.get(offset)
        if expected_value is None or not _same_value(expected_value, actual_value, tolerance):
            issues.append(
                _issue_for_expected(
                    expected,
                    severity="critical",
                    category="db_value",
                    check_name="value_mismatch",
                    period_label=period_label,
                    expected_value="" if expected_value is None else expected_value,
                    actual_value=actual_value,
                    expected_unit=expected_unit,
                    actual_unit=actual_unit,
                    message=(
                        "Excel nonblank value differs from DB-derived expected value: "
                        f"expected={expected_value} actual={actual_value}"
                    ),
                    detail={"offset": offset, "actual_row_number": actual.row_number},
                )
            )

    actual_ratio = actual.ratios_by_offset.get(offset)
    if not _is_blank(actual_ratio):
        expected_ratio = expected.ratios_by_offset.get(offset)
        if expected_ratio is None or not _same_value(expected_ratio, actual_ratio, tolerance):
            issues.append(
                _issue_for_expected(
                    expected,
                    severity="critical",
                    category="db_value",
                    check_name="ratio_mismatch",
                    period_label=period_label,
                    expected_value="" if expected_ratio is None else expected_ratio,
                    actual_value=actual_ratio,
                    message=(
                        "Excel nonblank ratio differs from DB-derived expected ratio: "
                        f"expected={expected_ratio} actual={actual_ratio}"
                    ),
                    detail={"offset": offset, "actual_row_number": actual.row_number},
                )
            )

    expected_rank = _clean_text(expected.ranks_by_offset.get(offset, ""))
    actual_rank = _clean_text(actual.ranks_by_offset.get(offset, ""))
    if expected_rank != actual_rank and (expected_rank or actual_rank):
        issues.append(
            _issue_for_expected(
                expected,
                severity="warning",
                category="classification",
                check_name="rank_mismatch",
                period_label=period_label,
                expected_value=expected_rank,
                actual_value=actual_rank,
                message=f"rank differs from expected: expected={expected_rank} actual={actual_rank}",
                detail={"offset": offset, "actual_row_number": actual.row_number},
            )
        )


def _compare_matched_row(
    issues: list[ExcelAuditIssue],
    *,
    expected: MetricExcelRow,
    actual: _ActualExcelRow,
    period_offsets: tuple[int, ...],
    tolerance: float,
) -> None:
    for offset in period_offsets:
        _compare_period_cells(
            issues,
            expected=expected,
            actual=actual,
            offset=offset,
            tolerance=tolerance,
        )


def _is_half_disabled_actual(row: _ActualExcelRow) -> bool:
    return (
        row.decision_label == DECISION_QUARTER
        and row.metric_base in HALF_DISABLED_BASES
        and any(_clean_text(value).startswith("2Q") for value in row.periods_by_offset.values())
    )


def _compare_rows(
    *,
    expected_rows: list[MetricExcelRow],
    actual_rows: list[_ActualExcelRow],
    period_offsets: tuple[int, ...],
    tolerance: float,
) -> list[ExcelAuditIssue]:
    issues: list[ExcelAuditIssue] = []
    by_key, by_without_sheet, by_without_decision = _index_actual_rows(actual_rows)
    explained_actual_keys: set[_AuditRowKey] = set()

    for expected in expected_rows:
        key = _expected_key(expected)
        candidates = by_key.get(key) or []
        if not candidates:
            _add_missing_or_classification_issue(
                issues,
                expected,
                by_without_sheet=by_without_sheet,
                by_without_decision=by_without_decision,
                explained_actual_keys=explained_actual_keys,
            )
            continue
        actual = candidates.pop(0)
        explained_actual_keys.add(_actual_key(actual))
        _compare_matched_row(
            issues,
            expected=expected,
            actual=actual,
            period_offsets=period_offsets,
            tolerance=tolerance,
        )

    for candidates in by_key.values():
        for actual in candidates:
            key = _actual_key(actual)
            if key in explained_actual_keys:
                continue
            if _is_half_disabled_actual(actual):
                issues.append(
                    _issue_for_actual(
                        actual,
                        severity="critical",
                        category="row_presence",
                        check_name="half_disabled_metric_present",
                        message="half/2Q row contains a metric that should be suppressed",
                    )
                )
            else:
                issues.append(
                    _issue_for_actual(
                        actual,
                        severity="warning",
                        category="row_presence",
                        check_name="unexpected_row",
                        message="Excel contains a target row that was not expected from DB-derived rows",
                    )
                )
    return issues


def _counts_by_severity(issues: list[ExcelAuditIssue]) -> dict[str, int]:
    counts: dict[str, int] = {"critical": 0, "warning": 0, "info": 0}
    for issue in issues:
        counts[issue.severity] = counts.get(issue.severity, 0) + 1
    return counts


def _write_rows_sheet(ws, headers: list[str], rows: list[dict[str, Any]]) -> None:
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(fill_type="solid", fgColor="D9EAF7")
    for row in rows:
        ws.append([row.get(header, "") for header in headers])
    for column_cells in ws.columns:
        max_length = max(len(str(cell.value or "")) for cell in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = min(max(max_length + 2, 10), 70)


def _write_excel_report(result: ExcelAuditResult) -> None:
    wb = Workbook()
    summary = wb.active
    summary.title = "Summary"
    _write_rows_sheet(
        summary,
        ["key", "value"],
        [
            {"key": "audit_id", "value": result.audit_id},
            {"key": "generated_at", "value": result.generated_at},
            {"key": "excel_path", "value": str(result.excel_path)},
            {"key": "target_set", "value": result.target_set},
            {"key": "target_count", "value": len(result.targets)},
            {"key": "expected_rows", "value": result.expected_rows},
            {"key": "actual_rows", "value": result.actual_rows},
            {"key": "issue_count", "value": result.issue_count},
            {"key": "critical", "value": result.counts_by_severity.get("critical", 0)},
            {"key": "warning", "value": result.counts_by_severity.get("warning", 0)},
            {"key": "json_path", "value": str(result.json_path)},
        ],
    )

    issue_headers = [
        "severity",
        "category",
        "check_name",
        "security_code",
        "company_name",
        "sheet_name",
        "period_scope",
        "metric_base",
        "metric_label",
        "row_kind",
        "period_label",
        "expected_value",
        "actual_value",
        "expected_unit",
        "actual_unit",
        "source_table",
        "message",
        "detail_json",
    ]
    issue_rows = []
    for issue in sorted(
        result.issues,
        key=lambda item: (
            SEVERITY_ORDER.get(item.severity, 99),
            item.category,
            item.check_name,
            item.security_code,
            item.metric_label,
            item.period_label,
        ),
    ):
        row = issue.to_dict()
        row["detail_json"] = _json_dumps(row.pop("detail"))
        issue_rows.append(row)
    _write_rows_sheet(wb.create_sheet("Issues"), issue_headers, issue_rows)

    _write_rows_sheet(
        wb.create_sheet("Targets"),
        ["target_set", "security_code", "features"],
        [
            {
                "target_set": target.target_set,
                "security_code": target.security_code,
                "features": ",".join(target.features),
            }
            for target in result.targets
        ],
    )
    result.report_excel_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(result.report_excel_path)


def _write_json_report(result: ExcelAuditResult) -> None:
    payload = {
        "audit_id": result.audit_id,
        "generated_at": result.generated_at,
        "excel_path": str(result.excel_path),
        "target_set": result.target_set,
        "targets": [
            {
                "target_set": target.target_set,
                "security_code": target.security_code,
                "features": list(target.features),
            }
            for target in result.targets
        ],
        "expected_rows": result.expected_rows,
        "actual_rows": result.actual_rows,
        "issue_count": result.issue_count,
        "counts_by_severity": result.counts_by_severity,
        "errors": result.errors,
        "warnings": result.warnings,
        "issues": [issue.to_dict() for issue in result.issues],
        "json_path": str(result.json_path),
        "report_excel_path": str(result.report_excel_path),
    }
    result.json_path.parent.mkdir(parents=True, exist_ok=True)
    result.json_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def _build_output_paths(output_dir: Path, generated_at: str) -> tuple[str, Path, Path]:
    timestamp = generated_at.replace("-", "").replace(":", "").replace("T", "_")
    audit_id = f"metric_excel_audit_{timestamp}"
    return (
        audit_id,
        output_dir / f"{audit_id}.json",
        output_dir / f"{audit_id}.xlsx",
    )


def audit_metric_excel(conn: sqlite3.Connection, options: ExcelAuditOptions) -> ExcelAuditResult:
    targets = load_excel_audit_targets(options.target_config_path, target_set=options.target_set)
    target_codes = {target.security_code for target in targets}
    condition = MetricExcelCondition(
        security_codes=sorted(target_codes),
        period_scopes=list(options.period_scopes),
        period_offsets=list(options.period_offsets),
    )
    expected_rows, errors, build_warnings, _preview, _target_companies = build_metric_excel_rows(
        conn,
        condition,
    )
    actual_rows, read_warnings = read_metric_excel_rows(
        options.excel_path,
        target_codes=target_codes,
    )
    issues = _compare_rows(
        expected_rows=expected_rows,
        actual_rows=actual_rows,
        period_offsets=options.period_offsets,
        tolerance=options.value_tolerance,
    )
    generated_at = datetime.now().isoformat(timespec="seconds")
    audit_id, json_path, report_excel_path = _build_output_paths(Path(options.output_dir), generated_at)
    result = ExcelAuditResult(
        audit_id=audit_id,
        generated_at=generated_at,
        excel_path=Path(options.excel_path),
        json_path=json_path,
        report_excel_path=report_excel_path,
        target_set=options.target_set,
        targets=targets,
        expected_rows=len(expected_rows),
        actual_rows=len(actual_rows),
        issues=issues,
        errors=list(errors),
        warnings=list(build_warnings) + read_warnings,
        counts_by_severity=_counts_by_severity(issues),
    )
    _write_json_report(result)
    _write_excel_report(result)
    return result
