from __future__ import annotations

from dataclasses import dataclass, field
from datetime import date, datetime
import json
import math
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from edinet_monitor.config.settings import OPERATION_LOG_ROOT, PROJECT_ROOT
from edinet_monitor.services.metric_excel_audit_service import (
    HEADER_COMPANY_NAME,
    HEADER_CURRENT_PERIOD_END,
    HEADER_DECISION,
    HEADER_INDUSTRY,
    HEADER_MARKET,
    HEADER_METRIC,
    HEADER_ROW_KIND,
    HEADER_SECURITY_CODE,
    HEADER_SECURITY_CODE_ALIASES,
    HEADER_VALUE_KIND,
    PERIOD_PERIOD_SUFFIX,
    PERIOD_RANK_SUFFIX,
    PERIOD_RATIO_SUFFIX,
    PERIOD_UNIT_SUFFIX,
    PERIOD_VALUE_SUFFIX,
)
from edinet_monitor.services.metric_excel_export_service import (
    CONDITION_SHEET,
    PERIOD_LABEL_BY_OFFSET,
    SUMMARY_SHEET,
)


DEFAULT_GOLDEN_MASTER_DIR = PROJECT_ROOT / "config" / "excel" / "golden_master"
DEFAULT_GOLDEN_MASTER_DIFF_OUTPUT_DIR = OPERATION_LOG_ROOT / "golden_master_diff"
EXCLUDED_SUMMARY_KEYS = {"generated_at"}
SEVERITY_ORDER = {"critical": 0, "warning": 1, "info": 2}
REQUIRED_METRIC_HEADERS = {
    HEADER_COMPANY_NAME,
    HEADER_INDUSTRY,
    HEADER_MARKET,
    HEADER_DECISION,
    HEADER_METRIC,
}


@dataclass(frozen=True)
class MetricExcelNormalizedResult:
    excel_path: Path
    output_path: Path
    sheet_count: int
    row_count: int


@dataclass
class GoldenMasterDiffIssue:
    severity: str
    category: str
    check_name: str
    sheet_name: str = ""
    row_key: str = ""
    period_label: str = ""
    field_name: str = ""
    expected_value: Any = ""
    actual_value: Any = ""
    message: str = ""
    detail: dict[str, Any] = field(default_factory=dict)

    def to_dict(self) -> dict[str, Any]:
        return {
            "severity": self.severity,
            "category": self.category,
            "check_name": self.check_name,
            "sheet_name": self.sheet_name,
            "row_key": self.row_key,
            "period_label": self.period_label,
            "field_name": self.field_name,
            "expected_value": self.expected_value,
            "actual_value": self.actual_value,
            "message": self.message,
            "detail": self.detail,
        }


@dataclass(frozen=True)
class GoldenMasterDiffResult:
    comparison_id: str
    generated_at: str
    golden_json_path: Path
    actual_excel_path: Path | None
    actual_json_path: Path
    report_json_path: Path
    report_excel_path: Path
    issues: list[GoldenMasterDiffIssue]
    counts_by_severity: dict[str, int]

    @property
    def issue_count(self) -> int:
        return self.counts_by_severity.get("critical", 0) + self.counts_by_severity.get("warning", 0)


def _clean_text(value: Any) -> str:
    return str(value or "").strip()


def _is_blank(value: Any) -> bool:
    return value is None or str(value).strip() == ""


def _json_value(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.isoformat(timespec="seconds")
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, float):
        if math.isnan(value):
            return None
        return value
    return value


def _json_dumps(value: Any) -> str:
    return json.dumps(value, ensure_ascii=False, sort_keys=True, separators=(",", ":"))


def _load_json(path: str | Path) -> dict[str, Any]:
    loaded = json.loads(Path(path).read_text(encoding="utf-8"))
    if not isinstance(loaded, dict):
        raise ValueError(f"normalized JSON must contain an object: {path}")
    return loaded


def _header_map(header_row: tuple[Any, ...]) -> dict[str, int]:
    return {_clean_text(value): index for index, value in enumerate(header_row) if _clean_text(value)}


def _value_kind_header_index(headers: dict[str, int]) -> int | None:
    if HEADER_VALUE_KIND in headers:
        return headers[HEADER_VALUE_KIND]
    return headers.get(HEADER_ROW_KIND)


def _header_index(headers: dict[str, int], *names: str) -> int | None:
    for name in names:
        if name in headers:
            return headers[name]
    return None


def _value_at(values: tuple[Any, ...], index: int | None) -> Any:
    if index is None or index >= len(values):
        return None
    return values[index]


def _period_columns(headers: dict[str, int]) -> dict[int, dict[str, int]]:
    columns: dict[int, dict[str, int]] = {}
    for offset, label in PERIOD_LABEL_BY_OFFSET.items():
        prefix = f"{label}_"
        mapping: dict[str, int] = {}
        for key, suffix in [
            ("period", PERIOD_PERIOD_SUFFIX),
            ("value", PERIOD_VALUE_SUFFIX),
            ("unit", PERIOD_UNIT_SUFFIX),
            ("ratio", PERIOD_RATIO_SUFFIX),
            ("rank", PERIOD_RANK_SUFFIX),
        ]:
            header = f"{prefix}{suffix}"
            if header in headers:
                mapping[key] = headers[header]
        if mapping:
            columns[offset] = mapping
    return columns


def _ordered_period_offsets(period_columns: dict[int, dict[str, int]]) -> list[int]:
    return sorted(period_columns, key=lambda offset: min(period_columns[offset].values()))


def _normalize_summary_sheet(workbook: Any) -> dict[str, Any]:
    if SUMMARY_SHEET not in workbook.sheetnames:
        return {}
    summary: dict[str, Any] = {}
    ws = workbook[SUMMARY_SHEET]
    for row in ws.iter_rows(values_only=True):
        if not row:
            continue
        key = _clean_text(row[0] if len(row) > 0 else "")
        if not key or key in EXCLUDED_SUMMARY_KEYS:
            continue
        value = row[1] if len(row) > 1 else None
        if _is_blank(value):
            continue
        summary[key] = _json_value(value)
    return summary


def _base_row_values(values: tuple[Any, ...], headers: dict[str, int]) -> dict[str, Any]:
    fields = [
        ("company_name", HEADER_COMPANY_NAME),
        ("industry", HEADER_INDUSTRY),
        ("market", HEADER_MARKET),
        ("decision_label", HEADER_DECISION),
        ("current_period_end", HEADER_CURRENT_PERIOD_END),
        ("metric_label", HEADER_METRIC),
    ]
    row: dict[str, Any] = {}
    security_code = _value_at(values, _header_index(headers, *HEADER_SECURITY_CODE_ALIASES))
    if not _is_blank(security_code):
        row["security_code"] = _json_value(security_code)
    for key, header in fields:
        value = _value_at(values, headers.get(header))
        if not _is_blank(value):
            row[key] = _json_value(value)
    row_kind = _value_at(values, _value_kind_header_index(headers))
    if not _is_blank(row_kind):
        row["row_kind"] = _json_value(row_kind)
    return row


def _period_values(
    values: tuple[Any, ...],
    period_columns: dict[int, dict[str, int]],
) -> list[dict[str, Any]]:
    periods: list[dict[str, Any]] = []
    for offset in _ordered_period_offsets(period_columns):
        columns = period_columns[offset]
        item: dict[str, Any] = {
            "offset": offset,
            "label": PERIOD_LABEL_BY_OFFSET[offset],
        }
        for key in ["period", "value", "unit", "ratio", "rank"]:
            value = _value_at(values, columns.get(key))
            if not _is_blank(value):
                item[key] = _json_value(value)
        if len(item) > 2:
            periods.append(item)
    return periods


def _normalize_metric_sheet(sheet_name: str, ws: Any) -> dict[str, Any] | None:
    iterator = ws.iter_rows(values_only=True)
    try:
        header_row = next(iterator)
    except StopIteration:
        return None
    headers = _header_map(header_row)
    if (
        not REQUIRED_METRIC_HEADERS.issubset(headers)
        or _value_kind_header_index(headers) is None
        or _header_index(headers, *HEADER_SECURITY_CODE_ALIASES) is None
    ):
        return None
    period_columns = _period_columns(headers)
    rows: list[dict[str, Any]] = []
    for values in iterator:
        row = _base_row_values(values, headers)
        if not row.get("metric_label") and not row.get("row_kind"):
            continue
        periods = _period_values(values, period_columns)
        if periods:
            row["periods"] = periods
        rows.append(row)
    return {
        "sheet_name": sheet_name,
        "rows": rows,
    }


def normalize_metric_excel_workbook(excel_path: str | Path) -> dict[str, Any]:
    path = Path(excel_path)
    workbook = load_workbook(path, data_only=True, read_only=True)
    try:
        sheets: list[dict[str, Any]] = []
        for sheet_name in workbook.sheetnames:
            if sheet_name in {SUMMARY_SHEET, CONDITION_SHEET}:
                continue
            normalized = _normalize_metric_sheet(sheet_name, workbook[sheet_name])
            if normalized is not None:
                sheets.append(normalized)
        return {
            "format_version": 1,
            "source_excel_name": path.name,
            "summary": _normalize_summary_sheet(workbook),
            "sheets": sheets,
        }
    finally:
        workbook.close()


def default_normalized_json_path(excel_path: str | Path, output_dir: str | Path) -> Path:
    path = Path(excel_path)
    return Path(output_dir) / f"{path.stem}.normalized.json"


def write_metric_excel_normalized_json(
    excel_path: str | Path,
    *,
    output_path: str | Path | None = None,
    output_dir: str | Path = DEFAULT_GOLDEN_MASTER_DIR,
) -> MetricExcelNormalizedResult:
    normalized = normalize_metric_excel_workbook(excel_path)
    destination = Path(output_path) if output_path is not None else default_normalized_json_path(
        excel_path,
        output_dir,
    )
    destination.parent.mkdir(parents=True, exist_ok=True)
    destination.write_text(
        json.dumps(normalized, ensure_ascii=False, indent=2, sort_keys=True) + "\n",
        encoding="utf-8",
    )
    row_count = sum(len(sheet.get("rows", [])) for sheet in normalized.get("sheets", []))
    return MetricExcelNormalizedResult(
        excel_path=Path(excel_path),
        output_path=destination,
        sheet_count=len(normalized.get("sheets", [])),
        row_count=row_count,
    )


def _row_key_parts(row: dict[str, Any]) -> tuple[str, ...]:
    return (
        _clean_text(row.get("security_code")),
        _clean_text(row.get("decision_label")),
        _clean_text(row.get("row_kind")),
        _clean_text(row.get("metric_label")),
    )


def _row_key_text(sheet_name: str, row: dict[str, Any], occurrence: int) -> str:
    return "|".join((sheet_name, *_row_key_parts(row), str(occurrence)))


def _sheet_rows_by_key(sheet: dict[str, Any]) -> dict[str, dict[str, Any]]:
    rows_by_key: dict[str, dict[str, Any]] = {}
    occurrences: dict[tuple[str, ...], int] = {}
    sheet_name = _clean_text(sheet.get("sheet_name"))
    for row in sheet.get("rows", []):
        if not isinstance(row, dict):
            continue
        parts = _row_key_parts(row)
        occurrence = occurrences.get(parts, 0)
        occurrences[parts] = occurrence + 1
        rows_by_key[_row_key_text(sheet_name, row, occurrence)] = row
    return rows_by_key


def _sheets_by_name(payload: dict[str, Any]) -> dict[str, dict[str, Any]]:
    sheets: dict[str, dict[str, Any]] = {}
    for sheet in payload.get("sheets", []):
        if not isinstance(sheet, dict):
            continue
        name = _clean_text(sheet.get("sheet_name"))
        if name:
            sheets[name] = sheet
    return sheets


def _periods_by_offset(row: dict[str, Any]) -> dict[int, dict[str, Any]]:
    periods: dict[int, dict[str, Any]] = {}
    for period in row.get("periods", []):
        if not isinstance(period, dict):
            continue
        try:
            offset = int(period.get("offset"))
        except (TypeError, ValueError):
            continue
        periods[offset] = period
    return periods


def _equal_value(expected: Any, actual: Any) -> bool:
    if isinstance(expected, (int, float)) and isinstance(actual, (int, float)):
        return abs(float(expected) - float(actual)) <= 1e-9
    return expected == actual


def _diff_issue(
    *,
    check_name: str,
    category: str,
    sheet_name: str = "",
    row_key: str = "",
    period_label: str = "",
    field_name: str = "",
    expected_value: Any = "",
    actual_value: Any = "",
    message: str,
    severity: str = "warning",
    detail: dict[str, Any] | None = None,
) -> GoldenMasterDiffIssue:
    return GoldenMasterDiffIssue(
        severity=severity,
        category=category,
        check_name=check_name,
        sheet_name=sheet_name,
        row_key=row_key,
        period_label=period_label,
        field_name=field_name,
        expected_value=expected_value,
        actual_value=actual_value,
        message=message,
        detail=detail or {},
    )


def _compare_summary(
    expected: dict[str, Any],
    actual: dict[str, Any],
    issues: list[GoldenMasterDiffIssue],
) -> None:
    expected_summary = expected.get("summary", {})
    actual_summary = actual.get("summary", {})
    if not isinstance(expected_summary, dict) or not isinstance(actual_summary, dict):
        issues.append(
            _diff_issue(
                check_name="summary_invalid",
                category="summary",
                severity="critical",
                message="summary payload is not an object",
            )
        )
        return
    for key in sorted(set(expected_summary) | set(actual_summary)):
        expected_value = expected_summary.get(key, "")
        actual_value = actual_summary.get(key, "")
        if _equal_value(expected_value, actual_value):
            continue
        issues.append(
            _diff_issue(
                check_name="summary_value_changed",
                category="summary",
                field_name=key,
                expected_value=expected_value,
                actual_value=actual_value,
                message=f"summary value changed: {key}",
            )
        )


def _compare_row_fields(
    expected_row: dict[str, Any],
    actual_row: dict[str, Any],
    *,
    sheet_name: str,
    row_key: str,
    issues: list[GoldenMasterDiffIssue],
) -> None:
    ignored_fields = {"periods"}
    for field_name in sorted((set(expected_row) | set(actual_row)) - ignored_fields):
        expected_value = expected_row.get(field_name, "")
        actual_value = actual_row.get(field_name, "")
        if _equal_value(expected_value, actual_value):
            continue
        issues.append(
            _diff_issue(
                check_name="row_field_changed",
                category="row",
                sheet_name=sheet_name,
                row_key=row_key,
                field_name=field_name,
                expected_value=expected_value,
                actual_value=actual_value,
                message=f"row field changed: {field_name}",
            )
        )


def _compare_period_fields(
    expected_row: dict[str, Any],
    actual_row: dict[str, Any],
    *,
    sheet_name: str,
    row_key: str,
    issues: list[GoldenMasterDiffIssue],
) -> None:
    expected_periods = _periods_by_offset(expected_row)
    actual_periods = _periods_by_offset(actual_row)
    for offset in sorted(set(expected_periods) | set(actual_periods)):
        expected_period = expected_periods.get(offset)
        actual_period = actual_periods.get(offset)
        label = PERIOD_LABEL_BY_OFFSET.get(offset, str(offset))
        if expected_period is None:
            issues.append(
                _diff_issue(
                    check_name="period_added",
                    category="period",
                    sheet_name=sheet_name,
                    row_key=row_key,
                    period_label=label,
                    actual_value=_json_dumps(actual_period),
                    message=f"period was added: {label}",
                )
            )
            continue
        if actual_period is None:
            issues.append(
                _diff_issue(
                    check_name="period_removed",
                    category="period",
                    sheet_name=sheet_name,
                    row_key=row_key,
                    period_label=label,
                    expected_value=_json_dumps(expected_period),
                    message=f"period was removed: {label}",
                )
            )
            continue
        for field_name in sorted(set(expected_period) | set(actual_period)):
            expected_value = expected_period.get(field_name, "")
            actual_value = actual_period.get(field_name, "")
            if _equal_value(expected_value, actual_value):
                continue
            issues.append(
                _diff_issue(
                    check_name="period_field_changed",
                    category="period",
                    sheet_name=sheet_name,
                    row_key=row_key,
                    period_label=label,
                    field_name=field_name,
                    expected_value=expected_value,
                    actual_value=actual_value,
                    message=f"period field changed: {label} {field_name}",
                )
            )


def _compare_sheet_rows(
    expected_sheet: dict[str, Any],
    actual_sheet: dict[str, Any],
    *,
    sheet_name: str,
    issues: list[GoldenMasterDiffIssue],
) -> None:
    expected_rows = _sheet_rows_by_key(expected_sheet)
    actual_rows = _sheet_rows_by_key(actual_sheet)
    for row_key in sorted(set(expected_rows) | set(actual_rows)):
        expected_row = expected_rows.get(row_key)
        actual_row = actual_rows.get(row_key)
        if expected_row is None:
            issues.append(
                _diff_issue(
                    check_name="row_added",
                    category="row",
                    sheet_name=sheet_name,
                    row_key=row_key,
                    actual_value=_json_dumps(actual_row),
                    message="row was added",
                )
            )
            continue
        if actual_row is None:
            issues.append(
                _diff_issue(
                    check_name="row_removed",
                    category="row",
                    sheet_name=sheet_name,
                    row_key=row_key,
                    expected_value=_json_dumps(expected_row),
                    message="row was removed",
                )
            )
            continue
        _compare_row_fields(
            expected_row,
            actual_row,
            sheet_name=sheet_name,
            row_key=row_key,
            issues=issues,
        )
        _compare_period_fields(
            expected_row,
            actual_row,
            sheet_name=sheet_name,
            row_key=row_key,
            issues=issues,
        )


def compare_metric_excel_normalized_payloads(
    expected: dict[str, Any],
    actual: dict[str, Any],
) -> list[GoldenMasterDiffIssue]:
    issues: list[GoldenMasterDiffIssue] = []
    if expected.get("format_version") != actual.get("format_version"):
        issues.append(
            _diff_issue(
                check_name="format_version_changed",
                category="schema",
                severity="critical",
                expected_value=expected.get("format_version", ""),
                actual_value=actual.get("format_version", ""),
                message="normalized JSON format_version changed",
            )
        )
    _compare_summary(expected, actual, issues)
    expected_sheets = _sheets_by_name(expected)
    actual_sheets = _sheets_by_name(actual)
    for sheet_name in sorted(set(expected_sheets) | set(actual_sheets)):
        expected_sheet = expected_sheets.get(sheet_name)
        actual_sheet = actual_sheets.get(sheet_name)
        if expected_sheet is None:
            issues.append(
                _diff_issue(
                    check_name="sheet_added",
                    category="sheet",
                    sheet_name=sheet_name,
                    message=f"sheet was added: {sheet_name}",
                )
            )
            continue
        if actual_sheet is None:
            issues.append(
                _diff_issue(
                    check_name="sheet_removed",
                    category="sheet",
                    sheet_name=sheet_name,
                    message=f"sheet was removed: {sheet_name}",
                )
            )
            continue
        _compare_sheet_rows(
            expected_sheet,
            actual_sheet,
            sheet_name=sheet_name,
            issues=issues,
        )
    return issues


def _counts_by_severity(issues: list[GoldenMasterDiffIssue]) -> dict[str, int]:
    counts: dict[str, int] = {"critical": 0, "warning": 0, "info": 0}
    for issue in issues:
        counts[issue.severity] = counts.get(issue.severity, 0) + 1
    return counts


def _write_rows_sheet(ws: Any, headers: list[str], rows: list[dict[str, Any]]) -> None:
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(fill_type="solid", fgColor="D9EAF7")
    for row in rows:
        ws.append([row.get(header, "") for header in headers])
    for column_cells in ws.columns:
        max_length = max(len(str(cell.value or "")) for cell in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = min(max(max_length + 2, 10), 80)


def _write_diff_excel_report(result: GoldenMasterDiffResult) -> None:
    workbook = Workbook()
    summary = workbook.active
    summary.title = "Summary"
    _write_rows_sheet(
        summary,
        ["key", "value"],
        [
            {"key": "comparison_id", "value": result.comparison_id},
            {"key": "generated_at", "value": result.generated_at},
            {"key": "golden_json_path", "value": str(result.golden_json_path)},
            {"key": "actual_excel_path", "value": str(result.actual_excel_path or "")},
            {"key": "actual_json_path", "value": str(result.actual_json_path)},
            {"key": "issue_count", "value": result.issue_count},
            {"key": "critical", "value": result.counts_by_severity.get("critical", 0)},
            {"key": "warning", "value": result.counts_by_severity.get("warning", 0)},
        ],
    )
    headers = [
        "severity",
        "category",
        "check_name",
        "sheet_name",
        "row_key",
        "period_label",
        "field_name",
        "expected_value",
        "actual_value",
        "message",
        "detail_json",
    ]
    rows: list[dict[str, Any]] = []
    for issue in sorted(
        result.issues,
        key=lambda item: (
            SEVERITY_ORDER.get(item.severity, 99),
            item.category,
            item.check_name,
            item.sheet_name,
            item.row_key,
            item.period_label,
            item.field_name,
        ),
    ):
        row = issue.to_dict()
        row["detail_json"] = _json_dumps(row.pop("detail"))
        rows.append(row)
    _write_rows_sheet(workbook.create_sheet("Diffs"), headers, rows)
    result.report_excel_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(result.report_excel_path)


def _write_diff_json_report(result: GoldenMasterDiffResult) -> None:
    payload = {
        "comparison_id": result.comparison_id,
        "generated_at": result.generated_at,
        "golden_json_path": str(result.golden_json_path),
        "actual_excel_path": str(result.actual_excel_path or ""),
        "actual_json_path": str(result.actual_json_path),
        "report_json_path": str(result.report_json_path),
        "report_excel_path": str(result.report_excel_path),
        "issue_count": result.issue_count,
        "counts_by_severity": result.counts_by_severity,
        "issues": [issue.to_dict() for issue in result.issues],
    }
    result.report_json_path.parent.mkdir(parents=True, exist_ok=True)
    result.report_json_path.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2, sort_keys=True) + "\n",
        encoding="utf-8",
    )


def _build_comparison_paths(
    output_dir: Path,
    *,
    actual_stem: str,
    generated_at: str,
) -> tuple[str, Path, Path, Path]:
    timestamp = generated_at.replace("-", "").replace(":", "").replace("T", "_")
    comparison_id = f"metric_excel_golden_master_diff_{actual_stem}_{timestamp}"
    return (
        comparison_id,
        output_dir / f"{comparison_id}.actual.normalized.json",
        output_dir / f"{comparison_id}.json",
        output_dir / f"{comparison_id}.xlsx",
    )


def compare_metric_excel_golden_master(
    *,
    golden_json_path: str | Path,
    actual_excel_path: str | Path | None = None,
    actual_json_path: str | Path | None = None,
    output_dir: str | Path = DEFAULT_GOLDEN_MASTER_DIFF_OUTPUT_DIR,
) -> GoldenMasterDiffResult:
    if actual_excel_path is None and actual_json_path is None:
        raise ValueError("actual_excel_path or actual_json_path is required")
    generated_at = datetime.now().isoformat(timespec="seconds")
    actual_stem = Path(actual_excel_path or actual_json_path or "actual").stem
    comparison_id, generated_actual_json, report_json, report_excel = _build_comparison_paths(
        Path(output_dir),
        actual_stem=actual_stem,
        generated_at=generated_at,
    )
    if actual_json_path is None:
        actual_json_path = generated_actual_json
        write_metric_excel_normalized_json(
            actual_excel_path,
            output_path=actual_json_path,
        )
    expected = _load_json(golden_json_path)
    actual = _load_json(actual_json_path)
    issues = compare_metric_excel_normalized_payloads(expected, actual)
    result = GoldenMasterDiffResult(
        comparison_id=comparison_id,
        generated_at=generated_at,
        golden_json_path=Path(golden_json_path),
        actual_excel_path=Path(actual_excel_path) if actual_excel_path is not None else None,
        actual_json_path=Path(actual_json_path),
        report_json_path=report_json,
        report_excel_path=report_excel,
        issues=issues,
        counts_by_severity=_counts_by_severity(issues),
    )
    _write_diff_json_report(result)
    _write_diff_excel_report(result)
    return result
