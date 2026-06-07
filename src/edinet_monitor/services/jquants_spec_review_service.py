from __future__ import annotations

from dataclasses import dataclass, field
from datetime import datetime
import json
from pathlib import Path
import sqlite3
import subprocess
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from edinet_monitor.config.settings import OPERATION_LOG_ROOT, PROJECT_ROOT
from edinet_monitor.services.jquants_official_cli_compare_service import (
    OfficialCliCompareError,
    Runner,
    _resolve_official_cli,
    run_jquants_official_cli_compare,
)


SUPPORTED_SPEC_REVIEW_ENDPOINTS = ("fins.summary", "eq.daily")
UNSUPPORTED_STANDARD_ENDPOINTS = {"fins.details"}
DEFAULT_JQUANTS_SCHEMA_BASELINE_DIR = PROJECT_ROOT / "config" / "jquants" / "schema_baseline"
DEFAULT_JQUANTS_SPEC_REVIEW_OUTPUT_DIR = OPERATION_LOG_ROOT / "jquants_spec_review"
SEVERITY_ORDER = {"critical": 0, "warning": 1, "info": 2}


@dataclass(frozen=True)
class JQuantsSpecReviewOptions:
    endpoints: tuple[str, ...] = SUPPORTED_SPEC_REVIEW_ENDPOINTS
    date_value: str = ""
    code: str = ""
    baseline_dir: Path = DEFAULT_JQUANTS_SCHEMA_BASELINE_DIR
    output_dir: Path = DEFAULT_JQUANTS_SPEC_REVIEW_OUTPUT_DIR
    official_cli: str | None = None
    update_baseline: bool = False


@dataclass
class JQuantsSpecReviewIssue:
    severity: str
    category: str
    check_name: str
    endpoint: str = ""
    field_name: str = ""
    row_key: str = ""
    expected_value: Any = ""
    actual_value: Any = ""
    message: str = ""
    detail: dict[str, Any] = field(default_factory=dict)

    def to_dict(self) -> dict[str, Any]:
        return {
            "severity": self.severity,
            "category": self.category,
            "check_name": self.check_name,
            "endpoint": self.endpoint,
            "field_name": self.field_name,
            "row_key": self.row_key,
            "expected_value": self.expected_value,
            "actual_value": self.actual_value,
            "message": self.message,
            "detail": self.detail,
        }


@dataclass(frozen=True)
class JQuantsSpecReviewResult:
    review_id: str
    generated_at: str
    status: str
    json_path: Path
    excel_path: Path
    counts_by_severity: dict[str, int]
    issues: list[JQuantsSpecReviewIssue]
    summary: dict[str, Any]
    schema_diff: list[dict[str, Any]]
    raw_compare: list[dict[str, Any]]
    official_cli_commands: list[dict[str, Any]]

    @property
    def issue_count(self) -> int:
        return self.counts_by_severity.get("critical", 0) + self.counts_by_severity.get("warning", 0)


def _now() -> str:
    return datetime.now().isoformat(timespec="seconds")


def _timestamp_for_filename(generated_at: str) -> str:
    return generated_at.replace("-", "").replace(":", "").replace("T", "_")


def _safe_endpoint(endpoint: str) -> str:
    return endpoint.replace(".", "_")


def _baseline_path(baseline_dir: Path, endpoint: str) -> Path:
    return baseline_dir / f"{endpoint}.schema.json"


def _jsonify(value: Any) -> Any:
    if isinstance(value, Path):
        return str(value)
    if isinstance(value, dict):
        return {str(key): _jsonify(item) for key, item in value.items()}
    if isinstance(value, (list, tuple)):
        return [_jsonify(item) for item in value]
    return value


def _json_dumps(value: Any) -> str:
    return json.dumps(_jsonify(value), ensure_ascii=False, sort_keys=True, separators=(",", ":"))


def _run_command(runner: Runner | None, command: list[str]) -> subprocess.CompletedProcess[str]:
    if runner is not None:
        return runner(command)
    return subprocess.run(
        command,
        capture_output=True,
        text=True,
        encoding="utf-8",
        errors="replace",
        check=False,
    )


def _load_json_text(text: str) -> Any:
    return json.loads(text)


def _field_name_from_item(item: dict[str, Any]) -> str:
    for key in ("name", "field", "field_name", "column", "id", "Name", "Field", "FieldName"):
        value = str(item.get(key) or "").strip()
        if value:
            return value
    return ""


def _field_type_from_item(item: dict[str, Any]) -> str:
    for key in ("type", "data_type", "dtype", "Type", "DataType"):
        if key in item and item[key] not in (None, ""):
            value = item[key]
            if isinstance(value, (dict, list)):
                return _json_dumps(value)
            return str(value).strip()
    return ""


def _normalize_field_list(value: Any) -> dict[str, dict[str, Any]]:
    fields: dict[str, dict[str, Any]] = {}
    if isinstance(value, list):
        for item in value:
            if isinstance(item, str):
                name = item.strip()
                field_type = ""
                raw: Any = item
            elif isinstance(item, dict):
                name = _field_name_from_item(item)
                field_type = _field_type_from_item(item)
                raw = item
            else:
                continue
            if name:
                fields[name] = {"name": name, "type": field_type, "raw": raw}
    elif isinstance(value, dict):
        for name, item in value.items():
            clean_name = str(name or "").strip()
            if not clean_name:
                continue
            if isinstance(item, dict):
                field_type = _field_type_from_item(item)
                raw = item
            elif item in (None, ""):
                field_type = ""
                raw = item
            else:
                field_type = str(item).strip()
                raw = item
            fields[clean_name] = {"name": clean_name, "type": field_type, "raw": raw}
    return fields


def normalize_schema_fields(payload: Any, endpoint: str) -> dict[str, dict[str, Any]]:
    """Extract a flexible field map from official jquants schema JSON."""
    if isinstance(payload, dict):
        if isinstance(payload.get("fields"), (list, dict)):
            return _normalize_field_list(payload["fields"])
        if isinstance(payload.get("columns"), (list, dict)):
            return _normalize_field_list(payload["columns"])
        if isinstance(payload.get("properties"), dict):
            return _normalize_field_list(payload["properties"])
        if endpoint in payload:
            nested = normalize_schema_fields(payload[endpoint], endpoint)
            if nested:
                return nested
        for key in ("schema", "data", "items"):
            if key in payload:
                nested = normalize_schema_fields(payload[key], endpoint)
                if nested:
                    return nested
        return _normalize_field_list(payload)
    return _normalize_field_list(payload)


def _load_baseline_fields(path: Path, endpoint: str) -> dict[str, dict[str, Any]] | None:
    if not path.exists():
        return None
    payload = _load_json_text(path.read_text(encoding="utf-8"))
    if isinstance(payload, dict) and isinstance(payload.get("fields"), list):
        return _normalize_field_list(payload["fields"])
    return normalize_schema_fields(payload, endpoint)


def _write_baseline(
    *,
    path: Path,
    endpoint: str,
    generated_at: str,
    command: list[str],
    jquants_version: str,
    schema_payload: Any,
    fields: dict[str, dict[str, Any]],
) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    payload = {
        "endpoint": endpoint,
        "generated_at": generated_at,
        "jquants_version": jquants_version,
        "command": command,
        "schema_payload": schema_payload,
        "fields": [fields[name] for name in sorted(fields)],
    }
    path.write_text(json.dumps(_jsonify(payload), ensure_ascii=False, indent=2), encoding="utf-8")


def _compare_schema_fields(
    *,
    endpoint: str,
    baseline_fields: dict[str, dict[str, Any]],
    current_fields: dict[str, dict[str, Any]],
    issues: list[JQuantsSpecReviewIssue],
) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    baseline_names = set(baseline_fields)
    current_names = set(current_fields)
    for name in sorted(baseline_names - current_names):
        row = {
            "endpoint": endpoint,
            "change_type": "removed",
            "field_name": name,
            "expected_type": baseline_fields[name].get("type", ""),
            "actual_type": "",
            "severity": "critical",
        }
        rows.append(row)
        issues.append(
            JQuantsSpecReviewIssue(
                severity="critical",
                category="schema",
                check_name="removed_field",
                endpoint=endpoint,
                field_name=name,
                expected_value=baseline_fields[name].get("type", ""),
                message=f"{endpoint} schema field was removed: {name}",
                detail=row,
            )
        )
    for name in sorted(current_names - baseline_names):
        row = {
            "endpoint": endpoint,
            "change_type": "added",
            "field_name": name,
            "expected_type": "",
            "actual_type": current_fields[name].get("type", ""),
            "severity": "warning",
        }
        rows.append(row)
        issues.append(
            JQuantsSpecReviewIssue(
                severity="warning",
                category="schema",
                check_name="added_field",
                endpoint=endpoint,
                field_name=name,
                actual_value=current_fields[name].get("type", ""),
                message=f"{endpoint} schema field was added: {name}",
                detail=row,
            )
        )
    for name in sorted(baseline_names & current_names):
        expected_type = str(baseline_fields[name].get("type") or "")
        actual_type = str(current_fields[name].get("type") or "")
        if expected_type and actual_type and expected_type != actual_type:
            row = {
                "endpoint": endpoint,
                "change_type": "type_changed",
                "field_name": name,
                "expected_type": expected_type,
                "actual_type": actual_type,
                "severity": "critical",
            }
            rows.append(row)
            issues.append(
                JQuantsSpecReviewIssue(
                    severity="critical",
                    category="schema",
                    check_name="field_type_changed",
                    endpoint=endpoint,
                    field_name=name,
                    expected_value=expected_type,
                    actual_value=actual_type,
                    message=f"{endpoint} schema field type changed: {name}",
                    detail=row,
                )
            )
    return rows


def _counts_by_severity(issues: list[JQuantsSpecReviewIssue]) -> dict[str, int]:
    counts = {"critical": 0, "warning": 0, "info": 0}
    for issue in issues:
        severity = issue.severity if issue.severity in counts else "warning"
        counts[severity] += 1
    return counts


def _status_from_counts(counts: dict[str, int]) -> str:
    if counts.get("critical", 0) or counts.get("warning", 0):
        return "review_required"
    return "ok"


def _excel_value(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, (dict, list, tuple)):
        return _json_dumps(value)
    if isinstance(value, Path):
        return str(value)
    return value


def _write_table(workbook: Workbook, *, title: str, headers: list[str], rows: list[dict[str, Any]]) -> None:
    worksheet = workbook.create_sheet(title)
    worksheet.append(headers)
    fill = PatternFill("solid", fgColor="D9EAF7")
    for cell in worksheet[1]:
        cell.font = Font(bold=True)
        cell.fill = fill
    for row in rows:
        worksheet.append([_excel_value(row.get(header, "")) for header in headers])
    for column_cells in worksheet.columns:
        max_length = max(len(str(cell.value or "")) for cell in column_cells)
        worksheet.column_dimensions[column_cells[0].column_letter].width = min(max(max_length + 2, 10), 80)


def _write_json_report(result: JQuantsSpecReviewResult) -> None:
    payload = {
        "review_id": result.review_id,
        "generated_at": result.generated_at,
        "status": result.status,
        "summary": result.summary,
        "issues": [issue.to_dict() for issue in result.issues],
        "schema_diff": result.schema_diff,
        "raw_compare": result.raw_compare,
        "official_cli_commands": result.official_cli_commands,
        "json_path": result.json_path,
        "excel_path": result.excel_path,
    }
    result.json_path.parent.mkdir(parents=True, exist_ok=True)
    result.json_path.write_text(json.dumps(_jsonify(payload), ensure_ascii=False, indent=2), encoding="utf-8")


def _write_excel_report(result: JQuantsSpecReviewResult) -> None:
    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "Summary"
    summary_sheet.append(["key", "value"])
    fill = PatternFill("solid", fgColor="D9EAF7")
    for cell in summary_sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = fill
    for key, value in result.summary.items():
        summary_sheet.append([key, _excel_value(value)])
    summary_sheet.column_dimensions["A"].width = 34
    summary_sheet.column_dimensions["B"].width = 90

    _write_table(
        workbook,
        title="Issues",
        headers=[
            "severity",
            "category",
            "check_name",
            "endpoint",
            "field_name",
            "row_key",
            "expected_value",
            "actual_value",
            "message",
            "detail",
        ],
        rows=[issue.to_dict() for issue in sorted(result.issues, key=lambda item: (SEVERITY_ORDER.get(item.severity, 99), item.endpoint, item.category, item.check_name, item.field_name))],
    )
    _write_table(
        workbook,
        title="Schema_Diff",
        headers=["endpoint", "change_type", "field_name", "expected_type", "actual_type", "severity", "baseline_path"],
        rows=result.schema_diff,
    )
    _write_table(
        workbook,
        title="Raw_Compare",
        headers=[
            "endpoint",
            "status",
            "official_rows",
            "db_rows",
            "matched_rows",
            "missing_in_db",
            "extra_in_db",
            "field_diff_rows",
            "diff_count",
            "txt_path",
            "tsv_path",
            "error",
        ],
        rows=result.raw_compare,
    )
    _write_table(
        workbook,
        title="Official_CLI",
        headers=["kind", "endpoint", "command", "returncode", "stdout", "stderr", "output_path"],
        rows=result.official_cli_commands,
    )
    result.excel_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(result.excel_path)


def _record_command(
    rows: list[dict[str, Any]],
    *,
    kind: str,
    endpoint: str = "",
    command: list[str],
    completed: subprocess.CompletedProcess[str] | None = None,
    output_path: Path | None = None,
) -> None:
    rows.append(
        {
            "kind": kind,
            "endpoint": endpoint,
            "command": " ".join(command),
            "returncode": "" if completed is None else completed.returncode,
            "stdout": "" if completed is None else str(completed.stdout or "")[:2000],
            "stderr": "" if completed is None else str(completed.stderr or "")[:2000],
            "output_path": output_path or "",
        }
    )


def build_jquants_spec_review(
    conn: sqlite3.Connection,
    options: JQuantsSpecReviewOptions,
    *,
    runner: Runner | None = None,
) -> JQuantsSpecReviewResult:
    generated_at = _now()
    review_id = f"jquants_spec_review_{_timestamp_for_filename(generated_at)}"
    output_dir = Path(options.output_dir)
    json_path = output_dir / f"{review_id}.json"
    excel_path = output_dir / f"{review_id}.xlsx"
    issues: list[JQuantsSpecReviewIssue] = []
    schema_diff: list[dict[str, Any]] = []
    raw_compare: list[dict[str, Any]] = []
    official_cli_commands: list[dict[str, Any]] = []
    endpoints = tuple(dict.fromkeys(endpoint.strip().lower() for endpoint in options.endpoints if endpoint.strip()))
    version_text = ""

    cli_available = True
    try:
        executable = _resolve_official_cli(options.official_cli)
    except OfficialCliCompareError as exc:
        cli_available = False
        issues.append(
            JQuantsSpecReviewIssue(
                severity="critical",
                category="official_cli",
                check_name="official_cli_not_found",
                message=str(exc),
            )
        )
        executable = options.official_cli or "jquants"

    version_command = [executable, "--version"]
    if cli_available or runner is not None:
        completed_version = _run_command(runner, version_command)
        _record_command(official_cli_commands, kind="version", command=version_command, completed=completed_version)
        if completed_version.returncode == 0:
            version_text = str(completed_version.stdout or "").strip()
        else:
            issues.append(
                JQuantsSpecReviewIssue(
                    severity="warning",
                    category="official_cli",
                    check_name="official_version_unavailable",
                    message=str(completed_version.stderr or "jquants --version failed").strip(),
                )
            )
    else:
        _record_command(official_cli_commands, kind="version", command=version_command)

    for endpoint in endpoints:
        if endpoint in UNSUPPORTED_STANDARD_ENDPOINTS:
            issues.append(
                JQuantsSpecReviewIssue(
                    severity="critical",
                    category="plan_policy",
                    check_name="unsupported_endpoint",
                    endpoint=endpoint,
                    message=f"{endpoint} is unsupported for the current Standard-plan operation.",
                )
            )
            continue
        if endpoint not in SUPPORTED_SPEC_REVIEW_ENDPOINTS:
            issues.append(
                JQuantsSpecReviewIssue(
                    severity="critical",
                    category="plan_policy",
                    check_name="unsupported_endpoint",
                    endpoint=endpoint,
                    message=f"Unsupported endpoint: {endpoint}. Supported: {', '.join(SUPPORTED_SPEC_REVIEW_ENDPOINTS)}",
                )
            )
            continue

        if not cli_available and runner is None:
            raw_compare.append(
                {
                    "endpoint": endpoint,
                    "status": "error",
                    "official_rows": 0,
                    "db_rows": 0,
                    "matched_rows": 0,
                    "missing_in_db": 0,
                    "extra_in_db": 0,
                    "field_diff_rows": 0,
                    "diff_count": 0,
                    "txt_path": "",
                    "tsv_path": "",
                    "error": "jquants executable was not found.",
                }
            )
            continue

        schema_command = [executable, "--output", "json", "schema", endpoint]
        completed_schema = _run_command(runner, schema_command)
        _record_command(official_cli_commands, kind="schema", endpoint=endpoint, command=schema_command, completed=completed_schema)
        schema_payload: Any | None = None
        current_fields: dict[str, dict[str, Any]] = {}
        schema_text_path = output_dir / f"jquants_schema_{_safe_endpoint(endpoint)}_{_timestamp_for_filename(generated_at)}.txt"
        if completed_schema.returncode != 0:
            schema_text_path.parent.mkdir(parents=True, exist_ok=True)
            schema_text_path.write_text(str(completed_schema.stdout or completed_schema.stderr or ""), encoding="utf-8")
            _record_command(official_cli_commands, kind="schema_text", endpoint=endpoint, command=schema_command, output_path=schema_text_path)
            issues.append(
                JQuantsSpecReviewIssue(
                    severity="critical",
                    category="schema",
                    check_name="official_schema_unavailable",
                    endpoint=endpoint,
                    message=str(completed_schema.stderr or "official jquants schema command failed").strip(),
                    detail={"text_path": str(schema_text_path)},
                )
            )
        else:
            try:
                schema_payload = _load_json_text(str(completed_schema.stdout or ""))
                current_fields = normalize_schema_fields(schema_payload, endpoint)
            except json.JSONDecodeError as exc:
                schema_text_path.parent.mkdir(parents=True, exist_ok=True)
                schema_text_path.write_text(str(completed_schema.stdout or ""), encoding="utf-8")
                _record_command(official_cli_commands, kind="schema_text", endpoint=endpoint, command=schema_command, output_path=schema_text_path)
                issues.append(
                    JQuantsSpecReviewIssue(
                        severity="critical",
                        category="schema",
                        check_name="official_schema_unavailable",
                        endpoint=endpoint,
                        message=f"official jquants schema did not return valid JSON: {exc}",
                        detail={"text_path": str(schema_text_path)},
                    )
                )

        baseline_path = _baseline_path(Path(options.baseline_dir), endpoint)
        if current_fields:
            baseline_fields = _load_baseline_fields(baseline_path, endpoint)
            if baseline_fields is None:
                row = {
                    "endpoint": endpoint,
                    "change_type": "missing_baseline",
                    "field_name": "",
                    "expected_type": "",
                    "actual_type": "",
                    "severity": "warning",
                    "baseline_path": str(baseline_path),
                }
                schema_diff.append(row)
                issues.append(
                    JQuantsSpecReviewIssue(
                        severity="warning",
                        category="baseline",
                        check_name="missing_baseline",
                        endpoint=endpoint,
                        message=f"Schema baseline does not exist: {baseline_path}",
                        detail=row,
                    )
                )
            else:
                diff_rows = _compare_schema_fields(
                    endpoint=endpoint,
                    baseline_fields=baseline_fields,
                    current_fields=current_fields,
                    issues=issues,
                )
                for row in diff_rows:
                    row["baseline_path"] = str(baseline_path)
                schema_diff.extend(diff_rows)
            if options.update_baseline and schema_payload is not None:
                _write_baseline(
                    path=baseline_path,
                    endpoint=endpoint,
                    generated_at=generated_at,
                    command=schema_command,
                    jquants_version=version_text,
                    schema_payload=schema_payload,
                    fields=current_fields,
                )

        try:
            compare_result = run_jquants_official_cli_compare(
                conn,
                endpoint=endpoint,
                date_value=options.date_value or None,
                code=options.code or None,
                output_dir=output_dir,
                official_cli=executable,
                runner=runner,
            )
            raw_row = {
                "endpoint": endpoint,
                "status": "ok" if compare_result.diff_count == 0 and compare_result.missing_in_db == 0 and compare_result.extra_in_db == 0 else "review_required",
                "official_rows": compare_result.official_rows,
                "db_rows": compare_result.db_rows,
                "matched_rows": compare_result.matched_rows,
                "missing_in_db": compare_result.missing_in_db,
                "extra_in_db": compare_result.extra_in_db,
                "field_diff_rows": compare_result.field_diff_rows,
                "diff_count": compare_result.diff_count,
                "txt_path": str(compare_result.txt_path),
                "tsv_path": str(compare_result.tsv_path),
                "error": "",
            }
            raw_compare.append(raw_row)
            if compare_result.missing_in_db:
                issues.append(
                    JQuantsSpecReviewIssue(
                        severity="critical",
                        category="raw_compare",
                        check_name="missing_in_db",
                        endpoint=endpoint,
                        actual_value=compare_result.missing_in_db,
                        message=f"{endpoint} official CLI rows are missing in local raw DB.",
                        detail=raw_row,
                    )
                )
            if compare_result.extra_in_db:
                issues.append(
                    JQuantsSpecReviewIssue(
                        severity="warning",
                        category="raw_compare",
                        check_name="extra_in_db",
                        endpoint=endpoint,
                        actual_value=compare_result.extra_in_db,
                        message=f"{endpoint} local raw DB has rows not found in official CLI output.",
                        detail=raw_row,
                    )
                )
            if compare_result.diff_count:
                issues.append(
                    JQuantsSpecReviewIssue(
                        severity="warning",
                        category="raw_compare",
                        check_name="field_value_diff",
                        endpoint=endpoint,
                        actual_value=compare_result.diff_count,
                        message=f"{endpoint} official CLI output differs from local raw DB fields.",
                        detail=raw_row,
                    )
                )
        except OfficialCliCompareError as exc:
            raw_row = {
                "endpoint": endpoint,
                "status": "error",
                "official_rows": 0,
                "db_rows": 0,
                "matched_rows": 0,
                "missing_in_db": 0,
                "extra_in_db": 0,
                "field_diff_rows": 0,
                "diff_count": 0,
                "txt_path": "",
                "tsv_path": "",
                "error": str(exc),
            }
            raw_compare.append(raw_row)
            issues.append(
                JQuantsSpecReviewIssue(
                    severity="critical",
                    category="raw_compare",
                    check_name="official_cli_failed",
                    endpoint=endpoint,
                    message=str(exc),
                    detail=raw_row,
                )
            )

    counts = _counts_by_severity(issues)
    status = _status_from_counts(counts)
    summary = {
        "review_id": review_id,
        "generated_at": generated_at,
        "status": status,
        "endpoints": list(endpoints),
        "date": options.date_value,
        "code": options.code,
        "baseline_dir": Path(options.baseline_dir),
        "output_dir": output_dir,
        "update_baseline": bool(options.update_baseline),
        "jquants_version": version_text,
        "critical": counts.get("critical", 0),
        "warning": counts.get("warning", 0),
        "info": counts.get("info", 0),
        "issue_count": counts.get("critical", 0) + counts.get("warning", 0),
    }
    result = JQuantsSpecReviewResult(
        review_id=review_id,
        generated_at=generated_at,
        status=status,
        json_path=json_path,
        excel_path=excel_path,
        counts_by_severity=counts,
        issues=issues,
        summary=summary,
        schema_diff=schema_diff,
        raw_compare=raw_compare,
        official_cli_commands=official_cli_commands,
    )
    _write_json_report(result)
    _write_excel_report(result)
    return result
