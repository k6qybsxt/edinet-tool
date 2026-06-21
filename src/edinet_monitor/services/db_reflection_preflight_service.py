from __future__ import annotations

from dataclasses import dataclass, field
from datetime import datetime
import json
from pathlib import Path
import re
import sqlite3
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from edinet_monitor.config.settings import OPERATION_LOG_ROOT
from edinet_monitor.services.prevention_catalog_service import (
    ACTIVE_PREVENTION_STATUSES,
    DEFAULT_PREVENTION_CATALOG_PATH,
    PreventionCatalogItem,
    load_prevention_catalog,
)


DEFAULT_DB_REFLECTION_PREFLIGHT_OUTPUT_DIR = OPERATION_LOG_ROOT / "db_reflection_preflight"
SEVERITY_ORDER = {"critical": 0, "warning": 1, "info": 2}
HEAVY_DB_SIZE_WARNING_BYTES = 100 * 1024 * 1024 * 1024
HEAVY_DB_SIZE_WARNING_GB = 100
LONG_DATE_RANGE_WARNING_DAYS = 365 * 5
METRIC_TABLES_WITH_CALC_STATUS = (
    "derived_metrics",
    "market_derived_metrics",
    "quarter_standalone_metrics",
    "segment_metrics",
    "jquants_financial_metrics",
    "industry_aggregate_metrics",
)
SCOPE_OPTIONS = (
    "--date-from",
    "--date-to",
    "--date",
    "--codes",
    "--code",
    "--doc-id",
    "--item-id",
    "--security-code",
    "--form-codes",
    "--industry-33",
    "--target-date",
)
CONTROL_SCOPE_OPTIONS = (
    "--batch-size",
    "--db-insert-chunk-size",
    "--db-doc-id-chunk-size",
    "--limit",
)
TARGET_COUNT_KEYWORDS = (
    "count(",
    "target_count",
    "target_total",
    "target_docs",
    "\u5bfe\u8c61\u4ef6\u6570",
    "\u4ef6\u6570",
)
DANGEROUS_FULL_SCOPE_MARKERS = (
    "--run-all",
    "--download-run-all",
    "run_all",
    "--codes all",
    "--codes=all",
    "--periods all",
    "--periods=all",
    "--period-scopes all",
    "--period-scopes=all",
    "\u5168\u4ef6",
)


@dataclass(frozen=True)
class DbReflectionPreflightItem:
    item_id: int
    title: str
    category: str
    status: str
    description: str
    notes: str
    required_commands: list[str]
    verification_sql: list[str]
    source_key: str
    created_at: str
    updated_at: str

    def to_dict(self) -> dict[str, Any]:
        return {
            "item_id": self.item_id,
            "title": self.title,
            "category": self.category,
            "status": self.status,
            "description": self.description,
            "notes": self.notes,
            "required_commands": self.required_commands,
            "verification_sql": self.verification_sql,
            "source_key": self.source_key,
            "created_at": self.created_at,
            "updated_at": self.updated_at,
        }


@dataclass
class DbReflectionPreflightIssue:
    severity: str
    category: str
    check_name: str
    item_id: int | str = ""
    title: str = ""
    message: str = ""
    detail: dict[str, Any] = field(default_factory=dict)

    def to_dict(self) -> dict[str, Any]:
        return {
            "severity": self.severity,
            "category": self.category,
            "check_name": self.check_name,
            "item_id": self.item_id,
            "title": self.title,
            "message": self.message,
            "detail": self.detail,
        }


@dataclass(frozen=True)
class DbReflectionPreflightOptions:
    db_path: Path | None = None
    catalog_path: Path = DEFAULT_PREVENTION_CATALOG_PATH
    item_id: int | None = None
    output_dir: Path = DEFAULT_DB_REFLECTION_PREFLIGHT_OUTPUT_DIR
    pipeline_failure_policy: str = "report_only"
    guard_cli_name: str = ""
    command_names: tuple[str, ...] = ()
    catalog_areas: tuple[str, ...] = ()
    catalog_triggers: tuple[str, ...] = ()


@dataclass(frozen=True)
class DbReflectionPreflightResult:
    preflight_id: str
    generated_at: str
    status: str
    json_path: Path
    excel_path: Path
    summary: dict[str, Any]
    pending_items: list[DbReflectionPreflightItem]
    catalog_items: list[PreventionCatalogItem]
    issues: list[DbReflectionPreflightIssue]
    counts_by_severity: dict[str, int]

    @property
    def issue_count(self) -> int:
        return self.counts_by_severity.get("critical", 0) + self.counts_by_severity.get("warning", 0)


def _now() -> str:
    return datetime.now().isoformat(timespec="seconds")


def _timestamp_for_filename(generated_at: str) -> str:
    return generated_at.replace("-", "").replace(":", "").replace("T", "_")


def _jsonify(value: Any) -> Any:
    if isinstance(value, Path):
        return str(value)
    if isinstance(value, dict):
        return {str(key): _jsonify(item) for key, item in value.items()}
    if isinstance(value, (list, tuple)):
        return [_jsonify(item) for item in value]
    return value


def _json_dumps(value: Any) -> str:
    return json.dumps(_jsonify(value), ensure_ascii=False, sort_keys=True, separators=(",", ":"))


def _table_exists(conn: sqlite3.Connection, table_name: str) -> bool:
    row = conn.execute(
        """
        SELECT 1
        FROM sqlite_master
        WHERE type = 'table' AND name = ?
        """,
        (table_name,),
    ).fetchone()
    return row is not None


def _table_columns(conn: sqlite3.Connection, table_name: str) -> list[str]:
    rows = conn.execute(f"PRAGMA table_info({table_name})").fetchall()
    return [str(row["name"] if isinstance(row, sqlite3.Row) else row[1]) for row in rows]


def _row_value(row: sqlite3.Row, column: str, default: Any = "") -> Any:
    if column not in row.keys():
        return default
    return row[column]


def _parse_json_list(value: Any, *, item_id: int | str, title: str, column_name: str, issues: list[DbReflectionPreflightIssue]) -> list[str]:
    if value in (None, ""):
        return []
    try:
        parsed = json.loads(str(value))
    except Exception as exc:
        issues.append(
            DbReflectionPreflightIssue(
                severity="critical",
                category="db_reflection_item",
                check_name=f"{column_name}_invalid_json",
                item_id=item_id,
                title=title,
                message=f"{column_name} is not valid JSON: {exc}",
            )
        )
        return []
    if not isinstance(parsed, list):
        issues.append(
            DbReflectionPreflightIssue(
                severity="critical",
                category="db_reflection_item",
                check_name=f"{column_name}_not_list",
                item_id=item_id,
                title=title,
                message=f"{column_name} must contain a JSON list.",
            )
        )
        return []
    return [str(item).strip() for item in parsed if str(item).strip()]


def _load_pending_items(
    conn: sqlite3.Connection,
    *,
    item_id: int | None,
    issues: list[DbReflectionPreflightIssue],
) -> list[DbReflectionPreflightItem]:
    conn.row_factory = sqlite3.Row
    if not _table_exists(conn, "db_reflection_items"):
        if item_id is not None:
            issues.append(
                DbReflectionPreflightIssue(
                    severity="critical",
                    category="db_reflection_table",
                    check_name="db_reflection_items_missing",
                    item_id=item_id,
                    message="db_reflection_items table does not exist.",
                )
            )
        return []

    columns = set(_table_columns(conn, "db_reflection_items"))
    required_columns = {"title", "category", "description", "required_commands_json", "verification_sql_json"}
    missing_columns = sorted(required_columns - columns)
    if "item_id" not in columns and "id" not in columns:
        missing_columns.append("item_id")
    for column in missing_columns:
        issues.append(
            DbReflectionPreflightIssue(
                severity="critical",
                category="db_reflection_table",
                check_name="required_column_missing",
                message=f"db_reflection_items is missing required column: {column}",
            )
        )
    if missing_columns:
        return []

    id_column = "item_id" if "item_id" in columns else "id"
    clauses: list[str] = []
    params: list[Any] = []
    if "status" in columns:
        clauses.append("LOWER(COALESCE(status, 'pending')) NOT IN ('complete', 'completed', 'done')")
    if item_id is not None:
        clauses.append(f"{id_column} = ?")
        params.append(item_id)
    where_sql = f"WHERE {' AND '.join(clauses)}" if clauses else ""
    order_columns = [column for column in ("created_at", id_column) if column in columns]
    order_sql = f"ORDER BY {', '.join(order_columns)}" if order_columns else ""
    rows = conn.execute(f"SELECT * FROM db_reflection_items {where_sql} {order_sql}", params).fetchall()
    if item_id is not None and not rows:
        issues.append(
            DbReflectionPreflightIssue(
                severity="critical",
                category="db_reflection_item",
                check_name="item_not_found",
                item_id=item_id,
                message=f"pending db_reflection_items row was not found: {item_id}",
            )
        )

    items: list[DbReflectionPreflightItem] = []
    for row in rows:
        current_item_id = int(_row_value(row, id_column, 0) or 0)
        title = str(_row_value(row, "title", "") or "")
        required_commands = _parse_json_list(
            _row_value(row, "required_commands_json", ""),
            item_id=current_item_id,
            title=title,
            column_name="required_commands_json",
            issues=issues,
        )
        verification_sql = _parse_json_list(
            _row_value(row, "verification_sql_json", ""),
            item_id=current_item_id,
            title=title,
            column_name="verification_sql_json",
            issues=issues,
        )
        items.append(
            DbReflectionPreflightItem(
                item_id=current_item_id,
                title=title,
                category=str(_row_value(row, "category", "") or ""),
                status=str(_row_value(row, "status", "pending") or "pending"),
                description=str(_row_value(row, "description", "") or ""),
                notes=str(_row_value(row, "notes", "") or ""),
                required_commands=required_commands,
                verification_sql=verification_sql,
                source_key=str(_row_value(row, "source_key", "") or ""),
                created_at=str(_row_value(row, "created_at", "") or ""),
                updated_at=str(_row_value(row, "updated_at", "") or ""),
            )
        )
    return items


def _lower_text(*parts: Any) -> str:
    return "\n".join(str(part or "") for part in parts).lower()


def _normalized_tuple(values: tuple[str, ...] | list[str] | None) -> tuple[str, ...]:
    if not values:
        return ()
    return tuple(str(value).strip() for value in values if str(value).strip())


def _filter_pending_items_by_command(
    items: list[DbReflectionPreflightItem],
    command_names: tuple[str, ...],
) -> list[DbReflectionPreflightItem]:
    normalized_command_names = tuple(name.lower() for name in _normalized_tuple(command_names))
    if not normalized_command_names:
        return list(items)
    matched_items: list[DbReflectionPreflightItem] = []
    for item in items:
        command_text = "\n".join(item.required_commands).lower()
        if any(command_name in command_text for command_name in normalized_command_names):
            matched_items.append(item)
    return matched_items


def _has_target_count_check(item: DbReflectionPreflightItem) -> bool:
    text = _lower_text(item.description, item.notes, *item.required_commands, *item.verification_sql)
    return any(keyword in text for keyword in TARGET_COUNT_KEYWORDS)


def _is_dangerous_full_scope_command(command: str) -> bool:
    lower = command.lower()
    return any(marker in lower for marker in DANGEROUS_FULL_SCOPE_MARKERS)


def _has_scope_option(command: str) -> bool:
    lower = command.lower()
    return any(option in lower for option in SCOPE_OPTIONS)


def _has_control_scope_option(command: str) -> bool:
    lower = command.lower()
    return any(option in lower for option in CONTROL_SCOPE_OPTIONS)


def _option_value(command: str, option: str) -> str:
    escaped = re.escape(option)
    patterns = (
        rf"{escaped}=([^\s]+)",
        rf"{escaped}\s+([^\s]+)",
    )
    for pattern in patterns:
        match = re.search(pattern, command, flags=re.IGNORECASE)
        if match:
            return match.group(1).strip().strip("\"'")
    return ""


def _date_range_days(command: str) -> int | None:
    date_from = _option_value(command, "--date-from")
    date_to = _option_value(command, "--date-to")
    if not date_from or not date_to:
        return None
    try:
        start = datetime.strptime(date_from, "%Y-%m-%d")
        end = datetime.strptime(date_to, "%Y-%m-%d")
    except ValueError:
        return None
    return abs((end - start).days)


def _db_file_size_bytes(db_path: Path | None) -> int | None:
    if db_path is None:
        return None
    try:
        path = Path(db_path)
        if not path.exists():
            return None
        return path.stat().st_size
    except OSError:
        return None


def _normalized_sql(sql: str) -> str:
    return re.sub(r"\s+", " ", str(sql or "").strip().lower())


def _metric_table_with_calc_status(sql: str) -> str:
    normalized = _normalized_sql(sql)
    for table_name in METRIC_TABLES_WITH_CALC_STATUS:
        if table_name in normalized:
            return table_name
    return ""


def _add_item_preflight_issues(
    item: DbReflectionPreflightItem,
    issues: list[DbReflectionPreflightIssue],
    *,
    db_size_bytes: int | None = None,
) -> None:
    has_target_count = _has_target_count_check(item)
    has_full_scope_command = any(_is_dangerous_full_scope_command(command) for command in item.required_commands)
    if not item.required_commands:
        issues.append(
            DbReflectionPreflightIssue(
                severity="critical",
                category="db_reflection_item",
                check_name="missing_required_command",
                item_id=item.item_id,
                title=item.title,
                message="required command is missing.",
            )
        )
    if not item.verification_sql:
        issues.append(
            DbReflectionPreflightIssue(
                severity="critical",
                category="db_reflection_item",
                check_name="missing_verification_sql",
                item_id=item.item_id,
                title=item.title,
                message="verification SQL is missing.",
            )
        )
    if not has_target_count and not has_full_scope_command:
        issues.append(
            DbReflectionPreflightIssue(
                severity="warning",
                category="db_reflection_item",
                check_name="target_count_not_confirmed",
                item_id=item.item_id,
                title=item.title,
                message="target count confirmation was not found in commands, SQL, description, or notes.",
            )
        )

    for command in item.required_commands:
        full_scope = _is_dangerous_full_scope_command(command)
        if full_scope and not has_target_count:
            issues.append(
                DbReflectionPreflightIssue(
                    severity="critical",
                    category="db_reflection_command",
                    check_name="full_scope_command_without_target_count",
                    item_id=item.item_id,
                    title=item.title,
                    message="full-scope command was found without target count confirmation.",
                    detail={"command": command},
                )
            )
        if _is_dangerous_full_scope_command(command) and not _has_scope_option(command):
            issues.append(
                DbReflectionPreflightIssue(
                    severity="warning",
                    category="db_reflection_command",
                    check_name="full_scope_command_without_scope",
                    item_id=item.item_id,
                    title=item.title,
                    message="full-scope command marker was found without an explicit scope option.",
                    detail={"command": command},
                )
            )
        if full_scope and db_size_bytes is not None and db_size_bytes >= HEAVY_DB_SIZE_WARNING_BYTES:
            issues.append(
                DbReflectionPreflightIssue(
                    severity="warning",
                    category="db_performance",
                    check_name="large_db_full_scope_command",
                    item_id=item.item_id,
                    title=item.title,
                    message="full-scope command will run against a large DB; confirm the target range and runtime before apply.",
                    detail={
                        "command": command,
                        "db_size_bytes": db_size_bytes,
                        "threshold_bytes": HEAVY_DB_SIZE_WARNING_BYTES,
                    },
                )
            )
        range_days = _date_range_days(command)
        if (
            range_days is not None
            and range_days >= LONG_DATE_RANGE_WARNING_DAYS
            and db_size_bytes is not None
            and db_size_bytes >= HEAVY_DB_SIZE_WARNING_BYTES
        ):
            issues.append(
                DbReflectionPreflightIssue(
                    severity="warning",
                    category="db_performance",
                    check_name="large_db_long_date_range",
                    item_id=item.item_id,
                    title=item.title,
                    message="long date range command will run against a large DB; consider splitting by period.",
                    detail={
                        "command": command,
                        "date_range_days": range_days,
                        "threshold_days": LONG_DATE_RANGE_WARNING_DAYS,
                        "db_size_bytes": db_size_bytes,
                    },
                )
            )
        if full_scope and not _has_control_scope_option(command):
            issues.append(
                DbReflectionPreflightIssue(
                    severity="warning",
                    category="db_performance",
                    check_name="broad_scope_without_batch_or_limit",
                    item_id=item.item_id,
                    title=item.title,
                    message="broad-scope command has no batch, chunk, or limit option.",
                    detail={"command": command},
                )
            )

    for sql in item.verification_sql:
        normalized = _normalized_sql(sql)
        if not normalized.startswith(("select ", "with ")):
            issues.append(
                DbReflectionPreflightIssue(
                    severity="warning",
                    category="verification_sql",
                    check_name="verification_sql_not_read_only_select",
                    item_id=item.item_id,
                    title=item.title,
                    message="verification SQL should start with SELECT or WITH.",
                    detail={"sql": sql},
                )
            )
        table_name = _metric_table_with_calc_status(sql)
        if table_name:
            if "calc_status" not in normalized or ("'ok'" not in normalized and '"ok"' not in normalized):
                issues.append(
                    DbReflectionPreflightIssue(
                        severity="warning",
                        category="verification_sql",
                        check_name="verification_calc_status_missing",
                        item_id=item.item_id,
                        title=item.title,
                        message=f"verification SQL for {table_name} should check calc_status = 'ok'.",
                        detail={"sql": sql},
                    )
                )
            if "value_num" not in normalized or "value_num is not null" not in normalized:
                issues.append(
                    DbReflectionPreflightIssue(
                        severity="warning",
                        category="verification_sql",
                        check_name="verification_value_num_condition_missing",
                        item_id=item.item_id,
                        title=item.title,
                        message=f"verification SQL for {table_name} should check value_num IS NOT NULL.",
                        detail={"sql": sql},
                    )
                )


def _load_db_reflection_catalog_items(
    catalog_path: Path,
    issues: list[DbReflectionPreflightIssue],
    *,
    areas: tuple[str, ...] = (),
    triggers: tuple[str, ...] = (),
) -> list[PreventionCatalogItem]:
    try:
        items = load_prevention_catalog(catalog_path)
    except Exception as exc:
        issues.append(
            DbReflectionPreflightIssue(
                severity="critical",
                category="prevention_catalog",
                check_name="catalog_load_failed",
                message=str(exc),
                detail={"catalog_path": str(catalog_path)},
            )
        )
        return []
    active_statuses = set(ACTIVE_PREVENTION_STATUSES)
    filter_areas = set(_normalized_tuple(areas))
    filter_triggers = set(_normalized_tuple(triggers))
    if not filter_areas and not filter_triggers:
        filter_areas = {"db_reflection"}
        filter_triggers = {"pre_db_reflection"}
    return [
        item
        for item in items
        if item.status in active_statuses
        and (
            bool(filter_areas.intersection(item.areas))
            or bool(filter_triggers.intersection(item.triggers))
        )
    ]


def _counts_by_severity(issues: list[DbReflectionPreflightIssue]) -> dict[str, int]:
    counts = {"critical": 0, "warning": 0, "info": 0}
    for issue in issues:
        severity = issue.severity if issue.severity in counts else "warning"
        counts[severity] += 1
    return counts


def _status_from_counts(counts: dict[str, int]) -> str:
    if counts.get("critical", 0) or counts.get("warning", 0):
        return "review_required"
    return "ok"


def _excel_value(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, (dict, list, tuple)):
        return _json_dumps(value)
    if isinstance(value, Path):
        return str(value)
    return value


def _write_rows_sheet(workbook: Workbook, *, title: str, headers: list[str], rows: list[dict[str, Any]]) -> None:
    worksheet = workbook.create_sheet(title)
    worksheet.append(headers)
    fill = PatternFill("solid", fgColor="D9EAF7")
    for cell in worksheet[1]:
        cell.font = Font(bold=True)
        cell.fill = fill
    for row in rows:
        worksheet.append([_excel_value(row.get(header, "")) for header in headers])
    for column_cells in worksheet.columns:
        max_length = max(len(str(cell.value or "")) for cell in column_cells)
        worksheet.column_dimensions[column_cells[0].column_letter].width = min(max(max_length + 2, 10), 80)


def _write_json_report(result: DbReflectionPreflightResult) -> None:
    payload = {
        "preflight_id": result.preflight_id,
        "generated_at": result.generated_at,
        "status": result.status,
        "summary": result.summary,
        "issues": [issue.to_dict() for issue in result.issues],
        "pending_items": [item.to_dict() for item in result.pending_items],
        "catalog_items": [item.to_dict() for item in result.catalog_items],
        "json_path": result.json_path,
        "excel_path": result.excel_path,
    }
    result.json_path.parent.mkdir(parents=True, exist_ok=True)
    result.json_path.write_text(json.dumps(_jsonify(payload), ensure_ascii=False, indent=2), encoding="utf-8")


def _write_excel_report(result: DbReflectionPreflightResult) -> None:
    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "Summary"
    summary_sheet.append(["key", "value"])
    fill = PatternFill("solid", fgColor="D9EAF7")
    for cell in summary_sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = fill
    for key, value in result.summary.items():
        summary_sheet.append([key, _excel_value(value)])
    summary_sheet.column_dimensions["A"].width = 34
    summary_sheet.column_dimensions["B"].width = 80

    _write_rows_sheet(
        workbook,
        title="Issues",
        headers=["severity", "category", "check_name", "item_id", "title", "message", "detail"],
        rows=[
            issue.to_dict()
            for issue in sorted(
                result.issues,
                key=lambda row: (SEVERITY_ORDER.get(row.severity, 99), str(row.item_id), row.category, row.check_name),
            )
        ],
    )
    _write_rows_sheet(
        workbook,
        title="Pending_Items",
        headers=[
            "item_id",
            "title",
            "category",
            "status",
            "required_commands",
            "verification_sql",
            "description",
            "notes",
            "source_key",
            "created_at",
            "updated_at",
        ],
        rows=[item.to_dict() for item in result.pending_items],
    )
    _write_rows_sheet(
        workbook,
        title="Catalog_Items",
        headers=[
            "id",
            "severity",
            "status",
            "title",
            "areas",
            "triggers",
            "problem",
            "prevention",
            "review_points",
        ],
        rows=[item.to_dict() for item in result.catalog_items],
    )
    result.excel_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(result.excel_path)


def build_db_reflection_preflight(
    conn: sqlite3.Connection,
    options: DbReflectionPreflightOptions,
) -> DbReflectionPreflightResult:
    generated_at = _now()
    preflight_id = f"db_reflection_preflight_{_timestamp_for_filename(generated_at)}"
    output_dir = Path(options.output_dir)
    json_path = output_dir / f"{preflight_id}.json"
    excel_path = output_dir / f"{preflight_id}.xlsx"
    issues: list[DbReflectionPreflightIssue] = []
    load_issues: list[DbReflectionPreflightIssue] = []
    db_size_bytes = _db_file_size_bytes(Path(options.db_path) if options.db_path else None)
    all_pending_items = _load_pending_items(conn, item_id=options.item_id, issues=load_issues)
    command_names = _normalized_tuple(options.command_names)
    pending_items = _filter_pending_items_by_command(all_pending_items, command_names)
    matched_item_ids = {item.item_id for item in pending_items}
    for issue in load_issues:
        if not command_names or issue.item_id in ("", 0) or issue.item_id in matched_item_ids:
            issues.append(issue)
    catalog_areas = _normalized_tuple(options.catalog_areas)
    catalog_triggers = _normalized_tuple(options.catalog_triggers)
    catalog_items = _load_db_reflection_catalog_items(
        Path(options.catalog_path),
        issues,
        areas=catalog_areas,
        triggers=catalog_triggers,
    )
    for item in pending_items:
        _add_item_preflight_issues(item, issues, db_size_bytes=db_size_bytes)

    counts = _counts_by_severity(issues)
    status = _status_from_counts(counts)
    db_reflection_blocked = (
        options.pipeline_failure_policy == "block_on_critical"
        and counts.get("critical", 0) > 0
    )
    summary = {
        "preflight_id": preflight_id,
        "generated_at": generated_at,
        "status": status,
        "pipeline_failure_policy": options.pipeline_failure_policy,
        "db_reflection_blocked": db_reflection_blocked,
        "guard_cli_name": options.guard_cli_name,
        "command_names": command_names,
        "catalog_areas": catalog_areas,
        "catalog_triggers": catalog_triggers,
        "db_path": options.db_path or "",
        "db_size_bytes": db_size_bytes if db_size_bytes is not None else "",
        "db_size_gb": round(db_size_bytes / (1024 * 1024 * 1024), 3) if db_size_bytes is not None else "",
        "heavy_db_size_warning_threshold_gb": HEAVY_DB_SIZE_WARNING_GB,
        "long_date_range_warning_days": LONG_DATE_RANGE_WARNING_DAYS,
        "catalog_path": Path(options.catalog_path),
        "item_id": options.item_id or "",
        "pending_count": len(all_pending_items),
        "matched_pending_count": len(pending_items),
        "catalog_item_count": len(catalog_items),
        "critical": counts.get("critical", 0),
        "warning": counts.get("warning", 0),
        "info": counts.get("info", 0),
    }
    result = DbReflectionPreflightResult(
        preflight_id=preflight_id,
        generated_at=generated_at,
        status=status,
        json_path=json_path,
        excel_path=excel_path,
        summary=summary,
        pending_items=pending_items,
        catalog_items=catalog_items,
        issues=issues,
        counts_by_severity=counts,
    )
    _write_json_report(result)
    _write_excel_report(result)
    return result
