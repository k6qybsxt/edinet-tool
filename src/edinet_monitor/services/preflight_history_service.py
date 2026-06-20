from __future__ import annotations

from dataclasses import dataclass
from datetime import datetime, timedelta
import json
from pathlib import Path
import sqlite3
from typing import Any
from uuid import uuid4

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from edinet_monitor.config.settings import OPERATION_LOG_ROOT
from edinet_monitor.db.migrations import apply_schema_migrations
from edinet_monitor.services.db_reflection_preflight_service import DbReflectionPreflightResult


DEFAULT_PREFLIGHT_HISTORY_REVIEW_OUTPUT_DIR = OPERATION_LOG_ROOT / "preflight_history_review"
PREFLIGHT_HISTORY_STATUSES = {
    "blocked",
    "passed",
    "passed_with_warnings",
    "completed",
    "report_only",
}


@dataclass(frozen=True)
class PreflightHistorySaveResult:
    preflight_id: str
    status: str
    history_saved: bool
    issue_count: int


@dataclass(frozen=True)
class PreflightHistoryReviewOptions:
    db_path: Path | None = None
    days: int = 7
    cli_name: str = ""
    blocked_only: bool = False
    warnings_only: bool = False
    output_dir: Path = DEFAULT_PREFLIGHT_HISTORY_REVIEW_OUTPUT_DIR
    limit_preview: int = 20


@dataclass(frozen=True)
class PreflightHistoryReviewResult:
    review_id: str
    generated_at: str
    status: str
    json_path: Path
    excel_path: Path
    summary: dict[str, Any]
    runs: list[dict[str, Any]]
    issues: list[dict[str, Any]]


@dataclass(frozen=True)
class PreflightHistoryCleanupOptions:
    keep_days: int = 180
    keep_critical_days: int = 730
    apply: bool = False
    vacuum: bool = False
    limit_preview: int = 20


@dataclass(frozen=True)
class PreflightHistoryCleanupResult:
    mode: str
    keep_days: int
    keep_critical_days: int
    target_count: int
    deleted_count: int
    issue_deleted_count: int
    vacuumed: bool
    preview: list[dict[str, Any]]


def _now() -> str:
    return datetime.now().isoformat(timespec="seconds")


def _timestamp_for_filename(generated_at: str) -> str:
    return generated_at.replace("-", "").replace(":", "").replace("T", "_")


def _json_dumps(value: Any) -> str:
    return json.dumps(_jsonify(value), ensure_ascii=False, sort_keys=True, separators=(",", ":"))


def _json_loads(value: Any) -> Any:
    if value in (None, ""):
        return None
    try:
        return json.loads(str(value))
    except Exception:
        return str(value)


def _jsonify(value: Any) -> Any:
    if isinstance(value, Path):
        return str(value)
    if isinstance(value, dict):
        return {str(key): _jsonify(item) for key, item in value.items()}
    if isinstance(value, (list, tuple)):
        return [_jsonify(item) for item in value]
    return value


def _excel_value(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, (dict, list, tuple)):
        return _json_dumps(value)
    if isinstance(value, Path):
        return str(value)
    return value


def _table_exists(conn: sqlite3.Connection, table_name: str) -> bool:
    row = conn.execute(
        """
        SELECT 1
        FROM sqlite_master
        WHERE type = 'table' AND name = ?
        """,
        (table_name,),
    ).fetchone()
    return row is not None


def _row_to_dict(row: sqlite3.Row) -> dict[str, Any]:
    return {key: row[key] for key in row.keys()}


def _normalize_run_row(row: sqlite3.Row | dict[str, Any]) -> dict[str, Any]:
    data = _row_to_dict(row) if isinstance(row, sqlite3.Row) else dict(row)
    data["command_names_json"] = _json_loads(data.get("command_names_json")) or []
    data["db_reflection_blocked"] = bool(data.get("db_reflection_blocked"))
    return data


def _status_for_preflight_result(result: DbReflectionPreflightResult, *, report_only: bool = False) -> str:
    if report_only:
        return "report_only"
    if bool(result.summary.get("db_reflection_blocked", False)):
        return "blocked"
    if result.counts_by_severity.get("warning", 0) > 0:
        return "passed_with_warnings"
    return "passed"


def _validate_status(status: str) -> str:
    normalized = str(status or "").strip()
    if normalized not in PREFLIGHT_HISTORY_STATUSES:
        raise ValueError(f"Unsupported preflight history status: {status}")
    return normalized


def save_preflight_history(
    conn: sqlite3.Connection,
    result: DbReflectionPreflightResult,
    *,
    status: str | None = None,
    report_only: bool = False,
) -> PreflightHistorySaveResult:
    conn.row_factory = sqlite3.Row
    apply_schema_migrations(conn)
    resolved_status = _validate_status(status or _status_for_preflight_result(result, report_only=report_only))
    now = _now()
    summary = result.summary
    command_names = tuple(str(item) for item in (summary.get("command_names") or ()) if str(item).strip())
    cli_name = str(summary.get("guard_cli_name") or "")
    if not cli_name and command_names:
        cli_name = command_names[0]
    conn.execute(
        """
        INSERT INTO preflight_runs (
            preflight_id, generated_at, cli_name, command_names_json,
            pipeline_failure_policy, db_reflection_blocked, status,
            pending_count, matched_pending_count, critical_count, warning_count,
            json_path, excel_path, completed_at, created_at, updated_at
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, NULL, ?, ?)
        ON CONFLICT(preflight_id) DO UPDATE SET
            generated_at = excluded.generated_at,
            cli_name = excluded.cli_name,
            command_names_json = excluded.command_names_json,
            pipeline_failure_policy = excluded.pipeline_failure_policy,
            db_reflection_blocked = excluded.db_reflection_blocked,
            status = excluded.status,
            pending_count = excluded.pending_count,
            matched_pending_count = excluded.matched_pending_count,
            critical_count = excluded.critical_count,
            warning_count = excluded.warning_count,
            json_path = excluded.json_path,
            excel_path = excluded.excel_path,
            completed_at = CASE
                WHEN excluded.status = 'completed' THEN COALESCE(preflight_runs.completed_at, excluded.updated_at)
                ELSE NULL
            END,
            updated_at = excluded.updated_at
        """,
        (
            result.preflight_id,
            result.generated_at,
            cli_name,
            _json_dumps(command_names),
            str(summary.get("pipeline_failure_policy") or ""),
            1 if bool(summary.get("db_reflection_blocked", False)) else 0,
            resolved_status,
            int(summary.get("pending_count") or len(result.pending_items)),
            int(summary.get("matched_pending_count") or len(result.pending_items)),
            int(result.counts_by_severity.get("critical", 0)),
            int(result.counts_by_severity.get("warning", 0)),
            str(result.json_path),
            str(result.excel_path),
            now,
            now,
        ),
    )
    conn.execute(
        """
        DELETE FROM preflight_run_issues
        WHERE preflight_id = ?
        """,
        (result.preflight_id,),
    )
    issue_rows = [
        (
            result.preflight_id,
            issue.severity,
            issue.category,
            issue.check_name,
            str(issue.item_id),
            issue.title,
            issue.message,
            _json_dumps(issue.detail),
            now,
        )
        for issue in result.issues
    ]
    conn.executemany(
        """
        INSERT INTO preflight_run_issues (
            preflight_id, severity, category, check_name, item_id,
            title, message, detail_json, created_at
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        issue_rows,
    )
    conn.commit()
    return PreflightHistorySaveResult(
        preflight_id=result.preflight_id,
        status=resolved_status,
        history_saved=True,
        issue_count=len(issue_rows),
    )


def mark_preflight_history_completed(
    conn: sqlite3.Connection,
    *,
    preflight_id: str,
) -> bool:
    conn.row_factory = sqlite3.Row
    apply_schema_migrations(conn)
    now = _now()
    cursor = conn.execute(
        """
        UPDATE preflight_runs
        SET status = 'completed',
            completed_at = COALESCE(completed_at, ?),
            updated_at = ?
        WHERE preflight_id = ?
        """,
        (now, now, preflight_id),
    )
    conn.commit()
    return cursor.rowcount > 0


def _review_where(options: PreflightHistoryReviewOptions, generated_at: str) -> tuple[str, list[Any]]:
    clauses: list[str] = []
    params: list[Any] = []
    if int(options.days) > 0:
        cutoff = datetime.fromisoformat(generated_at) - timedelta(days=int(options.days))
        clauses.append("generated_at >= ?")
        params.append(cutoff.isoformat(timespec="seconds"))
    if str(options.cli_name or "").strip():
        clauses.append("cli_name = ?")
        params.append(str(options.cli_name).strip())
    if options.blocked_only:
        clauses.append("(db_reflection_blocked = 1 OR status = 'blocked')")
    if options.warnings_only:
        clauses.append("warning_count > 0")
    where_sql = f"WHERE {' AND '.join(clauses)}" if clauses else ""
    return where_sql, params


def _cli_counts(runs: list[dict[str, Any]]) -> dict[str, int]:
    counts: dict[str, int] = {}
    for run in runs:
        cli_name = str(run.get("cli_name") or "(blank)")
        counts[cli_name] = counts.get(cli_name, 0) + 1
    return counts


def _write_rows_sheet(workbook: Workbook, *, title: str, headers: list[str], rows: list[dict[str, Any]]) -> None:
    worksheet = workbook.create_sheet(title)
    worksheet.append(headers)
    fill = PatternFill("solid", fgColor="D9EAF7")
    for cell in worksheet[1]:
        cell.font = Font(bold=True)
        cell.fill = fill
    for row in rows:
        worksheet.append([_excel_value(row.get(header, "")) for header in headers])
    for column_cells in worksheet.columns:
        max_length = max(len(str(cell.value or "")) for cell in column_cells)
        worksheet.column_dimensions[column_cells[0].column_letter].width = min(max(max_length + 2, 10), 80)


def _write_review_json(result: PreflightHistoryReviewResult) -> None:
    payload = {
        "review_id": result.review_id,
        "generated_at": result.generated_at,
        "status": result.status,
        "summary": result.summary,
        "runs": result.runs,
        "issues": result.issues,
        "json_path": result.json_path,
        "excel_path": result.excel_path,
    }
    result.json_path.parent.mkdir(parents=True, exist_ok=True)
    result.json_path.write_text(json.dumps(_jsonify(payload), ensure_ascii=False, indent=2), encoding="utf-8")


def _write_review_excel(result: PreflightHistoryReviewResult) -> None:
    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "Summary"
    summary_sheet.append(["key", "value"])
    fill = PatternFill("solid", fgColor="D9EAF7")
    for cell in summary_sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = fill
    for key, value in result.summary.items():
        summary_sheet.append([key, _excel_value(value)])
    summary_sheet.column_dimensions["A"].width = 34
    summary_sheet.column_dimensions["B"].width = 80
    _write_rows_sheet(
        workbook,
        title="Runs",
        headers=[
            "preflight_id",
            "generated_at",
            "cli_name",
            "command_names_json",
            "status",
            "db_reflection_blocked",
            "critical_count",
            "warning_count",
            "pending_count",
            "matched_pending_count",
            "completed_at",
            "json_path",
            "excel_path",
        ],
        rows=result.runs,
    )
    _write_rows_sheet(
        workbook,
        title="Issues",
        headers=[
            "preflight_id",
            "severity",
            "category",
            "check_name",
            "item_id",
            "title",
            "message",
            "detail_json",
        ],
        rows=result.issues,
    )
    result.excel_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(result.excel_path)


def build_preflight_history_review(
    conn: sqlite3.Connection,
    *,
    options: PreflightHistoryReviewOptions,
) -> PreflightHistoryReviewResult:
    conn.row_factory = sqlite3.Row
    generated_at = _now()
    review_id = f"preflight_history_review_{_timestamp_for_filename(generated_at)}_{uuid4().hex[:8]}"
    output_dir = Path(options.output_dir)
    json_path = output_dir / f"{review_id}.json"
    excel_path = output_dir / f"{review_id}.xlsx"
    if not _table_exists(conn, "preflight_runs"):
        summary = {
            "review_id": review_id,
            "generated_at": generated_at,
            "status": "no_table",
            "db_path": options.db_path or "",
            "run_count": 0,
            "blocked_count": 0,
            "warning_run_count": 0,
            "incomplete_count": 0,
            "cli_counts": {},
        }
        result = PreflightHistoryReviewResult(
            review_id=review_id,
            generated_at=generated_at,
            status="no_table",
            json_path=json_path,
            excel_path=excel_path,
            summary=summary,
            runs=[],
            issues=[],
        )
        _write_review_json(result)
        _write_review_excel(result)
        return result

    where_sql, params = _review_where(options, generated_at)
    run_rows = conn.execute(
        f"""
        SELECT *
        FROM preflight_runs
        {where_sql}
        ORDER BY generated_at DESC, id DESC
        """,
        params,
    ).fetchall()
    runs = [_normalize_run_row(row) for row in run_rows]
    preflight_ids = [str(run["preflight_id"]) for run in runs]
    issues: list[dict[str, Any]] = []
    if preflight_ids and _table_exists(conn, "preflight_run_issues"):
        placeholders = ",".join("?" for _ in preflight_ids)
        issue_rows = conn.execute(
            f"""
            SELECT *
            FROM preflight_run_issues
            WHERE preflight_id IN ({placeholders})
            ORDER BY
                CASE severity
                    WHEN 'critical' THEN 0
                    WHEN 'warning' THEN 1
                    WHEN 'info' THEN 2
                    ELSE 9
                END,
                preflight_id,
                category,
                check_name
            LIMIT ?
            """,
            [*preflight_ids, max(int(options.limit_preview), 0)],
        ).fetchall()
        for issue_row in issue_rows:
            issue = _row_to_dict(issue_row)
            issue["detail_json"] = _json_loads(issue.get("detail_json"))
            issues.append(issue)
    blocked_count = sum(1 for run in runs if bool(run.get("db_reflection_blocked")) or run.get("status") == "blocked")
    warning_run_count = sum(1 for run in runs if int(run.get("warning_count") or 0) > 0)
    incomplete_count = sum(
        1
        for run in runs
        if run.get("status") in {"passed", "passed_with_warnings"} and not run.get("completed_at")
    )
    status = "review_required" if blocked_count or warning_run_count or incomplete_count else "ok"
    summary = {
        "review_id": review_id,
        "generated_at": generated_at,
        "status": status,
        "db_path": options.db_path or "",
        "days": int(options.days),
        "cli_name": options.cli_name,
        "blocked_only": bool(options.blocked_only),
        "warnings_only": bool(options.warnings_only),
        "run_count": len(runs),
        "blocked_count": blocked_count,
        "warning_run_count": warning_run_count,
        "incomplete_count": incomplete_count,
        "critical_count": sum(int(run.get("critical_count") or 0) for run in runs),
        "warning_count": sum(int(run.get("warning_count") or 0) for run in runs),
        "cli_counts": _cli_counts(runs),
    }
    result = PreflightHistoryReviewResult(
        review_id=review_id,
        generated_at=generated_at,
        status=status,
        json_path=json_path,
        excel_path=excel_path,
        summary=summary,
        runs=runs,
        issues=issues,
    )
    _write_review_json(result)
    _write_review_excel(result)
    return result


def _cleanup_targets(
    conn: sqlite3.Connection,
    *,
    options: PreflightHistoryCleanupOptions,
    generated_at: str,
) -> list[dict[str, Any]]:
    if not _table_exists(conn, "preflight_runs"):
        return []
    keep_cutoff = datetime.fromisoformat(generated_at) - timedelta(days=max(int(options.keep_days), 0))
    critical_cutoff = datetime.fromisoformat(generated_at) - timedelta(days=max(int(options.keep_critical_days), 0))
    rows = conn.execute(
        """
        SELECT *
        FROM preflight_runs
        WHERE (
            (db_reflection_blocked = 1 OR critical_count > 0 OR status = 'blocked')
            AND generated_at < ?
        ) OR (
            NOT (db_reflection_blocked = 1 OR critical_count > 0 OR status = 'blocked')
            AND generated_at < ?
        )
        ORDER BY generated_at ASC, id ASC
        """,
        (critical_cutoff.isoformat(timespec="seconds"), keep_cutoff.isoformat(timespec="seconds")),
    ).fetchall()
    return [_normalize_run_row(row) for row in rows]


def cleanup_preflight_history(
    conn: sqlite3.Connection,
    *,
    options: PreflightHistoryCleanupOptions,
) -> PreflightHistoryCleanupResult:
    conn.row_factory = sqlite3.Row
    apply_schema_migrations(conn)
    generated_at = _now()
    targets = _cleanup_targets(conn, options=options, generated_at=generated_at)
    preflight_ids = [str(row["preflight_id"]) for row in targets]
    issue_deleted_count = 0
    deleted_count = 0
    vacuumed = False
    if options.apply and preflight_ids:
        placeholders = ",".join("?" for _ in preflight_ids)
        issue_cursor = conn.execute(
            f"""
            DELETE FROM preflight_run_issues
            WHERE preflight_id IN ({placeholders})
            """,
            preflight_ids,
        )
        run_cursor = conn.execute(
            f"""
            DELETE FROM preflight_runs
            WHERE preflight_id IN ({placeholders})
            """,
            preflight_ids,
        )
        issue_deleted_count = int(issue_cursor.rowcount or 0)
        deleted_count = int(run_cursor.rowcount or 0)
    conn.commit()
    if options.apply and options.vacuum and deleted_count > 0:
        conn.execute("VACUUM")
        vacuumed = True
    return PreflightHistoryCleanupResult(
        mode="apply" if options.apply else "dry_run",
        keep_days=int(options.keep_days),
        keep_critical_days=int(options.keep_critical_days),
        target_count=len(targets),
        deleted_count=deleted_count,
        issue_deleted_count=issue_deleted_count,
        vacuumed=vacuumed,
        preview=targets[: max(int(options.limit_preview), 0)],
    )
