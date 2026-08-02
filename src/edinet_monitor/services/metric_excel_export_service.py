from __future__ import annotations

import calendar
import json
from bisect import bisect_right
from dataclasses import dataclass, field, replace
from datetime import datetime
from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path
import re
import sqlite3
from time import perf_counter
from typing import Any, Callable

from openpyxl import Workbook, load_workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.comments import Comment
from openpyxl.styles import Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from edinet_pipeline.domain.metric_labels import (
    BANK_INDUSTRY_LABEL,
    INSURANCE_INDUSTRY_LABEL,
    METRIC_BASE_LABELS,
    SECURITIES_INDUSTRY_LABEL,
    metric_base_to_display_name,
)
from edinet_monitor.domain.accounting_standard import is_ifrs_or_us_gaap
from edinet_monitor.domain.issuer_flags import tenbagger_learning_mark
from edinet_monitor.services.industry_aggregate_metric_service import (
    INDUSTRY_AGGREGATE_PERIOD_SCOPE,
    INDUSTRY_AGGREGATE_ROW_BASES,
    industry_aggregate_table_exists,
)
from edinet_monitor.services.market_derived_metric_service import (
    MARKET_METRIC_BASES,
    market_derived_table_exists,
)
from edinet_monitor.services.obsolete_quarter_metric_service import (
    FCF_GROWTH_BASE,
)
from edinet_monitor.services.quarter_standalone_metric_service import (
    FLOW_BASES as QUARTER_STANDALONE_FLOW_BASES,
    GROWTH_BASE_BY_FLOW_BASE as QUARTER_STANDALONE_GROWTH_BASE_BY_FLOW_BASE,
    QUARTER_STANDALONE_PERIOD_SCOPE,
    quarter_standalone_table_exists,
)
from edinet_monitor.services.quarter_source_policy_service import (
    EDINET_SEGMENT_QUARTERS,
    JQUANTS_ACTUAL_FINANCIAL_QUARTERS,
)
from edinet_monitor.services.segment_metric_service import (
    SEGMENT_EXCEL_METRIC_LABELS,
    segment_metrics_table_exists,
)
from edinet_monitor.services.segment_name_normalize_service import (
    SegmentNameCandidate,
    canonical_segment_key,
    preferred_segment_name_map,
)


SpanRecorder = Callable[[str, str, float, int, dict[str, Any]], None]


GENERAL_SHEET = "一般企業"
SUMMARY_SHEET = "summary"
CONDITION_SHEET = "条件"

SHEET_ORDER = [
    GENERAL_SHEET,
    BANK_INDUSTRY_LABEL,
    SECURITIES_INDUSTRY_LABEL,
    INSURANCE_INDUSTRY_LABEL,
]

PERIOD_LABEL_BY_OFFSET = {
    0: "当期",
    1: "前期",
    2: "2期前",
    3: "3期前",
    4: "4期前",
    5: "5期前",
    6: "6期前",
    7: "7期前",
    8: "8期前",
    9: "9期前",
    10: "10期前",
}
OFFSET_BY_PERIOD_LABEL = {label: offset for offset, label in PERIOD_LABEL_BY_OFFSET.items()}

ABSORBED_RATIO_BASE_BY_ROW_BASE = {
    "NetSales": None,
    "GrossProfit": "GrossProfitMargin",
    "CostOfSales": "CostOfSalesRatio",
    "SellingExpenses": "SellingExpensesRatio",
    "OperatingIncome": "OperatingMargin",
    "OrdinaryIncome": "OrdinaryIncomeMargin",
    "ProfitLoss": "EstimatedNetMargin",
    "TotalAssets": None,
    "NetAssets": "EquityRatio",
}

CONSTANT_RATIO_BY_ROW_BASE = {
    "NetSales": 1.0,
    "TotalAssets": 1.0,
}

ROW_BASE_BY_RATIO_BASE = {
    ratio_base: row_base
    for row_base, ratio_base in ABSORBED_RATIO_BASE_BY_ROW_BASE.items()
    if ratio_base
}

ABSORBED_RATIO_BASES = set(ROW_BASE_BY_RATIO_BASE)

PERIOD_SCOPE_BY_FORM_TYPE = {
    "030000": "annual",
    "043A00": "quarter:2Q",
    "043000": "quarter:2Q",
}
FORM_TYPES_BY_PERIOD_SCOPE = {
    "annual": ("030000",),
    "quarter": ("043A00", "043000"),
}
PERIOD_SCOPE_LABEL_BY_FORM_TYPE = {
    "030000": "4Q",
    "043A00": "2Q",
    "043000": "2Q",
}
ALL_PERIOD_SCOPES = ["annual", "quarter", "quarter_standalone", "forecast"]
SEGMENT_MODES = {"none", "all", "region", "business"}
SEGMENT_METRIC_BASES = tuple(SEGMENT_EXCEL_METRIC_LABELS)
SEGMENT_VALUE_KIND_PRIORITY = {
    "external": 0,
    "total": 1,
    "operating_profit": 0,
    "profit_before_tax": 0,
    "profit_loss": 0,
    "segment_profit": 0,
}
QUARTER_SUPPORTED_BASES = {
    "NetSales",
    "NetSalesGrowthRate",
    "CostOfSalesAndSellingGeneralAndAdministrativeExpenses",
    "CostOfSalesAndSellingGeneralAndAdministrativeExpensesGrowthRate",
    "OperatingIncome",
    "OperatingIncomeGrowthRate",
    "OrdinaryIncome",
    "OrdinaryIncomeGrowthRate",
    "ProfitLoss",
    "ProfitLossGrowthRate",
    "EstimatedNetIncome",
    "EstimatedNetIncomeGrowthRate",
    "FCF",
    "InvestmentCashToNetSalesRatio",
    "InvestmentCashToOperatingCashRatio",
    "EPS",
    "TotalAssets",
    "NetAssets",
    "BPS",
    "OperatingCash",
    "InvestmentCash",
    "FinancingCash",
    "CashAndCashEquivalents",
    "CashBalanceGrowthRate",
    "IssuedShares",
    "TreasuryShares",
    "OutstandingShares",
    "EquityRatio",
    "EPSGrowthRate",
    "BPSGrowthRate",
    "MarketCapitalization",
    "StockPrice",
    "StockPriceGrowthRate",
    "PBR",
    "PER",
    "TheoreticalSharePrice",
    "TheoreticalSharePriceGrowthRate",
    "TheoreticalPBR",
    "TheoreticalPER",
}
QUARTER_CUMULATIVE_GROWTH_SOURCE_BY_BASE = {
    "NetSalesGrowthRate": "NetSales",
    "CostOfSalesAndSellingGeneralAndAdministrativeExpensesGrowthRate": "CostOfSalesAndSellingGeneralAndAdministrativeExpenses",
    "OperatingIncomeGrowthRate": "OperatingIncome",
    "OrdinaryIncomeGrowthRate": "OrdinaryIncome",
    "ProfitLossGrowthRate": "ProfitLoss",
    "EstimatedNetIncomeGrowthRate": "EstimatedNetIncome",
    "CashBalanceGrowthRate": "CashAndCashEquivalents",
    "EPSGrowthRate": "EPS",
    "BPSGrowthRate": "BPS",
    "StockPriceGrowthRate": "StockPrice",
}
FORECAST_SUPPORTED_BASES = {
    "NetSales",
    "OperatingIncome",
    "OrdinaryIncome",
    "ProfitBeforeTax",
    "ProfitLoss",
}
QUARTER_STANDALONE_SUPPORTED_BASES = set(QUARTER_STANDALONE_FLOW_BASES) | set(
    QUARTER_STANDALONE_GROWTH_BASE_BY_FLOW_BASE.values()
)
FORECAST_PROGRESS_BASES = {"NetSales", "OperatingIncome", "OrdinaryIncome", "ProfitLoss"}
JQUANTS_QUARTER_TYPES = JQUANTS_ACTUAL_FINANCIAL_QUARTERS
QUARTER_STANDALONE_EXCEL_QUARTERS = ("1Q", "2Q", "3Q", "4Q", "1~2Q", "3~4Q")
QUARTER_STANDALONE_HALF_CASHFLOW_BASES = {
    "OperatingCash",
    "InvestmentCash",
    "FinancingCash",
    "FCF",
    "OperatingCashGrowthRate",
    "InvestmentCashGrowthRate",
    "FinancingCashGrowthRate",
}
JQUANTS_FORECAST_STAGES = ("initial", "1Q", "2Q", "3Q")
JQUANTS_FORECAST_STAGE_LABELS = {
    "initial": "0Q",
    "1Q": "1Q",
    "2Q": "2Q",
    "3Q": "3Q",
}
FORECAST_REVISION_UP_KIND = "forecast_revision_up"
FORECAST_REVISION_DOWN_KIND = "forecast_revision_down"
FORECAST_PROGRESS_RATIO_KIND = "forecast_progress"
DATE_POINT_PERIOD_BASES = {
    "MarketCapitalization",
    "StockPrice",
    "PBR",
    "PER",
    "PCFR",
    "TheoreticalSharePrice",
    "TheoreticalSharePriceGrowthRate",
    "TheoreticalPBR",
    "TheoreticalPER",
    "TheoreticalPCFR",
}
HALF_ONLY_BASES = {
    "HalfNetSalesProgressRate",
    "HalfOrdinaryIncomeProgressRate",
    "HalfProfitProgressRate",
}
HALF_DISABLED_BASES = {
    "NetSalesGrowthRate5Year",
    "NetSalesGrowthRate10Year",
    "OrdinaryIncomeGrowthRate5Year",
    "OrdinaryIncomeGrowthRate10Year",
    "CashBalanceGrowthRate5Year",
    "CashBalanceGrowthRate10Year",
    "EPSGrowthRate5Year",
    "EPSGrowthRate10Year",
    "BPSGrowthRate5Year",
    "BPSGrowthRate10Year",
    "OutstandingSharesGrowthRate",
    "OutstandingSharesGrowthRate5Year",
    "OutstandingSharesGrowthRate10Year",
    "StockPriceGrowthRate5Year",
    "StockPriceGrowthRate10Year",
    "TheoreticalSharePriceGrowthRate5Year",
    "TheoreticalSharePriceGrowthRate10Year",
    "NumberOfEmployees",
    "AverageAge",
    "AverageLengthOfService",
    "AverageAnnualSalary",
    "InterestBearingDebt",
    "ROIC",
}
HALF_PREFIX = "\u534a\u671f "
HALF_YOY_PREFIX = "\u534a\u671f\u524d\u5e74\u6bd4 "
INDUSTRY_ONLY_TOKEN = "\u696d\u7a2e\u306e\u307f"
ROW_KIND_DETAIL = "\u660e\u7d30"
ROW_KIND_CHANGE_RATE = "\u5897\u6e1b\u7387"
ROW_KIND_AVERAGE = "\u5e73\u5747\u5024"
ROW_KIND_MEDIAN = "\u4e2d\u592e\u5024"
VALUE_KIND_BASE = "\u57fa\u6e96\u5024"
VALUE_KIND_CALCULATED = "\u8a08\u7b97\u5024"
DETAIL_ROW_KINDS = {ROW_KIND_DETAIL, ROW_KIND_CHANGE_RATE}
SUPPRESSED_EXCEL_BASES = set(HALF_ONLY_BASES) | {"ProfitBeforeTax", "SegmentProfit", FCF_GROWTH_BASE}
QUARTER_SUPPRESSED_EXCEL_BASES = {
    "OutstandingSharesGrowthRate",
    "AssetsPerShare",
    "LiabilitiesPerShare",
    "InterestBearingDebt",
    "ROIC",
}
QUARTER_STANDALONE_SUPPRESSED_BY_QUARTER = {
    "1Q": {
        "OperatingCash",
        "InvestmentCash",
        "FinancingCash",
        "FCF",
        "OperatingCashGrowthRate",
        "InvestmentCashGrowthRate",
        "FinancingCashGrowthRate",
        "FCFGrowthRate",
    },
    "2Q": {
        "OperatingCash",
        "InvestmentCash",
        "FinancingCash",
        "FCF",
        "OperatingCashGrowthRate",
        "InvestmentCashGrowthRate",
        "FinancingCashGrowthRate",
        "FCFGrowthRate",
    },
    "3Q": {
        "OperatingCash",
        "InvestmentCash",
        "FinancingCash",
        "FCF",
        "OperatingCashGrowthRate",
        "InvestmentCashGrowthRate",
        "FinancingCashGrowthRate",
        "FCFGrowthRate",
    },
    "4Q": {
        "OperatingCash",
        "InvestmentCash",
        "FinancingCash",
        "FCF",
        "OperatingCashGrowthRate",
        "InvestmentCashGrowthRate",
        "FinancingCashGrowthRate",
        "FCFGrowthRate",
    },
}
for _quarter_type in ("1Q", "2Q", "3Q", "4Q"):
    QUARTER_STANDALONE_SUPPRESSED_BY_QUARTER.setdefault(_quarter_type, set()).update(
        {FCF_GROWTH_BASE}
    )
for _quarter_type in ("1~2Q", "3~4Q"):
    QUARTER_STANDALONE_SUPPRESSED_BY_QUARTER.setdefault(_quarter_type, set()).update(
        QUARTER_STANDALONE_SUPPORTED_BASES - QUARTER_STANDALONE_HALF_CASHFLOW_BASES
    )
PERIOD_BLOCK_FILL_COLORS = ("EAF4FF", "FFFFFF")
CURRENT_PERIOD_BLOCK_FILL_COLOR = "D9EAF7"
PERIOD_BLOCK_BORDER_COLOR = "9FBAD0"
FORECAST_PROGRESS_FILL_COLOR = "FFF2CC"
FORECAST_REVISION_UP_FONT_COLOR = "FF0000"
FORECAST_REVISION_DOWN_FONT_COLOR = "00B0F0"

MONETARY_BASES = {
    "NetSales",
    "CostOfSales",
    "CostOfSalesAndSellingGeneralAndAdministrativeExpenses",
    "SellingExpenses",
    "GeneralAndAdministrativeExpenses",
    "SellingExpensesOnly",
    "OperatingIncome",
    "OrdinaryIncome",
    "ProfitBeforeTax",
    "ProfitLoss",
    "SegmentProfit",
    "OperatingCash",
    "InvestmentCash",
    "FinancingCash",
    "TotalAssets",
    "NetAssets",
    "BeginningCashBalance",
    "CashAndCashEquivalents",
    "GrossProfit",
    "EstimatedNetIncome",
    "FCF",
    "FundingIncome",
    "FeesAndCommissionsIncome",
    "InsuranceClaimsPayments",
    "PolicyReserveProvision",
    "InvestmentExpenses",
    "ProjectExpenses",
    "IncomeTaxes",
    "InterestBearingDebt",
    "MarketCapitalization",
}

PERCENT_VALUE_BASES = {
    "ROA",
    "ROE",
    "ROIC",
    "EquityRatio",
    "InvestmentCashToNetSalesRatio",
    "InvestmentCashToOperatingCashRatio",
    "StockPriceGrowthRate",
    "StockPriceGrowthRate5Year",
    "StockPriceGrowthRate10Year",
} | ABSORBED_RATIO_BASES
RATIO_VALUE_BASES = {
    "PBR",
    "PER",
    "PCFR",
    "TheoreticalPBR",
    "TheoreticalPER",
    "TheoreticalPCFR",
}

PER_SHARE_BASES = {
    "AssetsPerShare",
    "LiabilitiesPerShare",
    "OperatingCashPerShare",
    "InvestmentCashPerShare",
    "FinancingCashPerShare",
    "FCFPerShare",
    "AssetValue",
    "BusinessValue",
    "StockPrice",
    "TheoreticalSharePrice",
    "UpperBoundTheoreticalSharePrice",
}

ONE_DECIMAL_VALUE_BASES = {
    "EPS",
    "BPS",
    "TheoreticalPBR",
    "TheoreticalPER",
    "PBR",
    "PER",
    "PCFR",
    "TheoreticalPCFR",
    "AverageAge",
    "AverageLengthOfService",
}

INTEGER_VALUE_BASES = {
    "OutstandingShares",
    "AssetsPerShare",
    "LiabilitiesPerShare",
    "AssetValue",
    "BusinessValue",
    "TheoreticalSharePrice",
    "UpperBoundTheoreticalSharePrice",
    "MarketCapitalization",
    "OperatingCashPerShare",
    "InvestmentCashPerShare",
    "FinancingCashPerShare",
    "FCFPerShare",
    "NumberOfEmployees",
    "AverageAnnualSalary",
}

GROWTH_RATIO_BASES = {
    "NetSalesGrowthRate",
    "CostOfSalesAndSellingGeneralAndAdministrativeExpensesGrowthRate",
    "OperatingIncomeGrowthRate",
    "NetSalesGrowthRate5Year",
    "NetSalesGrowthRate10Year",
    "OrdinaryIncomeGrowthRate",
    "OrdinaryIncomeGrowthRate5Year",
    "OrdinaryIncomeGrowthRate10Year",
    "ProfitLossGrowthRate",
    "EstimatedNetIncomeGrowthRate",
    "OperatingCashGrowthRate",
    "InvestmentCashGrowthRate",
    "FinancingCashGrowthRate",
    "FCFGrowthRate",
    "CashBalanceGrowthRate",
    "CashBalanceGrowthRate5Year",
    "CashBalanceGrowthRate10Year",
    "OutstandingSharesGrowthRate",
    "OutstandingSharesGrowthRate5Year",
    "OutstandingSharesGrowthRate10Year",
    "EPSGrowthRate",
    "EPSGrowthRate5Year",
    "EPSGrowthRate10Year",
    "BPSGrowthRate",
    "BPSGrowthRate5Year",
    "BPSGrowthRate10Year",
    "TheoreticalSharePriceGrowthRate",
    "TheoreticalSharePriceGrowthRate5Year",
    "TheoreticalSharePriceGrowthRate10Year",
    "StockPriceGrowthRate",
    "StockPriceGrowthRate5Year",
    "StockPriceGrowthRate10Year",
    "HalfNetSalesProgressRate",
    "HalfOrdinaryIncomeProgressRate",
    "HalfProfitProgressRate",
    "InvestmentCashToNetSalesRatio",
    "InvestmentCashToOperatingCashRatio",
}

CALCULATED_VALUE_KIND_BASES = (
    GROWTH_RATIO_BASES
    | PERCENT_VALUE_BASES
    | RATIO_VALUE_BASES
    | PER_SHARE_BASES
    | {
        "EstimatedNetIncome",
        "EPS",
        "BPS",
        "InterestBearingDebt",
        "MarketCapitalization",
    }
)

INDUSTRY_AGGREGATE_VALUE_BASES = set(INDUSTRY_AGGREGATE_ROW_BASES)

CHANGE_RATE_ROW_KIND_BASES = {
    "BPSGrowthRate",
    "BPSGrowthRate5Year",
    "BPSGrowthRate10Year",
    "CashBalanceGrowthRate",
    "CashBalanceGrowthRate5Year",
    "CashBalanceGrowthRate10Year",
    "EPSGrowthRate",
    "EPSGrowthRate5Year",
    "EPSGrowthRate10Year",
    "EstimatedNetIncomeGrowthRate",
    "FCFGrowthRate",
    "FinancingCashGrowthRate",
    "InvestmentCashGrowthRate",
    "CostOfSalesAndSellingGeneralAndAdministrativeExpensesGrowthRate",
    "NetSalesGrowthRate",
    "NetSalesGrowthRate5Year",
    "NetSalesGrowthRate10Year",
    "OperatingCashGrowthRate",
    "OperatingIncomeGrowthRate",
    "OrdinaryIncomeGrowthRate",
    "OrdinaryIncomeGrowthRate5Year",
    "OrdinaryIncomeGrowthRate10Year",
    "OutstandingSharesGrowthRate",
    "OutstandingSharesGrowthRate5Year",
    "OutstandingSharesGrowthRate10Year",
    "ProfitLossGrowthRate",
    "StockPriceGrowthRate",
    "StockPriceGrowthRate5Year",
    "StockPriceGrowthRate10Year",
    "TheoreticalSharePriceGrowthRate",
    "TheoreticalSharePriceGrowthRate5Year",
    "TheoreticalSharePriceGrowthRate10Year",
}

SPARSE_PERIOD_OFFSETS_BY_BASE: dict[str, set[int]] = {}

FIXED_ROW_BASE_ORDER = [
    "OutstandingShares",
    "TotalAssets",
    "NetAssets",
    "NetSales",
    "CostOfSales",
    "GrossProfit",
    "SellingExpenses",
    "GeneralAndAdministrativeExpenses",
    "SellingExpensesOnly",
    "CostOfSalesAndSellingGeneralAndAdministrativeExpenses",
    "OperatingIncome",
    "OrdinaryIncome",
    "ProfitBeforeTax",
    "ProfitLoss",
    "SegmentProfit",
    "EstimatedNetIncome",
    "OperatingCash",
    "InvestmentCash",
    "FinancingCash",
    "BeginningCashBalance",
    "CashAndCashEquivalents",
    "FCF",
    "InvestmentCashToNetSalesRatio",
    "InvestmentCashToOperatingCashRatio",
    "InterestBearingDebt",
    "EPS",
    "EPSGrowthRate",
    "EPSGrowthRate5Year",
    "EPSGrowthRate10Year",
    "BPS",
    "BPSGrowthRate",
    "BPSGrowthRate5Year",
    "BPSGrowthRate10Year",
    "HalfNetSalesProgressRate",
    "HalfOrdinaryIncomeProgressRate",
    "HalfProfitProgressRate",
    "NetSalesGrowthRate",
    "CostOfSalesAndSellingGeneralAndAdministrativeExpensesGrowthRate",
    "OperatingIncomeGrowthRate",
    "NetSalesGrowthRate5Year",
    "NetSalesGrowthRate10Year",
    "OrdinaryIncomeGrowthRate",
    "OrdinaryIncomeGrowthRate5Year",
    "OrdinaryIncomeGrowthRate10Year",
    "ProfitLossGrowthRate",
    "EstimatedNetIncomeGrowthRate",
    "OperatingCashGrowthRate",
    "InvestmentCashGrowthRate",
    "FinancingCashGrowthRate",
    "FCFGrowthRate",
    "CashBalanceGrowthRate",
    "CashBalanceGrowthRate5Year",
    "CashBalanceGrowthRate10Year",
    "OutstandingSharesGrowthRate",
    "OutstandingSharesGrowthRate5Year",
    "OutstandingSharesGrowthRate10Year",
    "AssetsPerShare",
    "LiabilitiesPerShare",
    "ROA",
    "ROE",
    "ROIC",
    "EquityRatio",
    "MarketCapitalization",
    "StockPrice",
    "StockPriceGrowthRate",
    "StockPriceGrowthRate5Year",
    "StockPriceGrowthRate10Year",
    "PBR",
    "PER",
    "AssetValue",
    "BusinessValue",
    "TheoreticalSharePrice",
    "UpperBoundTheoreticalSharePrice",
    "TheoreticalSharePriceGrowthRate",
    "TheoreticalSharePriceGrowthRate5Year",
    "TheoreticalSharePriceGrowthRate10Year",
    "TheoreticalPBR",
    "TheoreticalPER",
    "OperatingCashPerShare",
    "InvestmentCashPerShare",
    "FinancingCashPerShare",
    "FCFPerShare",
    "PCFR",
    "TheoreticalPCFR",
    "NumberOfEmployees",
    "AverageAge",
    "AverageLengthOfService",
    "AverageAnnualSalary",
]

ROW_BASE_ORDER_INDEX = {
    metric_base: index for index, metric_base in enumerate(FIXED_ROW_BASE_ORDER)
}

EXCEL_METRIC_LABEL_OVERRIDES = {
    "CostOfSalesAndSellingGeneralAndAdministrativeExpensesGrowthRate": "\u8cbb\u7528\u5408\u8a08\u5897\u6e1b\u7387(\u524d\u671f\u6bd4)",
    "ProfitBeforeTax": "\u7a0e\u5f15\u524d\u5229\u76ca",
    "OutstandingShares": "実質発行株数",
    "OutstandingSharesGrowthRate": "実質発行株数増加率(前期比)",
    "OutstandingSharesGrowthRate5Year": "実質発行株数増加率(５年)",
    "OutstandingSharesGrowthRate10Year": "実質発行株数増加率(10年)",
    "CostOfSales": "├売上原価",
    "SellingExpenses": "└販管費",
    "GeneralAndAdministrativeExpenses": "　├ 一般管理費",
    "SellingExpensesOnly": "　└ 販売費",
    "OperatingCash": "営業CF",
    "InvestmentCash": "投資CF",
    "FinancingCash": "財務CF",
    "EstimatedNetIncome": "統一純利益",
    "NetSalesGrowthRate": "売上高増収率(前期比)",
    "OperatingIncomeGrowthRate": "営業利益増益率(前期比)",
    "ProfitLossGrowthRate": "純利益増益率(前期比)",
    "EstimatedNetIncomeGrowthRate": "統一純利益増益率(前期比)",
    "OperatingCashGrowthRate": "営業CF増加率(前期比)",
    "InvestmentCashGrowthRate": "投資CF増加率(前期比)",
    "FinancingCashGrowthRate": "財務CF増加率(前期比)",
    "FCFGrowthRate": "FCF増加率(前期比)",
    "EPSGrowthRate": "EPS増加率(前期比)",
    "BPSGrowthRate": "BPS増加率(前期比)",
    "OrdinaryIncomeGrowthRate": "経常利益増益率(前期比)",
    "CashBalanceGrowthRate": "期末残高増加率(前期比)",
    "CashBalanceGrowthRate5Year": "期末残高増加率(５年)",
    "CashBalanceGrowthRate10Year": "期末残高増加率(10年)",
    "StockPriceGrowthRate": "株価上昇率(前期比)",
    "TheoreticalSharePriceGrowthRate": "理論株価上昇率",
    "TheoreticalSharePriceGrowthRate5Year": "理論株価上昇率(５年)",
    "TheoreticalSharePriceGrowthRate10Year": "理論株価上昇率(10年)",
}

INDUSTRY_EXCEL_METRIC_LABEL_OVERRIDES = {
    BANK_INDUSTRY_LABEL: {
        "CostOfSales": "├資金調達費用",
        "SellingExpenses": "└営業経費",
        "FeesAndCommissionsIncome": "役務取引等収益",
    },
    SECURITIES_INDUSTRY_LABEL: {
        "GrossProfit": "純収益",
        "CostOfSales": "├金融費用",
        "SellingExpenses": "└販管費",
        "GeneralAndAdministrativeExpenses": "　├ 一般管理費",
        "SellingExpensesOnly": "　└ 販売費",
    },
    INSURANCE_INDUSTRY_LABEL: {
        "InsuranceClaimsPayments": "├保険金等支払金",
        "PolicyReserveProvision": "├責任準備金等繰入額",
        "InvestmentExpenses": "├資産運用費用",
        "ProjectExpenses": "└ 事業費",
    },
}

COMMON_TAIL_BASES = [
    base
    for base in FIXED_ROW_BASE_ORDER
    if base
    not in {
        "NetSales",
        "GrossProfit",
        "CostOfSalesAndSellingGeneralAndAdministrativeExpenses",
        "CostOfSales",
        "SellingExpenses",
        "GeneralAndAdministrativeExpenses",
        "SellingExpensesOnly",
        "OperatingIncome",
        "OrdinaryIncome",
        "ProfitLoss",
        "TotalAssets",
        "NetAssets",
        "BeginningCashBalance",
        "CashAndCashEquivalents",
        "OperatingCash",
        "InvestmentCash",
        "FinancingCash",
        "FCF",
    }
]

POST_BPS_TAIL_BASES = FIXED_ROW_BASE_ORDER[FIXED_ROW_BASE_ORDER.index("BPS") + 1 :]
POST_BPS_TAIL_BASES_FOR_FINANCIAL = [base for base in POST_BPS_TAIL_BASES if base != "ROIC"]

DEFAULT_BASES_BY_SHEET = {
    GENERAL_SHEET: list(FIXED_ROW_BASE_ORDER),
    BANK_INDUSTRY_LABEL: [
        "NetSales",
        "CostOfSalesAndSellingGeneralAndAdministrativeExpenses",
        "CostOfSales",
        "SellingExpenses",
        "FeesAndCommissionsIncome",
        "OperatingIncome",
        "OrdinaryIncome",
        "ProfitLoss",
        "EstimatedNetIncome",
        "TotalAssets",
        "NetAssets",
        "BeginningCashBalance",
        "CashAndCashEquivalents",
        "OperatingCash",
        "InvestmentCash",
        "FinancingCash",
        "FCF",
        "OutstandingShares",
        "EPS",
        "EPSGrowthRate",
        "BPS",
    ]
    + POST_BPS_TAIL_BASES_FOR_FINANCIAL,
    SECURITIES_INDUSTRY_LABEL: [
        "NetSales",
        "GrossProfit",
        "CostOfSalesAndSellingGeneralAndAdministrativeExpenses",
        "CostOfSales",
        "SellingExpenses",
        "GeneralAndAdministrativeExpenses",
        "SellingExpensesOnly",
        "OperatingIncome",
        "OrdinaryIncome",
        "ProfitLoss",
        "EstimatedNetIncome",
        "TotalAssets",
        "NetAssets",
        "BeginningCashBalance",
        "CashAndCashEquivalents",
        "OperatingCash",
        "InvestmentCash",
        "FinancingCash",
        "FCF",
        "OutstandingShares",
        "EPS",
        "EPSGrowthRate",
        "BPS",
    ]
    + POST_BPS_TAIL_BASES_FOR_FINANCIAL,
    INSURANCE_INDUSTRY_LABEL: [
        "NetSales",
        "GrossProfit",
        "CostOfSalesAndSellingGeneralAndAdministrativeExpenses",
        "InsuranceClaimsPayments",
        "PolicyReserveProvision",
        "InvestmentExpenses",
        "ProjectExpenses",
        "OperatingIncome",
        "OrdinaryIncome",
        "ProfitLoss",
        "EstimatedNetIncome",
        "TotalAssets",
        "NetAssets",
        "BeginningCashBalance",
        "CashAndCashEquivalents",
        "OperatingCash",
        "InvestmentCash",
        "FinancingCash",
        "FCF",
        "OutstandingShares",
        "EPS",
        "EPSGrowthRate",
        "BPS",
    ]
    + POST_BPS_TAIL_BASES_FOR_FINANCIAL,
}

ROW_BASE_ORDER_INDEX_BY_SHEET = {
    sheet_name: {metric_base: index for index, metric_base in enumerate(metric_bases)}
    for sheet_name, metric_bases in DEFAULT_BASES_BY_SHEET.items()
}

CONDITION_KEYS = {
    "業種": "industries",
    "証券コード": "security_codes",
    "企業名": "company_names",
    "指標": "metrics",
    "決算種別": "period_scopes",
    "期間": "periods",
    "期間Start": "period_start",
    "期間End": "period_end",
    "連続増減": "trend",
    "連続増減判定": "trend",
    "連続増減指標": "trend_metrics",
    "連続増減期間": "trend_periods",
    "連続増減期間Start": "trend_period_start",
    "連続増減期間End": "trend_period_end",
    "連続増減下限": "trend_min",
    "連続増減以上": "trend_min",
    "連続増減上限": "trend_max",
    "連続増減以下": "trend_max",
    "増減判定": "trend",
    "増減判定指標": "trend_metrics",
    "増減判定期間": "trend_periods",
    "増減判定下限": "trend_min",
    "増減判定以上": "trend_min",
    "増減判定上限": "trend_max",
    "増減判定以下": "trend_max",
    "％条件指標": "percent_filter_metrics",
    "%条件指標": "percent_filter_metrics",
    "比率条件指標": "percent_filter_metrics",
    "比率条件": "percent_filter_metrics",
    "比率条件期間": "percent_filter_periods",
    "比率条件期間Start": "percent_filter_period_start",
    "比率条件期間End": "percent_filter_period_end",
    "％条件期間": "percent_filter_periods",
    "%条件期間": "percent_filter_periods",
    "比率条件下限": "percent_filter_min",
    "比率条件以上": "percent_filter_min",
    "比率条件上限": "percent_filter_max",
    "比率条件以下": "percent_filter_max",
    "％条件下限": "percent_filter_min",
    "%条件下限": "percent_filter_min",
    "％条件以上": "percent_filter_min",
    "%条件以上": "percent_filter_min",
    "％条件上限": "percent_filter_max",
    "%条件上限": "percent_filter_max",
    "％条件以下": "percent_filter_max",
    "%条件以下": "percent_filter_max",
    "％下限": "percent_filter_min",
    "%下限": "percent_filter_min",
    "比率下限": "percent_filter_min",
    "下限": "percent_filter_min",
    "以上": "percent_filter_min",
    "％上限": "percent_filter_max",
    "%上限": "percent_filter_max",
    "比率上限": "percent_filter_max",
    "上限": "percent_filter_max",
    "以下": "percent_filter_max",
    "セグメント": "segment_mode",
}

EXCEL_PERCENT_RATIO_PREFIX = "__excel_percent_ratio__:"

INDUSTRY_ALIASES = {
    "証券・商品先物取引業": SECURITIES_INDUSTRY_LABEL,
    "証券商品先物取引業": SECURITIES_INDUSTRY_LABEL,
}


@dataclass(frozen=True)
class MetricExcelCondition:
    industries: list[str] = field(default_factory=list)
    industry_only: bool = False
    security_codes: list[str] = field(default_factory=list)
    company_names: list[str] = field(default_factory=list)
    metric_labels: list[str] = field(default_factory=list)
    period_scopes: list[str] = field(default_factory=lambda: list(ALL_PERIOD_SCOPES))
    period_offsets: list[int] = field(default_factory=lambda: list(range(10, -1, -1)))
    trend: str = "none"
    trend_metric_labels: list[str] = field(default_factory=list)
    trend_period_offsets: list[int] = field(default_factory=list)
    trend_min: float | None = None
    trend_max: float | None = None
    percent_filter_metric_labels: list[str] = field(default_factory=list)
    percent_filter_period_offsets: list[int] = field(default_factory=lambda: [0])
    percent_filter_min: float | None = None
    percent_filter_max: float | None = None
    segment_mode: str = "none"
    raw_values: dict[str, str] = field(default_factory=dict)


@dataclass
class MetricExcelRow:
    sheet_name: str
    security_code: str
    company_name: str
    industry_33: str
    market: str
    period_scope: str
    current_period_end: str
    metric_base: str
    metric_label: str
    periods_by_offset: dict[int, str]
    values_by_offset: dict[int, float | None]
    units_by_offset: dict[int, str]
    ratios_by_offset: dict[int, float | None]
    row_kind: str = ROW_KIND_DETAIL
    raw_values_by_offset: dict[int, float | None] = field(default_factory=dict)
    ranks_by_offset: dict[int, str] = field(default_factory=dict)
    ratio_kinds_by_offset: dict[int, str] = field(default_factory=dict)
    value_kinds_by_offset: dict[int, str] = field(default_factory=dict)
    fiscal_years_by_offset: dict[int, int | None] = field(default_factory=dict)
    segment_kind: str = ""
    segment_name: str = ""
    segment_order: int | None = None
    accounting_standard: str = ""


@dataclass(frozen=True)
class MetricExcelExportResult:
    output_path: Path
    target_companies: int
    output_rows: int
    errors: list[str]
    warnings: list[str]
    preview_rows: list[dict[str, Any]]


@dataclass(frozen=True)
class _JQuantsLookupIndexes:
    max_fiscal_year_by_period: dict[tuple[str, str, str], int]
    forecast_values_by_as_of_key: dict[tuple[str, int, str], list[tuple[str, str, float]]]
    forecast_value_by_stage: dict[tuple[str, int, str, str], float]


@dataclass(frozen=True)
class _QuarterStandaloneLookupIndexes:
    max_fiscal_year_by_code: dict[str, int]
    period_end_by_code_year_quarter: dict[tuple[str, int, str], str]


def _record_span(
    span_recorder: SpanRecorder | None,
    category: str,
    name: str,
    started: float,
    *,
    count: int = 0,
    detail: dict[str, Any] | None = None,
) -> None:
    if span_recorder is None:
        return
    span_recorder(
        category,
        name,
        max(perf_counter() - started, 0.0),
        count,
        dict(detail or {}),
    )


def _normalize_text(value: Any) -> str:
    text = str(value or "").replace("\u3000", " ").replace("\xa0", " ").strip()
    return re.sub(r"\s+", "", text)


def _split_multi(value: str | None) -> list[str]:
    text = str(value or "").strip()
    if not text or text.upper() == "ALL":
        return []
    parts = re.split(r"[,、\n\r]+", text)
    return [part.strip() for part in parts if part.strip()]


def _normalize_industry(value: str) -> str:
    text = str(value or "").strip()
    return INDUSTRY_ALIASES.get(text, text)


def _split_industries(value: str | None) -> list[str]:
    text = str(value or "").strip()
    if not text or text.upper() == "ALL":
        return []

    # The official securities industry name contains "、", so protect it before comma splitting.
    protected = text
    for alias in [SECURITIES_INDUSTRY_LABEL, *INDUSTRY_ALIASES]:
        protected = protected.replace(alias, "__SECURITIES_INDUSTRY__")

    parts = re.split(r"[,、\n\r]+", protected)
    industries: list[str] = []
    for part in parts:
        item = part.strip()
        if not item:
            continue
        if item == "__SECURITIES_INDUSTRY__":
            item = SECURITIES_INDUSTRY_LABEL
        industries.append(_normalize_industry(item))
    return industries


def _normalize_security_code(value: str) -> str:
    text = str(value or "").strip().replace("-", "")
    if len(text) == 5 and text.endswith("0"):
        return text[:-1]
    return text


def _security_code_candidates(security_codes: list[str]) -> list[str]:
    candidates: set[str] = set()
    for value in security_codes:
        text = str(value or "").strip().replace("-", "")
        if not text:
            continue
        normalized = _normalize_security_code(text)
        candidates.add(normalized)
        if len(normalized) == 4:
            candidates.add(f"{normalized}0")
    return sorted(candidates)


def _jquants_local_code_candidates(security_codes: list[str]) -> list[str]:
    candidates: set[str] = set()
    for value in security_codes:
        text = str(value or "").strip().replace("-", "")
        if not text:
            continue
        normalized = _normalize_security_code(text)
        candidates.add(text if len(text) == 5 else f"{normalized}0")
    return sorted(candidates)


def _index_exists(conn: sqlite3.Connection, index_name: str) -> bool:
    row = conn.execute(
        """
        SELECT 1
        FROM sqlite_master
        WHERE type = 'index' AND name = ?
        """,
        (index_name,),
    ).fetchone()
    return row is not None


def _parse_period_token(token: str) -> int:
    text = _normalize_text(token)
    if text in {"当期", "最新"}:
        return 0
    if text in {"前期", "1期前", "1年前"}:
        return 1
    match = re.fullmatch(r"(\d+)(?:期前|年前)", text)
    if match:
        offset = int(match.group(1))
        if 0 <= offset <= 10:
            return offset
    raise ValueError(f"Unsupported period: {token}")


def _parse_periods(value: str | None) -> list[int]:
    text = str(value or "").strip()
    if not text or text.upper() == "ALL":
        return list(range(10, -1, -1))

    normalized = text.replace("～", "-").replace("－", "-").replace("〜", "-")
    if "-" in normalized and "," not in normalized and "、" not in normalized:
        left, right = normalized.split("-", 1)
        start = _parse_period_token(left)
        end = _parse_period_token(right)
        hi = max(start, end)
        lo = min(start, end)
        return list(range(hi, lo - 1, -1))

    offsets = sorted({_parse_period_token(part) for part in _split_multi(normalized)}, reverse=True)
    if not offsets:
        return list(range(10, -1, -1))
    return offsets


def _period_range_text(
    raw: dict[str, str],
    *,
    combined_key: str,
    start_key: str,
    end_key: str,
) -> str | None:
    combined = str(raw.get(combined_key) or "").strip()
    if combined:
        return combined

    start = str(raw.get(start_key) or "").strip()
    end = str(raw.get(end_key) or "").strip()
    if not start and not end:
        return None
    if not start or not end:
        return start or end
    if start.upper() == "ALL" or end.upper() == "ALL":
        return "all"
    return f"{start}-{end}"


def _condition_cell_value(cell: Any, key: str) -> str:
    value = getattr(cell, "value", None)
    if value is None:
        return ""
    if (
        key in {"percent_filter_min", "percent_filter_max", "trend_min", "trend_max"}
        and isinstance(value, (int, float))
        and "%" in str(getattr(cell, "number_format", ""))
    ):
        # Excel stores 500% as numeric 5.0 when the cell uses percent formatting.
        return f"{EXCEL_PERCENT_RATIO_PREFIX}{value}"
    return str(value).strip()


def _parse_percent_threshold(value: str | None) -> float | None:
    text = str(value or "").strip()
    if not text:
        return None
    if text.startswith(EXCEL_PERCENT_RATIO_PREFIX):
        return float(text.removeprefix(EXCEL_PERCENT_RATIO_PREFIX))
    has_percent_sign = "%" in text or "％" in text
    normalized = text.replace("%", "").replace("％", "").replace(",", "").strip()
    if not normalized:
        return None
    threshold = float(normalized)
    if has_percent_sign or abs(threshold) > 1:
        threshold /= 100
    return threshold


def _parse_period_scopes(value: str | None) -> list[str]:
    text = str(value or "").strip()
    if not text:
        return list(ALL_PERIOD_SCOPES)
    if text.upper() == "ALL":
        return list(ALL_PERIOD_SCOPES)
    mapping = {
        "\u901a\u671f": "annual",
        "annual": "annual",
        "030000": "annual",
        "\u534a\u671f": "quarter",
        "half": "quarter",
        "2q": "quarter",
        "4q": "annual",
        "043000": "quarter",
        "043a00": "quarter",
        "\u56db\u534a\u671f": "quarter",
        "quarter": "quarter",
        "1q": "quarter",
        "3q": "quarter",
        "\u56db\u534a\u671f\u5358\u72ec": "quarter_standalone",
        "\u56db\u534a\u671f\u5358\u4f53": "quarter_standalone",
        "quarter_standalone": "quarter_standalone",
        "standalone": "quarter_standalone",
        "\u4e88\u60f3": "forecast",
        "forecast": "forecast",
    }
    scopes: list[str] = []
    seen: set[str] = set()
    for item in _split_multi(text):
        scope = mapping.get(_normalize_text(item).lower()) or mapping.get(str(item).strip())
        if not scope:
            raise ValueError(f"Unsupported period scope: {item}")
        if scope not in seen:
            scopes.append(scope)
            seen.add(scope)
    return scopes or ["annual"]


def _parse_segment_mode(value: str | None) -> str:
    text = _normalize_text(value).lower()
    if not text or text in {"none", "なし", "不要", "no"}:
        return "none"
    mapping = {
        "all": "all",
        "すべて": "all",
        "全部": "all",
        "地域": "region",
        "region": "region",
        "地域別": "region",
        "部門": "business",
        "部門別": "business",
        "事業": "business",
        "事業別": "business",
        "business": "business",
    }
    if text not in mapping:
        raise ValueError("セグメントは none / all / 地域 / 部門 のいずれかで指定してください。")
    return mapping[text]


def read_metric_excel_condition(condition_xlsx: str | Path) -> MetricExcelCondition:
    path = Path(condition_xlsx)
    workbook = load_workbook(path, data_only=True)
    if CONDITION_SHEET not in workbook.sheetnames:
        raise ValueError("条件シートがありません。条件シートを追加してください。")

    ws = workbook[CONDITION_SHEET]
    raw: dict[str, str] = {}
    for row in ws.iter_rows():
        cell_texts = ["" if cell.value is None else str(cell.value).strip() for cell in row]
        for idx, cell_text in enumerate(cell_texts):
            key = CONDITION_KEYS.get(_normalize_text(cell_text))
            if key is None:
                continue
            value = ""
            for candidate in row[idx + 1 :]:
                candidate_value = _condition_cell_value(candidate, key)
                if candidate_value:
                    value = candidate_value
                    break
            raw[key] = value

    trend = str(raw.get("trend") or "none").strip().lower()
    if trend in {"", "all"}:
        trend = "none"
    if trend not in {"none", "increase", "decrease"}:
        raise ValueError("連続増減は none / increase / decrease のいずれかで指定してください。")

    period_text = _period_range_text(
        raw,
        combined_key="periods",
        start_key="period_start",
        end_key="period_end",
    )
    trend_period_text = _period_range_text(
        raw,
        combined_key="trend_periods",
        start_key="trend_period_start",
        end_key="trend_period_end",
    )
    percent_filter_period_text = _period_range_text(
        raw,
        combined_key="percent_filter_periods",
        start_key="percent_filter_period_start",
        end_key="percent_filter_period_end",
    )

    period_offsets = _parse_periods(period_text)
    trend_period_offsets = _parse_periods(trend_period_text) if trend_period_text else period_offsets
    trend_min = _parse_percent_threshold(raw.get("trend_min"))
    trend_max = _parse_percent_threshold(raw.get("trend_max"))
    percent_filter_period_offsets = _parse_periods(percent_filter_period_text) if percent_filter_period_text else [0]
    percent_filter_min = _parse_percent_threshold(raw.get("percent_filter_min"))
    percent_filter_max = _parse_percent_threshold(raw.get("percent_filter_max"))
    industries = _split_industries(raw.get("industries"))
    industry_only = any(_normalize_text(item) == _normalize_text(INDUSTRY_ONLY_TOKEN) for item in industries)
    industries = [
        item
        for item in industries
        if _normalize_text(item) != _normalize_text(INDUSTRY_ONLY_TOKEN)
    ]

    return MetricExcelCondition(
        industries=industries,
        industry_only=industry_only,
        security_codes=[_normalize_security_code(code) for code in _split_multi(raw.get("security_codes"))],
        company_names=_split_multi(raw.get("company_names")),
        metric_labels=_split_multi(raw.get("metrics")),
        period_scopes=_parse_period_scopes(raw.get("period_scopes")),
        period_offsets=period_offsets,
        trend=trend,
        trend_metric_labels=_split_multi(raw.get("trend_metrics")),
        trend_period_offsets=trend_period_offsets,
        trend_min=trend_min,
        trend_max=trend_max,
        percent_filter_metric_labels=_split_multi(raw.get("percent_filter_metrics")),
        percent_filter_period_offsets=percent_filter_period_offsets,
        percent_filter_min=percent_filter_min,
        percent_filter_max=percent_filter_max,
        segment_mode=_parse_segment_mode(raw.get("segment_mode")),
        raw_values=raw,
    )


def _sheet_name_for_industry(industry_33: str | None) -> str:
    industry = str(industry_33 or "").strip()
    if industry in {BANK_INDUSTRY_LABEL, SECURITIES_INDUSTRY_LABEL, INSURANCE_INDUSTRY_LABEL}:
        return industry
    return GENERAL_SHEET


def _base_metric_label_for_excel(metric_base: str, industry_33: str | None = None) -> str:
    industry = str(industry_33 or "").strip()
    industry_overrides = INDUSTRY_EXCEL_METRIC_LABEL_OVERRIDES.get(industry, {})
    if metric_base in industry_overrides:
        return industry_overrides[metric_base]
    return EXCEL_METRIC_LABEL_OVERRIDES.get(
        metric_base,
        metric_base_to_display_name(metric_base, industry_33),
    )


def _display_base_for_accounting_standard(metric_base: str, accounting_standard: str | None) -> str:
    if metric_base == "OrdinaryIncome" and is_ifrs_or_us_gaap(accounting_standard):
        return "ProfitBeforeTax"
    return metric_base


def _metric_label_base_for_excel(
    metric_base: str,
    industry_33: str | None,
    accounting_standard: str | None = None,
) -> str:
    if metric_base == "OrdinaryIncomeGrowthRate" and is_ifrs_or_us_gaap(accounting_standard):
        return f"{_base_metric_label_for_excel('ProfitBeforeTax', industry_33)}\u5897\u76ca\u7387(\u524d\u671f\u6bd4)"
    display_base = _display_base_for_accounting_standard(metric_base, accounting_standard)
    return _base_metric_label_for_excel(display_base, industry_33)


def _quarter_standalone_label(
    metric_base: str,
    industry_33: str | None,
    quarter: str,
    accounting_standard: str | None = None,
) -> str:
    reverse_growth_map = {
        growth_base: source_base
        for source_base, growth_base in QUARTER_STANDALONE_GROWTH_BASE_BY_FLOW_BASE.items()
    }
    if metric_base in reverse_growth_map:
        source_base = reverse_growth_map[metric_base]
        source_base = _display_base_for_accounting_standard(source_base, accounting_standard)
        source_label = _base_metric_label_for_excel(source_base, industry_33)
        if source_base in {"NetSales", "CashAndCashEquivalents"}:
            growth_word = "\u5897\u53ce\u7387"
        elif source_base in {"OperatingCash", "InvestmentCash", "FinancingCash", "FCF"}:
            growth_word = "\u5897\u52a0\u7387"
        else:
            growth_word = "\u5897\u76ca\u7387"
        return f"{quarter} {source_label}{growth_word} \u5358\u72ec(\u524d\u671f\u6bd4)"
    display_base = _display_base_for_accounting_standard(metric_base, accounting_standard)
    label = _base_metric_label_for_excel(display_base, industry_33)
    return f"{quarter} {label} \u5358\u72ec"


def _metric_label_for_excel(
    metric_base: str,
    industry_33: str | None = None,
    *,
    period_scope: str = "annual",
    accounting_standard: str | None = None,
) -> str:
    label = _metric_label_base_for_excel(metric_base, industry_33, accounting_standard)
    if period_scope == "annual":
        return f"4Q {label}"
    if period_scope.startswith("quarter:"):
        quarter = period_scope.split(":", 1)[1]
        return f"{quarter} {label}"
    if period_scope.startswith("quarter_standalone:"):
        quarter = period_scope.split(":", 1)[1]
        return _quarter_standalone_label(metric_base, industry_33, quarter, accounting_standard)
    if period_scope.startswith("forecast:"):
        stage = period_scope.split(":", 1)[1]
        stage_label = JQUANTS_FORECAST_STAGE_LABELS.get(stage, stage)
        return f"{stage_label} {label} \u4e88\u60f3"
    if period_scope == "half":
        return f"2Q {label}"
    if period_scope != "half":
        return label
    if metric_base in HALF_ONLY_BASES or label.startswith(HALF_PREFIX):
        return label
    prefix = HALF_YOY_PREFIX if metric_base in GROWTH_RATIO_BASES else HALF_PREFIX
    return f"{prefix}{label}"


def _add_half_label_aliases(mapping: dict[str, str], label: str, base: str) -> None:
    mapping[_normalize_text(f"{HALF_PREFIX}{label}")] = base
    mapping[_normalize_text(f"{HALF_YOY_PREFIX}{label}")] = base


def _build_label_to_base_map(sheet_name: str) -> dict[str, str]:
    mapping: dict[str, str] = {}
    candidate_bases = set(METRIC_BASE_LABELS) | set(DEFAULT_BASES_BY_SHEET[sheet_name]) | ABSORBED_RATIO_BASES
    for base in candidate_bases:
        mapping[_normalize_text(metric_base_to_display_name(base))] = base
        mapping[_normalize_text(metric_base_to_display_name(base, sheet_name))] = base
        mapping[_normalize_text(_base_metric_label_for_excel(base, sheet_name))] = base
        mapping[_normalize_text(_metric_label_for_excel(base, sheet_name))] = base
        if base == "NetSalesGrowthRate":
            mapping[_normalize_text("\u58f2\u4e0a\u9ad8\u5897\u53ce\u7387")] = base
        _add_half_label_aliases(mapping, metric_base_to_display_name(base), base)
        _add_half_label_aliases(mapping, metric_base_to_display_name(base, sheet_name), base)
        _add_half_label_aliases(mapping, _base_metric_label_for_excel(base, sheet_name), base)
        for quarter in ("1Q", "2Q", "3Q", "4Q"):
            mapping[_normalize_text(f"{quarter} {metric_base_to_display_name(base)}")] = base
            mapping[_normalize_text(f"{quarter} {_base_metric_label_for_excel(base, sheet_name)}")] = base
            mapping[_normalize_text(_quarter_standalone_label(base, sheet_name, quarter))] = base
            if base == "OrdinaryIncomeGrowthRate":
                mapping[_normalize_text(f"{quarter} 税引前利益増益率(前期比)")] = base
                mapping[_normalize_text(f"{quarter} 税引前利益増益率 単独(前期比)")] = base
        mapping[_normalize_text(f"2Q {metric_base_to_display_name(base)}")] = base
        mapping[_normalize_text(f"2Q {_base_metric_label_for_excel(base, sheet_name)}")] = base
        if base == "OrdinaryIncomeGrowthRate":
            mapping[_normalize_text("税引前利益増益率(前期比)")] = base
        for forecast_stage, stage_label in JQUANTS_FORECAST_STAGE_LABELS.items():
            mapping[_normalize_text(f"{stage_label} {metric_base_to_display_name(base)} \u4e88\u60f3")] = base
            mapping[_normalize_text(f"{stage_label} {_base_metric_label_for_excel(base, sheet_name)} \u4e88\u60f3")] = base
            mapping[_normalize_text(f"{stage_label} {metric_base_to_display_name(base)}(\u4e88\u60f3)")] = base
            mapping[_normalize_text(f"{stage_label} {_base_metric_label_for_excel(base, sheet_name)}(\u4e88\u60f3)")] = base
            if forecast_stage == "initial":
                mapping[_normalize_text(f"\u5f53\u671f {metric_base_to_display_name(base)} \u4e88\u60f3")] = base
                mapping[_normalize_text(f"\u5f53\u671f {_base_metric_label_for_excel(base, sheet_name)} \u4e88\u60f3")] = base
                mapping[_normalize_text(f"\u5f53\u671f {metric_base_to_display_name(base)}(\u4e88\u60f3)")] = base
                mapping[_normalize_text(f"\u5f53\u671f {_base_metric_label_for_excel(base, sheet_name)}(\u4e88\u60f3)")] = base
                mapping[_normalize_text(f"\u5f53\u521d {metric_base_to_display_name(base)} \u4e88\u60f3")] = base
                mapping[_normalize_text(f"\u5f53\u521d {_base_metric_label_for_excel(base, sheet_name)} \u4e88\u60f3")] = base
                mapping[_normalize_text(f"\u5f53\u521d {metric_base_to_display_name(base)}(\u4e88\u60f3)")] = base
                mapping[_normalize_text(f"\u5f53\u521d {_base_metric_label_for_excel(base, sheet_name)}(\u4e88\u60f3)")] = base
        for industry in SHEET_ORDER:
            mapping[_normalize_text(metric_base_to_display_name(base, industry))] = base
            mapping[_normalize_text(_metric_label_for_excel(base, industry))] = base
            _add_half_label_aliases(mapping, metric_base_to_display_name(base, industry), base)
            _add_half_label_aliases(mapping, _base_metric_label_for_excel(base, industry), base)
        mapping[_normalize_text(base)] = base

    for ratio_base, row_base in ROW_BASE_BY_RATIO_BASE.items():
        mapping[_normalize_text(metric_base_to_display_name(ratio_base))] = row_base
        mapping[_normalize_text(ratio_base)] = row_base
    mapping[_normalize_text("経常利益")] = "OrdinaryIncome"
    mapping[_normalize_text("税引前利益")] = "OrdinaryIncome"
    mapping[_normalize_text("税引き前利益")] = "OrdinaryIncome"
    mapping[_normalize_text("経常利益相当")] = "OrdinaryIncome"
    return mapping


def _build_label_to_value_base_map(sheet_name: str) -> dict[str, str]:
    mapping = _build_label_to_base_map(sheet_name)
    for ratio_base in ABSORBED_RATIO_BASES:
        mapping[_normalize_text(metric_base_to_display_name(ratio_base))] = ratio_base
        mapping[_normalize_text(ratio_base)] = ratio_base
    return mapping


def _resolve_row_bases(sheet_name: str, metric_labels: list[str], errors: list[str]) -> list[str]:
    if not metric_labels:
        return list(DEFAULT_BASES_BY_SHEET[sheet_name])

    mapping = _build_label_to_base_map(sheet_name)
    bases: list[str] = []
    seen: set[str] = set()
    for label in metric_labels:
        base = mapping.get(_normalize_text(label))
        if not base:
            errors.append(f"unknown_metric sheet={sheet_name} label={label}")
            continue
        if base in ABSORBED_RATIO_BASES:
            base = ROW_BASE_BY_RATIO_BASE[base]
        if base not in seen:
            bases.append(base)
            seen.add(base)
    return bases


def _resolve_value_bases(sheet_name: str, metric_labels: list[str], errors: list[str]) -> list[str]:
    mapping = _build_label_to_value_base_map(sheet_name)
    bases: list[str] = []
    seen: set[str] = set()
    for label in metric_labels:
        base = mapping.get(_normalize_text(label))
        if not base:
            errors.append(f"unknown_trend_metric sheet={sheet_name} label={label}")
            continue
        if base not in seen:
            bases.append(base)
            seen.add(base)
    return bases


def _fetch_ranked_filings(
    conn: sqlite3.Connection,
    condition: MetricExcelCondition,
) -> list[sqlite3.Row]:
    params: list[Any] = []
    form_types: list[str] = []
    for scope in condition.period_scopes:
        form_types.extend(FORM_TYPES_BY_PERIOD_SCOPE.get(scope, ()))
    if not form_types:
        return []
    form_type_placeholders = ",".join("?" for _ in form_types)
    where = [
        f"f.form_type IN ({form_type_placeholders})",
        "f.parse_status = 'derived_metrics_saved'",
        "coalesce(im.is_listed, 0) = 1",
        "coalesce(im.exchange, '') = 'TSE'",
    ]
    params.extend(form_types)

    if condition.industries:
        placeholders = ",".join("?" for _ in condition.industries)
        where.append(f"im.industry_33 IN ({placeholders})")
        params.extend(condition.industries)

    if condition.company_names:
        placeholders = ",".join("?" for _ in condition.company_names)
        where.append(f"im.company_name IN ({placeholders})")
        params.extend(condition.company_names)

    quarter_form_type_values = ",".join(
        "'" + form_type.replace("'", "''") + "'" for form_type in FORM_TYPES_BY_PERIOD_SCOPE["quarter"]
    )
    sql = f"""
    WITH base AS (
      SELECT
        f.doc_id,
        f.edinet_code,
        f.security_code,
        f.form_type,
        f.period_end,
        f.submit_date,
        f.accounting_standard,
        f.document_display_unit,
        im.company_name,
        im.industry_33,
        im.market,
        im.security_code AS issuer_security_code,
        CASE
          WHEN f.form_type IN ({quarter_form_type_values}) THEN 'quarter:2Q'
          ELSE 'annual'
        END AS period_scope_key,
        CASE
          WHEN f.form_type IN ({quarter_form_type_values})
            THEN COALESCE(date(f.period_end, 'start of month', '+6 months', '+1 month', '-1 day'), f.period_end)
          ELSE f.period_end
        END AS period_bucket_end
      FROM filings f
      JOIN issuer_master im
        ON im.edinet_code = f.edinet_code
      WHERE {" AND ".join(where)}
    ),
    ranked AS (
      SELECT
        *,
        -- Half filings are bucketed with their corresponding annual fiscal year.
        DENSE_RANK() OVER (
          PARTITION BY f.edinet_code, f.period_scope_key
          ORDER BY period_bucket_end DESC
        ) - 1 AS period_offset
      FROM base f
    )
    SELECT *
    FROM ranked
    WHERE period_offset BETWEEN 0 AND 10
    ORDER BY edinet_code, period_offset, form_type
    """

    rows = conn.execute(sql, params).fetchall()
    if not condition.security_codes:
        return rows

    allowed = set(condition.security_codes)
    filtered = []
    for row in rows:
        code = _normalize_security_code(row["issuer_security_code"] or row["security_code"] or "")
        if code in allowed:
            filtered.append(row)
    return filtered


def _metric_key(base: str) -> str:
    return f"{base}Current"


def _chunked(items: list[str], size: int) -> list[list[str]]:
    return [items[index : index + size] for index in range(0, len(items), size)]


def _fetch_metric_values(
    conn: sqlite3.Connection,
    *,
    doc_ids: list[str],
    metric_bases: list[str],
) -> dict[tuple[str, str], float | None]:
    if not doc_ids or not metric_bases:
        return {}

    metric_keys = [_metric_key(base) for base in metric_bases]
    key_placeholders = ",".join("?" for _ in metric_keys)

    values: dict[tuple[str, str], float | None] = {}
    unique_doc_ids = sorted(set(doc_ids))

    # SQLite builds often cap bound parameters, so large all-company exports read doc_ids in chunks.
    for doc_chunk in _chunked(unique_doc_ids, 800):
        doc_placeholders = ",".join("?" for _ in doc_chunk)
        normalized_rows = conn.execute(
            f"""
            SELECT doc_id, metric_key, value_num
            FROM normalized_metrics
            WHERE doc_id IN ({doc_placeholders})
              AND metric_key IN ({key_placeholders})
            """,
            [*doc_chunk, *metric_keys],
        ).fetchall()
        for row in normalized_rows:
            values[(str(row["doc_id"]), str(row["metric_key"]))] = row["value_num"]

    for doc_chunk in _chunked(unique_doc_ids, 800):
        doc_placeholders = ",".join("?" for _ in doc_chunk)
        derived_rows = conn.execute(
            f"""
            SELECT doc_id, metric_key, value_num, calc_status
            FROM derived_metrics
            WHERE doc_id IN ({doc_placeholders})
              AND metric_key IN ({key_placeholders})
            """,
            [*doc_chunk, *metric_keys],
        ).fetchall()
        for row in derived_rows:
            key = (str(row["doc_id"]), str(row["metric_key"]))
            if key in values and values[key] is not None:
                continue
            if str(row["calc_status"] or "") == "missing_input":
                values[key] = None
            else:
                values[key] = row["value_num"]
    if market_derived_table_exists(conn):
        for doc_chunk in _chunked(unique_doc_ids, 800):
            doc_placeholders = ",".join("?" for _ in doc_chunk)
            market_rows = conn.execute(
                f"""
                SELECT source_id AS doc_id, metric_key, value_num, calc_status
                FROM market_derived_metrics
                WHERE source_type = 'edinet'
                  AND source_id IN ({doc_placeholders})
                  AND metric_key IN ({key_placeholders})
                """,
                [*doc_chunk, *metric_keys],
            ).fetchall()
            for row in market_rows:
                key = (str(row["doc_id"]), str(row["metric_key"]))
                if str(row["calc_status"] or "") == "missing_input":
                    values[key] = None
                else:
                    values[key] = row["value_num"]
    return values


def _fetch_industry_aggregate_max_year(
    conn: sqlite3.Connection,
    industries: list[str],
) -> int | None:
    where = ["period_scope = ?"]
    params: list[Any] = [INDUSTRY_AGGREGATE_PERIOD_SCOPE]
    if industries:
        placeholders = ",".join("?" for _ in industries)
        where.append(f"industry_33 IN ({placeholders})")
        params.extend(industries)
    row = conn.execute(
        f"""
        SELECT MAX(fiscal_year) AS max_fiscal_year
        FROM industry_aggregate_metrics
        WHERE {" AND ".join(where)}
        """,
        params,
    ).fetchone()
    if row is None or row["max_fiscal_year"] is None:
        return None
    return int(row["max_fiscal_year"])


def _fetch_industry_aggregate_rows(
    conn: sqlite3.Connection,
    *,
    industries: list[str],
    fiscal_years: list[int],
    metric_bases: list[str],
) -> list[sqlite3.Row]:
    if not fiscal_years or not metric_bases:
        return []
    where = [
        "period_scope = ?",
        f"fiscal_year IN ({','.join('?' for _ in fiscal_years)})",
        f"metric_base IN ({','.join('?' for _ in metric_bases)})",
    ]
    params: list[Any] = [INDUSTRY_AGGREGATE_PERIOD_SCOPE, *fiscal_years, *metric_bases]
    if industries:
        where.append(f"industry_33 IN ({','.join('?' for _ in industries)})")
        params.extend(industries)
    rows = conn.execute(
        f"""
        SELECT
          industry_33,
          fiscal_year,
          period_bucket_end,
          metric_key,
          metric_base,
          value_num,
          value_unit,
          calc_status,
          source_company_count
        FROM industry_aggregate_metrics
        WHERE {" AND ".join(where)}
        ORDER BY industry_33, fiscal_year, metric_base
        """,
        params,
    ).fetchall()
    return rows


def _scale_value(metric_base: str, value: float | None) -> float | None:
    return _scale_value_for_document_unit(metric_base, value, "")


def _round_half_up(value: float) -> int:
    return int(Decimal(str(value)).quantize(Decimal("1"), rounding=ROUND_HALF_UP))


def _scale_value_for_document_unit(
    metric_base: str,
    value: float | None,
    document_display_unit: str | None,
) -> float | None:
    if value is None:
        return None
    if metric_base == "MarketCapitalization":
        return _round_half_up(value / 100_000_000)
    if metric_base == "OutstandingShares":
        return _round_half_up(value / 1_000)
    if metric_base in MONETARY_BASES:
        display_unit = str(document_display_unit or "").strip()
        if display_unit == "千円":
            return value / 1_000
        if display_unit == "百万円":
            return value / 1_000_000
    return value


def _display_unit_for_metric(metric_base: str, document_display_unit: str | None) -> str:
    if metric_base == "NumberOfEmployees":
        return "\u4eba"
    if metric_base == "AverageAge":
        return "\u6b73"
    if metric_base == "AverageLengthOfService":
        return "\u5e74"
    if metric_base == "AverageAnnualSalary":
        return "\u5186"
    if metric_base == "MarketCapitalization":
        return "\u5104\u5186"
    if metric_base == "StockPrice":
        return "\u5186"
    if metric_base in MONETARY_BASES:
        display_unit = str(document_display_unit or "").strip()
        if display_unit in {"百万円", "千円"}:
            return display_unit
        return "円"
    if metric_base == "OutstandingShares":
        return "\u5343\u682a"
    if metric_base in GROWTH_RATIO_BASES or metric_base in PERCENT_VALUE_BASES:
        return "%"
    if metric_base in RATIO_VALUE_BASES:
        return "倍"
    if metric_base in PER_SHARE_BASES or metric_base in ONE_DECIMAL_VALUE_BASES or metric_base in INTEGER_VALUE_BASES:
        return "円"
    return ""


def _period_display_for_filing(filing: sqlite3.Row | dict[str, Any] | None) -> str:
    if filing is None:
        return ""
    form_type = str(filing["form_type"] if isinstance(filing, sqlite3.Row) else filing.get("form_type") or "")
    scope_label = PERIOD_SCOPE_LABEL_BY_FORM_TYPE.get(form_type, form_type)
    period_end = str(filing["period_end"] if isinstance(filing, sqlite3.Row) else filing.get("period_end") or "")
    period_month = period_end[:7] if len(period_end) >= 7 else period_end
    if scope_label and period_month:
        return f"{scope_label} {period_month}"
    return period_month


def _period_point_display_for_filing(
    filing: sqlite3.Row | dict[str, Any] | None,
    period_scope: str,
) -> str:
    if filing is None:
        return ""
    period_end = str(filing["period_end"] if isinstance(filing, sqlite3.Row) else filing.get("period_end") or "")
    if not period_end:
        return ""
    if period_scope == "quarter:2Q":
        return f"2Q {period_end}\u6642\u70b9"
    form_type = str(filing["form_type"] if isinstance(filing, sqlite3.Row) else filing.get("form_type") or "")
    scope_label = PERIOD_SCOPE_LABEL_BY_FORM_TYPE.get(form_type, form_type)
    return f"{scope_label} {period_end}\u6642\u70b9" if scope_label else f"{period_end}\u6642\u70b9"


def _period_scope_label(period_scope: str) -> str:
    if period_scope == "half":
        return "\u56db\u534a\u671f"
    if period_scope.startswith("quarter_standalone"):
        return "\u56db\u534a\u671f\u5358\u72ec"
    if period_scope.startswith("quarter"):
        return "\u56db\u534a\u671f"
    if period_scope.startswith("forecast"):
        return "\u4e88\u60f3"
    return "\u901a\u671f"


def _decision_label_for_row(row: MetricExcelRow) -> str:
    if row.segment_kind:
        return _segment_decision_label(row.segment_kind)
    return _period_scope_label(row.period_scope)


def _quarter_label_from_period_scope(period_scope: str) -> str:
    if period_scope == "annual":
        return "4Q"
    if period_scope in {"half", "quarter"}:
        return "2Q"
    for prefix in ("quarter:", "quarter_standalone:"):
        if period_scope.startswith(prefix):
            quarter = period_scope[len(prefix):]
            if quarter in {"1Q", "2Q", "3Q", "4Q", "1~2Q", "3~4Q"}:
                return quarter
    return ""


def decision_label_for_excel(row: MetricExcelRow) -> str:
    if row.segment_kind:
        return _segment_decision_label(row.segment_kind)
    if row.period_scope.startswith("forecast"):
        return "\u4e88\u60f3"
    quarter_label = _quarter_label_from_period_scope(row.period_scope)
    if quarter_label:
        return quarter_label
    return _period_scope_label(row.period_scope)


def value_kind_label_for_excel(row: MetricExcelRow) -> str:
    if row.row_kind in {ROW_KIND_AVERAGE, ROW_KIND_MEDIAN}:
        return row.row_kind
    if row.segment_kind or row.period_scope.startswith("forecast"):
        return VALUE_KIND_BASE
    if row.row_kind == ROW_KIND_CHANGE_RATE:
        return VALUE_KIND_CALCULATED
    if row.period_scope.startswith("quarter_standalone"):
        return VALUE_KIND_CALCULATED
    if row.metric_base in CALCULATED_VALUE_KIND_BASES:
        return VALUE_KIND_CALCULATED
    return VALUE_KIND_BASE


def _row_kind_for_metric_base(metric_base: str) -> str:
    if metric_base in CHANGE_RATE_ROW_KIND_BASES:
        return ROW_KIND_CHANGE_RATE
    return ROW_KIND_DETAIL


def _is_detail_row(row: MetricExcelRow) -> bool:
    return row.row_kind in DETAIL_ROW_KINDS


def _source_offset_for_display(period_scope: str, display_offset: int) -> int | None:
    if period_scope == "annual":
        if display_offset == 0:
            return None
        return display_offset - 1
    return display_offset


def _parse_iso_date(value: Any) -> datetime | None:
    text = str(value or "").strip()[:10]
    if not text:
        return None
    try:
        return datetime.strptime(text, "%Y-%m-%d")
    except ValueError:
        return None


def _date_text(value: Any) -> str:
    parsed = _parse_iso_date(value)
    return parsed.strftime("%Y-%m-%d") if parsed is not None else ""


def _add_months_to_period_end(value: Any, months: int) -> str:
    parsed = _parse_iso_date(value)
    if parsed is None:
        return ""
    month_index = parsed.month - 1 + months
    year = parsed.year + month_index // 12
    month = month_index % 12 + 1
    day = min(parsed.day, calendar.monthrange(year, month)[1])
    return f"{year:04d}-{month:02d}-{day:02d}"


def _add_years_to_period_end(value: Any, years: int = 1) -> str:
    parsed = _parse_iso_date(value)
    if parsed is None:
        return ""
    try:
        return parsed.replace(year=parsed.year + years).strftime("%Y-%m-%d")
    except ValueError:
        # Leap-day fiscal years should remain at month-end.
        year = parsed.year + years
        day = calendar.monthrange(year, parsed.month)[1]
        return f"{year:04d}-{parsed.month:02d}-{day:02d}"


def _annual_current_period_end_from_latest_actual(value: Any) -> str:
    return _add_years_to_period_end(value, 1)


def _current_period_end_for_scope(value: Any, period_scope: str) -> str:
    if period_scope == "annual":
        return _annual_current_period_end_from_latest_actual(value)
    return _date_text(value)


def _calendar_year_bucket(period_bucket_end: str | None) -> int | None:
    text = str(period_bucket_end or "").strip()
    if len(text) < 4:
        return None
    try:
        year = int(text[:4])
    except ValueError:
        return None
    return year


def _fiscal_year_anchor_by_security_code(filings: list[sqlite3.Row]) -> dict[str, int]:
    anchors: dict[str, int] = {}
    for filing in filings:
        security_code = _normalize_security_code(
            filing["issuer_security_code"] or filing["security_code"] or ""
        )
        fiscal_year = _calendar_year_bucket(filing["period_bucket_end"])
        if not security_code or fiscal_year is None:
            continue
        anchors[security_code] = max(anchors.get(security_code, fiscal_year), fiscal_year)
    return anchors


def _fetch_jquants_anchor_filings(
    conn: sqlite3.Connection,
    condition: MetricExcelCondition,
    current_filings: list[sqlite3.Row],
) -> list[sqlite3.Row]:
    if not ({"quarter", "quarter_standalone"} & set(condition.period_scopes)):
        return current_filings
    if {"annual", "quarter"}.issubset(set(condition.period_scopes)):
        return current_filings
    anchor_condition = replace(condition, period_scopes=["annual", "quarter"])
    return [*current_filings, *_fetch_ranked_filings(conn, anchor_condition)]


def _aggregate_period_display(period_scope: str, calendar_year: int | None) -> str:
    if calendar_year is None:
        return ""
    return f"{_period_scope_label(period_scope)} {calendar_year}\u5e74\u6c7a\u7b97"


def _scale_industry_aggregate_value(metric_base: str, value: float | None) -> tuple[float | None, str]:
    if value is None:
        return None, ""
    if metric_base == "MarketCapitalization":
        return _round_half_up(value / 100_000_000), "\u5104\u5186"
    if metric_base == "OutstandingShares":
        return _round_half_up(value / 1_000), "\u5343\u682a"
    if metric_base in MONETARY_BASES:
        return value / 100_000_000, "\u5104\u5186"
    if metric_base == "NumberOfEmployees":
        return value, "\u4eba"
    return value, _display_unit_for_metric(metric_base, "")


def _aggregate_ratio_for_row_base(metric_base: str, sums: dict[str, float]) -> float | None:
    def ratio(numerator_base: str, denominator_base: str, *, positive_denominator: bool = True) -> float | None:
        numerator = sums.get(numerator_base)
        denominator = sums.get(denominator_base)
        if numerator is None or denominator is None:
            return None
        if positive_denominator and denominator <= 0:
            return None
        if denominator == 0:
            return None
        return numerator / denominator

    if metric_base in CONSTANT_RATIO_BY_ROW_BASE:
        return 1.0 if sums.get(metric_base) is not None else None
    if metric_base == "GrossProfit":
        return ratio("GrossProfit", "NetSales")
    if metric_base == "CostOfSales":
        return ratio("CostOfSales", "NetSales")
    if metric_base == "SellingExpenses":
        return ratio("SellingExpenses", "NetSales")
    if metric_base == "OperatingIncome":
        return ratio("OperatingIncome", "NetSales")
    if metric_base == "OrdinaryIncome":
        return ratio("OrdinaryIncome", "NetSales")
    if metric_base == "ProfitLoss":
        return ratio("ProfitLoss", "NetSales")
    if metric_base == "NetAssets":
        return ratio("NetAssets", "TotalAssets", positive_denominator=False)
    return None


def _append_warning_once(warnings: list[str], warning: str) -> None:
    if warning not in warnings:
        warnings.append(warning)


def _filter_excel_visible_bases(
    bases: list[str],
    *,
    warnings: list[str],
    explicit_metric_request: bool,
) -> list[str]:
    suppressed = [base for base in bases if base in SUPPRESSED_EXCEL_BASES]
    if suppressed and explicit_metric_request:
        labels = ", ".join(_base_metric_label_for_excel(base) for base in suppressed)
        _append_warning_once(warnings, f"excel_suppressed_metrics={labels}")
    return [base for base in bases if base not in SUPPRESSED_EXCEL_BASES]


def _resolve_segment_bases(metric_labels: list[str]) -> list[str]:
    if not metric_labels:
        return list(SEGMENT_METRIC_BASES)
    label_map = {
        _normalize_text(label): base
        for base, label in SEGMENT_EXCEL_METRIC_LABELS.items()
    }
    label_map.update(
        {
            _normalize_text("売上高"): "NetSales",
            _normalize_text("営業利益"): "OperatingIncome",
            _normalize_text("経常利益"): "ProfitBeforeTax",
            _normalize_text("経常利益相当"): "ProfitBeforeTax",
            _normalize_text("税引前利益"): "ProfitBeforeTax",
            _normalize_text("税引き前利益"): "ProfitBeforeTax",
            _normalize_text("純利益"): "ProfitLoss",
            _normalize_text("セグメント利益"): "SegmentProfit",
        }
    )
    bases: list[str] = []
    seen: set[str] = set()
    for label in metric_labels:
        base = label_map.get(_normalize_text(label))
        if base and base not in seen:
            bases.append(base)
            seen.add(base)
    return bases


def _segment_kind_matches(mode: str, kind: str) -> bool:
    if mode == "all":
        return kind in {"region", "business", "total"}
    if mode == "region":
        return kind in {"region", "total"}
    if mode == "business":
        return kind in {"business", "total"}
    return False


def _segment_value_priority(row: sqlite3.Row) -> int:
    return SEGMENT_VALUE_KIND_PRIORITY.get(str(row["value_kind"] or ""), 9)


def _segment_order_from_row(row: sqlite3.Row) -> int | None:
    try:
        detail = json.loads(str(row["source_detail_json"] or "{}"))
    except Exception:
        return None
    if not isinstance(detail, dict):
        return None
    try:
        value = detail.get("segment_order")
        return int(value) if value is not None else None
    except (TypeError, ValueError):
        return None


def _segment_scale_value(
    value: float | None,
    value_unit: str,
    *,
    document_display_unit: str | None = None,
) -> tuple[float | None, str]:
    if value is None:
        return None, ""
    if str(value_unit or "").lower() == "yen":
        if str(document_display_unit or "").strip() == "千円":
            return value / 1_000, "千円"
        return value / 1_000_000, "百万円"
    return value, str(value_unit or "")


def _segment_profit_display_key(row: sqlite3.Row) -> tuple[str, str, str, str, str]:
    member_qname = str(row["member_qname"] or "")
    if str(row["segment_kind"] or "") == "total":
        member_qname = "TOTAL"
    return (
        str(row["doc_id"] or ""),
        str(row["segment_kind"] or ""),
        member_qname,
        str(row["period_start"] or ""),
        str(row["period_end"] or ""),
    )


def _hide_redundant_segment_profit_rows(rows: list[sqlite3.Row]) -> list[sqlite3.Row]:
    operating_income_keys = {
        _segment_profit_display_key(row)
        for row in rows
        if str(row["metric_base"] or "") == "OperatingIncome"
    }
    return [
        row
        for row in rows
        if not (
            str(row["metric_base"] or "") == "SegmentProfit"
            and _segment_profit_display_key(row) in operating_income_keys
        )
    ]


def _segment_period_display(row: sqlite3.Row) -> str:
    period_end = str(row["period_end"] or "")
    if str(row["period_scope"] or "") == "quarter":
        quarter = str(row["quarter_type"] or "2Q")
        return f"{quarter} {period_end}" if period_end else quarter
    period_month = period_end[:7] if len(period_end) >= 7 else period_end
    return f"4Q {period_month}" if period_month else "4Q"


def _normalize_segment_name_for_excel(segment_name: Any) -> str:
    text = str(segment_name or "").strip()
    normalized = _normalize_text(text).lower()
    if (
        "operatingsegmentsnotincludedinreportablesegmentsandotherrevenuegeneratingbusinessactivities"
        in normalized
        or normalized == "otherreportablesegmentsmember"
    ):
        return "\u305d\u306e\u4ed6"
    return text


def _segment_decision_label(segment_kind: str) -> str:
    normalized = str(segment_kind or "").strip()
    return {
        "region": "地域別",
        "地域": "地域別",
        "地域別": "地域別",
        "business": "部門別",
        "部門": "部門別",
        "部門別": "部門別",
        "total": "合計",
        "合計": "合計",
    }.get(normalized, normalized)


def _segment_metric_label(
    *,
    metric_base: str,
    period_scope: str,
    quarter_type: str,
    segment_name: str,
) -> str:
    base_label = SEGMENT_EXCEL_METRIC_LABELS.get(metric_base, metric_base)
    prefix = (quarter_type or "2Q") if period_scope == "quarter" else "4Q"
    name = str(segment_name or "").strip() or "合計"
    return f"{prefix} {base_label} <{name}>"


def _fetch_segment_metric_rows(
    conn: sqlite3.Connection,
    condition: MetricExcelCondition,
    *,
    metric_bases: list[str],
) -> list[sqlite3.Row]:
    if condition.segment_mode == "none" or not metric_bases:
        return []
    scope_filters: list[str] = []
    if "annual" in condition.period_scopes:
        scope_filters.append("annual")
    if "quarter" in condition.period_scopes:
        scope_filters.append("quarter")
    if not scope_filters:
        return []

    where = [
        "sm.calc_status = 'ok'",
        "coalesce(sm.source_detail_json, '') NOT LIKE '%operating_income_segment_profit_fallback%'",
        f"sm.metric_base IN ({','.join('?' for _ in metric_bases)})",
        f"sm.period_scope IN ({','.join('?' for _ in scope_filters)})",
    ]
    params: list[Any] = [*metric_bases, *scope_filters]
    if "quarter" in scope_filters:
        where.append(
            "(sm.period_scope <> 'quarter' OR sm.quarter_type IN ("
            + ",".join("?" for _ in EDINET_SEGMENT_QUARTERS)
            + "))"
        )
        params.extend(EDINET_SEGMENT_QUARTERS)
    segment_kinds = [
        kind
        for kind in ("region", "business", "total")
        if _segment_kind_matches(condition.segment_mode, kind)
    ]
    where.append(f"sm.segment_kind IN ({','.join('?' for _ in segment_kinds)})")
    params.extend(segment_kinds)
    if condition.industries:
        where.append(f"im.industry_33 IN ({','.join('?' for _ in condition.industries)})")
        params.extend(condition.industries)
    if condition.security_codes:
        security_code_candidates = _security_code_candidates(condition.security_codes)
        where.append(f"sm.security_code IN ({','.join('?' for _ in security_code_candidates)})")
        params.extend(security_code_candidates)
    if condition.company_names:
        where.append(f"im.company_name IN ({','.join('?' for _ in condition.company_names)})")
        params.extend(condition.company_names)

    index_hint = (
        " INDEXED BY idx_segment_metrics_code_period"
        if condition.security_codes and _index_exists(conn, "idx_segment_metrics_code_period")
        else ""
    )
    rows = conn.execute(
        f"""
        SELECT
          sm.*,
          f.document_display_unit AS filing_document_display_unit,
          im.company_name,
          im.industry_33,
          im.market
        FROM segment_metrics AS sm{index_hint}
        LEFT JOIN filings f
          ON f.doc_id = sm.doc_id
        LEFT JOIN issuer_master im
          ON im.edinet_code = sm.edinet_code
        WHERE {' AND '.join(where)}
        ORDER BY sm.security_code, sm.period_scope, sm.quarter_type, sm.fiscal_year,
                 sm.segment_kind, sm.segment_name, sm.metric_base
        """,
        params,
    ).fetchall()
    return _hide_redundant_segment_profit_rows(rows)


def _build_segment_metric_excel_rows(
    conn: sqlite3.Connection,
    condition: MetricExcelCondition,
    warnings: list[str],
    *,
    span_recorder: SpanRecorder | None = None,
) -> list[MetricExcelRow]:
    if condition.segment_mode == "none":
        return []
    if not segment_metrics_table_exists(conn):
        _append_warning_once(warnings, "segment_metrics_not_ready")
        return []
    metric_bases = _resolve_segment_bases(condition.metric_labels)
    if not metric_bases:
        _append_warning_once(warnings, "segment_metrics_no_matching_metric")
        return []
    started = perf_counter()
    raw_rows = _fetch_segment_metric_rows(conn, condition, metric_bases=metric_bases)
    _record_span(
        span_recorder,
        "db_read",
        "fetch_segment_metric_rows",
        started,
        count=len(raw_rows),
    )
    if not raw_rows:
        _append_warning_once(warnings, "segment_metrics_empty")
        return []

    started = perf_counter()
    preferred_segment_names = preferred_segment_name_map(
        SegmentNameCandidate(
            edinet_code=str(raw["edinet_code"] or ""),
            segment_kind=str(raw["segment_kind"] or ""),
            member_qname=str(raw["member_qname"] or ""),
            segment_name=_normalize_segment_name_for_excel(raw["segment_name"]),
            period_end=_date_text(raw["period_end"]),
        )
        for raw in raw_rows
    )
    max_year_by_scope: dict[tuple[str, str, str, str], int] = {}
    current_period_end_by_scope: dict[tuple[str, str, str, str], str] = {}
    grouped: dict[tuple[str, str, str, str, str, str, str], dict[int, sqlite3.Row]] = {}
    segment_order_by_key: dict[tuple[str, str, str, str, str, str, str], int] = {}
    segment_name_by_key: dict[tuple[str, str, str, str, str, str, str], str] = {}
    for raw in raw_rows:
        fiscal_year = raw["fiscal_year"]
        if fiscal_year is None:
            continue
        security_code = _normalize_security_code(raw["security_code"])
        period_scope = str(raw["period_scope"] or "")
        quarter_type = str(raw["quarter_type"] or "")
        edinet_code = str(raw["edinet_code"] or "")
        scope_key = (security_code, period_scope, quarter_type, edinet_code)
        year = int(fiscal_year)
        period_end = _date_text(raw["period_end"])
        if year > max_year_by_scope.get(scope_key, -1):
            max_year_by_scope[scope_key] = year
            current_period_end_by_scope[scope_key] = (
                _annual_current_period_end_from_latest_actual(period_end)
                if period_scope == "annual"
                else period_end
            )
        elif year == max_year_by_scope.get(scope_key, -1) and period_end:
            current_period_end_by_scope[scope_key] = (
                _annual_current_period_end_from_latest_actual(period_end)
                if period_scope == "annual"
                else max(period_end, current_period_end_by_scope.get(scope_key, ""))
            )
        key = (
            security_code,
            period_scope,
            quarter_type,
            str(raw["segment_kind"] or ""),
            canonical_segment_key(str(raw["member_qname"] or ""), str(raw["segment_name"] or "")),
            str(raw["metric_base"] or ""),
            edinet_code,
        )
        preferred_name = preferred_segment_names.get(
            (
                edinet_code,
                str(raw["segment_kind"] or ""),
                key[4],
            ),
            _normalize_segment_name_for_excel(raw["segment_name"]),
        )
        segment_name_by_key[key] = preferred_name
        segment_order = _segment_order_from_row(raw)
        if segment_order is not None:
            segment_order_by_key[key] = min(segment_order_by_key.get(key, segment_order), segment_order)
        by_year = grouped.setdefault(key, {})
        previous = by_year.get(year)
        if previous is None or _segment_value_priority(raw) < _segment_value_priority(previous):
            by_year[year] = raw

    rows: list[MetricExcelRow] = []
    for key, by_year in grouped.items():
        security_code, period_scope, quarter_type, segment_kind, _segment_key, metric_base, edinet_code = key
        segment_name = segment_name_by_key.get(key, _segment_key)
        scope_key = (security_code, period_scope, quarter_type, edinet_code)
        max_year = max_year_by_scope.get(scope_key, max(by_year))
        periods_by_offset: dict[int, str] = {}
        values_by_offset: dict[int, float | None] = {}
        units_by_offset: dict[int, str] = {}
        ratios_by_offset: dict[int, float | None] = {}
        raw_values_by_offset: dict[int, float | None] = {}
        fiscal_years_by_offset: dict[int, int | None] = {}
        display_scope = f"quarter:{quarter_type or '2Q'}" if period_scope == "quarter" else "annual"
        current_period_end = current_period_end_by_scope.get(scope_key, "")
        sample_row = by_year.get(max_year) or by_year[max(by_year)]
        for offset in condition.period_offsets:
            source_offset = _source_offset_for_display(display_scope, offset)
            if source_offset is None:
                periods_by_offset[offset] = ""
                values_by_offset[offset] = None
                units_by_offset[offset] = ""
                ratios_by_offset[offset] = None
                raw_values_by_offset[offset] = None
                fiscal_years_by_offset[offset] = None
                continue
            source_year = max_year - source_offset
            fiscal_years_by_offset[offset] = source_year
            raw = by_year.get(source_year)
            if raw is None:
                periods_by_offset[offset] = ""
                values_by_offset[offset] = None
                units_by_offset[offset] = ""
                ratios_by_offset[offset] = None
                raw_values_by_offset[offset] = None
                continue
            raw_value = raw["value_num"]
            scaled_value, unit = _segment_scale_value(
                raw_value,
                str(raw["value_unit"] or ""),
                document_display_unit=str(raw["filing_document_display_unit"] or ""),
            )
            periods_by_offset[offset] = _segment_period_display(raw)
            values_by_offset[offset] = scaled_value
            units_by_offset[offset] = unit
            ratios_by_offset[offset] = None
            raw_values_by_offset[offset] = raw_value

        rows.append(
            MetricExcelRow(
                sheet_name=_sheet_name_for_industry(sample_row["industry_33"]),
                security_code=security_code,
                company_name=str(sample_row["company_name"] or ""),
                industry_33=str(sample_row["industry_33"] or ""),
                market=str(sample_row["market"] or ""),
                period_scope=display_scope,
                row_kind=_row_kind_for_metric_base(metric_base),
                current_period_end=current_period_end,
                metric_base=metric_base,
                metric_label=_segment_metric_label(
                    metric_base=metric_base,
                    period_scope=period_scope,
                    quarter_type=quarter_type,
                    segment_name=segment_name,
                ),
                periods_by_offset=periods_by_offset,
                values_by_offset=values_by_offset,
                units_by_offset=units_by_offset,
                ratios_by_offset=ratios_by_offset,
                raw_values_by_offset=raw_values_by_offset,
                fiscal_years_by_offset=fiscal_years_by_offset,
                segment_kind=_segment_decision_label(segment_kind),
                segment_name=segment_name,
                segment_order=segment_order_by_key.get(key),
            )
        )
    _record_span(
        span_recorder,
        "compute",
        "build_segment_rows",
        started,
        count=len(rows),
        detail={"segment_metric_rows": len(raw_rows)},
    )
    return rows


def _mean(values: list[float]) -> float | None:
    if not values:
        return None
    return sum(values) / len(values)


def _median(values: list[float]) -> float | None:
    if not values:
        return None
    ordered = sorted(values)
    midpoint = len(ordered) // 2
    if len(ordered) % 2:
        return ordered[midpoint]
    return (ordered[midpoint - 1] + ordered[midpoint]) / 2


def _stat_value_and_unit(
    metric_base: str,
    value: float | None,
    *,
    industry_only: bool,
) -> tuple[float | None, str]:
    if value is None:
        return None, ""
    if metric_base in MONETARY_BASES:
        if industry_only:
            return value / 100_000_000, "\u5104\u5186"
        return value / 1_000_000, "\u767e\u4e07\u5186"
    return value, _display_unit_for_metric(metric_base, "")


def _period_text_for_stats(rows: list[MetricExcelRow], offset: int) -> str:
    for row in rows:
        text = row.periods_by_offset.get(offset, "")
        if text:
            return text
    return ""


def _fiscal_year_for_stats(rows: list[MetricExcelRow], offset: int) -> int | None:
    for row in rows:
        fiscal_year = row.fiscal_years_by_offset.get(offset)
        if fiscal_year is not None:
            return fiscal_year
    return None


def _raw_value_for_export_stats(row: MetricExcelRow, offset: int) -> float | None:
    value = row.raw_values_by_offset.get(offset)
    if value is not None:
        return value
    return row.values_by_offset.get(offset)


def _assign_ranks(rows: list[MetricExcelRow], period_offsets: list[int]) -> None:
    groups: dict[tuple[str, str, str, int], list[tuple[int, float]]] = {}
    group_sizes: dict[tuple[str, str, str, int], int] = {}
    for index, row in enumerate(rows):
        if not _is_detail_row(row):
            continue
        for offset in period_offsets:
            key = (row.sheet_name, row.period_scope, row.metric_base, offset)
            group_sizes[key] = group_sizes.get(key, 0) + 1
            value = _raw_value_for_export_stats(row, offset)
            if value is None:
                continue
            groups.setdefault(key, []).append((index, float(value)))

    for key, values in groups.items():
        denominator = group_sizes.get(key, len(values))
        previous_value: float | None = None
        previous_rank = 0
        for position, (row_index, value) in enumerate(
            sorted(values, key=lambda item: item[1], reverse=True),
            start=1,
        ):
            if previous_value is None or value != previous_value:
                previous_rank = position
                previous_value = value
            rows[row_index].ranks_by_offset[key[3]] = f"{previous_rank}/{denominator}"


def _append_stat_rows(
    rows: list[MetricExcelRow],
    period_offsets: list[int],
    *,
    industry_only: bool,
) -> list[MetricExcelRow]:
    detail_rows = [row for row in rows if _is_detail_row(row)]
    grouped: dict[tuple[str, str, str], list[MetricExcelRow]] = {}
    for row in detail_rows:
        grouped.setdefault((row.sheet_name, row.period_scope, row.metric_base), []).append(row)

    stat_rows: list[MetricExcelRow] = []
    for (_sheet, _scope, metric_base), group_rows in grouped.items():
        for row_kind, aggregator in (
            (ROW_KIND_AVERAGE, _mean),
            (ROW_KIND_MEDIAN, _median),
        ):
            periods_by_offset: dict[int, str] = {}
            values_by_offset: dict[int, float | None] = {}
            units_by_offset: dict[int, str] = {}
            ratios_by_offset: dict[int, float | None] = {}
            raw_values_by_offset: dict[int, float | None] = {}
            fiscal_years_by_offset: dict[int, int | None] = {}

            for offset in period_offsets:
                raw_values = [
                    float(raw_value)
                    for row in group_rows
                    if (raw_value := _raw_value_for_export_stats(row, offset)) is not None
                ]
                ratio_values = [
                    float(row.ratios_by_offset[offset])
                    for row in group_rows
                    if row.ratios_by_offset.get(offset) is not None
                ]
                raw_stat = aggregator(raw_values)
                display_stat, unit = _stat_value_and_unit(
                    metric_base,
                    raw_stat,
                    industry_only=industry_only,
                )
                periods_by_offset[offset] = _period_text_for_stats(group_rows, offset)
                values_by_offset[offset] = display_stat
                units_by_offset[offset] = unit if raw_stat is not None else ""
                ratios_by_offset[offset] = aggregator(ratio_values)
                raw_values_by_offset[offset] = raw_stat
                fiscal_years_by_offset[offset] = _fiscal_year_for_stats(group_rows, offset)

            first = group_rows[0]
            stat_rows.append(
                MetricExcelRow(
                    sheet_name=first.sheet_name,
                    security_code="",
                    company_name="",
                    industry_33="",
                    market="",
                    period_scope=first.period_scope,
                    row_kind=row_kind,
                    current_period_end=first.current_period_end,
                    metric_base=metric_base,
                    metric_label=first.metric_label,
                    periods_by_offset=periods_by_offset,
                    values_by_offset=values_by_offset,
                    units_by_offset=units_by_offset,
                    ratios_by_offset=ratios_by_offset,
                    raw_values_by_offset=raw_values_by_offset,
                    fiscal_years_by_offset=fiscal_years_by_offset,
                )
            )
    return [*rows, *stat_rows]


def _passes_trend(values: list[float | None], direction: str) -> bool:
    if direction == "none":
        return True
    if len(values) < 2 or any(value is None for value in values):
        return False
    pairs = zip(values, values[1:])
    if direction == "increase":
        return all(float(left) < float(right) for left, right in pairs)
    return all(float(left) > float(right) for left, right in pairs)


def _passes_value_thresholds(
    values: list[float | None],
    *,
    min_value: float | None,
    max_value: float | None,
) -> bool:
    if min_value is None and max_value is None:
        return True
    if not values or any(value is None for value in values):
        return False
    for value in values:
        numeric_value = float(value)
        if min_value is not None and numeric_value < min_value:
            return False
        if max_value is not None and numeric_value > max_value:
            return False
    return True


def _passes_percent_filters(
    *,
    filter_bases: list[str],
    filter_offsets: list[int],
    period_scope: str,
    by_offset: dict[int, sqlite3.Row],
    metric_values: dict[tuple[str, str], float | None],
    min_value: float | None,
    max_value: float | None,
) -> bool:
    if not filter_bases or (min_value is None and max_value is None):
        return True
    offsets = filter_offsets or [0]
    for base in filter_bases:
        for offset in offsets:
            source_offset = _source_offset_for_display(period_scope, offset)
            if source_offset is None:
                return False
            filing = by_offset.get(source_offset)
            if filing is None:
                return False
            doc_id = str(filing["doc_id"])
            value = _metric_value(
                metric_values,
                doc_id,
                base,
                str(filing["accounting_standard"] or "") if "accounting_standard" in filing.keys() else "",
            )
            if value is None:
                return False
            numeric_value = float(value)
            if min_value is not None and numeric_value < min_value:
                return False
            if max_value is not None and numeric_value > max_value:
                return False
    return True


def _metric_value(
    metric_values: dict[tuple[str, str], float | None],
    doc_id: str,
    metric_base: str,
    accounting_standard: str | None = None,
) -> float | None:
    value_base = _display_base_for_accounting_standard(metric_base, accounting_standard)
    value = metric_values.get((doc_id, _metric_key(value_base)))
    if value is None and metric_base == "OrdinaryIncome" and not accounting_standard:
        return metric_values.get((doc_id, _metric_key("ProfitBeforeTax")))
    return value


def _period_blocks_for_rows(
    rows: list[MetricExcelRow],
    period_offsets: list[int],
) -> list[tuple[str, int]]:
    fiscal_years: set[int] = set()
    for row in rows:
        for offset in period_offsets:
            fiscal_year = row.fiscal_years_by_offset.get(offset)
            if fiscal_year is not None:
                fiscal_years.add(int(fiscal_year))
    if fiscal_years:
        return [("year", fiscal_year) for fiscal_year in sorted(fiscal_years)]
    return [("offset", offset) for offset in period_offsets]


def _period_block_label(block: tuple[str, int]) -> str:
    kind, value = block
    if kind == "year":
        return str(value)
    return PERIOD_LABEL_BY_OFFSET[value]


def _offset_for_period_block(
    row: MetricExcelRow,
    block: tuple[str, int],
    period_offsets: list[int],
) -> int | None:
    kind, value = block
    if kind == "offset":
        return value
    for offset in period_offsets:
        if row.fiscal_years_by_offset.get(offset) == value:
            return offset
    return None


def _build_preview_rows(rows: list[MetricExcelRow], periods: list[int], limit: int) -> list[dict[str, Any]]:
    preview = []
    period_blocks = _period_blocks_for_rows(rows, periods)
    for row in rows[:limit]:
        item: dict[str, Any] = {
            "security_code": row.security_code,
            "company_name": row.company_name,
            "industry_33": row.industry_33,
            "market": row.market,
            "period_scope": decision_label_for_excel(row),
            "metric": row.metric_label,
        }
        if row.segment_kind or row.segment_name:
            item["segment_kind"] = row.segment_kind
            item["segment_name"] = row.segment_name
        for block in period_blocks:
            offset = _offset_for_period_block(row, block, periods)
            item[_period_block_label(block)] = (
                row.values_by_offset.get(offset)
                if offset is not None
                else None
            )
        preview.append(item)
    return preview


def _row_sort_key(row: MetricExcelRow) -> tuple[int, int, int, int, int, str, str, int, str]:
    sheet_order = SHEET_ORDER.index(row.sheet_name)
    segment_order = 1 if row.segment_kind or row.segment_name else 0
    sheet_metric_order = ROW_BASE_ORDER_INDEX_BY_SHEET.get(row.sheet_name, ROW_BASE_ORDER_INDEX)
    metric_order = sheet_metric_order.get(row.metric_base, len(sheet_metric_order))
    scope_order = {
        "quarter:1Q": 0,
        "quarter:2Q": 1,
        "quarter:3Q": 2,
        "annual": 3,
        "half": 1,
        "quarter_standalone:1Q": 4,
        "quarter_standalone:2Q": 5,
        "quarter_standalone:3Q": 6,
        "quarter_standalone:4Q": 7,
        "quarter_standalone:1~2Q": 8,
        "quarter_standalone:3~4Q": 9,
        "forecast:initial": 10,
        "forecast:1Q": 11,
        "forecast:2Q": 12,
        "forecast:3Q": 13,
    }.get(row.period_scope, 99)
    row_kind_order = {
        ROW_KIND_DETAIL: 0,
        ROW_KIND_CHANGE_RATE: 0,
        ROW_KIND_AVERAGE: 1,
        ROW_KIND_MEDIAN: 2,
    }.get(row.row_kind, 9)
    return (
        sheet_order,
        segment_order,
        metric_order,
        scope_order,
        row_kind_order,
        row.security_code,
        row.segment_kind,
        row.segment_order if row.segment_order is not None else 999999,
        row.segment_name,
    )


def _aggregate_required_bases(row_bases: list[str]) -> list[str]:
    required = set(row_bases)
    for base in row_bases:
        ratio_base = ABSORBED_RATIO_BASE_BY_ROW_BASE.get(base)
        if ratio_base:
            required.add(ratio_base)
        if base in {
            "GrossProfit",
            "CostOfSales",
            "SellingExpenses",
            "OperatingIncome",
            "OrdinaryIncome",
            "ProfitLoss",
        }:
            required.add("NetSales")
        if base == "OrdinaryIncome":
            required.add("ProfitBeforeTax")
        if base == "NetAssets":
            required.add("TotalAssets")
    return sorted(required)


def _build_industry_only_metric_excel_rows(
    conn: sqlite3.Connection,
    condition: MetricExcelCondition,
    *,
    preview_limit: int,
) -> tuple[list[MetricExcelRow], list[str], list[str], list[dict[str, Any]], int]:
    errors: list[str] = []
    warnings: list[str] = []
    if condition.period_scopes != ["annual"]:
        warnings.append("industry_only_mode_forced_to_annual")
    if condition.security_codes:
        warnings.append("industry_only_mode_security_codes_filter_applied")
    if condition.company_names:
        warnings.append("industry_only_mode_company_names_filter_applied")
    if not industry_aggregate_table_exists(conn):
        warnings.append("industry_aggregate_metrics_not_ready")
        return [], errors, warnings, [], 0

    selected_row_bases = _resolve_row_bases(GENERAL_SHEET, condition.metric_labels, errors)
    unsupported_bases = [
        base
        for base in selected_row_bases
        if base not in INDUSTRY_AGGREGATE_VALUE_BASES
    ]
    if unsupported_bases and condition.metric_labels:
        labels = ", ".join(_base_metric_label_for_excel(base) for base in unsupported_bases)
        _append_warning_once(warnings, f"industry_only_suppressed_metrics={labels}")
    selected_row_bases = [
        base
        for base in selected_row_bases
        if base in INDUSTRY_AGGREGATE_VALUE_BASES
    ]
    selected_row_bases = _filter_excel_visible_bases(
        selected_row_bases,
        warnings=warnings,
        explicit_metric_request=bool(condition.metric_labels),
    )
    if not selected_row_bases:
        return [], errors, warnings, [], 0

    ratio_bases = [
        ratio_base
        for base in selected_row_bases
        if (ratio_base := ABSORBED_RATIO_BASE_BY_ROW_BASE.get(base))
    ]
    metric_bases_to_fetch = sorted(set(selected_row_bases + ratio_bases))
    max_fiscal_year = _fetch_industry_aggregate_max_year(conn, condition.industries)
    if max_fiscal_year is None:
        warnings.append("industry_aggregate_metrics_empty")
        return [], errors, warnings, [], 0
    fiscal_years = sorted(
        {
            max_fiscal_year - source_offset
            for offset in condition.period_offsets
            if (source_offset := _source_offset_for_display("annual", offset)) is not None
        }
    )
    aggregate_rows = _fetch_industry_aggregate_rows(
        conn,
        industries=condition.industries,
        fiscal_years=fiscal_years,
        metric_bases=metric_bases_to_fetch,
    )
    aggregate_by_key = {
        (str(row["industry_33"] or ""), int(row["fiscal_year"]), str(row["metric_base"] or "")): row
        for row in aggregate_rows
    }
    industries = sorted({str(row["industry_33"] or "") for row in aggregate_rows})
    rows: list[MetricExcelRow] = []
    for industry in industries:
        for base in selected_row_bases:
            periods_by_offset: dict[int, str] = {}
            values_by_offset: dict[int, float | None] = {}
            units_by_offset: dict[int, str] = {}
            ratios_by_offset: dict[int, float | None] = {}
            raw_values_by_offset: dict[int, float | None] = {}
            fiscal_years_by_offset: dict[int, int | None] = {}
            for offset in condition.period_offsets:
                source_offset = _source_offset_for_display("annual", offset)
                fiscal_year = max_fiscal_year - source_offset if source_offset is not None else None
                fiscal_years_by_offset[offset] = fiscal_year
                row = (
                    aggregate_by_key.get((industry, fiscal_year, base))
                    if fiscal_year is not None
                    else None
                )
                raw_value = (
                    float(row["value_num"])
                    if row is not None
                    and str(row["calc_status"] or "") == "ok"
                    and row["value_num"] is not None
                    else None
                )
                periods_by_offset[offset] = _aggregate_period_display(
                    INDUSTRY_AGGREGATE_PERIOD_SCOPE,
                    fiscal_year,
                )
                scaled_value, unit = _scale_industry_aggregate_value(base, raw_value)
                values_by_offset[offset] = scaled_value
                raw_values_by_offset[offset] = raw_value
                units_by_offset[offset] = unit if raw_value is not None else ""
                ratio_base = ABSORBED_RATIO_BASE_BY_ROW_BASE.get(base)
                if base in CONSTANT_RATIO_BY_ROW_BASE:
                    ratios_by_offset[offset] = 1.0 if raw_value is not None else None
                elif ratio_base:
                    ratio_row = aggregate_by_key.get((industry, fiscal_year, ratio_base))
                    ratios_by_offset[offset] = (
                        float(ratio_row["value_num"])
                        if ratio_row is not None
                        and str(ratio_row["calc_status"] or "") == "ok"
                        and ratio_row["value_num"] is not None
                        else None
                    )
                else:
                    ratios_by_offset[offset] = None
            rows.append(
                MetricExcelRow(
                    sheet_name=GENERAL_SHEET,
                    security_code="",
                    company_name="",
                    industry_33=industry,
                    market="",
                    period_scope=INDUSTRY_AGGREGATE_PERIOD_SCOPE,
                    row_kind=_row_kind_for_metric_base(base),
                    current_period_end=_aggregate_period_display(
                        INDUSTRY_AGGREGATE_PERIOD_SCOPE,
                        max_fiscal_year,
                    ),
                    metric_base=base,
                    metric_label=_metric_label_for_excel(
                        base,
                        industry,
                        period_scope=INDUSTRY_AGGREGATE_PERIOD_SCOPE,
                    ),
                    periods_by_offset=periods_by_offset,
                    values_by_offset=values_by_offset,
                    units_by_offset=units_by_offset,
                    ratios_by_offset=ratios_by_offset,
                    raw_values_by_offset=raw_values_by_offset,
                    fiscal_years_by_offset=fiscal_years_by_offset,
                )
            )

    _assign_ranks(rows, condition.period_offsets)
    rows = _append_stat_rows(rows, condition.period_offsets, industry_only=True)
    rows.sort(key=_row_sort_key)
    preview_rows = _build_preview_rows(rows, condition.period_offsets, preview_limit)
    return rows, errors, warnings, preview_rows, len(industries)


def _jquants_table_exists(conn: sqlite3.Connection, table_name: str) -> bool:
    row = conn.execute(
        """
        SELECT name
        FROM sqlite_master
        WHERE type = 'table' AND name = ?
        """,
        (table_name,),
    ).fetchone()
    return row is not None


def _fetch_jquants_companies(
    conn: sqlite3.Connection,
    condition: MetricExcelCondition,
) -> list[sqlite3.Row]:
    where = [
        "coalesce(im.is_listed, 0) = 1",
        "coalesce(im.exchange, '') = 'TSE'",
    ]
    params: list[Any] = []
    if condition.industries:
        placeholders = ",".join("?" for _ in condition.industries)
        where.append(f"im.industry_33 IN ({placeholders})")
        params.extend(condition.industries)
    if condition.company_names:
        placeholders = ",".join("?" for _ in condition.company_names)
        where.append(f"im.company_name IN ({placeholders})")
        params.extend(condition.company_names)
    if condition.security_codes:
        codes = [_normalize_security_code(code) for code in condition.security_codes]
        placeholders = ",".join("?" for _ in codes)
        where.append(
            f"(substr(coalesce(im.security_code, ''), 1, 4) IN ({placeholders}) "
            f"OR coalesce(im.security_code, '') IN ({placeholders}))"
        )
        params.extend([*codes, *codes])
    rows = conn.execute(
        f"""
        SELECT
          im.edinet_code,
          im.security_code,
          im.company_name,
          im.industry_33,
          im.market,
          (
            SELECT f.accounting_standard
            FROM filings f
            WHERE f.edinet_code = im.edinet_code
              AND f.accounting_standard IS NOT NULL
              AND f.accounting_standard <> ''
            ORDER BY f.period_end DESC, f.submit_date DESC
            LIMIT 1
          ) AS accounting_standard,
          (
            SELECT f.document_display_unit
            FROM filings f
            WHERE f.edinet_code = im.edinet_code
              AND f.document_display_unit IS NOT NULL
              AND f.document_display_unit <> ''
            ORDER BY f.period_end DESC, f.submit_date DESC
            LIMIT 1
          ) AS document_display_unit
        FROM issuer_master im
        WHERE {" AND ".join(where)}
        ORDER BY im.security_code, im.edinet_code
        """,
        params,
    ).fetchall()
    return rows


def _security_code_for_jquants(row: sqlite3.Row) -> str:
    return _normalize_security_code(str(row["security_code"] or ""))


def _latest_jquants_metric_rows(
    rows: list[sqlite3.Row],
) -> dict[tuple[str, str, str, int, str, str], sqlite3.Row]:
    latest: dict[tuple[str, str, str, int, str, str], sqlite3.Row] = {}
    for row in rows:
        fiscal_year = row["fiscal_year"]
        if fiscal_year is None:
            continue
        key = (
            _normalize_security_code(row["security_code"] or row["local_code"] or ""),
            str(row["period_scope"] or ""),
            str(row["metric_base"] or ""),
            int(fiscal_year),
            str(row["period_key"] or ""),
            str(row["forecast_stage"] or ""),
        )
        existing = latest.get(key)
        if existing is None:
            latest[key] = row
            continue
        existing_order = (str(existing["disclosed_date"] or ""), str(existing["disclosed_time"] or ""))
        row_order = (str(row["disclosed_date"] or ""), str(row["disclosed_time"] or ""))
        if row_order >= existing_order:
            latest[key] = row
    return latest


def _fetch_jquants_metric_rows(
    conn: sqlite3.Connection,
    *,
    security_codes: list[str],
    metric_bases: list[str],
    min_fiscal_year: int | None = None,
) -> list[sqlite3.Row]:
    if not security_codes or not metric_bases:
        return []
    columns = {
        str(row[1])
        for row in conn.execute("PRAGMA table_info(jquants_financial_metrics)").fetchall()
    }
    forecast_stage_expr = "forecast_stage" if "forecast_stage" in columns else "NULL AS forecast_stage"
    local_codes = _jquants_local_code_candidates(security_codes)
    local_code_placeholders = ",".join("?" for _ in local_codes)
    base_placeholders = ",".join("?" for _ in metric_bases)
    where = [
        f"local_code IN ({local_code_placeholders})",
        f"metric_base IN ({base_placeholders})",
    ]
    params: list[Any] = [*local_codes, *metric_bases]
    if min_fiscal_year is not None:
        where.append("fiscal_year >= ?")
        params.append(min_fiscal_year)
    rows = conn.execute(
        f"""
        SELECT
          disclosure_number,
          local_code,
          security_code,
          metric_kind,
          period_scope,
          period_key,
          quarter_type,
          forecast_target,
          {forecast_stage_expr},
          fiscal_year,
          period_start,
          period_end,
          disclosed_date,
          disclosed_time,
          metric_key,
          metric_base,
          metric_group,
          value_num,
          value_unit,
          calc_status,
          source_field
        FROM jquants_financial_metrics
        WHERE {" AND ".join(where)}
        ORDER BY security_code, fiscal_year, period_key, forecast_stage, metric_base, disclosed_date, disclosed_time
        """,
        params,
    ).fetchall()
    if market_derived_table_exists(conn):
        market_metric_bases = [base for base in metric_bases if base in MARKET_METRIC_BASES]
        if market_metric_bases:
            code_placeholders = ",".join("?" for _ in security_codes)
            market_base_placeholders = ",".join("?" for _ in market_metric_bases)
            market_where = [
                "source_type = 'jquants'",
                f"security_code IN ({code_placeholders})",
                f"metric_base IN ({market_base_placeholders})",
            ]
            market_params: list[Any] = [*security_codes, *market_metric_bases]
            if min_fiscal_year is not None:
                market_where.append("fiscal_year >= ?")
                market_params.append(min_fiscal_year)
            rows.extend(
                conn.execute(
                    f"""
                    SELECT
                      source_id AS disclosure_number,
                      '' AS local_code,
                      security_code,
                      'market_derived' AS metric_kind,
                      period_scope,
                      period_key,
                      quarter_type,
                      NULL AS forecast_target,
                      NULL AS forecast_stage,
                      fiscal_year,
                      NULL AS period_start,
                      period_end,
                      period_end AS disclosed_date,
                      '' AS disclosed_time,
                      metric_key,
                      metric_base,
                      metric_group,
                      value_num,
                      value_unit,
                      calc_status,
                      formula_name AS source_field
                    FROM market_derived_metrics
                    WHERE {" AND ".join(market_where)}
                    ORDER BY security_code, fiscal_year, period_key, metric_base
                    """,
                    market_params,
                ).fetchall()
            )
    return rows


def _fetch_quarter_standalone_metric_rows(
    conn: sqlite3.Connection,
    *,
    security_codes: list[str],
    metric_bases: list[str],
) -> list[sqlite3.Row]:
    if not security_codes or not metric_bases or not quarter_standalone_table_exists(conn):
        return []
    code_placeholders = ",".join("?" for _ in security_codes)
    base_placeholders = ",".join("?" for _ in metric_bases)
    return conn.execute(
        f"""
        SELECT
          security_code,
          edinet_code,
          fiscal_year,
          quarter_type,
          period_end,
          metric_key,
          metric_base,
          metric_group,
          value_num,
          value_unit,
          calc_status
        FROM quarter_standalone_metrics
        WHERE security_code IN ({code_placeholders})
          AND metric_base IN ({base_placeholders})
        ORDER BY security_code, fiscal_year, quarter_type, metric_base
        """,
        [*security_codes, *metric_bases],
    ).fetchall()


def _latest_quarter_standalone_metric_rows(
    rows: list[sqlite3.Row],
) -> dict[tuple[str, str, str, int], sqlite3.Row]:
    latest: dict[tuple[str, str, str, int], sqlite3.Row] = {}
    for row in rows:
        if row["fiscal_year"] is None:
            continue
        key = (
            _normalize_security_code(row["security_code"]),
            str(row["quarter_type"] or ""),
            str(row["metric_base"] or ""),
            int(row["fiscal_year"]),
        )
        latest[key] = row
    return latest


def _build_jquants_lookup_indexes(rows: list[sqlite3.Row]) -> _JQuantsLookupIndexes:
    max_fiscal_year_by_period: dict[tuple[str, str, str], int] = {}
    forecast_values_by_as_of_key: dict[tuple[str, int, str], list[tuple[str, str, float]]] = {}
    forecast_value_by_stage: dict[tuple[str, int, str, str], float] = {}
    forecast_order_by_stage: dict[tuple[str, int, str, str], tuple[str, str]] = {}
    for row in rows:
        if row["fiscal_year"] is None:
            continue
        security_code = _normalize_security_code(row["security_code"] or row["local_code"] or "")
        fiscal_year = int(row["fiscal_year"])
        period_key = str(row["period_key"] or "")
        forecast_stage = str(row["forecast_stage"] or "")
        for stage_key in {"", forecast_stage}:
            key = (security_code, period_key, stage_key)
            max_fiscal_year_by_period[key] = max(max_fiscal_year_by_period.get(key, fiscal_year), fiscal_year)
        if (
            str(row["period_scope"] or "") != "forecast"
            or period_key != "forecast:FY"
            or str(row["calc_status"] or "") != "ok"
            or row["value_num"] is None
        ):
            continue
        metric_base = str(row["metric_base"] or "")
        disclosed_date = str(row["disclosed_date"] or "")
        disclosed_time = str(row["disclosed_time"] or "")
        value = float(row["value_num"])
        forecast_values_by_as_of_key.setdefault(
            (security_code, fiscal_year, metric_base),
            [],
        ).append((disclosed_date, disclosed_time, value))
        stage_key = (security_code, fiscal_year, metric_base, forecast_stage)
        row_order = (disclosed_date, disclosed_time)
        if row_order >= forecast_order_by_stage.get(stage_key, ("", "")):
            forecast_order_by_stage[stage_key] = row_order
            forecast_value_by_stage[stage_key] = value
    for values in forecast_values_by_as_of_key.values():
        values.sort()
    return _JQuantsLookupIndexes(
        max_fiscal_year_by_period=max_fiscal_year_by_period,
        forecast_values_by_as_of_key=forecast_values_by_as_of_key,
        forecast_value_by_stage=forecast_value_by_stage,
    )


def _build_quarter_standalone_lookup_indexes(
    rows: list[sqlite3.Row],
) -> _QuarterStandaloneLookupIndexes:
    max_fiscal_year_by_code: dict[str, int] = {}
    period_end_by_code_year_quarter: dict[tuple[str, int, str], str] = {}
    for row in rows:
        if row["fiscal_year"] is None:
            continue
        security_code = _normalize_security_code(row["security_code"] or "")
        fiscal_year = int(row["fiscal_year"])
        quarter_type = str(row["quarter_type"] or "")
        max_fiscal_year_by_code[security_code] = max(
            max_fiscal_year_by_code.get(security_code, fiscal_year),
            fiscal_year,
        )
        period_end = _date_text(row["period_end"])
        if not period_end:
            continue
        key = (security_code, fiscal_year, quarter_type)
        period_end_by_code_year_quarter[key] = max(
            period_end_by_code_year_quarter.get(key, ""),
            period_end,
        )
    return _QuarterStandaloneLookupIndexes(
        max_fiscal_year_by_code=max_fiscal_year_by_code,
        period_end_by_code_year_quarter=period_end_by_code_year_quarter,
    )


def _max_quarter_standalone_fiscal_year(
    rows: list[sqlite3.Row],
    security_code: str,
    *,
    lookup_indexes: _QuarterStandaloneLookupIndexes | None = None,
) -> int | None:
    if lookup_indexes is not None:
        return lookup_indexes.max_fiscal_year_by_code.get(security_code)
    years = [
        int(row["fiscal_year"])
        for row in rows
        if _normalize_security_code(row["security_code"]) == security_code
        and row["fiscal_year"] is not None
    ]
    return max(years) if years else None


def _max_jquants_fiscal_year(
    rows: list[sqlite3.Row],
    security_code: str,
    period_key: str,
    forecast_stage: str | None = None,
    *,
    lookup_indexes: _JQuantsLookupIndexes | None = None,
) -> int | None:
    if lookup_indexes is not None:
        return lookup_indexes.max_fiscal_year_by_period.get(
            (security_code, period_key, forecast_stage or ""),
        )
    years = [
        int(row["fiscal_year"])
        for row in rows
        if _normalize_security_code(row["security_code"] or row["local_code"] or "") == security_code
        and str(row["period_key"] or "") == period_key
        and (forecast_stage is None or str(row["forecast_stage"] or "") == forecast_stage)
        and row["fiscal_year"] is not None
    ]
    return max(years) if years else None


def _period_end_from_jquants_row(row: sqlite3.Row | None) -> str:
    if row is None:
        return ""
    return _date_text(row["period_end"])


def _quarter_standalone_period_end(
    rows: list[sqlite3.Row],
    *,
    security_code: str,
    fiscal_year: int,
    quarter: str,
    lookup_indexes: _QuarterStandaloneLookupIndexes | None = None,
) -> str:
    if lookup_indexes is not None:
        period_end = lookup_indexes.period_end_by_code_year_quarter.get(
            (security_code, fiscal_year, quarter),
            "",
        )
        if period_end:
            return period_end
    else:
        matching = [
            _date_text(row["period_end"])
            for row in rows
            if _normalize_security_code(row["security_code"] or "") == security_code
            and row["fiscal_year"] is not None
            and int(row["fiscal_year"]) == fiscal_year
            and str(row["quarter_type"] or "") == quarter
            and _date_text(row["period_end"])
        ]
        if matching:
            return max(matching)

    quarter_order = ["1Q", "2Q", "3Q", "4Q"]
    if quarter not in quarter_order:
        return ""
    target_index = quarter_order.index(quarter)
    ordered_fallback_quarters = sorted(
        (item for item in quarter_order if item != quarter),
        key=lambda item: (abs(target_index - quarter_order.index(item)), quarter_order.index(item) > target_index),
    )
    for other in ordered_fallback_quarters:
        if lookup_indexes is not None:
            other_period_end = lookup_indexes.period_end_by_code_year_quarter.get(
                (security_code, fiscal_year, other),
                "",
            )
            if not other_period_end:
                continue
        else:
            other_dates = [
                _date_text(row["period_end"])
                for row in rows
                if _normalize_security_code(row["security_code"] or "") == security_code
                and row["fiscal_year"] is not None
                and int(row["fiscal_year"]) == fiscal_year
                and str(row["quarter_type"] or "") == other
                and _date_text(row["period_end"])
            ]
            if not other_dates:
                continue
            other_period_end = max(other_dates)
        other_index = quarter_order.index(other)
        return _add_months_to_period_end(other_period_end, (target_index - other_index) * 3)
    return ""


def _quarter_standalone_period_display(
    rows: list[sqlite3.Row],
    *,
    security_code: str,
    fiscal_year: int,
    quarter: str,
    lookup_indexes: _QuarterStandaloneLookupIndexes | None = None,
) -> str:
    period_end = _quarter_standalone_period_end(
        rows,
        security_code=security_code,
        fiscal_year=fiscal_year,
        quarter=quarter,
        lookup_indexes=lookup_indexes,
    )
    return f"{quarter} {period_end}" if period_end else quarter


def _jquants_period_display(row: sqlite3.Row | None, period_scope: str, metric_base: str | None = None) -> str:
    if row is None:
        return ""
    period_end = str(row["period_end"] or "")
    if metric_base in DATE_POINT_PERIOD_BASES and period_scope.startswith("quarter:"):
        quarter = period_scope.split(":", 1)[1]
        return f"{quarter} {period_end}\u6642\u70b9" if period_end else quarter
    if metric_base in DATE_POINT_PERIOD_BASES and period_scope.startswith("quarter_standalone:"):
        quarter = period_scope.split(":", 1)[1]
        return f"{quarter} {period_end}\u6642\u70b9" if period_end else quarter
    period_month = period_end[:7] if len(period_end) >= 7 else period_end
    if period_scope.startswith("quarter:"):
        quarter = period_scope.split(":", 1)[1]
        return f"{quarter} {period_month}" if period_month else quarter
    if period_scope.startswith("quarter_standalone:"):
        quarter = period_scope.split(":", 1)[1]
        return f"{quarter} {period_month}" if period_month else quarter
    if period_scope.startswith("forecast:"):
        stage = period_scope.split(":", 1)[1]
        stage_label = f"{JQUANTS_FORECAST_STAGE_LABELS.get(stage, stage)}\u4e88\u60f3"
        return f"{stage_label} {period_month}" if period_month else stage_label
    return period_month


def _scale_jquants_value(
    metric_base: str,
    value: float | None,
    *,
    document_display_unit: str | None = None,
) -> tuple[float | None, str]:
    if value is None:
        return None, ""
    if metric_base == "MarketCapitalization":
        return _round_half_up(value / 100_000_000), "\u5104\u5186"
    if metric_base == "OutstandingShares":
        return _round_half_up(value / 1_000), "\u5343\u682a"
    if metric_base in MONETARY_BASES:
        if str(document_display_unit or "").strip() == "\u5343\u5186":
            return value / 1_000, "\u5343\u5186"
        return value / 1_000_000, "\u767e\u4e07\u5186"
    if metric_base in {"IssuedShares", "TreasuryShares"}:
        return value, "\u682a"
    if metric_base in GROWTH_RATIO_BASES or metric_base in PERCENT_VALUE_BASES:
        return value, "%"
    if metric_base in RATIO_VALUE_BASES:
        return value, "\u500d"
    if metric_base in PER_SHARE_BASES or metric_base in ONE_DECIMAL_VALUE_BASES:
        return value, "\u5186"
    return value, ""


def _jquants_metric_fetch_bases(requested_bases: set[str]) -> set[str]:
    fetch_bases = set(requested_bases)
    for growth_base, source_base in QUARTER_CUMULATIVE_GROWTH_SOURCE_BY_BASE.items():
        if growth_base in fetch_bases:
            fetch_bases.add(source_base)
    if {"OrdinaryIncome", "OrdinaryIncomeGrowthRate", "EstimatedNetIncome"} & fetch_bases:
        fetch_bases.update({"OrdinaryIncome", "ProfitBeforeTax"})
    if "FCF" in fetch_bases:
        fetch_bases.update({"OperatingCash", "InvestmentCash"})
    if "BPS" in fetch_bases:
        fetch_bases.update({"NetAssets", "OutstandingShares"})
    if "InvestmentCashToNetSalesRatio" in fetch_bases:
        fetch_bases.update({"InvestmentCash", "NetSales"})
    if "InvestmentCashToOperatingCashRatio" in fetch_bases:
        fetch_bases.update({"InvestmentCash", "OperatingCash"})
    return fetch_bases


def _jquants_latest_row(
    latest: dict[tuple[str, str, str, int, str, str], sqlite3.Row],
    *,
    security_code: str,
    period_scope: str,
    metric_base: str,
    fiscal_year: int,
    period_key: str,
    forecast_stage: str | None,
) -> sqlite3.Row | None:
    return latest.get(
        (
            security_code,
            "quarter" if period_scope.startswith("quarter") else "forecast",
            metric_base,
            fiscal_year,
            period_key,
            forecast_stage or "",
        )
    )


def _jquants_ok_value(row: sqlite3.Row | None) -> float | None:
    if (
        row is not None
        and str(row["calc_status"] or "") == "ok"
        and row["value_num"] is not None
    ):
        return float(row["value_num"])
    return None


def _jquants_display_row_and_value(
    latest: dict[tuple[str, str, str, int, str, str], sqlite3.Row],
    *,
    security_code: str,
    period_scope: str,
    metric_base: str,
    fiscal_year: int,
    period_key: str,
    forecast_stage: str | None,
    accounting_standard: str | None = None,
) -> tuple[sqlite3.Row | None, float | None]:
    if metric_base == "CashBalanceGrowthRate" and period_scope.startswith("quarter"):
        row = _jquants_latest_row(
            latest,
            security_code=security_code,
            period_scope=period_scope,
            metric_base=metric_base,
            fiscal_year=fiscal_year,
            period_key=period_key,
            forecast_stage=forecast_stage,
        )
        if row is not None:
            return row, _jquants_ok_value(row)
    if metric_base in QUARTER_CUMULATIVE_GROWTH_SOURCE_BY_BASE and period_scope.startswith("quarter"):
        source_base = QUARTER_CUMULATIVE_GROWTH_SOURCE_BY_BASE[metric_base]
        current_row, current_value = _jquants_display_row_and_value(
            latest,
            security_code=security_code,
            period_scope=period_scope,
            metric_base=source_base,
            fiscal_year=fiscal_year,
            period_key=period_key,
            forecast_stage=forecast_stage,
            accounting_standard=accounting_standard,
        )
        prior_row, prior_value = _jquants_display_row_and_value(
            latest,
            security_code=security_code,
            period_scope=period_scope,
            metric_base=source_base,
            fiscal_year=fiscal_year - 1,
            period_key=period_key,
            forecast_stage=forecast_stage,
            accounting_standard=accounting_standard,
        )
        value = (
            current_value / prior_value
            if current_value is not None and prior_value is not None and prior_value > 0
            else None
        )
        return current_row or prior_row, value
    if metric_base == "EstimatedNetIncome" and period_scope.startswith("quarter"):
        profit_base = _display_base_for_accounting_standard("OrdinaryIncome", accounting_standard)
        ordinary_row = _jquants_latest_row(
            latest,
            security_code=security_code,
            period_scope=period_scope,
            metric_base=profit_base,
            fiscal_year=fiscal_year,
            period_key=period_key,
            forecast_stage=forecast_stage,
        )
        ordinary_value = _jquants_ok_value(ordinary_row)
        value = ordinary_value * 0.7 if ordinary_value is not None else None
        return ordinary_row, value
    if metric_base in {"InvestmentCashToNetSalesRatio", "InvestmentCashToOperatingCashRatio"} and period_scope.startswith("quarter"):
        investment_row = _jquants_latest_row(
            latest,
            security_code=security_code,
            period_scope=period_scope,
            metric_base="InvestmentCash",
            fiscal_year=fiscal_year,
            period_key=period_key,
            forecast_stage=forecast_stage,
        )
        denominator_base = "NetSales" if metric_base == "InvestmentCashToNetSalesRatio" else "OperatingCash"
        denominator_row = _jquants_latest_row(
            latest,
            security_code=security_code,
            period_scope=period_scope,
            metric_base=denominator_base,
            fiscal_year=fiscal_year,
            period_key=period_key,
            forecast_stage=forecast_stage,
        )
        investment_value = _jquants_ok_value(investment_row)
        denominator_value = _jquants_ok_value(denominator_row)
        value = (
            abs(investment_value) / denominator_value
            if investment_value is not None and denominator_value and denominator_value > 0
            else None
        )
        return investment_row or denominator_row, value
    if metric_base == "FCF" and period_scope.startswith("quarter"):
        operating_row = _jquants_latest_row(
            latest,
            security_code=security_code,
            period_scope=period_scope,
            metric_base="OperatingCash",
            fiscal_year=fiscal_year,
            period_key=period_key,
            forecast_stage=forecast_stage,
        )
        investment_row = _jquants_latest_row(
            latest,
            security_code=security_code,
            period_scope=period_scope,
            metric_base="InvestmentCash",
            fiscal_year=fiscal_year,
            period_key=period_key,
            forecast_stage=forecast_stage,
        )
        operating_value = _jquants_ok_value(operating_row)
        investment_value = _jquants_ok_value(investment_row)
        value = (
            operating_value + investment_value
            if operating_value is not None and investment_value is not None
            else None
        )
        return operating_row or investment_row, value
    if metric_base == "BPS" and period_scope.startswith("quarter"):
        row = _jquants_latest_row(
            latest,
            security_code=security_code,
            period_scope=period_scope,
            metric_base=metric_base,
            fiscal_year=fiscal_year,
            period_key=period_key,
            forecast_stage=forecast_stage,
        )
        direct_value = _jquants_ok_value(row)
        if direct_value is not None:
            return row, direct_value
        net_assets_row = _jquants_latest_row(
            latest,
            security_code=security_code,
            period_scope=period_scope,
            metric_base="NetAssets",
            fiscal_year=fiscal_year,
            period_key=period_key,
            forecast_stage=forecast_stage,
        )
        shares_row = _jquants_latest_row(
            latest,
            security_code=security_code,
            period_scope=period_scope,
            metric_base="OutstandingShares",
            fiscal_year=fiscal_year,
            period_key=period_key,
            forecast_stage=forecast_stage,
        )
        net_assets_value = _jquants_ok_value(net_assets_row)
        shares_value = _jquants_ok_value(shares_row)
        value = (
            net_assets_value / shares_value
            if net_assets_value is not None and shares_value and shares_value > 0
            else None
        )
        return row or net_assets_row or shares_row, value
    lookup_base = _display_base_for_accounting_standard(metric_base, accounting_standard)
    row = _jquants_latest_row(
        latest,
        security_code=security_code,
        period_scope=period_scope,
        metric_base=lookup_base,
        fiscal_year=fiscal_year,
        period_key=period_key,
        forecast_stage=forecast_stage,
    )
    return row, _jquants_ok_value(row)


def _find_existing_detail_row(
    rows: list[MetricExcelRow],
    *,
    security_code: str,
    period_scope: str,
    metric_base: str,
    detail_rows_by_key: dict[tuple[str, str, str], MetricExcelRow] | None = None,
) -> MetricExcelRow | None:
    if detail_rows_by_key is not None:
        return detail_rows_by_key.get((security_code, period_scope, metric_base))
    for row in rows:
        if (
            _is_detail_row(row)
            and row.security_code == security_code
            and row.period_scope == period_scope
            and row.metric_base == metric_base
            and not row.segment_kind
        ):
            return row
    return None


def _merge_missing_offsets(
    target: MetricExcelRow,
    *,
    period_offsets: list[int],
    periods_by_offset: dict[int, str],
    values_by_offset: dict[int, float | None],
    units_by_offset: dict[int, str],
    ratios_by_offset: dict[int, float | None],
    raw_values_by_offset: dict[int, float | None],
    fiscal_years_by_offset: dict[int, int | None],
    ratio_kinds_by_offset: dict[int, str],
    value_kinds_by_offset: dict[int, str],
) -> None:
    for offset in period_offsets:
        if not target.periods_by_offset.get(offset) and periods_by_offset.get(offset):
            target.periods_by_offset[offset] = periods_by_offset[offset]
        if target.fiscal_years_by_offset.get(offset) is None and fiscal_years_by_offset.get(offset) is not None:
            target.fiscal_years_by_offset[offset] = fiscal_years_by_offset[offset]
        if target.raw_values_by_offset.get(offset) is None and raw_values_by_offset.get(offset) is not None:
            target.raw_values_by_offset[offset] = raw_values_by_offset[offset]
            target.values_by_offset[offset] = values_by_offset.get(offset)
            target.units_by_offset[offset] = units_by_offset.get(offset, "")
            target.ratios_by_offset[offset] = ratios_by_offset.get(offset)
        if offset not in target.ratio_kinds_by_offset and ratio_kinds_by_offset.get(offset):
            target.ratio_kinds_by_offset[offset] = ratio_kinds_by_offset[offset]
        if offset not in target.value_kinds_by_offset and value_kinds_by_offset.get(offset):
            target.value_kinds_by_offset[offset] = value_kinds_by_offset[offset]


def _append_jquants_rows(
    conn: sqlite3.Connection,
    condition: MetricExcelCondition,
    rows: list[MetricExcelRow],
    selected_row_bases_by_sheet: dict[str, list[str]],
    warnings: list[str],
    *,
    anchor_fiscal_year_by_security_code: dict[str, int] | None = None,
    span_recorder: SpanRecorder | None = None,
) -> dict[str, int]:
    needs_quarter = "quarter" in condition.period_scopes
    needs_quarter_standalone = "quarter_standalone" in condition.period_scopes
    needs_forecast = "forecast" in condition.period_scopes
    if not needs_quarter and not needs_quarter_standalone and not needs_forecast:
        return {}
    if (needs_quarter or needs_forecast) and not _jquants_table_exists(conn, "jquants_financial_metrics"):
        warnings.append("jquants_financial_metrics_not_ready")
        if not needs_quarter_standalone:
            return {}

    companies = _fetch_jquants_companies(conn, condition)
    if not companies:
        return {}
    security_codes = sorted({_security_code_for_jquants(company) for company in companies if _security_code_for_jquants(company)})
    requested_bases = sorted(
        {
            base
            for sheet, bases in selected_row_bases_by_sheet.items()
            for base in bases
            if base in (QUARTER_SUPPORTED_BASES | FORECAST_SUPPORTED_BASES)
            or base in QUARTER_STANDALONE_SUPPORTED_BASES
        }
    )
    if not requested_bases:
        return {}
    max_offset = max(condition.period_offsets or [0])
    metric_bases_to_fetch = _jquants_metric_fetch_bases(set(requested_bases) | FORECAST_PROGRESS_BASES)
    started = perf_counter()
    all_metric_rows = (
        _fetch_jquants_metric_rows(
            conn,
            security_codes=security_codes,
            metric_bases=sorted(metric_bases_to_fetch),
            min_fiscal_year=None,
        )
        if _jquants_table_exists(conn, "jquants_financial_metrics")
        else []
    )
    _record_span(
        span_recorder,
        "db_read",
        "fetch_jquants_metric_rows",
        started,
        count=len(all_metric_rows),
    )
    if not all_metric_rows and (needs_quarter or needs_forecast):
        warnings.append("jquants_metrics_not_found")
        if not needs_quarter_standalone:
            return {"jquants_metric_rows": 0}
    started = perf_counter()
    latest = _latest_jquants_metric_rows(all_metric_rows)
    lookup_indexes = _build_jquants_lookup_indexes(all_metric_rows)
    standalone_metric_bases = {
        base for base in requested_bases if base in QUARTER_STANDALONE_SUPPORTED_BASES
    }
    if "OrdinaryIncome" in standalone_metric_bases:
        standalone_metric_bases.add("ProfitBeforeTax")
    _record_span(
        span_recorder,
        "compute",
        "build_jquants_lookup_indexes",
        started,
        count=len(all_metric_rows),
    )
    started = perf_counter()
    standalone_metric_rows = _fetch_quarter_standalone_metric_rows(
        conn,
        security_codes=security_codes,
        metric_bases=sorted(standalone_metric_bases),
    )
    _record_span(
        span_recorder,
        "db_read",
        "fetch_quarter_standalone_metric_rows",
        started,
        count=len(standalone_metric_rows),
    )
    started = perf_counter()
    standalone_latest = _latest_quarter_standalone_metric_rows(standalone_metric_rows)
    standalone_lookup_indexes = _build_quarter_standalone_lookup_indexes(standalone_metric_rows)
    if needs_quarter_standalone and not standalone_metric_rows:
        warnings.append("quarter_standalone_metrics_not_ready")

    detail_rows_by_key = {
        (row.security_code, row.period_scope, row.metric_base): row
        for row in rows
        if _is_detail_row(row) and not row.segment_kind
    }
    rows_before = len(rows)
    started = perf_counter()
    for company in companies:
        security_code = _security_code_for_jquants(company)
        if not security_code:
            continue
        anchor_fiscal_year = (anchor_fiscal_year_by_security_code or {}).get(security_code)
        sheet_name = _sheet_name_for_industry(company["industry_33"])
        base_candidates = selected_row_bases_by_sheet[sheet_name]
        if needs_quarter:
            for quarter in JQUANTS_QUARTER_TYPES:
                _append_jquants_period_rows(
                    rows,
                    company=company,
                    security_code=security_code,
                    base_candidates=[
                        base
                        for base in base_candidates
                        if base in QUARTER_SUPPORTED_BASES and base not in QUARTER_SUPPRESSED_EXCEL_BASES
                    ],
                    period_scope=f"quarter:{quarter}",
                    period_key=f"actual:{quarter}",
                    latest=latest,
                    all_rows=all_metric_rows,
                    lookup_indexes=lookup_indexes,
                    detail_rows_by_key=detail_rows_by_key,
                    period_offsets=condition.period_offsets,
                    max_offset=max_offset,
                    with_progress_ratio=True,
                    anchor_fiscal_year=anchor_fiscal_year,
                )
        if needs_quarter_standalone:
            _append_quarter_standalone_period_rows(
                rows,
                company=company,
                security_code=security_code,
                base_candidates=[base for base in base_candidates if base in QUARTER_STANDALONE_SUPPORTED_BASES],
                latest=standalone_latest,
                all_rows=standalone_metric_rows,
                lookup_indexes=standalone_lookup_indexes,
                period_offsets=condition.period_offsets,
                max_offset=max_offset,
                anchor_fiscal_year=anchor_fiscal_year,
            )
        if needs_forecast:
            for forecast_stage in JQUANTS_FORECAST_STAGES:
                _append_jquants_period_rows(
                    rows,
                    company=company,
                    security_code=security_code,
                    base_candidates=[base for base in base_candidates if base in FORECAST_SUPPORTED_BASES],
                    period_scope=f"forecast:{forecast_stage}",
                    period_key="forecast:FY",
                    forecast_stage=forecast_stage,
                    latest=latest,
                    all_rows=all_metric_rows,
                    lookup_indexes=lookup_indexes,
                    detail_rows_by_key=detail_rows_by_key,
                    period_offsets=condition.period_offsets,
                    max_offset=max_offset,
                    with_progress_ratio=False,
                )
    appended_rows = len(rows) - rows_before
    _record_span(
        span_recorder,
        "compute",
        "append_jquants_rows",
        started,
        count=appended_rows,
        detail={
            "jquants_metric_rows": len(all_metric_rows),
            "quarter_standalone_metric_rows": len(standalone_metric_rows),
        },
    )
    return {
        "jquants_metric_rows": len(all_metric_rows),
        "quarter_standalone_metric_rows": len(standalone_metric_rows),
        "jquants_output_rows": appended_rows,
    }


def _append_quarter_standalone_period_rows(
    rows: list[MetricExcelRow],
    *,
    company: sqlite3.Row,
    security_code: str,
    base_candidates: list[str],
    latest: dict[tuple[str, str, str, int], sqlite3.Row],
    all_rows: list[sqlite3.Row],
    lookup_indexes: _QuarterStandaloneLookupIndexes | None,
    period_offsets: list[int],
    max_offset: int,
    anchor_fiscal_year: int | None = None,
) -> None:
    if not base_candidates:
        return
    latest_fiscal_year = _max_quarter_standalone_fiscal_year(
        all_rows,
        security_code,
        lookup_indexes=lookup_indexes,
    )
    if latest_fiscal_year is None:
        return
    max_fiscal_year = anchor_fiscal_year or latest_fiscal_year
    if max_fiscal_year is None:
        return
    min_year = max_fiscal_year - max_offset
    for quarter in QUARTER_STANDALONE_EXCEL_QUARTERS:
        period_scope = f"{QUARTER_STANDALONE_PERIOD_SCOPE}:{quarter}"
        for base in base_candidates:
            if base in QUARTER_STANDALONE_SUPPRESSED_BY_QUARTER.get(quarter, set()):
                continue
            periods_by_offset: dict[int, str] = {}
            values_by_offset: dict[int, float | None] = {}
            units_by_offset: dict[int, str] = {}
            ratios_by_offset: dict[int, float | None] = {}
            raw_values_by_offset: dict[int, float | None] = {}
            fiscal_years_by_offset: dict[int, int | None] = {}
            current_period_end = _quarter_standalone_period_end(
                all_rows,
                security_code=security_code,
                fiscal_year=max_fiscal_year,
                quarter=quarter,
                lookup_indexes=lookup_indexes,
            )
            for offset in period_offsets:
                fiscal_year = max_fiscal_year - offset
                fiscal_years_by_offset[offset] = fiscal_year
                if fiscal_year < min_year:
                    continue
                lookup_base = _display_base_for_accounting_standard(
                    base,
                    str(company["accounting_standard"] or ""),
                )
                row = latest.get((security_code, quarter, lookup_base, fiscal_year))
                raw_value = (
                    float(row["value_num"])
                    if row is not None
                    and str(row["calc_status"] or "") == "ok"
                    and row["value_num"] is not None
                    else None
                )
                display_value, display_unit = _scale_jquants_value(
                    base,
                    raw_value,
                    document_display_unit=str(company["document_display_unit"] or ""),
                )
                periods_by_offset[offset] = _quarter_standalone_period_display(
                    all_rows,
                    security_code=security_code,
                    fiscal_year=fiscal_year,
                    quarter=quarter,
                    lookup_indexes=lookup_indexes,
                )
                values_by_offset[offset] = display_value
                units_by_offset[offset] = display_unit if raw_value is not None else ""
                raw_values_by_offset[offset] = raw_value
                ratios_by_offset[offset] = None
            if not any(value is not None for value in raw_values_by_offset.values()) and not any(
                periods_by_offset.values()
            ):
                continue
            rows.append(
                MetricExcelRow(
                    sheet_name=_sheet_name_for_industry(company["industry_33"]),
                    security_code=security_code,
                    company_name=str(company["company_name"] or ""),
                    industry_33=str(company["industry_33"] or ""),
                    market=str(company["market"] or ""),
                    period_scope=period_scope,
                    row_kind=_row_kind_for_metric_base(base),
                    current_period_end=current_period_end,
                    metric_base=base,
                metric_label=_metric_label_for_excel(
                    base,
                    company["industry_33"],
                    period_scope=period_scope,
                    accounting_standard=str(company["accounting_standard"] or ""),
                ),
                    periods_by_offset=periods_by_offset,
                    values_by_offset=values_by_offset,
                    units_by_offset=units_by_offset,
                    ratios_by_offset=ratios_by_offset,
                    raw_values_by_offset=raw_values_by_offset,
                    fiscal_years_by_offset=fiscal_years_by_offset,
                    accounting_standard=str(company["accounting_standard"] or ""),
                )
            )


def _append_jquants_period_rows(
    rows: list[MetricExcelRow],
    *,
    company: sqlite3.Row,
    security_code: str,
    base_candidates: list[str],
    period_scope: str,
    period_key: str,
    latest: dict[tuple[str, str, str, int, str, str], sqlite3.Row],
    all_rows: list[sqlite3.Row],
    lookup_indexes: _JQuantsLookupIndexes | None,
    detail_rows_by_key: dict[tuple[str, str, str], MetricExcelRow],
    period_offsets: list[int],
    max_offset: int,
    with_progress_ratio: bool,
    forecast_stage: str | None = None,
    anchor_fiscal_year: int | None = None,
) -> None:
    if not base_candidates:
        return
    max_fiscal_year = anchor_fiscal_year or _max_jquants_fiscal_year(
        all_rows,
        security_code,
        period_key,
        forecast_stage=forecast_stage,
        lookup_indexes=lookup_indexes,
    )
    if max_fiscal_year is None:
        return
    min_year = max_fiscal_year - max_offset
    for base in base_candidates:
        periods_by_offset: dict[int, str] = {}
        values_by_offset: dict[int, float | None] = {}
        units_by_offset: dict[int, str] = {}
        ratios_by_offset: dict[int, float | None] = {}
        raw_values_by_offset: dict[int, float | None] = {}
        ratio_kinds_by_offset: dict[int, str] = {}
        value_kinds_by_offset: dict[int, str] = {}
        fiscal_years_by_offset: dict[int, int | None] = {}
        current_period_end = ""
        for offset in period_offsets:
            fiscal_year = max_fiscal_year - offset
            fiscal_years_by_offset[offset] = fiscal_year
            if fiscal_year < min_year:
                continue
            row, raw_value = _jquants_display_row_and_value(
                latest,
                security_code=security_code,
                period_scope=period_scope,
                metric_base=base,
                fiscal_year=fiscal_year,
                period_key=period_key,
                forecast_stage=forecast_stage,
                accounting_standard=str(company["accounting_standard"] or ""),
            )
            display_value, display_unit = _scale_jquants_value(
                base,
                raw_value,
                document_display_unit=str(company["document_display_unit"] or ""),
            )
            periods_by_offset[offset] = _jquants_period_display(row, period_scope, base)
            if offset == 0 and row is not None:
                current_period_end = _period_end_from_jquants_row(row)
            elif not current_period_end and row is not None:
                current_period_end = _period_end_from_jquants_row(row)
            values_by_offset[offset] = display_value
            units_by_offset[offset] = display_unit if raw_value is not None else ""
            raw_values_by_offset[offset] = raw_value
            ratios_by_offset[offset] = None
            if forecast_stage is not None and raw_value is not None:
                previous_forecast = _previous_forecast_value(
                    all_rows,
                    security_code=security_code,
                    fiscal_year=fiscal_year,
                    metric_base=base,
                    forecast_stage=forecast_stage,
                    lookup_indexes=lookup_indexes,
                )
                if previous_forecast is not None:
                    if raw_value > previous_forecast:
                        value_kinds_by_offset[offset] = FORECAST_REVISION_UP_KIND
                    elif raw_value < previous_forecast:
                        value_kinds_by_offset[offset] = FORECAST_REVISION_DOWN_KIND
            if with_progress_ratio and base in FORECAST_PROGRESS_BASES and row is not None:
                forecast_value = _latest_forecast_value_as_of(
                    all_rows,
                    security_code=security_code,
                    fiscal_year=fiscal_year,
                    metric_base=base,
                    disclosed_date=str(row["disclosed_date"] or ""),
                    lookup_indexes=lookup_indexes,
                )
                if forecast_value is not None and forecast_value > 0 and raw_value is not None:
                    ratios_by_offset[offset] = raw_value / forecast_value
                    ratio_kinds_by_offset[offset] = FORECAST_PROGRESS_RATIO_KIND
        if not any(value is not None for value in raw_values_by_offset.values()):
            continue
        existing_row = _find_existing_detail_row(
            rows,
            security_code=security_code,
            period_scope=period_scope,
            metric_base=base,
            detail_rows_by_key=detail_rows_by_key,
        )
        if existing_row is not None:
            _merge_missing_offsets(
                existing_row,
                period_offsets=period_offsets,
                periods_by_offset=periods_by_offset,
                values_by_offset=values_by_offset,
                units_by_offset=units_by_offset,
                ratios_by_offset=ratios_by_offset,
                raw_values_by_offset=raw_values_by_offset,
                fiscal_years_by_offset=fiscal_years_by_offset,
                ratio_kinds_by_offset=ratio_kinds_by_offset,
                value_kinds_by_offset=value_kinds_by_offset,
            )
            continue
        rows.append(
            MetricExcelRow(
                sheet_name=_sheet_name_for_industry(company["industry_33"]),
                security_code=security_code,
                company_name=str(company["company_name"] or ""),
                industry_33=str(company["industry_33"] or ""),
                market=str(company["market"] or ""),
                period_scope=period_scope,
                row_kind=_row_kind_for_metric_base(base),
                current_period_end=current_period_end,
                metric_base=base,
                metric_label=_metric_label_for_excel(
                    base,
                    company["industry_33"],
                    period_scope=period_scope,
                    accounting_standard=str(company["accounting_standard"] or ""),
                ),
                periods_by_offset=periods_by_offset,
                values_by_offset=values_by_offset,
                units_by_offset=units_by_offset,
                ratios_by_offset=ratios_by_offset,
                raw_values_by_offset=raw_values_by_offset,
                fiscal_years_by_offset=fiscal_years_by_offset,
                ratio_kinds_by_offset=ratio_kinds_by_offset,
                value_kinds_by_offset=value_kinds_by_offset,
                accounting_standard=str(company["accounting_standard"] or ""),
            )
        )
        detail_rows_by_key[(security_code, period_scope, base)] = rows[-1]


def _latest_forecast_value_as_of(
    rows: list[sqlite3.Row],
    *,
    security_code: str,
    fiscal_year: int,
    metric_base: str,
    disclosed_date: str,
    lookup_indexes: _JQuantsLookupIndexes | None = None,
) -> float | None:
    if lookup_indexes is not None:
        values = lookup_indexes.forecast_values_by_as_of_key.get(
            (security_code, fiscal_year, metric_base),
            [],
        )
        position = bisect_right(values, (disclosed_date, "\uffff", float("inf")))
        return values[position - 1][2] if position else None
    candidates = [
        row
        for row in rows
        if _normalize_security_code(row["security_code"] or row["local_code"] or "") == security_code
        and str(row["period_scope"] or "") == "forecast"
        and str(row["metric_base"] or "") == metric_base
        and row["fiscal_year"] is not None
        and int(row["fiscal_year"]) == fiscal_year
        and str(row["period_key"] or "") == "forecast:FY"
        and str(row["calc_status"] or "") == "ok"
        and row["value_num"] is not None
        and str(row["disclosed_date"] or "") <= disclosed_date
    ]
    if not candidates:
        return None
    candidates.sort(key=lambda row: (str(row["disclosed_date"] or ""), str(row["disclosed_time"] or "")))
    return float(candidates[-1]["value_num"])


def _previous_forecast_value(
    rows: list[sqlite3.Row],
    *,
    security_code: str,
    fiscal_year: int,
    metric_base: str,
    forecast_stage: str,
    lookup_indexes: _JQuantsLookupIndexes | None = None,
) -> float | None:
    stage_order = list(JQUANTS_FORECAST_STAGES)
    if forecast_stage not in stage_order:
        return None
    stage_index = stage_order.index(forecast_stage)
    if stage_index <= 0:
        return None
    previous_stage = stage_order[stage_index - 1]
    if lookup_indexes is not None:
        return lookup_indexes.forecast_value_by_stage.get(
            (security_code, fiscal_year, metric_base, previous_stage),
        )
    candidates = [
        row
        for row in rows
        if _normalize_security_code(row["security_code"] or row["local_code"] or "") == security_code
        and str(row["period_scope"] or "") == "forecast"
        and str(row["metric_base"] or "") == metric_base
        and row["fiscal_year"] is not None
        and int(row["fiscal_year"]) == fiscal_year
        and str(row["period_key"] or "") == "forecast:FY"
        and str(row["forecast_stage"] or "") == previous_stage
        and str(row["calc_status"] or "") == "ok"
        and row["value_num"] is not None
    ]
    if not candidates:
        return None
    candidates.sort(key=lambda row: (str(row["disclosed_date"] or ""), str(row["disclosed_time"] or "")))
    return float(candidates[-1]["value_num"])


def build_metric_excel_rows(
    conn: sqlite3.Connection,
    condition: MetricExcelCondition,
    *,
    preview_limit: int = 10,
    span_recorder: SpanRecorder | None = None,
) -> tuple[list[MetricExcelRow], list[str], list[str], list[dict[str, Any]], int]:
    if condition.industry_only:
        return _build_industry_only_metric_excel_rows(
            conn,
            condition,
            preview_limit=preview_limit,
        )

    errors: list[str] = []
    warnings: list[str] = []
    started = perf_counter()
    filings = _fetch_ranked_filings(conn, condition)
    _record_span(
        span_recorder,
        "db_read",
        "fetch_ranked_filings",
        started,
        count=len(filings),
    )

    filings_by_company_scope: dict[tuple[str, str], list[sqlite3.Row]] = {}
    for row in filings:
        period_scope = PERIOD_SCOPE_BY_FORM_TYPE.get(str(row["form_type"] or ""), "annual")
        filings_by_company_scope.setdefault((str(row["edinet_code"]), period_scope), []).append(row)
    anchor_fiscal_year_by_security_code = _fiscal_year_anchor_by_security_code(
        _fetch_jquants_anchor_filings(conn, condition, filings)
    )

    selected_row_bases_by_sheet = {
        sheet: _filter_excel_visible_bases(
            _resolve_row_bases(sheet, condition.metric_labels, errors),
            warnings=warnings,
            explicit_metric_request=bool(condition.metric_labels),
        )
        for sheet in SHEET_ORDER
    }

    selected_value_bases: set[str] = set()
    for bases in selected_row_bases_by_sheet.values():
        for base in bases:
            selected_value_bases.add(base)
            if base == "OrdinaryIncome":
                selected_value_bases.add("ProfitBeforeTax")
            ratio_base = ABSORBED_RATIO_BASE_BY_ROW_BASE.get(base)
            if ratio_base:
                selected_value_bases.add(ratio_base)
    if selected_value_bases & MARKET_METRIC_BASES and not market_derived_table_exists(conn):
        warnings.append("market_derived_metrics_not_ready")

    has_trend_threshold = condition.trend_min is not None or condition.trend_max is not None
    if condition.trend != "none" or has_trend_threshold:
        trend_labels = condition.trend_metric_labels or condition.metric_labels
        if trend_labels:
            for sheet in SHEET_ORDER:
                for base in _resolve_value_bases(sheet, trend_labels, errors):
                    selected_value_bases.add(base)
                    if base == "OrdinaryIncome":
                        selected_value_bases.add("ProfitBeforeTax")
        else:
            for bases in selected_row_bases_by_sheet.values():
                selected_value_bases.update(bases)
                if "OrdinaryIncome" in bases:
                    selected_value_bases.add("ProfitBeforeTax")

    percent_filter_bases_by_sheet: dict[str, list[str]] = {sheet: [] for sheet in SHEET_ORDER}
    if condition.percent_filter_metric_labels:
        for sheet in SHEET_ORDER:
            bases = _resolve_value_bases(sheet, condition.percent_filter_metric_labels, errors)
            percent_filter_bases_by_sheet[sheet] = bases
            selected_value_bases.update(bases)
            if "OrdinaryIncome" in bases:
                selected_value_bases.add("ProfitBeforeTax")

    doc_ids = [str(row["doc_id"]) for row in filings]
    started = perf_counter()
    metric_values = _fetch_metric_values(
        conn,
        doc_ids=doc_ids,
        metric_bases=sorted(selected_value_bases),
    )
    _record_span(
        span_recorder,
        "db_read",
        "fetch_edinet_metric_values",
        started,
        count=len(metric_values),
    )

    rows: list[MetricExcelRow] = []
    started = perf_counter()
    for (_edinet_code, current_period_scope), company_filings in filings_by_company_scope.items():
        by_offset = {int(row["period_offset"]): row for row in company_filings}
        current = by_offset.get(0) or min(
            company_filings,
            key=lambda row: int(row["period_offset"]),
            default=None,
        )
        if current is None:
            continue

        sheet_name = _sheet_name_for_industry(current["industry_33"])
        row_bases = selected_row_bases_by_sheet[sheet_name]
        if current_period_scope != "quarter:2Q":
            row_bases = [base for base in row_bases if base not in HALF_ONLY_BASES]
        else:
            row_bases = [base for base in row_bases if base not in HALF_DISABLED_BASES]

        if not _passes_percent_filters(
            filter_bases=percent_filter_bases_by_sheet.get(sheet_name, []),
            filter_offsets=condition.percent_filter_period_offsets,
            period_scope=current_period_scope,
            by_offset=by_offset,
            metric_values=metric_values,
            min_value=condition.percent_filter_min,
            max_value=condition.percent_filter_max,
        ):
            continue

        if condition.trend != "none" or has_trend_threshold:
            trend_labels = condition.trend_metric_labels or condition.metric_labels
            trend_bases = (
                _resolve_value_bases(sheet_name, trend_labels, errors)
                if trend_labels
                else row_bases
            )
            trend_ok = True
            for trend_base in trend_bases:
                trend_values = []
                for offset in sorted(condition.trend_period_offsets, reverse=True):
                    source_offset = _source_offset_for_display(current_period_scope, offset)
                    filing = by_offset.get(source_offset) if source_offset is not None else None
                    value = None
                    if filing is not None:
                        value = _metric_value(
                            metric_values,
                            str(filing["doc_id"]),
                            trend_base,
                            str(filing["accounting_standard"] or ""),
                        )
                    trend_values.append(value)
                if not _passes_trend(trend_values, condition.trend):
                    trend_ok = False
                    break
                if not _passes_value_thresholds(
                    trend_values,
                    min_value=condition.trend_min,
                    max_value=condition.trend_max,
                ):
                    trend_ok = False
                    break
            if not trend_ok:
                continue

        security_code = _normalize_security_code(current["issuer_security_code"] or current["security_code"] or "")
        for base in row_bases:
            periods_by_offset: dict[int, str] = {}
            values_by_offset: dict[int, float | None] = {}
            units_by_offset: dict[int, str] = {}
            ratios_by_offset: dict[int, float | None] = {}
            raw_values_by_offset: dict[int, float | None] = {}
            fiscal_years_by_offset: dict[int, int | None] = {}
            ratio_base = ABSORBED_RATIO_BASE_BY_ROW_BASE.get(base)
            allowed_offsets = SPARSE_PERIOD_OFFSETS_BY_BASE.get(base)
            for offset in condition.period_offsets:
                if allowed_offsets is not None and offset not in allowed_offsets:
                    periods_by_offset[offset] = ""
                    values_by_offset[offset] = None
                    units_by_offset[offset] = ""
                    ratios_by_offset[offset] = None
                    raw_values_by_offset[offset] = None
                    fiscal_years_by_offset[offset] = None
                    continue

                source_offset = _source_offset_for_display(current_period_scope, offset)
                filing = by_offset.get(source_offset) if source_offset is not None else None
                if filing is None:
                    periods_by_offset[offset] = ""
                    values_by_offset[offset] = None
                    units_by_offset[offset] = ""
                    ratios_by_offset[offset] = None
                    raw_values_by_offset[offset] = None
                    fiscal_years_by_offset[offset] = None
                    continue

                doc_id = str(filing["doc_id"])
                fiscal_years_by_offset[offset] = _calendar_year_bucket(filing["period_bucket_end"])
                periods_by_offset[offset] = (
                    _period_point_display_for_filing(filing, current_period_scope)
                    if base in DATE_POINT_PERIOD_BASES
                    else _period_display_for_filing(filing)
                )
                raw_value = _metric_value(
                    metric_values,
                    doc_id,
                    base,
                    str(filing["accounting_standard"] or ""),
                )
                raw_values_by_offset[offset] = raw_value
                values_by_offset[offset] = _scale_value_for_document_unit(
                    base,
                    raw_value,
                    str(filing["document_display_unit"] or ""),
                )
                units_by_offset[offset] = (
                    _display_unit_for_metric(base, str(filing["document_display_unit"] or ""))
                    if raw_value is not None
                    else ""
                )

                if base in CONSTANT_RATIO_BY_ROW_BASE:
                    ratios_by_offset[offset] = CONSTANT_RATIO_BY_ROW_BASE[base]
                elif ratio_base:
                    ratios_by_offset[offset] = _metric_value(
                        metric_values,
                        doc_id,
                        ratio_base,
                        str(filing["accounting_standard"] or ""),
                    )
                else:
                    ratios_by_offset[offset] = None

            rows.append(
                MetricExcelRow(
                    sheet_name=sheet_name,
                    security_code=security_code,
                    company_name=str(current["company_name"] or ""),
                    industry_33=str(current["industry_33"] or ""),
                    market=str(current["market"] or ""),
                    period_scope=current_period_scope,
                    row_kind=_row_kind_for_metric_base(base),
                    current_period_end=_current_period_end_for_scope(
                        current["period_end"],
                        current_period_scope,
                    ),
                    metric_base=base,
                    metric_label=_metric_label_for_excel(
                        base,
                        current["industry_33"],
                        period_scope=current_period_scope,
                        accounting_standard=str(current["accounting_standard"] or ""),
                    ),
                    periods_by_offset=periods_by_offset,
                    values_by_offset=values_by_offset,
                    units_by_offset=units_by_offset,
                    ratios_by_offset=ratios_by_offset,
                    raw_values_by_offset=raw_values_by_offset,
                    fiscal_years_by_offset=fiscal_years_by_offset,
                    accounting_standard=str(current["accounting_standard"] or ""),
                )
            )

    _record_span(
        span_recorder,
        "compute",
        "build_edinet_rows",
        started,
        count=len(rows),
    )
    jquants_summary = _append_jquants_rows(
        conn,
        condition,
        rows,
        selected_row_bases_by_sheet,
        warnings,
        anchor_fiscal_year_by_security_code=anchor_fiscal_year_by_security_code,
        span_recorder=span_recorder,
    )
    if any(
        FORECAST_PROGRESS_RATIO_KIND in row.ratio_kinds_by_offset.values()
        for row in rows
        if _is_detail_row(row)
    ):
        warnings.append("quarter_ratio_cells_show_latest_forecast_progress")
    target_companies = len({row.security_code for row in rows if _is_detail_row(row)})
    started = perf_counter()
    _assign_ranks(rows, condition.period_offsets)
    rows = _append_stat_rows(rows, condition.period_offsets, industry_only=False)
    _record_span(
        span_recorder,
        "compute",
        "assign_ranks_and_stats",
        started,
        count=len(rows),
    )
    segment_rows = _build_segment_metric_excel_rows(
        conn,
        condition,
        warnings,
        span_recorder=span_recorder,
    )
    rows.extend(segment_rows)
    started = perf_counter()
    rows.sort(key=_row_sort_key)
    preview_rows = _build_preview_rows(rows, condition.period_offsets, preview_limit)
    _record_span(
        span_recorder,
        "compute",
        "sort_and_preview",
        started,
        count=len(rows),
        detail={
            "row_builder_mode": "indexed_lookup",
            "segment_output_rows": len(segment_rows),
            **jquants_summary,
        },
    )
    return rows, errors, warnings, preview_rows, target_companies


def _write_summary_sheet(
    workbook: Workbook,
    *,
    condition: MetricExcelCondition,
    db_path: str | Path,
    output_rows: int,
    target_companies: int,
    errors: list[str],
    warnings: list[str],
) -> None:
    ws = workbook.create_sheet(SUMMARY_SHEET)
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 80
    rows = [
        ("generated_at", datetime.now().isoformat(timespec="seconds")),
        ("db_path", str(db_path)),
        ("target_companies", target_companies),
        ("output_rows", output_rows),
        ("errors", len(errors)),
        ("warnings", len(warnings)),
        ("industry_only", "1" if condition.industry_only else "0"),
        (
            "aggregation_basis_note",
            (
                "\u96c6\u8a08\u57fa\u6e96: "
                f"{datetime.now().isoformat(timespec='minutes')}"
                "\u6642\u70b9\u3067DB\u306b\u53cd\u6620\u6e08\u307f\u306e\u6c7a\u7b97\u671f\u672b\u65e5\u30d9\u30fc\u30b9"
            )
            if condition.industry_only
            else "",
        ),
        ("industries", ", ".join(condition.industries) if condition.industries else "ALL"),
        ("security_codes", ", ".join(condition.security_codes) if condition.security_codes else "ALL"),
        ("company_names", ", ".join(condition.company_names) if condition.company_names else "ALL"),
        ("metrics", ", ".join(condition.metric_labels) if condition.metric_labels else "ALL"),
        ("period_scopes", ", ".join(condition.period_scopes)),
        ("segment_mode", condition.segment_mode),
        (
            "periods",
            ", ".join(PERIOD_LABEL_BY_OFFSET[offset] for offset in condition.period_offsets),
        ),
        ("連続増減", condition.trend),
        (
            "連続増減指標",
            ", ".join(condition.trend_metric_labels) if condition.trend_metric_labels else "",
        ),
        (
            "連続増減期間",
            ", ".join(PERIOD_LABEL_BY_OFFSET[offset] for offset in condition.trend_period_offsets),
        ),
        (
            "連続増減下限",
            "" if condition.trend_min is None else condition.trend_min,
        ),
        (
            "連続増減上限",
            "" if condition.trend_max is None else condition.trend_max,
        ),
        (
            "比率条件指標",
            ", ".join(condition.percent_filter_metric_labels)
            if condition.percent_filter_metric_labels
            else "",
        ),
        (
            "比率条件期間",
            ", ".join(
                PERIOD_LABEL_BY_OFFSET[offset]
                for offset in condition.percent_filter_period_offsets
            ),
        ),
        (
            "比率条件下限",
            "" if condition.percent_filter_min is None else condition.percent_filter_min,
        ),
        (
            "比率条件上限",
            "" if condition.percent_filter_max is None else condition.percent_filter_max,
        ),
    ]
    for row in rows:
        ws.append(row)
    if errors:
        ws.append(("", ""))
        ws.append(("error_detail", ""))
        for error in errors:
            ws.append(("", error))
    if warnings:
        ws.append(("", ""))
        ws.append(("warning_detail", ""))
        for warning in warnings:
            ws.append(("", warning))


def _format_value_cell(cell: Any, metric_base: str) -> None:
    if metric_base in PERCENT_VALUE_BASES:
        cell.number_format = "0.0%"
    elif metric_base in GROWTH_RATIO_BASES:
        cell.number_format = "0.0%"
    elif metric_base in ONE_DECIMAL_VALUE_BASES:
        cell.number_format = "#,##0.0"
    elif metric_base in INTEGER_VALUE_BASES:
        cell.number_format = "#,##0"
    elif metric_base in PER_SHARE_BASES:
        cell.number_format = "#,##0.00"
    else:
        cell.number_format = "#,##0"


def _write_only_cell(
    ws: Any,
    value: Any,
    *,
    fill: PatternFill | None = None,
    font: Font | None = None,
    border: Border | None = None,
    number_format: str = "",
    comment: Comment | None = None,
) -> WriteOnlyCell:
    cell = WriteOnlyCell(ws, value=value)
    if fill is not None:
        cell.fill = fill
    if font is not None:
        cell.font = font
    if border is not None:
        cell.border = border
    if number_format:
        cell.number_format = number_format
    if comment is not None:
        cell.comment = comment
    return cell


def _period_block_styles(
    period_blocks: list[tuple[str, int]],
    rows: list[MetricExcelRow],
    period_offsets: list[int],
) -> list[tuple[PatternFill, Border]]:
    left_border = Border(left=Side(style="thin", color=PERIOD_BLOCK_BORDER_COLOR))
    styles: list[tuple[PatternFill, Border]] = []
    for block_index, block in enumerate(period_blocks):
        is_current_block = (
            block[0] == "offset" and block[1] == 0
        ) or (
            block[0] == "year"
            and any(
                _offset_for_period_block(row, block, period_offsets) == 0
                for row in rows
            )
        )
        fill_color = (
            CURRENT_PERIOD_BLOCK_FILL_COLOR
            if is_current_block
            else PERIOD_BLOCK_FILL_COLORS[block_index % len(PERIOD_BLOCK_FILL_COLORS)]
        )
        styles.append((PatternFill("solid", fgColor=fill_color), left_border))
    return styles


def _write_metric_sheet(
    workbook: Workbook,
    sheet_name: str,
    rows: list[MetricExcelRow],
    period_offsets: list[int],
) -> int:
    ws = workbook.create_sheet(sheet_name)
    include_segment_columns = False
    base_headers = [
        "\u30b3\u30fc\u30c9",
        "\u4f01\u696d\u540d",
        "\u30c6\u30f3\u30d0\u30ac\u30fc",
        "\u696d\u7a2e",
        "\u5e02\u5834\u533a\u5206",
    ]
    if include_segment_columns:
        base_headers.extend(["\u30bb\u30b0\u30e1\u30f3\u30c8\u533a\u5206", "\u30bb\u30b0\u30e1\u30f3\u30c8\u540d"])
    base_headers.extend(
        [
        "\u6c7a\u7b97\u7a2e\u5225",
        "\u5024\u7a2e\u5225",
        "\u6307\u6a19",
        ]
    )
    base_col_count = len(base_headers)
    headers = list(base_headers)
    period_blocks = _period_blocks_for_rows(rows, period_offsets)
    for block in period_blocks:
        label = _period_block_label(block)
        headers.extend(
            [
                f"{label}_\u671f\u9593",
                f"{label}_\u6570\u5024",
                f"{label}_\u5358\u4f4d",
                f"{label}_\u6bd4\u7387",
                f"{label}_\u9806\u4f4d",
            ]
        )
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    header_font = Font(bold=True)
    progress_fill = PatternFill("solid", fgColor=FORECAST_PROGRESS_FILL_COLOR)
    block_styles = _period_block_styles(period_blocks, rows, period_offsets)

    ws.freeze_panes = f"{get_column_letter(base_col_count + 1)}2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(rows) + 1}"
    base_widths = [12, 28, 12, 18, 12]
    if include_segment_columns:
        base_widths.extend([14, 24])
    base_widths.extend([12, 16, 24])
    for col_idx, width in enumerate(base_widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.column_dimensions["C"].hidden = True
    for col_idx in range(base_col_count + 1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 14

    header_cells: list[Any] = []
    for col_idx, header in enumerate(headers, start=1):
        if col_idx <= base_col_count:
            header_cells.append(_write_only_cell(ws, header, fill=header_fill, font=header_font))
            continue
        period_cell_index = col_idx - base_col_count - 1
        block_index = period_cell_index // 5
        cell_index_in_block = period_cell_index % 5
        fill, border = block_styles[block_index]
        header_cells.append(
            _write_only_cell(
                ws,
                header,
                fill=fill,
                font=header_font,
                border=border if cell_index_in_block == 0 else None,
            )
        )
    ws.append(header_cells)

    for row in rows:
        values: list[Any] = [
            row.security_code,
            row.company_name,
            tenbagger_learning_mark(row.security_code),
            row.industry_33,
            row.market,
        ]
        if include_segment_columns:
            values.extend([row.segment_kind, row.segment_name])
        values.extend(
            [
            decision_label_for_excel(row),
            value_kind_label_for_excel(row),
            row.metric_label,
            ]
        )
        for idx, block in enumerate(period_blocks):
            offset = _offset_for_period_block(row, block, period_offsets)
            fill, border = block_styles[idx]
            value_kind = row.value_kinds_by_offset.get(offset) if offset is not None else None
            value_font: Font | None = None
            if value_kind in {FORECAST_REVISION_UP_KIND, FORECAST_REVISION_DOWN_KIND}:
                value_font = Font(
                    color=(
                        FORECAST_REVISION_UP_FONT_COLOR
                        if value_kind == FORECAST_REVISION_UP_KIND
                        else FORECAST_REVISION_DOWN_FONT_COLOR
                    )
                )
            value_cell = _write_only_cell(
                ws,
                row.values_by_offset.get(offset) if offset is not None else None,
                fill=fill,
                font=value_font,
            )
            _format_value_cell(value_cell, row.metric_base)
            ratio_is_progress = (
                row.ratio_kinds_by_offset.get(offset) == FORECAST_PROGRESS_RATIO_KIND
                if offset is not None
                else False
            )
            ratio_cell = _write_only_cell(
                ws,
                row.ratios_by_offset.get(offset) if offset is not None else None,
                fill=progress_fill if ratio_is_progress else fill,
                number_format="0.0%",
                comment=(
                    Comment(
                        "\u3053\u306e\u6bd4\u7387\u306f\u3001\u56db\u534a\u671f\u5b9f\u7e3e \u00f7 \u540c\u4e00\u5e74\u5ea6\u306e\u6700\u65b0\u901a\u671f\u4e88\u60f3\u3067\u8a08\u7b97\u3057\u305f\u6700\u65b0\u4e88\u60f3\u9032\u6357\u7387\u3067\u3059\u3002",
                        "EDINET_MONITOR",
                    )
                    if ratio_is_progress
                    else None
                ),
            )
            values.extend(
                [
                    _write_only_cell(
                        ws,
                        row.periods_by_offset.get(offset, "") if offset is not None else "",
                        fill=fill,
                        border=border,
                    ),
                    value_cell,
                    _write_only_cell(
                        ws,
                        row.units_by_offset.get(offset, "") if offset is not None else "",
                        fill=fill,
                    ),
                    ratio_cell,
                    _write_only_cell(
                        ws,
                        row.ranks_by_offset.get(offset, "") if offset is not None else "",
                        fill=fill,
                    ),
                ]
            )
        ws.append(values)
    return len(rows)


def write_metric_excel(
    *,
    rows: list[MetricExcelRow],
    condition: MetricExcelCondition,
    output_path: str | Path,
    db_path: str | Path,
    errors: list[str],
    warnings: list[str],
    target_companies: int | None = None,
    span_recorder: Callable[[str, str, float, int, dict[str, Any]], None] | None = None,
) -> Path:
    def record_span(
        span_name: str,
        started_at: float,
        *,
        count_total: int = 0,
        detail: dict[str, Any] | None = None,
    ) -> None:
        if span_recorder is None:
            return
        span_recorder(
            "file_io",
            span_name,
            max(perf_counter() - started_at, 0.0),
            int(count_total),
            dict(detail or {}),
        )

    path = Path(output_path)
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook(write_only=True)

    started_at = perf_counter()
    _write_summary_sheet(
        workbook,
        condition=condition,
        db_path=db_path,
        output_rows=len(rows),
        target_companies=target_companies
        if target_companies is not None
        else len({row.security_code for row in rows}),
        errors=errors,
        warnings=warnings,
    )
    record_span("write_summary_sheet", started_at, count_total=1)

    started_at = perf_counter()
    rows_by_sheet = {sheet: [] for sheet in SHEET_ORDER}
    for row in rows:
        rows_by_sheet.setdefault(row.sheet_name, []).append(row)

    sheet_row_counts: dict[str, int] = {}
    for sheet_name in SHEET_ORDER:
        sheet_row_counts[sheet_name] = _write_metric_sheet(
            workbook,
            sheet_name,
            rows_by_sheet.get(sheet_name, []),
            condition.period_offsets,
        )
    record_span(
        "write_metric_sheets",
        started_at,
        count_total=sum(sheet_row_counts.values()),
        detail={"sheet_row_counts": sheet_row_counts},
    )

    started_at = perf_counter()
    workbook.save(path)
    record_span(
        "save_workbook",
        started_at,
        count_total=path.stat().st_size,
        detail={"file_size_bytes": path.stat().st_size},
    )
    return path


def export_metric_excel(
    conn: sqlite3.Connection,
    *,
    condition_xlsx: str | Path,
    output_path: str | Path,
    db_path: str | Path,
    preview_limit: int = 10,
) -> MetricExcelExportResult:
    condition = read_metric_excel_condition(condition_xlsx)
    rows, errors, warnings, preview_rows, target_companies = build_metric_excel_rows(
        conn,
        condition,
        preview_limit=preview_limit,
    )
    path = write_metric_excel(
        rows=rows,
        condition=condition,
        output_path=output_path,
        db_path=db_path,
        errors=errors,
        warnings=warnings,
        target_companies=target_companies,
    )
    return MetricExcelExportResult(
        output_path=path,
        target_companies=target_companies,
        output_rows=len(rows),
        errors=errors,
        warnings=warnings,
        preview_rows=preview_rows,
    )
