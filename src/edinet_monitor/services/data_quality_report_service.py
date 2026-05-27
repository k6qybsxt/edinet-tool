from __future__ import annotations

from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
import json
import sqlite3
import uuid
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from edinet_monitor.config.settings import OPERATION_LOG_ROOT
from edinet_monitor.db.migrations import apply_schema_migrations
from edinet_monitor.services.jquants.mapper import normalize_security_code
from edinet_monitor.services.jquants_quality_audit_service import build_jquants_quality_issues


MAJOR_METRIC_BASES = (
    "NetSales",
    "OperatingIncome",
    "OrdinaryIncome",
    "ProfitLoss",
    "TotalAssets",
    "NetAssets",
    "OperatingCash",
    "FCF",
    "OutstandingShares",
    "EPS",
    "BPS",
    "ROE",
    "ROA",
    "EquityRatio",
)

RATIO_METRIC_BASES = ("ROE", "ROA", "EquityRatio")
SEVERITY_ORDER = {"critical": 0, "warning": 1, "info": 2}


@dataclass(frozen=True)
class DataQualityReportOptions:
    date_from: str = ""
    date_to: str = ""
    codes: tuple[str, ...] = ()
    industry_33_list: tuple[str, ...] = ()
    output_dir: Path | None = None
    coverage_warning_threshold: float = 0.8
    extreme_ratio_threshold: float = 5.0


@dataclass
class DataQualityReportItem:
    category: str
    severity: str
    check_name: str
    subject: str
    current_value: float | None
    value_unit: str
    message: str
    detail: dict[str, Any] = field(default_factory=dict)
    previous_value: float | None = None
    delta_value: float | None = None
    item_key: str = ""

    def __post_init__(self) -> None:
        if not self.item_key:
            self.item_key = ":".join(
                [
                    self.category,
                    self.check_name,
                    self.subject,
                    self.value_unit,
                ]
            )


@dataclass(frozen=True)
class DataQualityReportResult:
    run_id: str
    generated_at: str
    date_from: str
    date_to: str
    previous_run_id: str
    excel_path: Path
    items: list[DataQualityReportItem]
    edinet_status_rows: list[dict[str, Any]]
    metric_coverage_rows: list[dict[str, Any]]
    jquants_quality_rows: list[dict[str, Any]]
    counts_by_severity: dict[str, int]

    @property
    def issue_count(self) -> int:
        return self.counts_by_severity.get("critical", 0) + self.counts_by_severity.get("warning", 0)


def _json_dumps(value: Any) -> str:
    return json.dumps(value, ensure_ascii=False, sort_keys=True, separators=(",", ":"))


def _normalized_codes(codes: tuple[str, ...] | list[str] | None) -> tuple[str, ...]:
    return tuple(sorted({code for code in (normalize_security_code(value) for value in (codes or ())) if code}))


def _code_variants(code: str) -> list[str]:
    normalized = normalize_security_code(code)
    if not normalized:
        return []
    variants = {normalized}
    if len(normalized) == 4:
        variants.add(f"{normalized}0")
    if len(normalized) == 5 and normalized.endswith("0"):
        variants.add(normalized[:-1])
    return sorted(variants)


def _table_or_view_exists(conn: sqlite3.Connection, name: str) -> bool:
    row = conn.execute(
        """
        SELECT 1
        FROM sqlite_master
        WHERE type IN ('table', 'view') AND name = ?
        """,
        (name,),
    ).fetchone()
    return row is not None


def _resolve_date_range(conn: sqlite3.Connection, date_from: str, date_to: str) -> tuple[str, str]:
    if date_from and date_to:
        return date_from, date_to
    if not _table_or_view_exists(conn, "filings"):
        return date_from or "0000-01-01", date_to or "9999-12-31"
    row = conn.execute(
        """
        SELECT MIN(period_end) AS min_period_end, MAX(period_end) AS max_period_end
        FROM filings
        WHERE form_type = '030000'
          AND COALESCE(period_end, '') <> ''
        """
    ).fetchone()
    min_period_end = str(row["min_period_end"] or "0000-01-01") if row else "0000-01-01"
    max_period_end = str(row["max_period_end"] or "9999-12-31") if row else "9999-12-31"
    return date_from or min_period_end, date_to or max_period_end


def _condition_key(options: DataQualityReportOptions, date_from: str, date_to: str) -> str:
    return _json_dumps(
        {
            "date_from": date_from,
            "date_to": date_to,
            "codes": list(_normalized_codes(options.codes)),
            "industry_33_list": sorted(options.industry_33_list),
        }
    )


def _build_scope_where(
    options: DataQualityReportOptions,
    date_from: str,
    date_to: str,
) -> tuple[str, list[Any]]:
    clauses = ["f.form_type = '030000'", "COALESCE(f.period_end, '') BETWEEN ? AND ?"]
    params: list[Any] = [date_from, date_to]
    codes = _normalized_codes(options.codes)
    if codes:
        variants: list[str] = []
        for code in codes:
            variants.extend(_code_variants(code))
        variants = sorted(set(variants))
        placeholders = ",".join("?" for _ in variants)
        clauses.append(
            f"COALESCE(NULLIF(f.security_code, ''), NULLIF(im.security_code, '')) IN ({placeholders})"
        )
        params.extend(variants)
    industries = [str(value).strip() for value in options.industry_33_list if str(value).strip()]
    if industries:
        placeholders = ",".join("?" for _ in industries)
        clauses.append(f"COALESCE(im.industry_33, '') IN ({placeholders})")
        params.extend(industries)
    return " AND ".join(clauses), params


def _fetch_edinet_status_rows(
    conn: sqlite3.Connection,
    *,
    where_sql: str,
    params: list[Any],
) -> list[dict[str, Any]]:
    if not _table_or_view_exists(conn, "filings"):
        return []
    rows = conn.execute(
        f"""
        SELECT
            COALESCE(f.download_status, '') AS download_status,
            COALESCE(f.parse_status, '') AS parse_status,
            COUNT(*) AS filing_count
        FROM filings f
        LEFT JOIN issuer_master im
            ON im.edinet_code = f.edinet_code
        WHERE {where_sql}
        GROUP BY COALESCE(f.download_status, ''), COALESCE(f.parse_status, '')
        ORDER BY filing_count DESC, download_status, parse_status
        """,
        params,
    ).fetchall()
    return [dict(row) for row in rows]


def _add_status_items(status_rows: list[dict[str, Any]], items: list[DataQualityReportItem]) -> None:
    total = sum(int(row["filing_count"] or 0) for row in status_rows)
    items.append(
        DataQualityReportItem(
            category="edinet_status",
            severity="info" if total > 0 else "critical",
            check_name="filing_count",
            subject="annual_filings",
            current_value=float(total),
            value_unit="count",
            message=f"annual filing count in scope={total}",
        )
    )
    parse_counts: dict[str, int] = {}
    download_counts: dict[str, int] = {}
    for row in status_rows:
        parse_status = str(row.get("parse_status") or "")
        download_status = str(row.get("download_status") or "")
        count = int(row.get("filing_count") or 0)
        parse_counts[parse_status] = parse_counts.get(parse_status, 0) + count
        download_counts[download_status] = download_counts.get(download_status, 0) + count

    derived_saved = parse_counts.get("derived_metrics_saved", 0)
    missing_derived = max(total - derived_saved, 0)
    if total:
        items.append(
            DataQualityReportItem(
                category="edinet_status",
                severity="warning" if missing_derived else "info",
                check_name="not_derived_metrics_saved",
                subject="annual_filings",
                current_value=float(missing_derived),
                value_unit="count",
                message=f"{missing_derived} filings are not derived_metrics_saved",
                detail={"derived_metrics_saved": derived_saved, "total": total},
            )
        )

    for status, count in sorted(parse_counts.items()):
        if status.endswith("_error") and count:
            items.append(
                DataQualityReportItem(
                    category="edinet_status",
                    severity="critical",
                    check_name="parse_status_error",
                    subject=status,
                    current_value=float(count),
                    value_unit="count",
                    message=f"parse_status={status} count={count}",
                )
            )
    for status, count in sorted(download_counts.items()):
        if status in {"error", "pending"} and count:
            items.append(
                DataQualityReportItem(
                    category="edinet_status",
                    severity="warning",
                    check_name="download_status_attention",
                    subject=status,
                    current_value=float(count),
                    value_unit="count",
                    message=f"download_status={status} count={count}",
                )
            )


def _add_scoped_data_quality_items(
    conn: sqlite3.Connection,
    *,
    where_sql: str,
    params: list[Any],
    items: list[DataQualityReportItem],
) -> None:
    if not _table_or_view_exists(conn, "filings"):
        return
    rows = conn.execute(
        f"""
        WITH scope AS (
            SELECT f.*
            FROM filings f
            LEFT JOIN issuer_master im
                ON im.edinet_code = f.edinet_code
            WHERE {where_sql}
        ),
        raw_fact_doc_counts AS (
            SELECT doc_id, COUNT(*) AS row_count
            FROM raw_facts
            GROUP BY doc_id
        ),
        normalized_doc_counts AS (
            SELECT doc_id, COUNT(*) AS row_count
            FROM normalized_metrics
            GROUP BY doc_id
        ),
        derived_doc_counts AS (
            SELECT
                doc_id,
                COUNT(*) AS row_count,
                SUM(CASE WHEN calc_status = 'ok' THEN 1 ELSE 0 END) AS ok_row_count
            FROM derived_metrics
            GROUP BY doc_id
        )
        SELECT 'filings_missing_accounting_standard' AS check_name, COUNT(*) AS affected_count
        FROM scope
        WHERE accounting_standard IS NULL OR accounting_standard = ''
        UNION ALL
        SELECT 'filings_missing_document_display_unit' AS check_name, COUNT(*) AS affected_count
        FROM scope
        WHERE document_display_unit IS NULL OR document_display_unit = ''
        UNION ALL
        SELECT 'downloaded_missing_zip_path' AS check_name, COUNT(*) AS affected_count
        FROM scope
        WHERE download_status = 'downloaded' AND (zip_path IS NULL OR zip_path = '')
        UNION ALL
        SELECT 'raw_facts_saved_without_raw_rows' AS check_name, COUNT(*) AS affected_count
        FROM scope s
        LEFT JOIN raw_fact_doc_counts rfd
            ON rfd.doc_id = s.doc_id
        WHERE s.parse_status = 'raw_facts_saved'
          AND COALESCE(rfd.row_count, 0) = 0
        UNION ALL
        SELECT 'normalized_saved_without_normalized_rows' AS check_name, COUNT(*) AS affected_count
        FROM scope s
        LEFT JOIN normalized_doc_counts ndc
            ON ndc.doc_id = s.doc_id
        WHERE s.parse_status IN ('normalized_metrics_saved', 'derived_metrics_saved')
          AND COALESCE(ndc.row_count, 0) = 0
        UNION ALL
        SELECT 'derived_saved_without_derived_rows' AS check_name, COUNT(*) AS affected_count
        FROM scope s
        LEFT JOIN derived_doc_counts ddc
            ON ddc.doc_id = s.doc_id
        WHERE s.parse_status = 'derived_metrics_saved'
          AND COALESCE(ddc.row_count, 0) = 0
        UNION ALL
        SELECT 'derived_saved_without_derived_ok_rows' AS check_name, COUNT(*) AS affected_count
        FROM scope s
        LEFT JOIN derived_doc_counts ddc
            ON ddc.doc_id = s.doc_id
        WHERE s.parse_status = 'derived_metrics_saved'
          AND COALESCE(ddc.ok_row_count, 0) = 0
        """,
        params,
    ).fetchall()
    critical_checks = {
        "raw_facts_saved_without_raw_rows",
        "normalized_saved_without_normalized_rows",
        "derived_saved_without_derived_rows",
        "derived_saved_without_derived_ok_rows",
    }
    for row in rows:
        count = int(row["affected_count"] or 0)
        if count <= 0:
            continue
        check_name = str(row["check_name"])
        items.append(
            DataQualityReportItem(
                category="edinet_quality",
                severity="critical" if check_name in critical_checks else "warning",
                check_name=check_name,
                subject="annual_filings",
                current_value=float(count),
                value_unit="count",
                message=f"{check_name} count={count}",
            )
        )


def _fetch_metric_coverage_rows(
    conn: sqlite3.Connection,
    *,
    where_sql: str,
    params: list[Any],
) -> list[dict[str, Any]]:
    expected_row = conn.execute(
        f"""
        SELECT COUNT(*) AS expected_docs
        FROM filings f
        LEFT JOIN issuer_master im
            ON im.edinet_code = f.edinet_code
        WHERE {where_sql}
          AND f.parse_status = 'derived_metrics_saved'
        """,
        params,
    ).fetchone()
    expected_docs = int(expected_row["expected_docs"] or 0) if expected_row else 0
    placeholders = ",".join("?" for _ in MAJOR_METRIC_BASES)
    rows = conn.execute(
        f"""
        WITH scope AS (
            SELECT f.doc_id
            FROM filings f
            LEFT JOIN issuer_master im
                ON im.edinet_code = f.edinet_code
            WHERE {where_sql}
              AND f.parse_status = 'derived_metrics_saved'
        )
        SELECT
            dm.metric_base,
            COUNT(DISTINCT dm.doc_id) AS any_doc_count,
            COUNT(DISTINCT CASE
                WHEN dm.calc_status = 'ok' AND dm.value_num IS NOT NULL THEN dm.doc_id
            END) AS ok_doc_count,
            MIN(dm.value_num) AS min_value,
            MAX(dm.value_num) AS max_value
        FROM derived_metrics dm
        INNER JOIN scope s
            ON s.doc_id = dm.doc_id
        WHERE dm.metric_base IN ({placeholders})
        GROUP BY dm.metric_base
        """,
        [*params, *MAJOR_METRIC_BASES],
    ).fetchall()
    by_base = {str(row["metric_base"]): row for row in rows}
    coverage_rows: list[dict[str, Any]] = []
    for metric_base in MAJOR_METRIC_BASES:
        row = by_base.get(metric_base)
        ok_doc_count = int(row["ok_doc_count"] or 0) if row else 0
        any_doc_count = int(row["any_doc_count"] or 0) if row else 0
        coverage_ratio = (ok_doc_count / expected_docs) if expected_docs else 0.0
        coverage_rows.append(
            {
                "metric_base": metric_base,
                "expected_docs": expected_docs,
                "ok_doc_count": ok_doc_count,
                "any_doc_count": any_doc_count,
                "missing_doc_count": max(expected_docs - ok_doc_count, 0),
                "coverage_ratio": coverage_ratio,
                "min_value": row["min_value"] if row else None,
                "max_value": row["max_value"] if row else None,
            }
        )
    return coverage_rows


def _add_metric_coverage_items(
    coverage_rows: list[dict[str, Any]],
    *,
    coverage_warning_threshold: float,
    items: list[DataQualityReportItem],
) -> None:
    for row in coverage_rows:
        expected_docs = int(row["expected_docs"])
        coverage_ratio = float(row["coverage_ratio"])
        metric_base = str(row["metric_base"])
        if expected_docs <= 0:
            continue
        if coverage_ratio >= coverage_warning_threshold:
            continue
        items.append(
            DataQualityReportItem(
                category="metric_coverage",
                severity="warning",
                check_name="major_metric_coverage_below_threshold",
                subject=metric_base,
                current_value=coverage_ratio,
                value_unit="ratio",
                message=(
                    f"{metric_base} coverage={coverage_ratio:.1%} "
                    f"below threshold={coverage_warning_threshold:.1%}"
                ),
                detail=row,
            )
        )


def _add_edinet_outlier_items(
    conn: sqlite3.Connection,
    *,
    where_sql: str,
    params: list[Any],
    extreme_ratio_threshold: float,
    items: list[DataQualityReportItem],
) -> None:
    ratio_placeholders = ",".join("?" for _ in RATIO_METRIC_BASES)
    ratio_rows = conn.execute(
        f"""
        WITH scope AS (
            SELECT f.doc_id
            FROM filings f
            LEFT JOIN issuer_master im
                ON im.edinet_code = f.edinet_code
            WHERE {where_sql}
        )
        SELECT
            dm.security_code,
            dm.edinet_code,
            dm.metric_base,
            dm.metric_key,
            dm.period_end,
            dm.fiscal_year,
            dm.value_num
        FROM derived_metrics dm
        INNER JOIN scope s
            ON s.doc_id = dm.doc_id
        WHERE dm.calc_status = 'ok'
          AND dm.value_num IS NOT NULL
          AND dm.metric_base IN ({ratio_placeholders})
          AND ABS(dm.value_num) > ?
        ORDER BY ABS(dm.value_num) DESC
        LIMIT 1000
        """,
        [*params, *RATIO_METRIC_BASES, extreme_ratio_threshold],
    ).fetchall()
    for row in ratio_rows:
        metric_base = str(row["metric_base"] or "")
        subject = f"{row['security_code'] or row['edinet_code']}:{row['period_end']}:{metric_base}"
        items.append(
            DataQualityReportItem(
                category="edinet_outlier",
                severity="warning",
                check_name="ratio_metric_extreme_value",
                subject=subject,
                current_value=float(row["value_num"]),
                value_unit="ratio",
                message=f"{metric_base} absolute value exceeds {extreme_ratio_threshold}",
                detail=dict(row),
            )
        )

    share_rows = conn.execute(
        f"""
        WITH scope AS (
            SELECT f.doc_id
            FROM filings f
            LEFT JOIN issuer_master im
                ON im.edinet_code = f.edinet_code
            WHERE {where_sql}
        )
        SELECT
            dm.security_code,
            dm.edinet_code,
            dm.metric_base,
            dm.metric_key,
            dm.period_end,
            dm.fiscal_year,
            dm.value_num
        FROM derived_metrics dm
        INNER JOIN scope s
            ON s.doc_id = dm.doc_id
        WHERE dm.calc_status = 'ok'
          AND dm.metric_base = 'OutstandingShares'
          AND dm.value_num IS NOT NULL
          AND dm.value_num <= 0
        ORDER BY dm.period_end DESC
        LIMIT 1000
        """,
        params,
    ).fetchall()
    for row in share_rows:
        subject = f"{row['security_code'] or row['edinet_code']}:{row['period_end']}:OutstandingShares"
        items.append(
            DataQualityReportItem(
                category="edinet_outlier",
                severity="critical",
                check_name="outstanding_shares_non_positive",
                subject=subject,
                current_value=float(row["value_num"]),
                value_unit="shares",
                message="OutstandingShares is <= 0",
                detail=dict(row),
            )
        )


def _add_jquants_items(
    jquants_rows: list[dict[str, Any]],
    items: list[DataQualityReportItem],
) -> None:
    for row in jquants_rows:
        severity = str(row.get("severity") or "warning")
        check_name = str(row.get("check_name") or "jquants_quality_issue")
        security_code = str(row.get("security_code") or "")
        period_key = str(row.get("period_key") or "")
        fiscal_year = str(row.get("fiscal_year") or "")
        metric_base = str(row.get("metric_base") or "")
        disclosure_number = str(row.get("disclosure_number") or "")
        subject = ":".join(part for part in (security_code, period_key, fiscal_year, metric_base) if part)
        value = row.get("value_num")
        current_value = None
        if value not in (None, ""):
            try:
                current_value = float(value)
            except (TypeError, ValueError):
                current_value = None
        item = DataQualityReportItem(
            category="jquants_quality",
            severity=severity,
            check_name=check_name,
            subject=subject or "jquants",
            current_value=current_value,
            value_unit="value",
            message=str(row.get("message") or ""),
            detail=row,
        )
        item.item_key = (
            f"jquants_quality:{check_name}:{security_code}:{period_key}:"
            f"{fiscal_year}:{metric_base}:{disclosure_number}"
        )
        items.append(item)


def _fetch_previous_run_id(conn: sqlite3.Connection, condition_key: str) -> str:
    if not _table_or_view_exists(conn, "data_quality_report_runs"):
        return ""
    row = conn.execute(
        """
        SELECT run_id
        FROM data_quality_report_runs
        WHERE condition_key = ?
        ORDER BY generated_at DESC, id DESC
        LIMIT 1
        """,
        (condition_key,),
    ).fetchone()
    return str(row["run_id"] or "") if row else ""


def _fetch_previous_values(conn: sqlite3.Connection, previous_run_id: str) -> dict[str, float | None]:
    if not previous_run_id or not _table_or_view_exists(conn, "data_quality_report_items"):
        return {}
    rows = conn.execute(
        """
        SELECT item_key, current_value
        FROM data_quality_report_items
        WHERE run_id = ?
        """,
        (previous_run_id,),
    ).fetchall()
    return {str(row["item_key"]): row["current_value"] for row in rows}


def _apply_previous_values(
    items: list[DataQualityReportItem],
    previous_values: dict[str, float | None],
) -> None:
    for item in items:
        if item.item_key not in previous_values:
            continue
        previous_value = previous_values[item.item_key]
        item.previous_value = previous_value
        if item.current_value is not None and previous_value is not None:
            item.delta_value = float(item.current_value) - float(previous_value)


def _counts_by_severity(items: list[DataQualityReportItem]) -> dict[str, int]:
    counts = {"critical": 0, "warning": 0, "info": 0}
    for item in items:
        severity = item.severity if item.severity in counts else "warning"
        counts[severity] += 1
    return counts


def _build_output_path(output_dir: Path, date_from: str, date_to: str, generated_at: str) -> Path:
    safe_generated = generated_at.replace("-", "").replace(":", "").replace("T", "_")
    return output_dir / f"data_quality_report_{date_from}_to_{date_to}_{safe_generated}.xlsx"


def _write_rows_sheet(ws, headers: list[str], rows: list[dict[str, Any]]) -> None:
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(fill_type="solid", fgColor="D9EAF7")
    for row in rows:
        ws.append([row.get(header, "") for header in headers])
    for column_cells in ws.columns:
        max_length = max(len(str(cell.value or "")) for cell in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = min(max(max_length + 2, 10), 60)


def _write_excel_report(
    *,
    result: DataQualityReportResult,
    options: DataQualityReportOptions,
    condition_key: str,
) -> None:
    wb = Workbook()
    summary = wb.active
    summary.title = "Summary"
    summary_rows = [
        {"key": "run_id", "value": result.run_id},
        {"key": "generated_at", "value": result.generated_at},
        {"key": "date_from", "value": result.date_from},
        {"key": "date_to", "value": result.date_to},
        {"key": "condition_key", "value": condition_key},
        {"key": "previous_run_id", "value": result.previous_run_id},
        {"key": "issue_count", "value": result.issue_count},
        {"key": "critical", "value": result.counts_by_severity.get("critical", 0)},
        {"key": "warning", "value": result.counts_by_severity.get("warning", 0)},
        {"key": "info", "value": result.counts_by_severity.get("info", 0)},
        {"key": "coverage_warning_threshold", "value": options.coverage_warning_threshold},
        {"key": "extreme_ratio_threshold", "value": options.extreme_ratio_threshold},
    ]
    _write_rows_sheet(summary, ["key", "value"], summary_rows)

    diff_rows = [
        _item_row(item)
        for item in result.items
        if item.previous_value is not None and item.delta_value not in (None, 0)
    ]
    _write_rows_sheet(
        wb.create_sheet("Diff"),
        [
            "severity",
            "category",
            "check_name",
            "subject",
            "current_value",
            "previous_value",
            "delta_value",
            "value_unit",
            "message",
        ],
        diff_rows,
    )
    _write_rows_sheet(
        wb.create_sheet("EDINET_Status"),
        ["download_status", "parse_status", "filing_count"],
        result.edinet_status_rows,
    )
    _write_rows_sheet(
        wb.create_sheet("Metric_Coverage"),
        [
            "metric_base",
            "expected_docs",
            "ok_doc_count",
            "any_doc_count",
            "missing_doc_count",
            "coverage_ratio",
            "min_value",
            "max_value",
        ],
        result.metric_coverage_rows,
    )
    _write_rows_sheet(
        wb.create_sheet("JQuants_Quality"),
        [
            "severity",
            "check_name",
            "security_code",
            "fiscal_year",
            "period_key",
            "metric_base",
            "value_num",
            "reference_value",
            "disclosure_number",
            "message",
        ],
        result.jquants_quality_rows,
    )
    issue_rows = [
        _item_row(item)
        for item in sorted(
            result.items,
            key=lambda item: (
                SEVERITY_ORDER.get(item.severity, 99),
                item.category,
                item.check_name,
                item.subject,
            ),
        )
    ]
    _write_rows_sheet(
        wb.create_sheet("Issues"),
        [
            "severity",
            "category",
            "check_name",
            "subject",
            "current_value",
            "previous_value",
            "delta_value",
            "value_unit",
            "message",
        ],
        issue_rows,
    )
    result.excel_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(result.excel_path)


def _item_row(item: DataQualityReportItem) -> dict[str, Any]:
    return {
        "severity": item.severity,
        "category": item.category,
        "check_name": item.check_name,
        "subject": item.subject,
        "current_value": item.current_value if item.current_value is not None else "",
        "previous_value": item.previous_value if item.previous_value is not None else "",
        "delta_value": item.delta_value if item.delta_value is not None else "",
        "value_unit": item.value_unit,
        "message": item.message,
    }


def _save_report_run(
    conn: sqlite3.Connection,
    *,
    result: DataQualityReportResult,
    condition_key: str,
    options: DataQualityReportOptions,
) -> None:
    counts = result.counts_by_severity
    summary = {
        "issue_count": result.issue_count,
        "counts_by_severity": counts,
        "coverage_warning_threshold": options.coverage_warning_threshold,
        "extreme_ratio_threshold": options.extreme_ratio_threshold,
    }
    conn.execute(
        """
        INSERT INTO data_quality_report_runs (
            run_id, generated_at, date_from, date_to, condition_key, codes_json,
            industry_33_json, output_path, previous_run_id, total_items,
            issue_count, critical_count, warning_count, info_count, summary_json,
            created_at
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (
            result.run_id,
            result.generated_at,
            result.date_from,
            result.date_to,
            condition_key,
            _json_dumps(list(_normalized_codes(options.codes))),
            _json_dumps(sorted(options.industry_33_list)),
            str(result.excel_path),
            result.previous_run_id,
            len(result.items),
            result.issue_count,
            counts.get("critical", 0),
            counts.get("warning", 0),
            counts.get("info", 0),
            _json_dumps(summary),
            result.generated_at,
        ),
    )
    rows = [
        (
            result.run_id,
            item.item_key,
            item.category,
            item.severity,
            item.check_name,
            item.subject,
            item.current_value,
            item.previous_value,
            item.delta_value,
            item.value_unit,
            item.message,
            _json_dumps(item.detail),
            result.generated_at,
        )
        for item in result.items
    ]
    conn.executemany(
        """
        INSERT INTO data_quality_report_items (
            run_id, item_key, category, severity, check_name, subject,
            current_value, previous_value, delta_value, value_unit, message,
            detail_json, created_at
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        rows,
    )
    _prune_data_quality_report_history(conn, keep_runs=20)
    conn.commit()


def _prune_data_quality_report_history(conn: sqlite3.Connection, *, keep_runs: int) -> None:
    rows = conn.execute(
        """
        SELECT run_id
        FROM data_quality_report_runs
        ORDER BY generated_at DESC, id DESC
        """
    ).fetchall()
    stale_run_ids = [str(row["run_id"] or "") for row in rows[max(keep_runs, 0) :] if row["run_id"]]
    if not stale_run_ids:
        return
    placeholders = ",".join("?" for _ in stale_run_ids)
    conn.execute(
        f"""
        DELETE FROM data_quality_report_items
        WHERE run_id IN ({placeholders})
        """,
        stale_run_ids,
    )
    conn.execute(
        f"""
        DELETE FROM data_quality_report_runs
        WHERE run_id IN ({placeholders})
        """,
        stale_run_ids,
    )


def export_data_quality_report(
    conn: sqlite3.Connection,
    *,
    options: DataQualityReportOptions,
) -> DataQualityReportResult:
    conn.row_factory = sqlite3.Row
    apply_schema_migrations(conn)
    conn.commit()

    date_from, date_to = _resolve_date_range(conn, options.date_from, options.date_to)
    condition_key = _condition_key(options, date_from, date_to)
    where_sql, params = _build_scope_where(options, date_from, date_to)
    generated_at = datetime.now().isoformat(timespec="seconds")
    run_id = f"dqr_{generated_at.replace('-', '').replace(':', '').replace('T', '_')}_{uuid.uuid4().hex[:8]}"
    output_dir = options.output_dir or (OPERATION_LOG_ROOT / "data_quality")
    excel_path = _build_output_path(Path(output_dir), date_from, date_to, generated_at)

    items: list[DataQualityReportItem] = []
    edinet_status_rows = _fetch_edinet_status_rows(conn, where_sql=where_sql, params=params)
    _add_status_items(edinet_status_rows, items)
    _add_scoped_data_quality_items(conn, where_sql=where_sql, params=params, items=items)
    metric_coverage_rows = _fetch_metric_coverage_rows(conn, where_sql=where_sql, params=params)
    _add_metric_coverage_items(
        metric_coverage_rows,
        coverage_warning_threshold=options.coverage_warning_threshold,
        items=items,
    )
    _add_edinet_outlier_items(
        conn,
        where_sql=where_sql,
        params=params,
        extreme_ratio_threshold=options.extreme_ratio_threshold,
        items=items,
    )
    jquants_quality_rows = build_jquants_quality_issues(
        conn,
        date_from=date_from,
        date_to=date_to,
        codes=list(_normalized_codes(options.codes)),
    )
    _add_jquants_items(jquants_quality_rows, items)

    previous_run_id = _fetch_previous_run_id(conn, condition_key)
    _apply_previous_values(items, _fetch_previous_values(conn, previous_run_id))
    counts = _counts_by_severity(items)
    result = DataQualityReportResult(
        run_id=run_id,
        generated_at=generated_at,
        date_from=date_from,
        date_to=date_to,
        previous_run_id=previous_run_id,
        excel_path=excel_path,
        items=items,
        edinet_status_rows=edinet_status_rows,
        metric_coverage_rows=metric_coverage_rows,
        jquants_quality_rows=jquants_quality_rows,
        counts_by_severity=counts,
    )
    _write_excel_report(result=result, options=options, condition_key=condition_key)
    _save_report_run(conn, result=result, condition_key=condition_key, options=options)
    return result
