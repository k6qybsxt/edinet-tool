from __future__ import annotations

from dataclasses import dataclass
from datetime import date, datetime
import json
from pathlib import Path
import re
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from edinet_monitor.config.settings import OPERATION_LOG_ROOT


DEFAULT_EXCEL_ISSUE_INTAKE_OUTPUT_DIR = OPERATION_LOG_ROOT / "excel_issue_intake"
HEADER_SCAN_ROWS = 20
DEFAULT_MAX_BLANK_DETAILS = 5000
HEADER_TOKEN_GROUPS: dict[str, tuple[str, ...]] = {
    "metric": ("指標", "項目", "metric"),
    "company": ("企業", "会社", "銘柄", "証券コード", "security", "company"),
    "period": ("年度", "期末", "期間", "決算期", "period", "fiscal", "date"),
    "quarter": ("決算種別", "四半期", "quarter", "0q", "1q", "2q", "3q", "4q"),
    "value": ("数値", "値", "金額", "比率", "value", "amount"),
}
CONTEXT_COLUMN_HINTS: dict[str, tuple[str, ...]] = {
    "security_code": ("証券コード", "security", "code"),
    "company_name": ("企業名", "会社名", "銘柄名", "company"),
    "decision": ("決算種別", "四半期", "quarter"),
    "metric": ("指標", "項目", "metric"),
    "period": ("年度", "期末", "期間", "period", "fiscal", "date"),
}


@dataclass(frozen=True)
class ExcelIssueSheetSummary:
    sheet_name: str
    max_row: int
    max_column: int
    header_row: int
    header_score: int
    priority_sheet: bool
    detected_headers: tuple[str, ...]
    blank_cell_count: int
    matched_issue_row_count: int

    def to_dict(self) -> dict[str, Any]:
        return {
            "sheet_name": self.sheet_name,
            "max_row": self.max_row,
            "max_column": self.max_column,
            "header_row": self.header_row,
            "header_score": self.header_score,
            "priority_sheet": self.priority_sheet,
            "detected_headers": list(self.detected_headers),
            "blank_cell_count": self.blank_cell_count,
            "matched_issue_row_count": self.matched_issue_row_count,
        }


@dataclass(frozen=True)
class ExcelIssueBlankCell:
    sheet_name: str
    row_number: int
    column_letter: str
    column_index: int
    header: str
    security_code: str = ""
    company_name: str = ""
    decision: str = ""
    period: str = ""
    metric: str = ""
    matched_issue_terms: tuple[str, ...] = ()

    def to_dict(self) -> dict[str, Any]:
        return {
            "sheet_name": self.sheet_name,
            "row_number": self.row_number,
            "column_letter": self.column_letter,
            "column_index": self.column_index,
            "header": self.header,
            "security_code": self.security_code,
            "company_name": self.company_name,
            "decision": self.decision,
            "period": self.period,
            "metric": self.metric,
            "matched_issue_terms": list(self.matched_issue_terms),
        }


@dataclass(frozen=True)
class ExcelIssueMatchedRow:
    sheet_name: str
    row_number: int
    matched_issue_terms: tuple[str, ...]
    security_code: str = ""
    company_name: str = ""
    decision: str = ""
    period: str = ""
    metric: str = ""
    row_preview: dict[str, Any] | None = None

    def to_dict(self) -> dict[str, Any]:
        return {
            "sheet_name": self.sheet_name,
            "row_number": self.row_number,
            "matched_issue_terms": list(self.matched_issue_terms),
            "security_code": self.security_code,
            "company_name": self.company_name,
            "decision": self.decision,
            "period": self.period,
            "metric": self.metric,
            "row_preview": self.row_preview or {},
        }


@dataclass(frozen=True)
class ExcelIssueIntakeOptions:
    excel_path: Path
    issue_text_path: Path | None = None
    output_dir: Path = DEFAULT_EXCEL_ISSUE_INTAKE_OUTPUT_DIR
    limit_preview: int = 10
    max_blank_details: int = DEFAULT_MAX_BLANK_DETAILS


@dataclass(frozen=True)
class ExcelIssueIntakeResult:
    intake_id: str
    generated_at: str
    status: str
    excel_path: Path
    issue_text_path: Path | None
    json_path: Path
    report_excel_path: Path
    summary: dict[str, Any]
    sheet_summaries: list[ExcelIssueSheetSummary]
    blank_cells: list[ExcelIssueBlankCell]
    matched_rows: list[ExcelIssueMatchedRow]
    issue_terms: tuple[str, ...]


def _now() -> str:
    return datetime.now().isoformat(timespec="seconds")


def _timestamp_for_filename(generated_at: str) -> str:
    return generated_at.replace("-", "").replace(":", "").replace("T", "_")


def _jsonify(value: Any) -> Any:
    if isinstance(value, Path):
        return str(value)
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    if isinstance(value, dict):
        return {str(key): _jsonify(item) for key, item in value.items()}
    if isinstance(value, (list, tuple)):
        return [_jsonify(item) for item in value]
    return value


def _json_dumps(value: Any) -> str:
    return json.dumps(_jsonify(value), ensure_ascii=False, sort_keys=True, separators=(",", ":"))


def _excel_value(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, (dict, list, tuple)):
        return _json_dumps(value)
    if isinstance(value, Path):
        return str(value)
    return value


def _is_blank(value: Any) -> bool:
    if value is None:
        return True
    if isinstance(value, str) and not value.strip():
        return True
    return False


def _clean_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _read_issue_text(path: Path | None) -> str:
    if path is None:
        return ""
    try:
        return path.read_text(encoding="utf-8")
    except UnicodeDecodeError:
        return path.read_text(encoding="cp932")


def _trim_issue_term(value: str) -> str:
    cleaned = value.strip(" \t-・*　")
    cleaned = re.split(r"(?:が|は|を|に|で|出力|未出力|空欄|ない|なし)", cleaned, maxsplit=1)[0]
    return cleaned.strip(" \t-・*　")


def _extract_issue_terms(issue_text: str) -> tuple[str, ...]:
    if not issue_text.strip():
        return ()
    terms: list[str] = []
    for match in re.finditer(r"(?:0Q|[1-4]Q)\s*[^、,，。\s]+", issue_text, re.IGNORECASE):
        terms.append(_trim_issue_term(match.group(0)))
    for raw_part in re.split(r"[\r\n、,，。]+", issue_text):
        cleaned = _trim_issue_term(raw_part)
        if 2 <= len(cleaned) <= 60 and not cleaned.endswith((":", "：")):
            terms.append(cleaned)
    unique_terms = tuple(dict.fromkeys(term for term in terms if term))
    return unique_terms


def _header_score(values: tuple[Any, ...]) -> int:
    text = " ".join(_clean_text(value).lower() for value in values if not _is_blank(value))
    score = 0
    for tokens in HEADER_TOKEN_GROUPS.values():
        if any(token.lower() in text for token in tokens):
            score += 1
    return score


def _detect_header_row(rows: list[tuple[Any, ...]]) -> tuple[int, int]:
    best_row_number = 1
    best_score = 0
    for index, row_values in enumerate(rows[:HEADER_SCAN_ROWS], start=1):
        score = _header_score(row_values)
        if score > best_score:
            best_row_number = index
            best_score = score
    return best_row_number, best_score


def _header_map(headers: tuple[Any, ...]) -> dict[int, str]:
    return {
        index: _clean_text(value)
        for index, value in enumerate(headers, start=1)
        if not _is_blank(value)
    }


def _context_columns(headers: dict[int, str]) -> dict[str, int]:
    result: dict[str, int] = {}
    for context_name, hints in CONTEXT_COLUMN_HINTS.items():
        for index, header in headers.items():
            normalized = header.lower()
            if any(hint.lower() in normalized for hint in hints):
                result[context_name] = index
                break
    return result


def _row_value(row_values: tuple[Any, ...], index: int | None) -> str:
    if index is None or index <= 0 or index > len(row_values):
        return ""
    return _clean_text(row_values[index - 1])


def _row_context(row_values: tuple[Any, ...], context_columns: dict[str, int]) -> dict[str, str]:
    return {
        "security_code": _row_value(row_values, context_columns.get("security_code")),
        "company_name": _row_value(row_values, context_columns.get("company_name")),
        "decision": _row_value(row_values, context_columns.get("decision")),
        "period": _row_value(row_values, context_columns.get("period")),
        "metric": _row_value(row_values, context_columns.get("metric")),
    }


def _matched_terms(row_values: tuple[Any, ...], issue_terms: tuple[str, ...]) -> tuple[str, ...]:
    if not issue_terms:
        return ()
    row_text = " ".join(_clean_text(value) for value in row_values if not _is_blank(value))
    return tuple(term for term in issue_terms if term and term in row_text)


def _row_preview(row_values: tuple[Any, ...], headers: dict[int, str], *, max_columns: int = 20) -> dict[str, Any]:
    preview: dict[str, Any] = {}
    for index in sorted(headers)[:max_columns]:
        if index <= len(row_values):
            preview[headers[index]] = row_values[index - 1]
    return preview


def _write_rows_sheet(workbook: Workbook, *, title: str, headers: list[str], rows: list[dict[str, Any]]) -> None:
    worksheet = workbook.create_sheet(title)
    worksheet.append(headers)
    fill = PatternFill("solid", fgColor="D9EAF7")
    for cell in worksheet[1]:
        cell.font = Font(bold=True)
        cell.fill = fill
    for row in rows:
        worksheet.append([_excel_value(row.get(header, "")) for header in headers])
    for column_cells in worksheet.columns:
        max_length = max(len(str(cell.value or "")) for cell in column_cells)
        worksheet.column_dimensions[column_cells[0].column_letter].width = min(max(max_length + 2, 10), 80)


def _write_json_report(result: ExcelIssueIntakeResult) -> None:
    payload = {
        "intake_id": result.intake_id,
        "generated_at": result.generated_at,
        "status": result.status,
        "excel_path": result.excel_path,
        "issue_text_path": result.issue_text_path,
        "summary": result.summary,
        "issue_terms": list(result.issue_terms),
        "sheet_summaries": [row.to_dict() for row in result.sheet_summaries],
        "blank_cells": [row.to_dict() for row in result.blank_cells],
        "matched_rows": [row.to_dict() for row in result.matched_rows],
        "json_path": result.json_path,
        "report_excel_path": result.report_excel_path,
    }
    result.json_path.parent.mkdir(parents=True, exist_ok=True)
    result.json_path.write_text(json.dumps(_jsonify(payload), ensure_ascii=False, indent=2), encoding="utf-8")


def _write_excel_report(result: ExcelIssueIntakeResult) -> None:
    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "Summary"
    summary_sheet.append(["key", "value"])
    fill = PatternFill("solid", fgColor="D9EAF7")
    for cell in summary_sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = fill
    for key, value in result.summary.items():
        summary_sheet.append([key, _excel_value(value)])
    summary_sheet.column_dimensions["A"].width = 34
    summary_sheet.column_dimensions["B"].width = 80

    _write_rows_sheet(
        workbook,
        title="Sheets",
        headers=[
            "sheet_name",
            "max_row",
            "max_column",
            "header_row",
            "header_score",
            "priority_sheet",
            "detected_headers",
            "blank_cell_count",
            "matched_issue_row_count",
        ],
        rows=[row.to_dict() for row in result.sheet_summaries],
    )
    _write_rows_sheet(
        workbook,
        title="Blank_Cells",
        headers=[
            "sheet_name",
            "row_number",
            "column_letter",
            "column_index",
            "header",
            "security_code",
            "company_name",
            "decision",
            "period",
            "metric",
            "matched_issue_terms",
        ],
        rows=[row.to_dict() for row in result.blank_cells],
    )
    _write_rows_sheet(
        workbook,
        title="Matched_Rows",
        headers=[
            "sheet_name",
            "row_number",
            "matched_issue_terms",
            "security_code",
            "company_name",
            "decision",
            "period",
            "metric",
            "row_preview",
        ],
        rows=[row.to_dict() for row in result.matched_rows],
    )
    _write_rows_sheet(
        workbook,
        title="Issue_Terms",
        headers=["term"],
        rows=[{"term": term} for term in result.issue_terms],
    )
    result.report_excel_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(result.report_excel_path)


def build_excel_issue_intake(options: ExcelIssueIntakeOptions) -> ExcelIssueIntakeResult:
    excel_path = Path(options.excel_path)
    issue_text_path = Path(options.issue_text_path) if options.issue_text_path else None
    generated_at = _now()
    intake_id = f"excel_issue_intake_{_timestamp_for_filename(generated_at)}"
    output_dir = Path(options.output_dir)
    json_path = output_dir / f"{intake_id}.json"
    report_excel_path = output_dir / f"{intake_id}.xlsx"
    issue_terms = _extract_issue_terms(_read_issue_text(issue_text_path))

    workbook = load_workbook(excel_path, read_only=True, data_only=True)
    sheet_summaries: list[ExcelIssueSheetSummary] = []
    blank_cells: list[ExcelIssueBlankCell] = []
    matched_rows: list[ExcelIssueMatchedRow] = []
    total_blank_cell_count = 0
    blank_detail_limit = max(int(options.max_blank_details), 0)
    try:
        for worksheet in workbook.worksheets:
            rows = list(
                worksheet.iter_rows(
                    min_row=1,
                    max_row=min(max(worksheet.max_row or 1, 1), HEADER_SCAN_ROWS),
                    values_only=True,
                )
            )
            header_row, header_score = _detect_header_row(rows)
            header_values = next(
                worksheet.iter_rows(min_row=header_row, max_row=header_row, values_only=True),
                (),
            )
            headers = _header_map(header_values)
            context_columns = _context_columns(headers)
            priority_sheet = header_score >= 2
            sheet_blank_count = 0
            sheet_matched_count = 0

            if priority_sheet:
                for row_number, row_values in enumerate(
                    worksheet.iter_rows(min_row=header_row + 1, values_only=True),
                    start=header_row + 1,
                ):
                    if all(_is_blank(value) for value in row_values):
                        continue
                    context = _row_context(row_values, context_columns)
                    matched = _matched_terms(row_values, issue_terms)
                    if matched:
                        sheet_matched_count += 1
                        matched_rows.append(
                            ExcelIssueMatchedRow(
                                sheet_name=worksheet.title,
                                row_number=row_number,
                                matched_issue_terms=matched,
                                row_preview=_row_preview(row_values, headers),
                                **context,
                            )
                        )
                    for column_index, header in headers.items():
                        value = row_values[column_index - 1] if column_index <= len(row_values) else None
                        if not _is_blank(value):
                            continue
                        sheet_blank_count += 1
                        total_blank_cell_count += 1
                        if len(blank_cells) >= blank_detail_limit:
                            continue
                        blank_cells.append(
                            ExcelIssueBlankCell(
                                sheet_name=worksheet.title,
                                row_number=row_number,
                                column_letter=get_column_letter(column_index),
                                column_index=column_index,
                                header=header,
                                matched_issue_terms=matched,
                                **context,
                            )
                        )

            sheet_summaries.append(
                ExcelIssueSheetSummary(
                    sheet_name=worksheet.title,
                    max_row=int(worksheet.max_row or 0),
                    max_column=int(worksheet.max_column or 0),
                    header_row=header_row,
                    header_score=header_score,
                    priority_sheet=priority_sheet,
                    detected_headers=tuple(headers.values()),
                    blank_cell_count=sheet_blank_count,
                    matched_issue_row_count=sheet_matched_count,
                )
            )
    finally:
        workbook.close()

    priority_sheet_count = sum(1 for row in sheet_summaries if row.priority_sheet)
    status = "review_required" if total_blank_cell_count else "ok"
    summary = {
        "intake_id": intake_id,
        "generated_at": generated_at,
        "status": status,
        "excel_path": excel_path,
        "issue_text_path": issue_text_path or "",
        "sheet_count": len(sheet_summaries),
        "priority_sheet_count": priority_sheet_count,
        "issue_term_count": len(issue_terms),
        "blank_cell_count": total_blank_cell_count,
        "blank_detail_count": len(blank_cells),
        "blank_detail_truncated": total_blank_cell_count > len(blank_cells),
        "matched_issue_row_count": len(matched_rows),
        "max_blank_details": blank_detail_limit,
    }
    result = ExcelIssueIntakeResult(
        intake_id=intake_id,
        generated_at=generated_at,
        status=status,
        excel_path=excel_path,
        issue_text_path=issue_text_path,
        json_path=json_path,
        report_excel_path=report_excel_path,
        summary=summary,
        sheet_summaries=sheet_summaries,
        blank_cells=blank_cells,
        matched_rows=matched_rows,
        issue_terms=issue_terms,
    )
    _write_json_report(result)
    _write_excel_report(result)
    return result
