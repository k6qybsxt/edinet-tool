from __future__ import annotations

from dataclasses import dataclass
from datetime import datetime, timedelta
import json
from pathlib import Path
import sqlite3
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from edinet_monitor.config.settings import OPERATION_LOG_ROOT, PROJECT_ROOT
from edinet_monitor.db.migrations import (
    SCHEMA_MIGRATIONS,
    get_applied_migration_map,
    get_schema_migration_statuses,
)
from edinet_monitor.services.metric_excel_audit_service import (
    DEFAULT_TARGET_CONFIG_PATH,
    ExcelAuditOptions,
    audit_metric_excel,
)
from edinet_monitor.services.metric_excel_golden_master_service import (
    DEFAULT_GOLDEN_MASTER_DIR,
    compare_metric_excel_golden_master,
)
from edinet_monitor.services.prevention_catalog_service import (
    DEFAULT_PREVENTION_CATALOG_PATH,
    load_prevention_catalog,
)


DEFAULT_DAILY_REVIEW_OUTPUT_DIR = OPERATION_LOG_ROOT / "daily_review"
DEFAULT_NORMAL_GOLDEN_JSON_PATH = DEFAULT_GOLDEN_MASTER_DIR / "normal_audit_set.normalized.json"
DEFAULT_KNOWN_ISSUE_GOLDEN_JSON_PATH = DEFAULT_GOLDEN_MASTER_DIR / "known_issue_audit_set.normalized.json"
DEFAULT_DAILY_REVIEW_RETENTION_COUNT = 20


@dataclass(frozen=True)
class DailyReviewExcelSet:
    label: str
    target_set: str
    excel_path: Path
    golden_json_path: Path


@dataclass(frozen=True)
class DailyReviewOptions:
    db_path: Path | None = None
    normal_excel_path: Path | None = None
    known_issue_excel_path: Path | None = None
    normal_golden_json_path: Path = DEFAULT_NORMAL_GOLDEN_JSON_PATH
    known_issue_golden_json_path: Path = DEFAULT_KNOWN_ISSUE_GOLDEN_JSON_PATH
    target_config_path: Path = DEFAULT_TARGET_CONFIG_PATH
    catalog_path: Path = DEFAULT_PREVENTION_CATALOG_PATH
    output_dir: Path = DEFAULT_DAILY_REVIEW_OUTPUT_DIR
    retention_count: int = DEFAULT_DAILY_REVIEW_RETENTION_COUNT
    issue_preview_limit: int = 20
    run_excel_audit: bool = True
    run_golden_master_diff: bool = True


@dataclass(frozen=True)
class DailyReviewResult:
    review_id: str
    generated_at: str
    status: str
    json_path: Path
    excel_path: Path
    summary: dict[str, Any]
    schema_migrations: dict[str, Any]
    db_reflection_items: dict[str, Any]
    preflight_history: dict[str, Any]
    data_quality_report: dict[str, Any]
    excel_audit_results: dict[str, Any]
    golden_master_diff_results: dict[str, Any]


def _now() -> str:
    return datetime.now().isoformat(timespec="seconds")


def _timestamp_for_filename(generated_at: str) -> str:
    return generated_at.replace("-", "").replace(":", "").replace("T", "_")


def _json_dumps(value: Any) -> str:
    return json.dumps(value, ensure_ascii=False, sort_keys=True, separators=(",", ":"))


def _json_loads(text: Any) -> Any:
    if text is None or text == "":
        return None
    try:
        return json.loads(str(text))
    except Exception:
        return str(text)


def _jsonify(value: Any) -> Any:
    if isinstance(value, Path):
        return str(value)
    if isinstance(value, dict):
        return {str(key): _jsonify(item) for key, item in value.items()}
    if isinstance(value, (list, tuple)):
        return [_jsonify(item) for item in value]
    return value


def _excel_value(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, (dict, list, tuple)):
        return _json_dumps(_jsonify(value))
    if isinstance(value, Path):
        return str(value)
    return value


def _table_exists(conn: sqlite3.Connection, table_name: str) -> bool:
    row = conn.execute(
        """
        SELECT 1
        FROM sqlite_master
        WHERE type = 'table' AND name = ?
        """,
        (table_name,),
    ).fetchone()
    return row is not None


def _table_columns(conn: sqlite3.Connection, table_name: str) -> list[str]:
    if not _table_exists(conn, table_name):
        return []
    rows = conn.execute(f"PRAGMA table_info({table_name})").fetchall()
    return [str(row["name"] if isinstance(row, sqlite3.Row) else row[1]) for row in rows]


def _row_to_dict(row: sqlite3.Row) -> dict[str, Any]:
    return {key: row[key] for key in row.keys()}


def _counts_by_key(rows: list[dict[str, Any]], key: str) -> dict[str, int]:
    counts: dict[str, int] = {}
    for row in rows:
        value = str(row.get(key) or "").strip() or "(blank)"
        counts[value] = counts.get(value, 0) + 1
    return counts


def _issue_preview(issues: list[Any], limit: int) -> list[dict[str, Any]]:
    preview: list[dict[str, Any]] = []
    for issue in issues[: max(limit, 0)]:
        if hasattr(issue, "to_dict"):
            preview.append(issue.to_dict())
        elif isinstance(issue, dict):
            preview.append(issue)
        else:
            preview.append({"message": str(issue)})
    return preview


def _excel_sets(options: DailyReviewOptions) -> list[DailyReviewExcelSet]:
    sets: list[DailyReviewExcelSet] = []
    if options.normal_excel_path is not None:
        sets.append(
            DailyReviewExcelSet(
                label="normal",
                target_set="normal",
                excel_path=Path(options.normal_excel_path),
                golden_json_path=Path(options.normal_golden_json_path),
            )
        )
    if options.known_issue_excel_path is not None:
        sets.append(
            DailyReviewExcelSet(
                label="known_issue",
                target_set="known_issue",
                excel_path=Path(options.known_issue_excel_path),
                golden_json_path=Path(options.known_issue_golden_json_path),
            )
        )
    return sets


def _build_schema_migrations_section(conn: sqlite3.Connection) -> dict[str, Any]:
    statuses = get_schema_migration_statuses(conn)
    rows = [
        {
            "migration_id": status.migration_id,
            "description": status.description,
            "applied": status.applied,
            "applied_at": status.applied_at,
        }
        for status in statuses
    ]
    missing = [row for row in rows if not row["applied"]]
    applied_count = len(rows) - len(missing)
    known_ids = {migration.migration_id for migration in SCHEMA_MIGRATIONS}
    applied_map = get_applied_migration_map(conn)
    extra_applied = [
        {"migration_id": migration_id, "applied_at": applied_at}
        for migration_id, applied_at in sorted(applied_map.items())
        if migration_id not in known_ids
    ]
    return {
        "status": "review_required" if missing else "ok",
        "known_count": len(rows),
        "applied_count": applied_count,
        "missing_count": len(missing),
        "extra_applied_count": len(extra_applied),
        "missing": missing,
        "extra_applied": extra_applied,
        "rows": rows,
    }


def _build_db_reflection_section(conn: sqlite3.Connection) -> dict[str, Any]:
    if not _table_exists(conn, "db_reflection_items"):
        return {
            "status": "no_table",
            "pending_count": 0,
            "total_count": 0,
            "category_counts": {},
            "rows": [],
        }
    columns = _table_columns(conn, "db_reflection_items")
    order_columns = [column for column in ("created_at", "item_id", "id") if column in columns]
    order_sql = f"ORDER BY {', '.join(order_columns)}" if order_columns else ""
    rows = [
        _row_to_dict(row)
        for row in conn.execute(f"SELECT * FROM db_reflection_items {order_sql}").fetchall()
    ]
    if "status" in columns:
        pending_rows = [
            row
            for row in rows
            if str(row.get("status") or "").strip().lower() not in {"complete", "completed", "done"}
        ]
    else:
        pending_rows = rows
        for row in rows:
            row.setdefault("status", "pending")
    normalized_rows = [
        {
            "item_id": row.get("item_id", row.get("id", "")),
            "title": row.get("title", ""),
            "category": row.get("category", ""),
            "status": row.get("status", "pending"),
            "description": row.get("description", row.get("summary", "")),
            "source_key": row.get("source_key", ""),
            "created_at": row.get("created_at", ""),
            "updated_at": row.get("updated_at", ""),
        }
        for row in rows
    ]
    return {
        "status": "review_required" if pending_rows else "ok",
        "pending_count": len(pending_rows),
        "total_count": len(rows),
        "category_counts": _counts_by_key(normalized_rows, "category"),
        "rows": normalized_rows,
    }


def _build_preflight_history_section(
    conn: sqlite3.Connection,
    *,
    generated_at: str,
    catalog_path: Path,
    preview_limit: int,
) -> dict[str, Any]:
    cutoff = datetime.fromisoformat(generated_at) - timedelta(days=1)
    runs: list[dict[str, Any]] = []
    issues: list[dict[str, Any]] = []
    if _table_exists(conn, "preflight_runs"):
        run_rows = conn.execute(
            """
            SELECT *
            FROM preflight_runs
            WHERE generated_at >= ?
            ORDER BY generated_at DESC, id DESC
            LIMIT 200
            """,
            (cutoff.isoformat(timespec="seconds"),),
        ).fetchall()
        for row in run_rows:
            run = _row_to_dict(row)
            run["command_names_json"] = _json_loads(run.get("command_names_json")) or []
            run["db_reflection_blocked"] = bool(run.get("db_reflection_blocked"))
            runs.append(run)
        preflight_ids = [str(run.get("preflight_id") or "") for run in runs if run.get("preflight_id")]
        if preflight_ids and _table_exists(conn, "preflight_run_issues"):
            placeholders = ",".join("?" for _ in preflight_ids)
            issue_rows = conn.execute(
                f"""
                SELECT preflight_id, severity, category, check_name, item_id,
                       title, message, detail_json, created_at
                FROM preflight_run_issues
                WHERE preflight_id IN ({placeholders})
                ORDER BY
                    CASE severity
                        WHEN 'critical' THEN 0
                        WHEN 'warning' THEN 1
                        WHEN 'info' THEN 2
                        ELSE 9
                    END,
                    preflight_id,
                    category,
                    check_name
                LIMIT ?
                """,
                [*preflight_ids, max(int(preview_limit), 0)],
            ).fetchall()
            for row in issue_rows:
                issue = _row_to_dict(row)
                issue["detail_json"] = _json_loads(issue.get("detail_json"))
                issues.append(issue)
        history_status = "ok"
    else:
        history_status = "no_table"

    triggered_all: list[dict[str, Any]] = []
    triggered_items: list[dict[str, Any]] = []
    catalog_error = ""
    try:
        catalog_items = load_prevention_catalog(Path(catalog_path))
        triggered_all = [
            item.to_dict()
            for item in catalog_items
            if item.status == "triggered"
        ]
        triggered_items = triggered_all[: max(int(preview_limit), 0)]
    except Exception as exc:
        catalog_error = str(exc)

    blocked_count = sum(1 for run in runs if bool(run.get("db_reflection_blocked")) or run.get("status") == "blocked")
    warning_run_count = sum(1 for run in runs if int(run.get("warning_count") or 0) > 0)
    incomplete_count = sum(
        1
        for run in runs
        if run.get("status") in {"passed", "passed_with_warnings"} and not run.get("completed_at")
    )
    cli_counts = _counts_by_key(runs, "cli_name")
    if catalog_error:
        status = "review_required"
    elif blocked_count or warning_run_count or incomplete_count or triggered_all:
        status = "review_required"
    else:
        status = history_status
    return {
        "status": status,
        "history_status": history_status,
        "run_count": len(runs),
        "blocked_count": blocked_count,
        "warning_run_count": warning_run_count,
        "incomplete_count": incomplete_count,
        "critical_count": sum(int(run.get("critical_count") or 0) for run in runs),
        "warning_count": sum(int(run.get("warning_count") or 0) for run in runs),
        "cli_counts": cli_counts,
        "runs": runs,
        "issues": issues,
        "catalog_path": Path(catalog_path),
        "catalog_triggered_count": len(triggered_all),
        "catalog_triggered_items": triggered_items,
        "catalog_error": catalog_error,
    }


def _build_data_quality_section(conn: sqlite3.Connection) -> dict[str, Any]:
    if not _table_exists(conn, "data_quality_report_runs"):
        return {
            "status": "no_table",
            "run": None,
            "critical_count": 0,
            "warning_count": 0,
            "info_count": 0,
            "issue_count": 0,
            "items": [],
        }
    run = conn.execute(
        """
        SELECT *
        FROM data_quality_report_runs
        ORDER BY generated_at DESC, id DESC
        LIMIT 1
        """
    ).fetchone()
    if run is None:
        return {
            "status": "no_report",
            "run": None,
            "critical_count": 0,
            "warning_count": 0,
            "info_count": 0,
            "issue_count": 0,
            "items": [],
        }

    run_dict = _row_to_dict(run)
    run_dict["codes_json"] = _json_loads(run_dict.get("codes_json"))
    run_dict["industry_33_json"] = _json_loads(run_dict.get("industry_33_json"))
    run_dict["summary_json"] = _json_loads(run_dict.get("summary_json"))
    items: list[dict[str, Any]] = []
    if _table_exists(conn, "data_quality_report_items"):
        item_rows = conn.execute(
            """
            SELECT severity, category, check_name, subject, current_value,
                   previous_value, delta_value, value_unit, message, detail_json
            FROM data_quality_report_items
            WHERE run_id = ?
            ORDER BY
                CASE severity
                    WHEN 'critical' THEN 0
                    WHEN 'warning' THEN 1
                    WHEN 'info' THEN 2
                    ELSE 9
                END,
                category,
                check_name,
                subject
            LIMIT 200
            """,
            (run_dict.get("run_id"),),
        ).fetchall()
        for item_row in item_rows:
            item = _row_to_dict(item_row)
            item["detail_json"] = _json_loads(item.get("detail_json"))
            items.append(item)

    critical = int(run_dict.get("critical_count") or 0)
    warning = int(run_dict.get("warning_count") or 0)
    issue_count = int(run_dict.get("issue_count") or 0)
    return {
        "status": "review_required" if critical or warning else "ok",
        "run": run_dict,
        "critical_count": critical,
        "warning_count": warning,
        "info_count": int(run_dict.get("info_count") or 0),
        "issue_count": issue_count,
        "items": items,
    }


def _run_excel_audits(
    conn: sqlite3.Connection,
    options: DailyReviewOptions,
    excel_sets: list[DailyReviewExcelSet],
) -> dict[str, Any]:
    if not options.run_excel_audit:
        return {
            "status": "skipped",
            "critical_count": 0,
            "warning_count": 0,
            "error_count": 0,
            "results": [],
        }
    if not excel_sets:
        return {
            "status": "not_configured",
            "critical_count": 0,
            "warning_count": 0,
            "error_count": 1,
            "results": [
                {
                    "status": "error",
                    "error": "No Excel paths were configured for Excel audit.",
                    "critical_count": 0,
                    "warning_count": 0,
                    "issue_count": 0,
                }
            ],
        }
    results: list[dict[str, Any]] = []
    critical_total = 0
    warning_total = 0
    error_total = 0
    for excel_set in excel_sets:
        if not excel_set.excel_path.exists():
            error_total += 1
            results.append(
                {
                    "label": excel_set.label,
                    "target_set": excel_set.target_set,
                    "status": "error",
                    "excel_path": excel_set.excel_path,
                    "error": f"Excel file not found: {excel_set.excel_path}",
                    "critical_count": 0,
                    "warning_count": 0,
                    "issue_count": 0,
                }
            )
            continue
        try:
            result = audit_metric_excel(
                conn,
                ExcelAuditOptions(
                    excel_path=excel_set.excel_path,
                    db_path=options.db_path,
                    target_set=excel_set.target_set,
                    target_config_path=options.target_config_path,
                    output_dir=Path(options.output_dir) / "excel_audit",
                ),
            )
        except Exception as exc:
            error_total += 1
            results.append(
                {
                    "label": excel_set.label,
                    "target_set": excel_set.target_set,
                    "status": "error",
                    "excel_path": excel_set.excel_path,
                    "error": str(exc),
                    "critical_count": 0,
                    "warning_count": 0,
                    "issue_count": 0,
                }
            )
            continue
        critical = int(result.counts_by_severity.get("critical", 0))
        warning = int(result.counts_by_severity.get("warning", 0))
        critical_total += critical
        warning_total += warning
        results.append(
            {
                "label": excel_set.label,
                "target_set": excel_set.target_set,
                "status": "review_required" if result.issue_count else "ok",
                "audit_id": result.audit_id,
                "generated_at": result.generated_at,
                "excel_path": result.excel_path,
                "json_path": result.json_path,
                "report_excel_path": result.report_excel_path,
                "target_count": len(result.targets),
                "expected_rows": result.expected_rows,
                "actual_rows": result.actual_rows,
                "issue_count": result.issue_count,
                "critical_count": critical,
                "warning_count": warning,
                "errors": result.errors,
                "warnings": result.warnings,
                "issue_preview": _issue_preview(result.issues, options.issue_preview_limit),
            }
        )
    return {
        "status": "review_required" if critical_total or warning_total or error_total else "ok",
        "critical_count": critical_total,
        "warning_count": warning_total,
        "error_count": error_total,
        "results": results,
    }


def _run_golden_master_diffs(
    options: DailyReviewOptions,
    excel_sets: list[DailyReviewExcelSet],
) -> dict[str, Any]:
    if not options.run_golden_master_diff:
        return {
            "status": "skipped",
            "critical_count": 0,
            "warning_count": 0,
            "error_count": 0,
            "results": [],
        }
    if not excel_sets:
        return {
            "status": "not_configured",
            "critical_count": 0,
            "warning_count": 0,
            "error_count": 1,
            "results": [
                {
                    "status": "error",
                    "error": "No Excel paths were configured for Golden Master diff.",
                    "critical_count": 0,
                    "warning_count": 0,
                    "issue_count": 0,
                }
            ],
        }
    results: list[dict[str, Any]] = []
    critical_total = 0
    warning_total = 0
    error_total = 0
    for excel_set in excel_sets:
        if not excel_set.excel_path.exists():
            error_total += 1
            results.append(
                {
                    "label": excel_set.label,
                    "target_set": excel_set.target_set,
                    "status": "error",
                    "excel_path": excel_set.excel_path,
                    "golden_json_path": excel_set.golden_json_path,
                    "error": f"Excel file not found: {excel_set.excel_path}",
                    "critical_count": 0,
                    "warning_count": 0,
                    "issue_count": 0,
                }
            )
            continue
        if not excel_set.golden_json_path.exists():
            error_total += 1
            results.append(
                {
                    "label": excel_set.label,
                    "target_set": excel_set.target_set,
                    "status": "error",
                    "excel_path": excel_set.excel_path,
                    "golden_json_path": excel_set.golden_json_path,
                    "error": f"Golden Master JSON not found: {excel_set.golden_json_path}",
                    "critical_count": 0,
                    "warning_count": 0,
                    "issue_count": 0,
                }
            )
            continue
        try:
            result = compare_metric_excel_golden_master(
                golden_json_path=excel_set.golden_json_path,
                actual_excel_path=excel_set.excel_path,
                output_dir=Path(options.output_dir) / "golden_master_diff",
            )
        except Exception as exc:
            error_total += 1
            results.append(
                {
                    "label": excel_set.label,
                    "target_set": excel_set.target_set,
                    "status": "error",
                    "excel_path": excel_set.excel_path,
                    "golden_json_path": excel_set.golden_json_path,
                    "error": str(exc),
                    "critical_count": 0,
                    "warning_count": 0,
                    "issue_count": 0,
                }
            )
            continue
        critical = int(result.counts_by_severity.get("critical", 0))
        warning = int(result.counts_by_severity.get("warning", 0))
        critical_total += critical
        warning_total += warning
        results.append(
            {
                "label": excel_set.label,
                "target_set": excel_set.target_set,
                "status": "review_required" if result.issue_count else "ok",
                "comparison_id": result.comparison_id,
                "generated_at": result.generated_at,
                "golden_json_path": result.golden_json_path,
                "actual_excel_path": result.actual_excel_path,
                "actual_json_path": result.actual_json_path,
                "report_json_path": result.report_json_path,
                "report_excel_path": result.report_excel_path,
                "issue_count": result.issue_count,
                "critical_count": critical,
                "warning_count": warning,
                "issue_preview": _issue_preview(result.issues, options.issue_preview_limit),
            }
        )
    return {
        "status": "review_required" if critical_total or warning_total or error_total else "ok",
        "critical_count": critical_total,
        "warning_count": warning_total,
        "error_count": error_total,
        "results": results,
    }


def _build_summary(
    *,
    review_id: str,
    generated_at: str,
    options: DailyReviewOptions,
    schema_migrations: dict[str, Any],
    db_reflection_items: dict[str, Any],
    preflight_history: dict[str, Any],
    data_quality_report: dict[str, Any],
    excel_audit_results: dict[str, Any],
    golden_master_diff_results: dict[str, Any],
) -> dict[str, Any]:
    review_error_count = int(excel_audit_results.get("error_count") or 0) + int(
        golden_master_diff_results.get("error_count") or 0
    )
    attention_count = sum(
        [
            int(schema_migrations.get("missing_count") or 0),
            int(db_reflection_items.get("pending_count") or 0),
            int(preflight_history.get("blocked_count") or 0),
            int(preflight_history.get("warning_run_count") or 0),
            int(preflight_history.get("incomplete_count") or 0),
            int(preflight_history.get("catalog_triggered_count") or 0),
            1 if preflight_history.get("catalog_error") else 0,
            int(data_quality_report.get("critical_count") or 0),
            int(data_quality_report.get("warning_count") or 0),
            int(excel_audit_results.get("critical_count") or 0),
            int(excel_audit_results.get("warning_count") or 0),
            int(golden_master_diff_results.get("critical_count") or 0),
            int(golden_master_diff_results.get("warning_count") or 0),
            review_error_count,
        ]
    )
    status = "review_required" if attention_count else "ok"
    return {
        "review_id": review_id,
        "generated_at": generated_at,
        "status": status,
        "pipeline_failure_policy": "report_only",
        "pipeline_failed": False,
        "project_root": PROJECT_ROOT,
        "db_path": options.db_path or "",
        "schema_missing_count": int(schema_migrations.get("missing_count") or 0),
        "schema_applied_count": int(schema_migrations.get("applied_count") or 0),
        "schema_known_count": int(schema_migrations.get("known_count") or 0),
        "db_reflection_pending_count": int(db_reflection_items.get("pending_count") or 0),
        "preflight_run_count": int(preflight_history.get("run_count") or 0),
        "preflight_blocked_count": int(preflight_history.get("blocked_count") or 0),
        "preflight_warning_run_count": int(preflight_history.get("warning_run_count") or 0),
        "preflight_incomplete_count": int(preflight_history.get("incomplete_count") or 0),
        "preflight_catalog_triggered_count": int(preflight_history.get("catalog_triggered_count") or 0),
        "data_quality_critical_count": int(data_quality_report.get("critical_count") or 0),
        "data_quality_warning_count": int(data_quality_report.get("warning_count") or 0),
        "excel_audit_critical_count": int(excel_audit_results.get("critical_count") or 0),
        "excel_audit_warning_count": int(excel_audit_results.get("warning_count") or 0),
        "golden_master_critical_count": int(golden_master_diff_results.get("critical_count") or 0),
        "golden_master_warning_count": int(golden_master_diff_results.get("warning_count") or 0),
        "review_error_count": review_error_count,
        "attention_count": attention_count,
    }


def _write_table(
    workbook: Workbook,
    *,
    title: str,
    headers: list[str],
    rows: list[dict[str, Any]],
) -> None:
    worksheet = workbook.create_sheet(title)
    worksheet.append(headers)
    fill = PatternFill("solid", fgColor="D9EAF7")
    for cell in worksheet[1]:
        cell.font = Font(bold=True)
        cell.fill = fill
    for row in rows:
        worksheet.append([_excel_value(row.get(header, "")) for header in headers])
    for column_cells in worksheet.columns:
        max_length = max(len(str(cell.value or "")) for cell in column_cells)
        worksheet.column_dimensions[column_cells[0].column_letter].width = min(max(max_length + 2, 10), 60)


def _write_json_report(result: DailyReviewResult) -> None:
    payload = {
        "review_id": result.review_id,
        "generated_at": result.generated_at,
        "status": result.status,
        "summary": result.summary,
        "schema_migrations": result.schema_migrations,
        "db_reflection_items": result.db_reflection_items,
        "preflight_history": result.preflight_history,
        "data_quality_report": result.data_quality_report,
        "excel_audit_results": result.excel_audit_results,
        "golden_master_diff_results": result.golden_master_diff_results,
        "json_path": result.json_path,
        "excel_path": result.excel_path,
    }
    result.json_path.parent.mkdir(parents=True, exist_ok=True)
    result.json_path.write_text(
        json.dumps(_jsonify(payload), ensure_ascii=False, indent=2),
        encoding="utf-8",
    )


def _write_excel_report(result: DailyReviewResult) -> None:
    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "Summary"
    summary_sheet.append(["key", "value"])
    fill = PatternFill("solid", fgColor="D9EAF7")
    for cell in summary_sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = fill
    for key, value in result.summary.items():
        summary_sheet.append([key, _excel_value(value)])
    summary_sheet.column_dimensions["A"].width = 34
    summary_sheet.column_dimensions["B"].width = 80

    _write_table(
        workbook,
        title="Schema_Migrations",
        headers=["migration_id", "applied", "applied_at", "description"],
        rows=result.schema_migrations.get("rows", []),
    )
    _write_table(
        workbook,
        title="DB_Reflection",
        headers=[
            "item_id",
            "title",
            "category",
            "status",
            "description",
            "source_key",
            "created_at",
            "updated_at",
        ],
        rows=result.db_reflection_items.get("rows", []),
    )
    preflight_rows: list[dict[str, Any]] = []
    for run in result.preflight_history.get("runs", []):
        preflight_rows.append({**run, "kind": "run"})
    for issue in result.preflight_history.get("issues", []):
        preflight_rows.append({**issue, "kind": "issue"})
    for item in result.preflight_history.get("catalog_triggered_items", []):
        preflight_rows.append(
            {
                "kind": "catalog_triggered",
                "preflight_id": item.get("id", ""),
                "status": item.get("status", ""),
                "severity": item.get("severity", ""),
                "category": ",".join(item.get("areas", [])),
                "check_name": ",".join(item.get("triggers", [])),
                "title": item.get("title", ""),
                "message": item.get("problem", ""),
            }
        )
    _write_table(
        workbook,
        title="Preflight_History",
        headers=[
            "kind",
            "preflight_id",
            "generated_at",
            "cli_name",
            "status",
            "db_reflection_blocked",
            "critical_count",
            "warning_count",
            "completed_at",
            "severity",
            "category",
            "check_name",
            "item_id",
            "title",
            "message",
            "json_path",
            "excel_path",
        ],
        rows=preflight_rows,
    )
    data_quality_rows: list[dict[str, Any]] = []
    run = result.data_quality_report.get("run")
    if run:
        data_quality_rows.append(
            {
                "kind": "run_summary",
                "severity": "",
                "category": "",
                "check_name": str(run.get("run_id") or ""),
                "subject": str(run.get("condition_key") or ""),
                "current_value": result.data_quality_report.get("issue_count", 0),
                "value_unit": "issue_count",
                "message": str(run.get("output_path") or ""),
                "detail_json": run,
            }
        )
    for item in result.data_quality_report.get("items", []):
        row = dict(item)
        row["kind"] = "issue"
        data_quality_rows.append(row)
    _write_table(
        workbook,
        title="Data_Quality",
        headers=[
            "kind",
            "severity",
            "category",
            "check_name",
            "subject",
            "current_value",
            "previous_value",
            "delta_value",
            "value_unit",
            "message",
            "detail_json",
        ],
        rows=data_quality_rows,
    )

    audit_rows: list[dict[str, Any]] = []
    for audit in result.excel_audit_results.get("results", []):
        audit_rows.append({**audit, "kind": "audit_result"})
        for issue in audit.get("issue_preview", []):
            audit_rows.append(
                {
                    "kind": "issue",
                    "label": audit.get("label", ""),
                    "target_set": audit.get("target_set", ""),
                    "severity": issue.get("severity", ""),
                    "category": issue.get("category", ""),
                    "check_name": issue.get("check_name", ""),
                    "security_code": issue.get("security_code", ""),
                    "metric_label": issue.get("metric_label", ""),
                    "period_label": issue.get("period_label", ""),
                    "message": issue.get("message", ""),
                }
            )
    _write_table(
        workbook,
        title="Excel_Audit",
        headers=[
            "kind",
            "label",
            "target_set",
            "status",
            "critical_count",
            "warning_count",
            "issue_count",
            "severity",
            "category",
            "check_name",
            "security_code",
            "metric_label",
            "period_label",
            "message",
            "json_path",
            "report_excel_path",
            "error",
        ],
        rows=audit_rows,
    )

    golden_rows: list[dict[str, Any]] = []
    for diff in result.golden_master_diff_results.get("results", []):
        golden_rows.append({**diff, "kind": "diff_result"})
        for issue in diff.get("issue_preview", []):
            golden_rows.append(
                {
                    "kind": "issue",
                    "label": diff.get("label", ""),
                    "target_set": diff.get("target_set", ""),
                    "severity": issue.get("severity", ""),
                    "category": issue.get("category", ""),
                    "check_name": issue.get("check_name", ""),
                    "sheet_name": issue.get("sheet_name", ""),
                    "row_key": issue.get("row_key", ""),
                    "period_label": issue.get("period_label", ""),
                    "field_name": issue.get("field_name", ""),
                    "message": issue.get("message", ""),
                }
            )
    _write_table(
        workbook,
        title="Golden_Master",
        headers=[
            "kind",
            "label",
            "target_set",
            "status",
            "critical_count",
            "warning_count",
            "issue_count",
            "severity",
            "category",
            "check_name",
            "sheet_name",
            "row_key",
            "period_label",
            "field_name",
            "message",
            "report_json_path",
            "report_excel_path",
            "actual_json_path",
            "error",
        ],
        rows=golden_rows,
    )

    result.excel_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(result.excel_path)


def _prune_old_reports(output_dir: Path, retention_count: int) -> None:
    keep = max(int(retention_count), 1)
    for pattern in ("daily_review_*.json", "daily_review_*.xlsx"):
        files = sorted(
            output_dir.glob(pattern),
            key=lambda path: (path.stat().st_mtime, path.name),
            reverse=True,
        )
        for old_file in files[keep:]:
            old_file.unlink(missing_ok=True)


def build_daily_review(conn: sqlite3.Connection, options: DailyReviewOptions) -> DailyReviewResult:
    conn.row_factory = sqlite3.Row
    generated_at = _now()
    review_id = f"daily_review_{_timestamp_for_filename(generated_at)}"
    output_dir = Path(options.output_dir)
    json_path = output_dir / f"{review_id}.json"
    excel_path = output_dir / f"{review_id}.xlsx"
    excel_sets = _excel_sets(options)

    schema_migrations = _build_schema_migrations_section(conn)
    db_reflection_items = _build_db_reflection_section(conn)
    preflight_history = _build_preflight_history_section(
        conn,
        generated_at=generated_at,
        catalog_path=Path(options.catalog_path),
        preview_limit=options.issue_preview_limit,
    )
    data_quality_report = _build_data_quality_section(conn)
    excel_audit_results = _run_excel_audits(conn, options, excel_sets)
    golden_master_diff_results = _run_golden_master_diffs(options, excel_sets)
    summary = _build_summary(
        review_id=review_id,
        generated_at=generated_at,
        options=options,
        schema_migrations=schema_migrations,
        db_reflection_items=db_reflection_items,
        preflight_history=preflight_history,
        data_quality_report=data_quality_report,
        excel_audit_results=excel_audit_results,
        golden_master_diff_results=golden_master_diff_results,
    )
    result = DailyReviewResult(
        review_id=review_id,
        generated_at=generated_at,
        status=str(summary["status"]),
        json_path=json_path,
        excel_path=excel_path,
        summary=summary,
        schema_migrations=schema_migrations,
        db_reflection_items=db_reflection_items,
        preflight_history=preflight_history,
        data_quality_report=data_quality_report,
        excel_audit_results=excel_audit_results,
        golden_master_diff_results=golden_master_diff_results,
    )
    _write_json_report(result)
    _write_excel_report(result)
    _prune_old_reports(output_dir, options.retention_count)
    return result
