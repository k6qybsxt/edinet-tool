from __future__ import annotations

from dataclasses import dataclass
from datetime import datetime
import json
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from edinet_monitor.config.settings import OPERATION_LOG_ROOT, PROJECT_ROOT


DEFAULT_PREVENTION_CATALOG_PATH = PROJECT_ROOT / "config" / "quality" / "prevention_catalog.json"
DEFAULT_PREVENTION_CATALOG_REVIEW_OUTPUT_DIR = OPERATION_LOG_ROOT / "prevention_catalog_review"
ALLOWED_PREVENTION_STATUSES = {"active", "triggered", "monitoring", "retired", "rejected"}
ACTIVE_PREVENTION_STATUSES = ("active", "triggered", "monitoring")
ALLOWED_PREVENTION_SEVERITIES = {"critical", "warning", "info"}
SEVERITY_ORDER = {"critical": 0, "warning": 1, "info": 2}


@dataclass(frozen=True)
class PreventionCatalogItem:
    item_id: str
    title: str
    status: str
    severity: str
    areas: tuple[str, ...]
    triggers: tuple[str, ...]
    problem: str
    prevention: str
    review_points: tuple[str, ...]

    def to_dict(self) -> dict[str, Any]:
        return {
            "id": self.item_id,
            "title": self.title,
            "status": self.status,
            "severity": self.severity,
            "areas": list(self.areas),
            "triggers": list(self.triggers),
            "problem": self.problem,
            "prevention": self.prevention,
            "review_points": list(self.review_points),
        }


@dataclass(frozen=True)
class PreventionCatalogReviewOptions:
    catalog_path: Path = DEFAULT_PREVENTION_CATALOG_PATH
    areas: tuple[str, ...] = ()
    triggers: tuple[str, ...] = ()
    statuses: tuple[str, ...] = ACTIVE_PREVENTION_STATUSES
    output_dir: Path = DEFAULT_PREVENTION_CATALOG_REVIEW_OUTPUT_DIR


@dataclass(frozen=True)
class PreventionCatalogReviewResult:
    review_id: str
    generated_at: str
    status: str
    json_path: Path
    excel_path: Path
    summary: dict[str, Any]
    matched_items: list[PreventionCatalogItem]
    counts_by_severity: dict[str, int]

    @property
    def issue_count(self) -> int:
        return self.counts_by_severity.get("critical", 0) + self.counts_by_severity.get("warning", 0)


@dataclass(frozen=True)
class PreventionCatalogStatusUpdateResult:
    catalog_path: Path
    target_status: str
    requested_ids: tuple[str, ...]
    updated_ids: tuple[str, ...]
    skipped_ids: tuple[str, ...]


def _now() -> str:
    return datetime.now().isoformat(timespec="seconds")


def _timestamp_for_filename(generated_at: str) -> str:
    return generated_at.replace("-", "").replace(":", "").replace("T", "_")


def _jsonify(value: Any) -> Any:
    if isinstance(value, Path):
        return str(value)
    if isinstance(value, dict):
        return {str(key): _jsonify(item) for key, item in value.items()}
    if isinstance(value, (list, tuple)):
        return [_jsonify(item) for item in value]
    return value


def _json_dumps(value: Any) -> str:
    return json.dumps(_jsonify(value), ensure_ascii=False, sort_keys=True, separators=(",", ":"))


def _clean_token(value: Any) -> str:
    return str(value or "").strip()


def _validate_string_field(item: dict[str, Any], field_name: str, item_ref: str, errors: list[str]) -> str:
    value = item.get(field_name)
    if not isinstance(value, str) or not value.strip():
        errors.append(f"{item_ref}.{field_name} must be a non-empty string")
        return ""
    return value.strip()


def _validate_string_list_field(
    item: dict[str, Any],
    field_name: str,
    item_ref: str,
    errors: list[str],
) -> tuple[str, ...]:
    value = item.get(field_name)
    if not isinstance(value, list) or not value:
        errors.append(f"{item_ref}.{field_name} must be a non-empty string list")
        return ()
    values: list[str] = []
    for index, raw in enumerate(value):
        clean = _clean_token(raw)
        if not clean:
            errors.append(f"{item_ref}.{field_name}[{index}] must be a non-empty string")
            continue
        values.append(clean)
    return tuple(dict.fromkeys(values))


def _build_item(raw: Any, index: int, errors: list[str]) -> PreventionCatalogItem | None:
    item_ref = f"items[{index}]"
    if not isinstance(raw, dict):
        errors.append(f"{item_ref} must be an object")
        return None

    item_id = _validate_string_field(raw, "id", item_ref, errors)
    title = _validate_string_field(raw, "title", item_ref, errors)
    status = _validate_string_field(raw, "status", item_ref, errors)
    severity = _validate_string_field(raw, "severity", item_ref, errors)
    areas = _validate_string_list_field(raw, "areas", item_ref, errors)
    triggers = _validate_string_list_field(raw, "triggers", item_ref, errors)
    problem = _validate_string_field(raw, "problem", item_ref, errors)
    prevention = _validate_string_field(raw, "prevention", item_ref, errors)
    review_points = _validate_string_list_field(raw, "review_points", item_ref, errors)

    if status and status not in ALLOWED_PREVENTION_STATUSES:
        errors.append(f"{item_ref}.status is invalid: {status}")
    if severity and severity not in ALLOWED_PREVENTION_SEVERITIES:
        errors.append(f"{item_ref}.severity is invalid: {severity}")
    if not item_id:
        return None
    return PreventionCatalogItem(
        item_id=item_id,
        title=title,
        status=status,
        severity=severity,
        areas=areas,
        triggers=triggers,
        problem=problem,
        prevention=prevention,
        review_points=review_points,
    )


def load_prevention_catalog(catalog_path: str | Path = DEFAULT_PREVENTION_CATALOG_PATH) -> list[PreventionCatalogItem]:
    path = Path(catalog_path)
    payload = json.loads(path.read_text(encoding="utf-8"))
    errors: list[str] = []
    if not isinstance(payload, dict):
        raise ValueError(f"prevention catalog must be an object: {path}")
    if payload.get("version") != 1:
        errors.append("version must be 1")
    raw_items = payload.get("items")
    if not isinstance(raw_items, list):
        errors.append("items must be a list")
        raw_items = []

    items: list[PreventionCatalogItem] = []
    seen_ids: set[str] = set()
    for index, raw_item in enumerate(raw_items):
        item = _build_item(raw_item, index, errors)
        if item is None:
            continue
        if item.item_id in seen_ids:
            errors.append(f"duplicate item id: {item.item_id}")
            continue
        seen_ids.add(item.item_id)
        items.append(item)

    if errors:
        raise ValueError("; ".join(errors))
    return items


def update_prevention_catalog_statuses(
    catalog_path: str | Path,
    *,
    item_ids: tuple[str, ...] | list[str],
    from_statuses: tuple[str, ...] | list[str],
    to_status: str,
) -> PreventionCatalogStatusUpdateResult:
    path = Path(catalog_path)
    requested_ids = tuple(dict.fromkeys(_clean_token(item_id) for item_id in item_ids if _clean_token(item_id)))
    from_status_set = set(_normalize_filter_values(from_statuses))
    clean_to_status = _clean_token(to_status)
    if clean_to_status not in ALLOWED_PREVENTION_STATUSES:
        raise ValueError(f"unknown prevention catalog status: {clean_to_status}")
    unknown_from_statuses = from_status_set - ALLOWED_PREVENTION_STATUSES
    if unknown_from_statuses:
        raise ValueError(f"unknown prevention catalog status: {', '.join(sorted(unknown_from_statuses))}")
    if not requested_ids:
        return PreventionCatalogStatusUpdateResult(
            catalog_path=path,
            target_status=clean_to_status,
            requested_ids=(),
            updated_ids=(),
            skipped_ids=(),
        )

    payload = json.loads(path.read_text(encoding="utf-8"))
    # Validate the original catalog before mutating the raw JSON payload.
    load_prevention_catalog(path)
    requested_id_set = set(requested_ids)
    updated_ids: list[str] = []
    for raw_item in payload.get("items", []):
        if not isinstance(raw_item, dict):
            continue
        item_id = _clean_token(raw_item.get("id"))
        if item_id in requested_id_set and _clean_token(raw_item.get("status")) in from_status_set:
            raw_item["status"] = clean_to_status
            updated_ids.append(item_id)
    skipped_ids = [item_id for item_id in requested_ids if item_id not in set(updated_ids)]
    if not updated_ids:
        return PreventionCatalogStatusUpdateResult(
            catalog_path=path,
            target_status=clean_to_status,
            requested_ids=requested_ids,
            updated_ids=(),
            skipped_ids=tuple(skipped_ids),
        )

    tmp_path = path.with_name(f".{path.name}.tmp")
    try:
        tmp_path.write_text(
            json.dumps(payload, ensure_ascii=False, indent=2, sort_keys=False) + "\n",
            encoding="utf-8",
        )
        # Validate the would-be final catalog before replacing the original file.
        load_prevention_catalog(tmp_path)
        tmp_path.replace(path)
    finally:
        if tmp_path.exists():
            tmp_path.unlink()
    return PreventionCatalogStatusUpdateResult(
        catalog_path=path,
        target_status=clean_to_status,
        requested_ids=requested_ids,
        updated_ids=tuple(updated_ids),
        skipped_ids=tuple(skipped_ids),
    )


def _normalize_filter_values(values: tuple[str, ...] | list[str] | None) -> tuple[str, ...]:
    return tuple(dict.fromkeys(_clean_token(value) for value in (values or ()) if _clean_token(value)))


def filter_prevention_catalog_items(
    items: list[PreventionCatalogItem],
    *,
    areas: tuple[str, ...] | list[str] | None = None,
    triggers: tuple[str, ...] | list[str] | None = None,
    statuses: tuple[str, ...] | list[str] | None = ACTIVE_PREVENTION_STATUSES,
) -> list[PreventionCatalogItem]:
    area_filter = set(_normalize_filter_values(areas))
    trigger_filter = set(_normalize_filter_values(triggers))
    status_filter = set(_normalize_filter_values(statuses)) if statuses is not None else set()
    unknown_statuses = status_filter - ALLOWED_PREVENTION_STATUSES
    if unknown_statuses:
        raise ValueError(f"unknown prevention catalog status: {', '.join(sorted(unknown_statuses))}")

    matched: list[PreventionCatalogItem] = []
    for item in items:
        if status_filter and item.status not in status_filter:
            continue
        if area_filter and not area_filter.intersection(item.areas):
            continue
        if trigger_filter and not trigger_filter.intersection(item.triggers):
            continue
        matched.append(item)
    return matched


def _counts_by_severity(items: list[PreventionCatalogItem]) -> dict[str, int]:
    counts = {"critical": 0, "warning": 0, "info": 0}
    for item in items:
        counts[item.severity] = counts.get(item.severity, 0) + 1
    return counts


def _status_from_counts(counts: dict[str, int]) -> str:
    if counts.get("critical", 0) or counts.get("warning", 0):
        return "review_required"
    return "ok"


def _excel_value(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, (dict, list, tuple)):
        return _json_dumps(value)
    if isinstance(value, Path):
        return str(value)
    return value


def _write_rows_sheet(workbook: Workbook, *, title: str, headers: list[str], rows: list[dict[str, Any]]) -> None:
    worksheet = workbook.create_sheet(title)
    worksheet.append(headers)
    fill = PatternFill("solid", fgColor="D9EAF7")
    for cell in worksheet[1]:
        cell.font = Font(bold=True)
        cell.fill = fill
    for row in rows:
        worksheet.append([_excel_value(row.get(header, "")) for header in headers])
    for column_cells in worksheet.columns:
        max_length = max(len(str(cell.value or "")) for cell in column_cells)
        worksheet.column_dimensions[column_cells[0].column_letter].width = min(max(max_length + 2, 10), 80)


def _write_json_report(result: PreventionCatalogReviewResult) -> None:
    payload = {
        "review_id": result.review_id,
        "generated_at": result.generated_at,
        "status": result.status,
        "summary": result.summary,
        "items": [item.to_dict() for item in result.matched_items],
        "json_path": result.json_path,
        "excel_path": result.excel_path,
    }
    result.json_path.parent.mkdir(parents=True, exist_ok=True)
    result.json_path.write_text(json.dumps(_jsonify(payload), ensure_ascii=False, indent=2), encoding="utf-8")


def _write_excel_report(result: PreventionCatalogReviewResult) -> None:
    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "Summary"
    summary_sheet.append(["key", "value"])
    fill = PatternFill("solid", fgColor="D9EAF7")
    for cell in summary_sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = fill
    for key, value in result.summary.items():
        summary_sheet.append([key, _excel_value(value)])
    summary_sheet.column_dimensions["A"].width = 34
    summary_sheet.column_dimensions["B"].width = 80

    _write_rows_sheet(
        workbook,
        title="Items",
        headers=[
            "id",
            "severity",
            "status",
            "title",
            "areas",
            "triggers",
            "problem",
            "prevention",
            "review_points",
        ],
        rows=[item.to_dict() for item in sorted(result.matched_items, key=lambda row: (SEVERITY_ORDER.get(row.severity, 99), row.item_id))],
    )
    result.excel_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(result.excel_path)


def review_prevention_catalog(options: PreventionCatalogReviewOptions) -> PreventionCatalogReviewResult:
    generated_at = _now()
    review_id = f"prevention_catalog_review_{_timestamp_for_filename(generated_at)}"
    output_dir = Path(options.output_dir)
    json_path = output_dir / f"{review_id}.json"
    excel_path = output_dir / f"{review_id}.xlsx"
    items = load_prevention_catalog(options.catalog_path)
    matched_items = filter_prevention_catalog_items(
        items,
        areas=options.areas,
        triggers=options.triggers,
        statuses=options.statuses,
    )
    counts = _counts_by_severity(matched_items)
    status = _status_from_counts(counts)
    summary = {
        "review_id": review_id,
        "generated_at": generated_at,
        "status": status,
        "catalog_path": Path(options.catalog_path),
        "areas": list(options.areas),
        "triggers": list(options.triggers),
        "statuses": list(options.statuses),
        "total_count": len(items),
        "matched_count": len(matched_items),
        "critical": counts.get("critical", 0),
        "warning": counts.get("warning", 0),
        "info": counts.get("info", 0),
    }
    result = PreventionCatalogReviewResult(
        review_id=review_id,
        generated_at=generated_at,
        status=status,
        json_path=json_path,
        excel_path=excel_path,
        summary=summary,
        matched_items=matched_items,
        counts_by_severity=counts,
    )
    _write_json_report(result)
    _write_excel_report(result)
    return result
