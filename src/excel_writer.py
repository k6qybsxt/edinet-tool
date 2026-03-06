import os
import re
import glob
import shutil
import openpyxl

from openpyxl.cell.cell import Cell


def safe_filename(s: str) -> str:
    """
    Windowsでファイル名に使えない文字を置換して安全にする
    """
    if s is None:
        return ""
    s = str(s).strip()
    s = re.sub(r'[\\/:*?"<>|]', '_', s)
    s = s.rstrip(". ").strip()
    return s


def rename_excel_file(original_path, security_code, company_name, period_end_date, logger):
    """
    original_path を「<code>_<name>_<date>.xlsx」にリネームし、パスを返す。
    同名があれば _1, _2... を付ける。
    """
    dir_path = os.path.dirname(original_path)

    code = safe_filename(security_code)
    name = safe_filename(company_name)
    date = safe_filename(period_end_date)

    base_name = f"{code}_{name}_{date}".strip("_")
    if not base_name:
        raise ValueError("リネーム用の情報が不足しています（code/name/date が空です）。")

    new_file_path = os.path.join(dir_path, f"{base_name}.xlsx")

    counter = 1
    while os.path.exists(new_file_path):
        new_file_path = os.path.join(dir_path, f"{base_name}_{counter}.xlsx")
        counter += 1

    try:
        os.rename(original_path, new_file_path)
    except PermissionError as e:
        raise PermissionError(
            f"リネームできません（ファイルが開かれている可能性があります）。\n"
            f"対象: {original_path}\n"
            f"対策: Excelで該当ファイルを閉じてから再実行してください。"
        ) from e

    logger.info(f"Excelファイルがリネームされました: {new_file_path}")
    return new_file_path


def find_available_excel_file(base_path, file_name, logger, max_copies=30):
    """
    1) コピー系を探し、あれば更新日時が最新を返す
    2) コピーが無ければオリジナルを探す
    3) オリジナルが見つかったら新しいコピーを作って返す
    4) それも無ければワイルドカードで最後の救済
    """
    exts = ["xlsx", "xlsm"]

    def newest(paths):
        paths = [p for p in paths if os.path.exists(p)]
        if not paths:
            return None
        paths.sort(key=lambda p: os.path.getmtime(p), reverse=True)
        return paths[0]

    def make_next_copy(original_path):
        root, ext = os.path.splitext(original_path)
        base = root

        candidate0 = f"{base} - コピー{ext}"
        if not os.path.exists(candidate0):
            shutil.copy2(original_path, candidate0)
            return candidate0

        for i in range(2, max_copies + 2):
            cand = f"{base} - コピー ({i}){ext}"
            if not os.path.exists(cand):
                shutil.copy2(original_path, cand)
                return cand

        raise RuntimeError("コピー上限に達しました（max_copies を増やしてください）")

    variants = {file_name}
    variants.add(file_name.replace("決算分析シート-", "決算分析シート_"))
    variants.add(file_name.replace("決算分析シート_", "決算分析シート-"))

    copy_hits = []
    for v in variants:
        for ext in exts:
            copy_hits.append(os.path.join(base_path, f"{v} - コピー.{ext}"))
            for i in range(2, max_copies + 2):
                copy_hits.append(os.path.join(base_path, f"{v} - コピー ({i}).{ext}"))

    hit = newest(copy_hits)
    if hit:
        return hit

    originals = []
    for v in variants:
        for ext in exts:
            originals.append(os.path.join(base_path, f"{v}.{ext}"))

    orig = newest(originals)
    if orig:
        try:
            new_copy = make_next_copy(orig)
            logger.info(f"テンプレから作業用コピーを作成しました: {new_copy}")
            return new_copy
        except Exception:
            logger.exception("作業用コピー作成に失敗しました")
            raise

    patterns = []
    for v in variants:
        patterns += [
            os.path.join(base_path, f"{v}*コピー*.xlsx"),
            os.path.join(base_path, f"{v}*コピー*.xlsm"),
        ]

    hits = []
    for pat in patterns:
        hits.extend(glob.glob(pat))

    hit2 = newest(hits)
    if hit2:
        return hit2

    logger.warning(f"{file_name} のいずれのバージョンも見つかりませんでした。")
    return None


_SUFFIXES = [
    "YTD",
    "Quarter",
    "Current",
    "Prior1",
    "Prior2",
    "Prior3",
    "Prior4",
]

_suffix_pat = re.compile(rf"^(.*)({'|'.join(_SUFFIXES)})$")


def to_namedrange_key(key: str) -> str:
    """
    例:
      NetSalesYTD     -> NetSales_YTD
      TotalAssetsPrior1 -> TotalAssets_Prior1
      SecurityCodeDEI -> そのまま
    """
    if not isinstance(key, str) or not key:
        return key

    if "_" in key:
        return key
    if key.endswith("DEI"):
        return key

    m = _suffix_pat.match(key)
    if not m:
        return key

    metric = m.group(1)
    suffix = m.group(2)
    return f"{metric}_{suffix}"


def transform_keys_for_namedranges(data: dict) -> dict:
    out = {}
    for k, v in data.items():
        nk = to_namedrange_key(k)
        out[nk] = v
    return out


def _iter_namedrange_cells(workbook: openpyxl.Workbook, range_name: str):
    """
    NamedRangeが指すセル（1セル/複数セル）を返す
    """
    dn = workbook.defined_names.get(range_name)
    if dn is None:
        return

    for sheet_name, ref in dn.destinations:
        if sheet_name not in workbook.sheetnames:
            continue

        ws = workbook[sheet_name]
        obj = ws[ref]

        if isinstance(obj, Cell):
            yield obj
            continue

        for row in obj:
            for cell in row:
                yield cell


def write_data_to_excel_namedranges(
    excel_file: str,
    data: dict,
    *,
    transform_keys: bool = True,
    skip_if_formula: bool = True,
    skip_values=("データなし", "", None),
    dry_run: bool = False,
):
    """
    data の key と同名の NamedRange に書き込む
    """
    wb = openpyxl.load_workbook(excel_file, keep_vba=excel_file.lower().endswith(".xlsm"))
    result = {"written": [], "skipped": [], "missing": []}

    payload = transform_keys_for_namedranges(data) if transform_keys else dict(data)

    for name, value in payload.items():
        if value in skip_values or (isinstance(value, str) and value.strip() in skip_values):
            result["skipped"].append((name, "empty"))
            continue

        cells = list(_iter_namedrange_cells(wb, name))
        if not cells:
            result["missing"].append(name)
            continue

        for cell in cells:
            if skip_if_formula and isinstance(cell.value, str) and cell.value.startswith("="):
                result["skipped"].append((name, f"formula@{cell.coordinate}"))
                continue
            cell.value = value
            result["written"].append((name, f"{cell.parent.title}!{cell.coordinate}"))

    if not dry_run:
        wb.save(excel_file)

    return result


def write_rows_to_raw_sheet(excel_file: str, rows: list[dict], raw_cols: list[str], *, sheet_name: str = "raw_edinet"):
    """
    raw_edinet シートに rows を書き込む
    """
    import datetime as _dt

    def _to_excel_date(v):
        if isinstance(v, _dt.datetime):
            return v.date()
        if isinstance(v, _dt.date):
            return v
        if isinstance(v, str):
            s = v.strip()
            try:
                return _dt.datetime.strptime(s, "%Y-%m-%d").date()
            except Exception:
                return v
        return v

    wb = openpyxl.load_workbook(
        excel_file,
        keep_vba=excel_file.lower().endswith(".xlsm")
    )
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"sheet not found: {sheet_name}")

    ws = wb[sheet_name]

    if ws.max_row >= 2:
        ws.delete_rows(2, ws.max_row - 1)

    r = 2
    for row in rows:
        for c, col in enumerate(raw_cols, start=1):
            v = row.get(col)
            if col in ("period_start", "period_end"):
                v = _to_excel_date(v)
            ws.cell(row=r, column=c).value = v
        r += 1

    wb.save(excel_file)