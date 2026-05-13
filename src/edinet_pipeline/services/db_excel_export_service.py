from __future__ import annotations

import shutil
import sqlite3
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.cell.cell import Cell
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.utils import absolute_coordinate

from edinet_pipeline.services.excel_service import safe_filename


INPUT_SHEET_NAME = "\u6c7a\u7b97\u5165\u529b"
PROJECT_ROOT = Path(__file__).resolve().parents[3]
DEFAULT_DB_PATH = Path(r"E:\EDINET_Data\edinet_monitor\db\edinet_monitor.db")
DEFAULT_CONDITION_XLSX_PATH = PROJECT_ROOT / "config" / "excel" / "DB抽出条件.xlsx"
DEFAULT_CONDITION_SHEET_NAME = "\u6761\u4ef6"
DB_EXPORT_UNUSED_SHEET_NAMES = ("raw_edinet", "raw異常検知メーター")

PERIOD_ORDER = ["Prior4", "Prior3", "Prior2", "Prior1", "Current"]
PERIOD_LABEL_TO_ROLE = {
    "4\u671f\u524d": "Prior4",
    "3\u671f\u524d": "Prior3",
    "2\u671f\u524d": "Prior2",
    "1\u671f\u524d": "Prior1",
    "\u6700\u65b0\u901a\u671f": "Current",
    "\u6700\u65b0": "Latest",
    # Backward compatible alias. The condition workbook should use "\u6700\u65b0" going forward.
    "\u5f53\u671f": "Latest",
}

ANNUAL_BASE_TO_TARGET = {
    "NetSales": "NetSales",
    "CostOfSales": "CostOfSales",
    "GrossProfit": "GrossProfit",
    "SellingExpenses": "SellingExpenses",
    "OperatingIncome": "OperatingIncome",
    "OrdinaryIncome": "OrdinaryIncome",
    "ProfitLoss": "ProfitLoss",
    "OutstandingShares": "TotalNumber",
    "TotalAssets": "TotalAssets",
    "NetAssets": "NetAssets",
    "OperatingCash": "OperatingCash",
    "InvestmentCash": "InvestmentCash",
    "FinancingCash": "FinancingCash",
    "CashAndCashEquivalents": "CashAndCashEquivalents",
    "StockPrice": "StockPrice",
}

METRIC_SOURCE_KEYS = {
    "OutstandingShares": ["OutstandingSharesCurrent", "TotalNumberCurrent"],
    **{
        base: [f"{base}Current"]
        for base in ANNUAL_BASE_TO_TARGET
        if base not in {"OutstandingShares", "StockPrice"}
    },
}

MONETARY_BASES = {
    "NetSales",
    "CostOfSales",
    "GrossProfit",
    "SellingExpenses",
    "OperatingIncome",
    "OrdinaryIncome",
    "ProfitLoss",
    "TotalAssets",
    "NetAssets",
    "OperatingCash",
    "InvestmentCash",
    "FinancingCash",
    "CashAndCashEquivalents",
}
SHARE_BASES = {"OutstandingShares"}

QUARTER_CELLS = {
    ("1Q", "NetSales"): ("NetSales_Q1", "G36"),
    ("1Q", "OrdinaryIncome"): ("OrdinaryIncome_Q1", "G37"),
    ("1Q", "ProfitLoss"): ("ProfitLoss_Q1", "G38"),
    ("1Q", "OutstandingShares"): ("TotalNumber_Q1", "G40"),
    ("1Q", "TotalAssets"): ("TotalAssets_Q1", "F44"),
    ("1Q", "NetAssets"): ("NetAssets_Q1", "G44"),
    ("1Q", "OperatingCash"): ("OperatingCash_Q1", "F48"),
    ("1Q", "InvestmentCash"): ("InvestmentCash_Q1", "G48"),
    ("1Q", "FinancingCash"): ("FinancingCash_Q1", "H48"),
    ("1Q", "CashAndCashEquivalents"): ("CashAndCashEquivalents_Q1", "G49"),
    ("1Q", "StockPrice"): ("StockPrice_Q1", "G53"),
    ("2Q", "NetSales"): ("NetSales_YTD", "J36"),
    ("2Q", "OrdinaryIncome"): ("OrdinaryIncome_YTD", "J37"),
    ("2Q", "ProfitLoss"): ("ProfitLoss_YTD", "J38"),
    ("2Q", "OutstandingShares"): ("TotalNumber_YTD", "J40"),
    ("2Q", "TotalAssets"): ("TotalAssets_YTD", "I44"),
    ("2Q", "NetAssets"): ("NetAssets_YTD", "J44"),
    ("2Q", "OperatingCash"): ("OperatingCash_YTD", "I48"),
    ("2Q", "InvestmentCash"): ("InvestmentCash_YTD", "J48"),
    ("2Q", "FinancingCash"): ("FinancingCash_YTD", "K48"),
    ("2Q", "CashAndCashEquivalents"): ("CashAndCashEquivalents_YTD", "J49"),
    ("2Q", "StockPrice"): ("StockPrice_Q2", "J53"),
    ("3Q", "NetSales"): ("NetSales_Q3", "M36"),
    ("3Q", "OrdinaryIncome"): ("OrdinaryIncome_Q3", "M37"),
    ("3Q", "ProfitLoss"): ("ProfitLoss_Q3", "M38"),
    ("3Q", "OutstandingShares"): ("TotalNumber_Q3", "M40"),
    ("3Q", "TotalAssets"): ("TotalAssets_Q3", "L44"),
    ("3Q", "NetAssets"): ("NetAssets_Q3", "M44"),
    ("3Q", "OperatingCash"): ("OperatingCash_Q3", "L48"),
    ("3Q", "InvestmentCash"): ("InvestmentCash_Q3", "M48"),
    ("3Q", "FinancingCash"): ("FinancingCash_Q3", "N48"),
    ("3Q", "CashAndCashEquivalents"): ("CashAndCashEquivalents_Q3", "M49"),
    ("3Q", "StockPrice"): ("StockPrice_Q3", "M53"),
}

FORECAST_CELLS = {
    ("initial", "NetSales"): ("NetSales_ForecastInitial", "D30"),
    ("initial", "OrdinaryIncome"): ("OrdinaryIncome_ForecastInitial", "D31"),
    ("initial", "ProfitLoss"): ("ProfitLoss_ForecastInitial", "D32"),
    ("1Q", "NetSales"): ("NetSales_ForecastQ1", "G30"),
    ("1Q", "OrdinaryIncome"): ("OrdinaryIncome_ForecastQ1", "G31"),
    ("1Q", "ProfitLoss"): ("ProfitLoss_ForecastQ1", "G32"),
    ("2Q", "NetSales"): ("NetSales_ForecastQ2", "J30"),
    ("2Q", "OrdinaryIncome"): ("OrdinaryIncome_ForecastQ2", "J31"),
    ("2Q", "ProfitLoss"): ("ProfitLoss_ForecastQ2", "J32"),
    ("3Q", "NetSales"): ("NetSales_ForecastQ3", "M30"),
    ("3Q", "OrdinaryIncome"): ("OrdinaryIncome_ForecastQ3", "M31"),
    ("3Q", "ProfitLoss"): ("ProfitLoss_ForecastQ3", "M32"),
}


@dataclass(frozen=True)
class AnalysisWorkbookCondition:
    security_codes: list[str]
    period_start: str = "Prior4"
    period_end: str = "Latest"


@dataclass(frozen=True)
class AnalysisWorkbookExportResult:
    output_paths: list[Path]
    target_companies: int
    errors: list[str]
    warnings: list[str]


@dataclass(frozen=True)
class LatestActualPeriod:
    kind: str
    period_end: str
    fiscal_year: int | None
    quarter_type: str | None = None


def normalize_security_code(value: Any) -> str:
    text = str(value or "").strip()
    if not text or text.lower() == "all":
        return ""
    if text.endswith(".0"):
        text = text[:-2]
    if text.endswith("0") and len(text) == 5:
        text = text[:-1]
    return text


def _split_condition_list(value: Any) -> list[str]:
    text = str(value or "").strip()
    if not text or text.lower() == "all":
        return []
    return [part.strip() for part in text.replace("\n", ",").split(",") if part.strip()]


def _normalize_period_label(value: Any, *, default: str) -> str:
    text = str(value or "").strip()
    if not text or text.lower() == "all":
        return default
    normalized = text.replace(" ", "")
    if normalized in PERIOD_LABEL_TO_ROLE:
        return PERIOD_LABEL_TO_ROLE[normalized]
    if normalized in PERIOD_ORDER:
        return normalized
    if normalized == "Latest":
        return "Latest"
    raise ValueError(f"unsupported period label: {text}")


def _selected_period_roles(start_role: str, end_role: str, latest_actual: LatestActualPeriod) -> list[str]:
    annual_end_role = "Prior1" if end_role == "Latest" and latest_actual.kind != "annual" else end_role
    if annual_end_role == "Latest":
        annual_end_role = "Current"
    start_index = PERIOD_ORDER.index(start_role)
    end_index = PERIOD_ORDER.index(annual_end_role)
    if start_index > end_index:
        raise ValueError(f"period start must be older than or equal to end: {start_role} > {annual_end_role}")
    return PERIOD_ORDER[start_index : end_index + 1]


def _include_current_quarters(end_role: str, latest_actual: LatestActualPeriod) -> bool:
    return end_role == "Latest" and latest_actual.kind in {"1Q", "2Q", "3Q"}


def read_analysis_workbook_condition(path: str | Path) -> AnalysisWorkbookCondition:
    workbook = openpyxl.load_workbook(path, data_only=True)
    try:
        if DEFAULT_CONDITION_SHEET_NAME not in workbook.sheetnames:
            raise ValueError("condition sheet not found: \u6761\u4ef6")
        ws = workbook[DEFAULT_CONDITION_SHEET_NAME]
        values: dict[str, Any] = {}
        for row in ws.iter_rows(min_row=1, max_col=2):
            key = row[0].value
            if key in (None, ""):
                continue
            values[str(key).strip()] = row[1].value
        security_codes = [
            code
            for code in (normalize_security_code(value) for value in _split_condition_list(values.get("\u8a3c\u5238\u30b3\u30fc\u30c9")))
            if code
        ]
        return AnalysisWorkbookCondition(
            security_codes=security_codes,
            period_start=_normalize_period_label(values.get("\u671f\u9593 Start"), default="Prior4"),
            period_end=_normalize_period_label(values.get("\u671f\u9593 End"), default="Latest"),
        )
    finally:
        workbook.close()


def _table_exists(conn: sqlite3.Connection, table_name: str) -> bool:
    row = conn.execute(
        "SELECT 1 FROM sqlite_master WHERE type = 'table' AND name = ?",
        (table_name,),
    ).fetchone()
    return row is not None


def _table_columns(conn: sqlite3.Connection, table_name: str) -> set[str]:
    if not _table_exists(conn, table_name):
        return set()
    return {str(row[1]) for row in conn.execute(f"PRAGMA table_info({table_name})").fetchall()}


def _to_float(value: Any) -> float | None:
    if value in (None, ""):
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _format_unit(value: Any) -> str:
    text = str(value or "").strip()
    return "\u5343\u5186" if text == "\u5343\u5186" else "\u767e\u4e07\u5186"


def _scale_value(metric_base: str, value: float | None, display_unit: str) -> float | int | None:
    if value is None:
        return None
    if metric_base in MONETARY_BASES:
        divisor = 1_000 if display_unit == "\u5343\u5186" else 1_000_000
        return int(round(value / divisor))
    if metric_base in SHARE_BASES:
        return int(round(value / 1_000))
    return value


def _fetch_company(conn: sqlite3.Connection, security_code: str) -> sqlite3.Row | None:
    code = normalize_security_code(security_code)
    return conn.execute(
        """
        SELECT edinet_code, security_code, company_name
        FROM issuer_master
        WHERE substr(coalesce(security_code, ''), 1, 4) = ?
           OR coalesce(security_code, '') = ?
        ORDER BY is_listed DESC, edinet_code
        LIMIT 1
        """,
        (code, code),
    ).fetchone()


def _fetch_annual_filings(conn: sqlite3.Connection, security_code: str) -> list[sqlite3.Row]:
    code = normalize_security_code(security_code)
    return conn.execute(
        """
        SELECT
          f.doc_id,
          f.edinet_code,
          COALESCE(f.security_code, im.security_code) AS security_code,
          COALESCE(im.company_name, '') AS company_name,
          f.period_end,
          f.submit_date,
          f.document_display_unit
        FROM filings f
        LEFT JOIN issuer_master im
          ON im.edinet_code = f.edinet_code
        WHERE f.form_type = '030000'
          AND (substr(coalesce(f.security_code, im.security_code, ''), 1, 4) = ?
               OR coalesce(f.security_code, im.security_code, '') = ?)
        ORDER BY f.period_end DESC, coalesce(f.submit_date, '') DESC, f.doc_id DESC
        LIMIT 5
        """,
        (code, code),
    ).fetchall()


def _detect_latest_actual_period(
    conn: sqlite3.Connection,
    security_code: str,
    annual_rows: list[sqlite3.Row],
) -> LatestActualPeriod:
    latest_annual = annual_rows[0] if annual_rows else None
    candidates: list[LatestActualPeriod] = []
    if latest_annual is not None and latest_annual["period_end"]:
        period_end = str(latest_annual["period_end"])
        fiscal_year = int(period_end[:4]) if len(period_end) >= 4 and period_end[:4].isdigit() else None
        candidates.append(LatestActualPeriod(kind="annual", period_end=period_end, fiscal_year=fiscal_year))

    if _table_exists(conn, "filings"):
        half = (
            _fetch_half_filing_between(
                conn,
                security_code,
                start_after=str(latest_annual["period_end"] or ""),
                end_on_or_before="9999-12-31",
            )
            if latest_annual is not None
            else None
        )
        if half is not None and half["period_end"]:
            period_end = str(half["period_end"])
            base_fiscal_year = candidates[0].fiscal_year if candidates else None
            candidates.append(
                LatestActualPeriod(
                    kind="2Q",
                    period_end=period_end,
                    fiscal_year=(base_fiscal_year + 1) if base_fiscal_year is not None else None,
                    quarter_type="2Q",
                )
            )

    if _table_exists(conn, "jquants_financial_metrics"):
        code = normalize_security_code(security_code)
        columns = _table_columns(conn, "jquants_financial_metrics")
        quarter_type_expr = "quarter_type" if "quarter_type" in columns else "NULL AS quarter_type"
        rows = conn.execute(
            f"""
            SELECT period_key, {quarter_type_expr}, fiscal_year, period_end
            FROM jquants_financial_metrics
            WHERE metric_kind = 'actual'
              AND period_key IN ('actual:1Q', 'actual:3Q')
              AND (security_code = ? OR local_code = ?)
              AND period_end IS NOT NULL
              AND period_end <> ''
            ORDER BY period_end DESC, disclosed_date DESC, disclosed_time DESC
            LIMIT 1
            """,
            (code, code),
        ).fetchall()
        for row in rows:
            quarter = str(row["quarter_type"] or "")
            if quarter not in {"1Q", "3Q"}:
                quarter = "1Q" if str(row["period_key"] or "") == "actual:1Q" else "3Q"
            candidates.append(
                LatestActualPeriod(
                    kind=quarter,
                    period_end=str(row["period_end"]),
                    fiscal_year=int(row["fiscal_year"]) if row["fiscal_year"] is not None else None,
                    quarter_type=quarter,
                )
            )

    if not candidates:
        raise ValueError(f"latest actual period not found: {security_code}")
    return max(candidates, key=lambda item: item.period_end)


def _fetch_half_filing_for_current_year(
    conn: sqlite3.Connection,
    security_code: str,
    current_annual: sqlite3.Row,
    prior_annual: sqlite3.Row | None,
) -> sqlite3.Row | None:
    code = normalize_security_code(security_code)
    where = [
        "f.form_type IN ('043A00', '043000')",
        "(substr(coalesce(f.security_code, im.security_code, ''), 1, 4) = ? OR coalesce(f.security_code, im.security_code, '') = ?)",
        "f.period_end <= ?",
    ]
    params: list[Any] = [code, code, current_annual["period_end"]]
    if prior_annual is not None and prior_annual["period_end"]:
        where.append("f.period_end > ?")
        params.append(prior_annual["period_end"])
    return conn.execute(
        f"""
        SELECT
          f.doc_id,
          f.edinet_code,
          COALESCE(f.security_code, im.security_code) AS security_code,
          COALESCE(im.company_name, '') AS company_name,
          f.period_end,
          f.submit_date,
          f.document_display_unit
        FROM filings f
        LEFT JOIN issuer_master im
          ON im.edinet_code = f.edinet_code
        WHERE {" AND ".join(where)}
        ORDER BY f.period_end DESC, coalesce(f.submit_date, '') DESC, f.doc_id DESC
        LIMIT 1
        """,
        params,
    ).fetchone()


def _fetch_half_filing_between(
    conn: sqlite3.Connection,
    security_code: str,
    *,
    start_after: str | None,
    end_on_or_before: str,
) -> sqlite3.Row | None:
    code = normalize_security_code(security_code)
    where = [
        "f.form_type IN ('043A00', '043000')",
        "(substr(coalesce(f.security_code, im.security_code, ''), 1, 4) = ? OR coalesce(f.security_code, im.security_code, '') = ?)",
        "f.period_end <= ?",
    ]
    params: list[Any] = [code, code, end_on_or_before]
    if start_after:
        where.append("f.period_end > ?")
        params.append(start_after)
    return conn.execute(
        f"""
        SELECT
          f.doc_id,
          f.edinet_code,
          COALESCE(f.security_code, im.security_code) AS security_code,
          COALESCE(im.company_name, '') AS company_name,
          f.period_end,
          f.submit_date,
          f.document_display_unit
        FROM filings f
        LEFT JOIN issuer_master im
          ON im.edinet_code = f.edinet_code
        WHERE {" AND ".join(where)}
        ORDER BY f.period_end DESC, coalesce(f.submit_date, '') DESC, f.doc_id DESC
        LIMIT 1
        """,
        params,
    ).fetchone()


def _fetch_metric_values_for_doc(conn: sqlite3.Connection, doc_id: str) -> dict[str, float | None]:
    keys = sorted({key for keys in METRIC_SOURCE_KEYS.values() for key in keys})
    if not keys:
        return {}
    placeholders = ",".join("?" for _ in keys)
    values_by_key: dict[str, float | None] = {}
    if _table_exists(conn, "normalized_metrics"):
        rows = conn.execute(
            f"""
            SELECT metric_key, value_num
            FROM normalized_metrics
            WHERE doc_id = ?
              AND metric_key IN ({placeholders})
            """,
            [doc_id, *keys],
        ).fetchall()
        for row in rows:
            values_by_key[str(row["metric_key"])] = _to_float(row["value_num"])
    if _table_exists(conn, "derived_metrics"):
        rows = conn.execute(
            f"""
            SELECT metric_key, value_num, calc_status
            FROM derived_metrics
            WHERE doc_id = ?
              AND metric_key IN ({placeholders})
            """,
            [doc_id, *keys],
        ).fetchall()
        for row in rows:
            key = str(row["metric_key"])
            if key in values_by_key and values_by_key[key] is not None:
                continue
            values_by_key[key] = None if str(row["calc_status"] or "") == "missing_input" else _to_float(row["value_num"])
    result: dict[str, float | None] = {}
    for base, source_keys in METRIC_SOURCE_KEYS.items():
        for key in source_keys:
            if key in values_by_key:
                result[base] = values_by_key[key]
                break
    return result


def _fetch_market_metric_for_source(conn: sqlite3.Connection, source_type: str, source_id: str, metric_base: str) -> float | None:
    if not _table_exists(conn, "market_derived_metrics"):
        return None
    row = conn.execute(
        """
        SELECT value_num, calc_status
        FROM market_derived_metrics
        WHERE source_type = ?
          AND source_id = ?
          AND metric_base = ?
        ORDER BY updated_at DESC, id DESC
        LIMIT 1
        """,
        (source_type, source_id, metric_base),
    ).fetchone()
    if row is None or str(row["calc_status"] or "") == "missing_input":
        return None
    return _to_float(row["value_num"])


def _fetch_quote_on_or_before(conn: sqlite3.Connection, security_code: str, period_end: str) -> float | None:
    if not _table_exists(conn, "jquants_daily_quotes"):
        return None
    code = normalize_security_code(security_code)
    row = conn.execute(
        """
        SELECT adjustment_close_rounded
        FROM jquants_daily_quotes
        WHERE (security_code = ? OR local_code = ?)
          AND trade_date <= ?
        ORDER BY trade_date DESC
        LIMIT 1
        """,
        (code, code, period_end),
    ).fetchone()
    return _to_float(row["adjustment_close_rounded"]) if row is not None else None


def _fetch_stock_price(conn: sqlite3.Connection, source_type: str, source_id: str, security_code: str, period_end: str) -> float | None:
    value = _fetch_market_metric_for_source(conn, source_type, source_id, "StockPrice")
    if value is not None:
        return value
    return _fetch_quote_on_or_before(conn, security_code, period_end)


def _fetch_jquants_actual_values(
    conn: sqlite3.Connection,
    security_code: str,
    period_key: str,
    *,
    start_after: str | None,
    end_on_or_before: str,
) -> tuple[str | None, str | None, dict[str, float | None]]:
    if not _table_exists(conn, "jquants_financial_metrics"):
        return None, None, {}
    code = normalize_security_code(security_code)
    where = [
        "metric_kind = 'actual'",
        "period_key = ?",
        "(security_code = ? OR local_code = ?)",
        "period_end <= ?",
    ]
    params: list[Any] = [period_key, code, code, end_on_or_before]
    if start_after:
        where.append("period_end > ?")
        params.append(start_after)
    rows = conn.execute(
        f"""
        SELECT disclosure_number, period_end, metric_base, value_num, calc_status
        FROM jquants_financial_metrics
        WHERE {" AND ".join(where)}
        ORDER BY period_end DESC, disclosed_date DESC, disclosed_time DESC, disclosure_number DESC
        """,
        params,
    ).fetchall()
    if not rows:
        return None, None, {}
    disclosure_number = str(rows[0]["disclosure_number"])
    period_end = str(rows[0]["period_end"] or "")
    values: dict[str, float | None] = {}
    for row in rows:
        if str(row["disclosure_number"]) != disclosure_number:
            continue
        values[str(row["metric_base"])] = None if str(row["calc_status"] or "") == "missing_input" else _to_float(row["value_num"])
    return disclosure_number, period_end, values


def _fetch_forecast_values(conn: sqlite3.Connection, security_code: str, fiscal_year: int, forecast_stage: str) -> dict[str, float | None]:
    if not _table_exists(conn, "jquants_financial_metrics"):
        return {}
    code = normalize_security_code(security_code)
    rows = conn.execute(
        """
        SELECT metric_base, value_num, calc_status
        FROM jquants_financial_metrics
        WHERE metric_kind = 'forecast'
          AND forecast_stage = ?
          AND fiscal_year = ?
          AND (security_code = ? OR local_code = ?)
          AND metric_base IN ('NetSales', 'OrdinaryIncome', 'ProfitLoss')
        ORDER BY disclosed_date DESC, disclosed_time DESC, disclosure_number DESC
        """,
        (forecast_stage, fiscal_year, code, code),
    ).fetchall()
    result: dict[str, float | None] = {}
    for row in rows:
        base = str(row["metric_base"])
        if base in result:
            continue
        result[base] = None if str(row["calc_status"] or "") == "missing_input" else _to_float(row["value_num"])
    return result


def _absolute_ref(sheet_name: str, cell_ref: str) -> str:
    return f"'{sheet_name}'!{absolute_coordinate(cell_ref)}"


def _ensure_defined_name(workbook: openpyxl.Workbook, name: str, cell_ref: str) -> None:
    if workbook.defined_names.get(name) is not None:
        return
    workbook.defined_names.add(
        DefinedName(name=name, attr_text=_absolute_ref(INPUT_SHEET_NAME, cell_ref))
    )


def _iter_defined_name_cells(workbook: openpyxl.Workbook, name: str) -> list[Cell]:
    defined_name = workbook.defined_names.get(name)
    if defined_name is None:
        return []
    cells: list[Cell] = []
    for sheet_name, ref in defined_name.destinations:
        if sheet_name not in workbook.sheetnames:
            continue
        obj = workbook[sheet_name][ref]
        if isinstance(obj, Cell):
            cells.append(obj)
        else:
            for row in obj:
                cells.extend(row)
    return cells


def _register_db_defined_names(workbook: openpyxl.Workbook) -> None:
    for name, cell_ref in [*QUARTER_CELLS.values(), *FORECAST_CELLS.values()]:
        _ensure_defined_name(workbook, name, cell_ref)


def _write_target(workbook: openpyxl.Workbook, target_name: str, cell_ref: str | None, value: Any) -> None:
    cells = _iter_defined_name_cells(workbook, target_name)
    if not cells and cell_ref:
        cells = [workbook[INPUT_SHEET_NAME][cell_ref]]
    for cell in cells:
        cell.value = value


def _all_data_targets() -> list[tuple[str, str | None]]:
    targets: list[tuple[str, str | None]] = []
    for base, target_base in ANNUAL_BASE_TO_TARGET.items():
        for role in PERIOD_ORDER:
            if base == "StockPrice":
                target_name = "StockPrice_Q4" if role == "Current" else f"StockPrice_{role}"
            else:
                target_name = f"{target_base}_{role}"
            targets.append((target_name, None))
    targets.extend((name, cell_ref) for name, cell_ref in QUARTER_CELLS.values())
    targets.extend((name, cell_ref) for name, cell_ref in FORECAST_CELLS.values())
    return targets


def _write_payload_to_workbook(
    workbook: openpyxl.Workbook,
    payload: dict[str, Any],
    direct_cells: dict[str, Any],
) -> None:
    if INPUT_SHEET_NAME not in workbook.sheetnames:
        raise ValueError(f"sheet not found: {INPUT_SHEET_NAME}")
    _register_db_defined_names(workbook)
    for target_name, cell_ref in _all_data_targets():
        _write_target(workbook, target_name, cell_ref, None)
    for target_name, value in payload.items():
        if value is None:
            continue
        _write_target(workbook, target_name, None, value)
    ws = workbook[INPUT_SHEET_NAME]
    for cell_ref, value in direct_cells.items():
        ws[cell_ref] = value


def _remove_db_export_unused_sheets(workbook: openpyxl.Workbook) -> None:
    for sheet_name in DB_EXPORT_UNUSED_SHEET_NAMES:
        if sheet_name in workbook.sheetnames and len(workbook.sheetnames) > 1:
            del workbook[sheet_name]


def _build_company_payload(
    conn: sqlite3.Connection,
    security_code: str,
    selected_roles: list[str],
    latest_actual: LatestActualPeriod,
    *,
    include_current_quarters: bool,
    shift_annual_periods: bool,
) -> tuple[dict[str, Any], dict[str, Any], list[str], str]:
    warnings: list[str] = []
    annual_rows = _fetch_annual_filings(conn, security_code)
    if not annual_rows:
        raise ValueError(f"annual filings not found: {security_code}")
    current = annual_rows[0]
    prior = annual_rows[1] if len(annual_rows) > 1 else None
    display_unit = _format_unit(current["document_display_unit"])
    period_end = str(current["period_end"] or "")
    annual_fiscal_year = int(period_end[:4]) if len(period_end) >= 4 and period_end[:4].isdigit() else 0
    fiscal_year = latest_actual.fiscal_year or annual_fiscal_year
    payload: dict[str, Any] = {}

    for index, row in enumerate(annual_rows[:5]):
        if not shift_annual_periods:
            role = "Current" if index == 0 else f"Prior{index}"
        else:
            role_index = index + 1
            if role_index > 4:
                continue
            role = f"Prior{role_index}"
        if role not in selected_roles:
            continue
        values = _fetch_metric_values_for_doc(conn, str(row["doc_id"]))
        values["StockPrice"] = _fetch_stock_price(conn, "edinet", str(row["doc_id"]), security_code, str(row["period_end"] or ""))
        for base, target_base in ANNUAL_BASE_TO_TARGET.items():
            value = _scale_value(base, values.get(base), display_unit)
            if base == "StockPrice":
                target_name = "StockPrice_Q4" if role == "Current" else f"StockPrice_{role}"
            else:
                target_name = f"{target_base}_{role}"
            payload[target_name] = value

    if include_current_quarters:
        half = _fetch_half_filing_between(
            conn,
            security_code,
            start_after=str(current["period_end"] or ""),
            end_on_or_before=latest_actual.period_end,
        )
        if half is not None:
            half_values = _fetch_metric_values_for_doc(conn, str(half["doc_id"]))
            half_values["StockPrice"] = _fetch_stock_price(conn, "edinet", str(half["doc_id"]), security_code, str(half["period_end"] or ""))
            for base, value in half_values.items():
                target = QUARTER_CELLS.get(("2Q", base))
                if target is not None:
                    payload[target[0]] = _scale_value(base, value, display_unit)

        for quarter, period_key in (("1Q", "actual:1Q"), ("3Q", "actual:3Q")):
            disclosure_number, quarter_period_end, quarter_values = _fetch_jquants_actual_values(
                conn,
                security_code,
                period_key,
                start_after=str(current["period_end"] or ""),
                end_on_or_before=latest_actual.period_end,
            )
            if disclosure_number and quarter_period_end:
                quarter_values["StockPrice"] = _fetch_stock_price(
                    conn,
                    "jquants",
                    disclosure_number,
                    security_code,
                    quarter_period_end,
                )
            for base, value in quarter_values.items():
                target = QUARTER_CELLS.get((quarter, base))
                if target is not None:
                    payload[target[0]] = _scale_value(base, value, display_unit)

        for forecast_stage in ("initial", "1Q", "2Q", "3Q"):
            for base, value in _fetch_forecast_values(conn, security_code, fiscal_year, forecast_stage).items():
                target = FORECAST_CELLS.get((forecast_stage, base))
                if target is not None:
                    payload[target[0]] = _scale_value(base, value, display_unit)

    company_name = str(current["company_name"] or "")
    code = normalize_security_code(current["security_code"] or security_code)
    year = str(fiscal_year) if fiscal_year else (period_end[:4] if len(period_end) >= 4 else "")
    month = period_end[5:7] if len(period_end) >= 7 else ""
    direct_cells = {
        "J2": display_unit,
        "K2": code,
        "L2": company_name,
        "N2": year,
        "O2": month,
        "P2": datetime.now().strftime("%Y-%m-%d"),
    }
    if len(annual_rows) < 5:
        warnings.append(f"{code}: annual filings less than 5 ({len(annual_rows)})")
    return payload, direct_cells, warnings, latest_actual.period_end


def _copy_template(template_path: Path, output_dir: Path, security_code: str, company_name: str, period_end: str) -> Path:
    output_dir.mkdir(parents=True, exist_ok=True)
    suffix = template_path.suffix or ".xlsm"
    base_name = "_".join(
        part
        for part in (
            safe_filename(security_code),
            safe_filename(company_name),
            safe_filename(period_end),
            "db",
        )
        if part
    )
    candidate = output_dir / f"{base_name}{suffix}"
    counter = 1
    while candidate.exists():
        candidate = output_dir / f"{base_name}_{counter}{suffix}"
        counter += 1
    shutil.copy2(template_path, candidate)
    return candidate


def export_analysis_workbooks_from_db(
    conn: sqlite3.Connection,
    *,
    condition_xlsx: str | Path | None = None,
    security_codes: list[str] | None = None,
    db_path: str | Path | None = None,
    template_path: str | Path,
    output_dir: str | Path,
) -> AnalysisWorkbookExportResult:
    _ = db_path
    if condition_xlsx:
        condition = read_analysis_workbook_condition(condition_xlsx)
        codes = condition.security_codes
        period_start = condition.period_start
        period_end = condition.period_end
    else:
        codes = [code for code in (normalize_security_code(value) for value in (security_codes or [])) if code]
        period_start = "Prior4"
        period_end = "Latest"
    if security_codes and not codes:
        codes = [code for code in (normalize_security_code(value) for value in security_codes) if code]
    if not codes:
        raise ValueError("security codes are required")

    template = Path(template_path)
    out_dir = Path(output_dir)
    output_paths: list[Path] = []
    errors: list[str] = []
    warnings: list[str] = []

    for code in codes:
        try:
            annual_rows = _fetch_annual_filings(conn, code)
            if not annual_rows:
                raise ValueError(f"annual filings not found: {code}")
            latest_actual = _detect_latest_actual_period(conn, code, annual_rows)
            selected_roles = _selected_period_roles(period_start, period_end, latest_actual)
            include_current_quarters = _include_current_quarters(period_end, latest_actual)
            shift_annual_periods = latest_actual.kind != "annual" and period_end != "Current"
            current = annual_rows[0]
            payload, direct_cells, company_warnings, output_period_end = _build_company_payload(
                conn,
                code,
                selected_roles,
                latest_actual,
                include_current_quarters=include_current_quarters,
                shift_annual_periods=shift_annual_periods,
            )
            warnings.extend(company_warnings)
            out_path = _copy_template(
                template,
                out_dir,
                normalize_security_code(current["security_code"] or code),
                str(current["company_name"] or ""),
                output_period_end,
            )
            workbook = openpyxl.load_workbook(
                out_path,
                keep_vba=out_path.suffix.lower() == ".xlsm",
            )
            try:
                _write_payload_to_workbook(workbook, payload, direct_cells)
                _remove_db_export_unused_sheets(workbook)
                workbook.save(out_path)
            finally:
                vba_archive = getattr(workbook, "vba_archive", None)
                if vba_archive is not None:
                    vba_archive.close()
                    workbook.vba_archive = None
                workbook.close()
            output_paths.append(out_path)
        except Exception as exc:
            errors.append(f"{code}: {exc}")

    return AnalysisWorkbookExportResult(
        output_paths=output_paths,
        target_companies=len(codes),
        errors=errors,
        warnings=warnings,
    )


def connect_db(db_path: str | Path | None = None) -> sqlite3.Connection:
    conn = sqlite3.connect(str(db_path or DEFAULT_DB_PATH))
    conn.row_factory = sqlite3.Row
    return conn
