import os
import requests
import certifi
import pandas as pd
import openpyxl

from io import StringIO
from calendar import monthrange
from datetime import datetime, timedelta

# requests session を1回だけ作って使い回す
_YF_SESSION = None

_STOCK_CACHE_DIR = os.path.join("data", "cache", "stock")
os.makedirs(_STOCK_CACHE_DIR, exist_ok=True)

_STOCK_PRICE_MAP_CACHE = {}


def _get_yf_session():
    global _YF_SESSION
    if _YF_SESSION is None:
        s = requests.Session()
        s.verify = certifi.where()
        _YF_SESSION = s
    return _YF_SESSION


def _to_stooq_symbol(stock_code: str) -> str:
    """
    '2206.T' / '2206' どちらが来ても stooq用に変換する
    stooqの日本株は '2206.JP' 形式
    """
    code = stock_code.strip().upper()
    if code.endswith(".T"):
        code = code[:-2]
    return f"{code}.JP"

def _shift_years(dt, years):
    y = dt.year + years
    d = min(dt.day, monthrange(y, dt.month)[1])
    return dt.replace(year=y, day=d)


def _build_quarter_end_dates(fiscal_year_end_date):
    prev_fy_end = _shift_years(fiscal_year_end_date, -1)
    fiscal_year_start = prev_fy_end + timedelta(days=1)

    q1 = (pd.Timestamp(fiscal_year_start) + pd.DateOffset(months=3) - pd.Timedelta(days=1)).date()
    q2 = (pd.Timestamp(fiscal_year_start) + pd.DateOffset(months=6) - pd.Timedelta(days=1)).date()
    q3 = (pd.Timestamp(fiscal_year_start) + pd.DateOffset(months=9) - pd.Timedelta(days=1)).date()
    q4 = fiscal_year_end_date

    return [q1, q2, q3, q4]


def build_stock_date_pairs_from_fiscal_year_end(fiscal_year_end: str):
    fy_end = str(fiscal_year_end or "").strip().replace("/", "-")
    if not fy_end:
        return []

    fiscal_year_end_date = datetime.strptime(fy_end, "%Y-%m-%d").date()

    out = []

    for years_back, name in [(4, "StockPrice_Prior4"), (3, "StockPrice_Prior3"), (2, "StockPrice_Prior2"), (1, "StockPrice_Prior1")]:
        d = _shift_years(fiscal_year_end_date, -years_back)
        out.append({
            "target_date": d.isoformat(),
            "name": name,
        })

    q_dates = _build_quarter_end_dates(fiscal_year_end_date)
    q_names = ["StockPrice_Q1", "StockPrice_Q2", "StockPrice_Q3", "StockPrice_Q4"]

    for d, name in zip(q_dates, q_names):
        out.append({
            "target_date": d.isoformat(),
            "name": name,
        })

    return out

def get_stock_price_map(stock_code, date_pairs, logger=None, buffer_days=7):
    """
    stooq から日足CSVを取得し、必要期間の Date->Close 辞書を返す。
    target_date 以前を buffer_days 日さかのぼって使う前提。
    """
    if not date_pairs:
        return {}

    start_date = min(
        datetime.strptime(item["target_date"], "%Y-%m-%d").date()
        for item in date_pairs
        if item.get("target_date")
    ) - timedelta(days=buffer_days)

    end_date = max(
        datetime.strptime(item["target_date"], "%Y-%m-%d").date()
        for item in date_pairs
        if item.get("target_date")
    ) + timedelta(days=1)

    code = stock_code.strip().upper()
    if code.endswith(".T"):
        code = code[:-2]

    cache_key = (code, start_date.isoformat(), end_date.isoformat())
    if cache_key in _STOCK_PRICE_MAP_CACHE:
        if logger:
            logger.debug(f"[stock cache hit memory] code={code} range={start_date}..{end_date}")
        return _STOCK_PRICE_MAP_CACHE[cache_key]

    csv_cache_path = os.path.join(_STOCK_CACHE_DIR, f"{code}.csv")

    df = None

    if os.path.exists(csv_cache_path):
        try:
            if logger:
                logger.debug(f"[stock cache hit disk] code={code} path={csv_cache_path}")
            df = pd.read_csv(csv_cache_path)
        except Exception:
            df = None

    need_download = True
    if df is not None and (not df.empty) and "Date" in df.columns and "Close" in df.columns:
        try:
            df["Date"] = pd.to_datetime(df["Date"]).dt.date
            min_have = df["Date"].min()
            max_have = df["Date"].max()
            if min_have <= start_date and max_have >= end_date:
                need_download = False
        except Exception:
            need_download = True

    if need_download:
        if logger:
            logger.debug(f"[stock cache miss download] code={code} range={start_date}..{end_date}")

        symbol = _to_stooq_symbol(code)
        url = f"https://stooq.com/q/d/l/?s={symbol}&i=d"

        session = _get_yf_session()
        r = session.get(url, timeout=20)
        r.raise_for_status()

        df = pd.read_csv(StringIO(r.text))
        if df.empty or "Date" not in df.columns or "Close" not in df.columns:
            _STOCK_PRICE_MAP_CACHE[cache_key] = {}
            return {}

        try:
            df.to_csv(csv_cache_path, index=False)
        except Exception:
            pass

        df["Date"] = pd.to_datetime(df["Date"]).dt.date

    df = df[(df["Date"] >= start_date) & (df["Date"] <= end_date)]
    if df.empty:
        _STOCK_PRICE_MAP_CACHE[cache_key] = {}
        return {}

    close_map = dict(zip(df["Date"], df["Close"]))
    _STOCK_PRICE_MAP_CACHE[cache_key] = close_map
    return close_map

def _find_price_from_map(close_map, target_date, buffer_days=7):
    """
    target_date 以前を最大 buffer_days 日さかのぼって価格を探す
    """
    if not close_map:
        return None

    check = datetime.strptime(target_date, "%Y-%m-%d").date()
    start = check - timedelta(days=buffer_days)

    while check >= start:
        if check in close_map and pd.notna(close_map[check]):
            return float(close_map[check])
        check -= timedelta(days=1)

    return None

def validate_stock_date_pairs(date_pairs):
    """
    date_pairs が name / target_date 前提になっているか検証する。
    """
    bad = []
    for i, item in enumerate(date_pairs):
        if "cell" in item:
            bad.append((i, "cell_is_not_allowed", item))
            continue

        if not item.get("name") or not item.get("target_date"):
            bad.append((i, "missing_required_keys", item))

    if bad:
        msg_lines = ["株価date_pairsが name / target_date 前提になっていません。"]
        for i, reason, item in bad[:10]:
            msg_lines.append(f"  - index={i}, reason={reason}, item={item}")
        raise ValueError("\n".join(msg_lines))


def _set_value_to_namedrange(workbook, range_name: str, value) -> bool:
    """
    NamedRangeが存在すれば、その参照先セル(群)に value を書く。成功なら True。
    """
    dn = workbook.defined_names.get(range_name)
    if dn is None:
        return False

    for sheet_name, ref in dn.destinations:
        if sheet_name not in workbook.sheetnames:
            continue

        ws = workbook[sheet_name]
        obj = ws[ref]

        if isinstance(obj, openpyxl.cell.cell.Cell):
            obj.value = value
        else:
            for row in obj:
                for cell in row:
                    cell.value = value
    return True

def write_stock_data_to_workbook(workbook, stock_code, date_pairs, logger):
    """
    株価は NamedRange 専用。Workbookを外から受け取る版。
    """
    result = {
        "written": 0,
        "miss": 0,
        "errors": 0,
        "missing_name": 0,
        "bad_input": 0,
    }

    try:
        close_map = get_stock_price_map(stock_code, date_pairs, logger=logger)
    except Exception:
        logger.exception(f"株価一覧取得で想定外エラー（続行） code={stock_code}")
        result["errors"] += len(date_pairs)
        return result

    for item in date_pairs:
        name = item.get("name")
        target_date = item.get("target_date")

        if not name or not target_date:
            logger.warning(f"[WARNING] 株価date_pairsの要素が不完全なのでスキップ: {item}")
            result["bad_input"] += 1
            continue

        try:
            price = _find_price_from_map(close_map, target_date)
        except Exception:
            logger.exception(
                f"株価探索で想定外エラー（続行） "
                f"code={stock_code} target={target_date}"
            )
            result["errors"] += 1
            continue

        if price is None:
            logger.warning(f"{target_date} 以前の株価が取得できませんでした（続行）: name={name} code={stock_code}")
            result["miss"] += 1
            continue

        v = float(price)

        wrote = _set_value_to_namedrange(workbook, name, v)

        if not wrote:
            logger.warning(f"NamedRangeが見つからず書けませんでした: {name} ({target_date})")
            result["missing_name"] += 1
            continue

        result["written"] += 1
        logger.debug(f"{target_date} の株価を書き込みました: {name}")

    logger.info(
        f"[stock summary] "
        f"written={result['written']} "
        f"miss={result['miss']} "
        f"errors={result['errors']} "
        f"missing_name={result['missing_name']} "
        f"bad_input={result['bad_input']} "
        f"(code={stock_code})"
    )

    if result["errors"] > 0 or result["missing_name"] > 0 or result["bad_input"] > 0:
        logger.warning(
            f"[stock warning] issues detected "
            f"errors={result['errors']} "
            f"missing_name={result['missing_name']} "
            f"bad_input={result['bad_input']} "
            f"(code={stock_code})"
        )

    return result

def write_stock_data_to_excel(excel_file, stock_code, date_pairs, logger):
    """
    株価は NamedRange 専用。
    """
    workbook = openpyxl.load_workbook(
        excel_file,
        keep_vba=excel_file.lower().endswith(".xlsm")
    )

    result = {
        "written": 0,
        "miss": 0,
        "errors": 0,
        "missing_name": 0,
        "bad_input": 0,
    }

    try:
        close_map = get_stock_price_map(stock_code, date_pairs, logger=logger)
    except Exception:
        logger.exception(f"株価一覧取得で想定外エラー（続行） code={stock_code}")
        result["errors"] += len(date_pairs)
        workbook.save(excel_file)
        return result

    try:
        for item in date_pairs:
            name = item.get("name")
            target_date = item.get("target_date")

            if not name or not target_date:
                logger.warning(f"[WARNING] 株価date_pairsの要素が不完全なのでスキップ: {item}")
                result["bad_input"] += 1
                continue

            try:
                price = _find_price_from_map(close_map, target_date)
            except Exception:
                logger.exception(
                    f"株価探索で想定外エラー（続行） "
                    f"code={stock_code} target={target_date}"
                )
                result["errors"] += 1
                continue

            if price is None:
                logger.warning(f"{target_date} 以前の株価が取得できませんでした（続行）: name={name} code={stock_code}")
                result["miss"] += 1
                continue

            v = float(price)

            wrote = _set_value_to_namedrange(workbook, name, v)
            if not wrote:
                logger.warning(f"NamedRangeが見つからず書けませんでした: {name} ({target_date})")
                result["missing_name"] += 1
                continue

            result["written"] += 1
            logger.debug(f"{target_date} の株価を書き込みました: {name}")

        logger.info(
            f"[stock summary] "
            f"written={result['written']} "
            f"miss={result['miss']} "
            f"errors={result['errors']} "
            f"missing_name={result['missing_name']} "
            f"bad_input={result['bad_input']} "
            f"(code={stock_code})"
        )

        if result["errors"] > 0 or result["missing_name"] > 0 or result["bad_input"] > 0:
            logger.warning(
                f"[stock warning] issues detected "
                f"errors={result['errors']} "
                f"missing_name={result['missing_name']} "
                f"bad_input={result['bad_input']} "
                f"(code={stock_code})"
            )

        workbook.save(excel_file)
        return result

    finally:
        workbook.close()