import os
import re
import shutil
import openpyxl

from openpyxl.cell.cell import Cell


INPUT_SHEET_NAME = "決算入力"


def safe_filename(s: str) -> str:
    """Replace characters that are not safe in Windows file names."""
    if s is None:
        return ""
    s = str(s).strip()
    s = re.sub(r'[\\/:*?"<>|]', "_", s)
    s = s.rstrip(". ").strip()
    return s


def rename_excel_file(original_path, security_code, company_name, period_end_date, logger):
    """Rename a workbook to <code>_<name>_<date>.xlsm and avoid collisions."""
    dir_path = os.path.dirname(original_path)

    code = safe_filename(security_code)
    name = safe_filename(company_name)
    date = safe_filename(period_end_date)

    base_name = f"{code}_{name}_{date}".strip("_")
    if not base_name:
        raise ValueError("rename target is empty: code/name/date are all blank")

    new_file_path = os.path.join(dir_path, f"{base_name}.xlsm")

    counter = 1
    while os.path.exists(new_file_path):
        new_file_path = os.path.join(dir_path, f"{base_name}_{counter}.xlsm")
        counter += 1

    try:
        os.rename(original_path, new_file_path)
    except PermissionError as e:
        raise PermissionError(
            "failed to rename workbook because the file is likely open.\n"
            f"source: {original_path}\n"
            "close the workbook in Excel and try again."
        ) from e

    logger.info("renamed workbook: %s", new_file_path)
    return new_file_path


def find_available_excel_file(base_path, file_name, logger, max_copies=30):
    """Create a working copy of a template workbook next to the original."""
    exts = ["xlsx", "xlsm"]

    def first_existing(paths):
        for path in paths:
            if os.path.exists(path):
                return path
        return None

    def make_next_copy(original_path):
        root, ext = os.path.splitext(original_path)
        candidate0 = f"{root} - copy{ext}"
        if not os.path.exists(candidate0):
            shutil.copy2(original_path, candidate0)
            return candidate0

        for i in range(2, max_copies + 2):
            candidate = f"{root} - copy ({i}){ext}"
            if not os.path.exists(candidate):
                shutil.copy2(original_path, candidate)
                return candidate

        raise RuntimeError("no copy slot available; increase max_copies")

    candidates = [os.path.join(base_path, f"{file_name}.{ext}") for ext in exts]
    original_path = first_existing(candidates)
    if not original_path:
        logger.warning("template file was not found: %s", file_name)
        return None

    try:
        new_copy = make_next_copy(original_path)
        logger.info("created workbook copy: %s", new_copy)
        return new_copy
    except Exception:
        logger.exception("failed to create workbook copy")
        raise


_SUFFIXES = [
    "YTD",
    "Quarter",
    "Current",
    "Prior1",
    "Prior2",
    "Prior3",
    "Prior4",
]

_suffix_pat = re.compile(rf"^(.*)({'|'.join(_SUFFIXES)})$")


def to_namedrange_key(key: str) -> str:
    """
    Normalize code-side keys to workbook named ranges.

    Examples:
      NetSalesYTD -> NetSales_YTD
      TotalAssetsPrior1 -> TotalAssets_Prior1
      SecurityCodeDEI -> unchanged
    """
    if not isinstance(key, str) or not key:
        return key

    if "_" in key:
        return key
    if key.endswith("DEI"):
        return key

    match = _suffix_pat.match(key)
    if not match:
        return key

    metric = match.group(1)
    suffix = match.group(2)
    return f"{metric}_{suffix}"


def transform_keys_for_namedranges(data: dict) -> dict:
    return {to_namedrange_key(k): v for k, v in data.items()}


_FINANCIAL_PREFIXES = (
    "NetSales_",
    "CostOfSales_",
    "GrossProfit_",
    "SellingExpenses_",
    "OperatingIncome_",
    "OrdinaryIncome_",
    "ProfitLoss_",
    "OperatingCash_",
    "InvestmentCash_",
    "FinancingCash_",
    "TotalAssets_",
    "NetAssets_",
    "CashAndCashEquivalents_",
)

_TOTALNUMBER_KEYS = {
    "TotalNumber_Current",
    "TotalNumber_Quarter",
    "TotalNumber_Prior1",
    "TotalNumber_Prior2",
    "TotalNumber_Prior3",
    "TotalNumber_Prior4",
}


def _to_number(v):
    if isinstance(v, (int, float)):
        return float(v)
    if isinstance(v, str):
        s = v.replace(",", "").strip()
        if not s:
            return None
        try:
            return float(s)
        except Exception:
            return None
    return None


def _scale_value_for_excel(name: str, value, display_unit: str):
    num = _to_number(value)
    if num is None:
        return value

    if name in _TOTALNUMBER_KEYS:
        return int(round(num / 1000))

    if name.startswith(_FINANCIAL_PREFIXES):
        if display_unit == "千円":
            return int(round(num / 1000))
        return int(round(num / 1_000_000))

    return value


def _apply_excel_scaling(payload: dict, display_unit: str) -> dict:
    return {k: _scale_value_for_excel(k, v, display_unit) for k, v in payload.items()}


def _iter_namedrange_cells(workbook: openpyxl.Workbook, range_name: str):
    """Yield all cells pointed to by a defined name."""
    defined_name = workbook.defined_names.get(range_name)
    if defined_name is None:
        return

    for sheet_name, ref in defined_name.destinations:
        if sheet_name not in workbook.sheetnames:
            continue

        worksheet = workbook[sheet_name]
        obj = worksheet[ref]

        if isinstance(obj, Cell):
            yield obj
            continue

        for row in obj:
            for cell in row:
                yield cell


def _iter_defined_name_objects(workbook: openpyxl.Workbook):
    container = workbook.defined_names

    if hasattr(container, "definedName"):
        return list(container.definedName)
    if hasattr(container, "values"):
        return list(container.values())
    return []


def get_defined_name_set(workbook: openpyxl.Workbook) -> set[str]:
    names: set[str] = set()
    for defined_name in _iter_defined_name_objects(workbook):
        name = getattr(defined_name, "name", None)
        attr_text = getattr(defined_name, "attr_text", None)
        if name and attr_text:
            names.add(str(name))
    return names


def build_namedrange_cache(workbook: openpyxl.Workbook) -> dict[str, list[Cell]]:
    cache: dict[str, list[Cell]] = {}
    for defined_name in _iter_defined_name_objects(workbook):
        name = getattr(defined_name, "name", None)
        attr_text = getattr(defined_name, "attr_text", None)

        if not name or not attr_text:
            continue

        cells = list(_iter_namedrange_cells(workbook, str(name)))
        if cells:
            cache[str(name)] = cells
    return cache


def write_data_to_workbook_namedranges(
    wb,
    data: dict,
    *,
    display_unit: str = "百万円",
    transform_keys: bool = True,
    skip_if_formula: bool = False,
    skip_values=("データなし", "", None),
):
    result = {"written": [], "skipped": [], "missing": []}

    namedrange_cache = build_namedrange_cache(wb)
    payload = transform_keys_for_namedranges(data) if transform_keys else dict(data)
    payload = _apply_excel_scaling(payload, display_unit)

    if INPUT_SHEET_NAME in wb.sheetnames:
        wb[INPUT_SHEET_NAME]["J2"] = display_unit

    write_queue = []

    for name, value in payload.items():
        if value in skip_values or (isinstance(value, str) and value.strip() in skip_values):
            result["skipped"].append((name, "empty"))
            continue

        cells = namedrange_cache.get(name)
        if not cells:
            result["missing"].append(name)
            continue

        for cell in cells:
            if skip_if_formula and isinstance(cell.value, str) and cell.value.startswith("="):
                result["skipped"].append((name, f"formula@{cell.coordinate}"))
                continue

            write_queue.append((cell, value))
            result["written"].append((name, f"{cell.parent.title}!{cell.coordinate}"))

    for cell, value in write_queue:
        cell.value = value

    return result


def write_rows_to_raw_sheet_workbook(wb, rows: list[dict], raw_cols: list[str], *, sheet_name: str = "raw_edinet"):
    import datetime as _dt

    def _to_excel_date(v):
        if isinstance(v, _dt.datetime):
            return v.date()
        if isinstance(v, _dt.date):
            return v
        if isinstance(v, str):
            s = v.strip()
            try:
                return _dt.datetime.strptime(s, "%Y-%m-%d").date()
            except Exception:
                return v
        return v

    if sheet_name not in wb.sheetnames:
        raise ValueError(f"sheet not found: {sheet_name}")

    ws = wb[sheet_name]

    if ws.max_row >= 2:
        ws.delete_rows(2, ws.max_row - 1)

    r = 2
    for row in rows:
        for c, col in enumerate(raw_cols, start=1):
            v = row.get(col)
            if col in ("period_start", "period_end"):
                v = _to_excel_date(v)
            ws.cell(row=r, column=c).value = v
        r += 1


def write_rows_to_raw_sheet(excel_file: str, rows: list[dict], raw_cols: list[str], *, sheet_name: str = "raw_edinet"):
    """Write raw_edinet rows directly into a workbook file."""
    wb = openpyxl.load_workbook(
        excel_file,
        keep_vba=excel_file.lower().endswith(".xlsm")
    )
    try:
        write_rows_to_raw_sheet_workbook(
            wb,
            rows,
            raw_cols,
            sheet_name=sheet_name,
        )
        wb.save(excel_file)
    finally:
        wb.close()
