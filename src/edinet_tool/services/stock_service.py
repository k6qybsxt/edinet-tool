import requests
import certifi
import pandas as pd
import openpyxl

from io import StringIO
from datetime import datetime, timedelta

# requests session を1回だけ作って使い回す
_YF_SESSION = None
_STOCK_PRICE_MAP_CACHE = {}


def _get_yf_session():
    global _YF_SESSION
    if _YF_SESSION is None:
        s = requests.Session()
        s.verify = certifi.where()
        _YF_SESSION = s
    return _YF_SESSION


def _to_stooq_symbol(stock_code: str) -> str:
    """
    '2206.T' / '2206' どちらが来ても stooq用に変換する
    stooqの日本株は '2206.JP' 形式
    """
    code = stock_code.strip().upper()
    if code.endswith(".T"):
        code = code[:-2]
    return f"{code}.JP"


def get_stock_price_map(stock_code, date_pairs, logger=None, buffer_days=3):
    """
    stooq から日足CSVを1回だけ取得し、必要期間の Date->Close 辞書を返す
    同一実行中は stock_code ごとにキャッシュする
    """
    if not date_pairs:
        return {}

    start_date = min(
        datetime.strptime(item["backup_date"], "%Y-%m-%d").date()
        for item in date_pairs
        if item.get("backup_date")
    ) - timedelta(days=buffer_days)

    end_date = max(
        datetime.strptime(item["target_date"], "%Y-%m-%d").date()
        for item in date_pairs
        if item.get("target_date")
    ) + timedelta(days=1)

    cache_key = (stock_code, start_date.isoformat(), end_date.isoformat())
    if cache_key in _STOCK_PRICE_MAP_CACHE:
        if logger:
            logger.debug(f"[stock cache hit] code={stock_code} range={start_date}..{end_date}")
        return _STOCK_PRICE_MAP_CACHE[cache_key]

    if logger:
        logger.debug(f"[stock cache miss] code={stock_code} range={start_date}..{end_date}")

    symbol = _to_stooq_symbol(stock_code)
    url = f"https://stooq.com/q/d/l/?s={symbol}&i=d"

    r = requests.get(url, timeout=20)
    r.raise_for_status()

    df = pd.read_csv(StringIO(r.text))
    if df.empty or "Date" not in df.columns or "Close" not in df.columns:
        _STOCK_PRICE_MAP_CACHE[cache_key] = {}
        return {}

    df["Date"] = pd.to_datetime(df["Date"]).dt.date
    df = df[(df["Date"] >= start_date) & (df["Date"] <= end_date)]
    if df.empty:
        _STOCK_PRICE_MAP_CACHE[cache_key] = {}
        return {}

    close_map = dict(zip(df["Date"], df["Close"]))
    _STOCK_PRICE_MAP_CACHE[cache_key] = close_map
    return close_map

def _find_price_from_map(close_map, target_date, backup_date, buffer_days=3):
    """
    target_date から backup_date-buffer_days まで遡って価格を探す
    """
    if not close_map:
        return None

    check = datetime.strptime(target_date, "%Y-%m-%d").date()
    start = datetime.strptime(backup_date, "%Y-%m-%d").date() - timedelta(days=buffer_days)

    while check >= start:
        if check in close_map and pd.notna(close_map[check]):
            return float(close_map[check])
        check -= timedelta(days=1)

    return None

def validate_stock_date_pairs(date_pairs):
    """
    JSONのdate_pairsが name前提になっているか検証する。
    """
    bad = []
    for i, item in enumerate(date_pairs):
        if "cell" in item:
            bad.append((i, "cell_is_not_allowed", item))
            continue

        if not item.get("name") or not item.get("target_date") or not item.get("backup_date"):
            bad.append((i, "missing_required_keys", item))

    if bad:
        msg_lines = ["株価date_pairsが name前提になっていません（JSONを修正してください）。"]
        for i, reason, item in bad[:10]:
            msg_lines.append(f"  - index={i}, reason={reason}, item={item}")
        raise ValueError("\n".join(msg_lines))


def _set_value_to_namedrange(workbook, range_name: str, value) -> bool:
    """
    NamedRangeが存在すれば、その参照先セル(群)に value を書く。成功なら True。
    """
    dn = workbook.defined_names.get(range_name)
    if dn is None:
        return False

    for sheet_name, ref in dn.destinations:
        if sheet_name not in workbook.sheetnames:
            continue

        ws = workbook[sheet_name]
        obj = ws[ref]

        if isinstance(obj, openpyxl.cell.cell.Cell):
            obj.value = value
        else:
            for row in obj:
                for cell in row:
                    cell.value = value
    return True

def write_stock_data_to_workbook(workbook, stock_code, date_pairs, logger):
    """
    株価は NamedRange 専用。Workbookを外から受け取る版。
    """
    result = {
        "written": 0,
        "miss": 0,
        "errors": 0,
        "missing_name": 0,
        "bad_input": 0,
    }

    try:
        close_map = get_stock_price_map(stock_code, date_pairs, logger=logger)
    except Exception:
        logger.exception(f"株価一覧取得で想定外エラー（続行） code={stock_code}")
        result["errors"] += len(date_pairs)
        return result

    for item in date_pairs:
        name = item.get("name")
        target_date = item.get("target_date")
        backup_date = item.get("backup_date")

        if not name or not target_date or not backup_date:
            logger.warning(f"[WARNING] 株価date_pairsの要素が不完全なのでスキップ: {item}")
            result["bad_input"] += 1
            continue

        try:
            price = _find_price_from_map(close_map, target_date, backup_date)
        except Exception:
            logger.exception(
                f"株価探索で想定外エラー（続行） "
                f"code={stock_code} target={target_date} backup={backup_date}"
            )
            result["errors"] += 1
            continue

        if price is None:
            logger.warning(f"{target_date} の株価が取得できませんでした（続行）: name={name} code={stock_code}")
            result["miss"] += 1
            continue

        v = float(price)

        wrote = _set_value_to_namedrange(workbook, name, v)

        if not wrote:
            logger.warning(f"NamedRangeが見つからず書けませんでした: {name} ({target_date})")
            result["missing_name"] += 1
            continue

        result["written"] += 1
        logger.debug(f"{target_date} の株価を書き込みました: {name}")

    logger.info(
        f"[stock summary] "
        f"written={result['written']} "
        f"miss={result['miss']} "
        f"errors={result['errors']} "
        f"missing_name={result['missing_name']} "
        f"bad_input={result['bad_input']} "
        f"(code={stock_code})"
    )

    if result["errors"] > 0 or result["missing_name"] > 0 or result["bad_input"] > 0:
        logger.warning(
            f"[stock warning] issues detected "
            f"errors={result['errors']} "
            f"missing_name={result['missing_name']} "
            f"bad_input={result['bad_input']} "
            f"(code={stock_code})"
        )

    return result

def write_stock_data_to_excel(excel_file, stock_code, date_pairs, logger):
    """
    株価は NamedRange 専用。
    """
    workbook = openpyxl.load_workbook(
        excel_file,
        keep_vba=excel_file.lower().endswith(".xlsm")
    )

    result = {
        "written": 0,
        "miss": 0,
        "errors": 0,
        "missing_name": 0,
        "bad_input": 0,
    }

    try:
        close_map = get_stock_price_map(stock_code, date_pairs, logger=logger)
    except Exception:
        logger.exception(f"株価一覧取得で想定外エラー（続行） code={stock_code}")
        result["errors"] += len(date_pairs)
        workbook.save(excel_file)
        return result

    try:
        for item in date_pairs:
            name = item.get("name")
            target_date = item.get("target_date")
            backup_date = item.get("backup_date")

            if not name or not target_date or not backup_date:
                logger.warning(f"[WARNING] 株価date_pairsの要素が不完全なのでスキップ: {item}")
                result["bad_input"] += 1
                continue

            try:
                price = _find_price_from_map(close_map, target_date, backup_date)
            except Exception:
                logger.exception(
                    f"株価探索で想定外エラー（続行） "
                    f"code={stock_code} target={target_date} backup={backup_date}"
                )
                result["errors"] += 1
                continue

            if price is None:
                logger.warning(f"{target_date} の株価が取得できませんでした（続行）: name={name} code={stock_code}")
                result["miss"] += 1
                continue

            v = float(price)

            wrote = _set_value_to_namedrange(workbook, name, v)
            if not wrote:
                logger.warning(f"NamedRangeが見つからず書けませんでした: {name} ({target_date})")
                result["missing_name"] += 1
                continue

            result["written"] += 1
            logger.debug(f"{target_date} の株価を書き込みました: {name}")

        logger.info(
            f"[stock summary] "
            f"written={result['written']} "
            f"miss={result['miss']} "
            f"errors={result['errors']} "
            f"missing_name={result['missing_name']} "
            f"bad_input={result['bad_input']} "
            f"(code={stock_code})"
        )

        if result["errors"] > 0 or result["missing_name"] > 0 or result["bad_input"] > 0:
            logger.warning(
                f"[stock warning] issues detected "
                f"errors={result['errors']} "
                f"missing_name={result['missing_name']} "
                f"bad_input={result['bad_input']} "
                f"(code={stock_code})"
            )

        workbook.save(excel_file)
        return result

    finally:
        workbook.close()

def clear_stock_price_cache():
    _STOCK_PRICE_MAP_CACHE.clear()